import os
import sys
import site
import csv
import io
import html
import re
import smtplib
import ssl
from email.message import EmailMessage
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from urllib.parse import quote_plus
from werkzeug.utils import secure_filename

from flask import Flask, jsonify, request, Response, redirect, make_response, send_from_directory


# ------------------------------------------------------------
# Azure / package setup
# ------------------------------------------------------------

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PACKAGES_DIR = os.path.join(BASE_DIR, ".python_packages", "lib", "site-packages")

if os.path.isdir(PACKAGES_DIR):
    site.addsitedir(PACKAGES_DIR)
    sys.path.append(PACKAGES_DIR)


app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024


APP_ENVIRONMENT = os.getenv("APP_ENVIRONMENT", "Not set")
SQL_SERVER_NAME = os.getenv("SQL_SERVER_NAME", "Not set")
SQL_DATABASE_NAME = os.getenv("SQL_DATABASE_NAME", "Not set")
ALLOWED_EMAIL_DOMAIN = os.getenv("ALLOWED_EMAIL_DOMAIN", "c-diving.com")

# Email notification settings. Configure these in Azure App Service > Configuration.
# If SMTP_HOST and EMAIL_FROM are not configured, emails are skipped safely and the app still works.
APP_BASE_URL = os.getenv("APP_BASE_URL", "").rstrip("/")
SMTP_HOST = os.getenv("SMTP_HOST", "")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587") or "587")
SMTP_USERNAME = os.getenv("SMTP_USERNAME", "")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD", "")
SMTP_USE_TLS = os.getenv("SMTP_USE_TLS", "true").lower() not in ["0", "false", "no"]
EMAIL_FROM = os.getenv("EMAIL_FROM", "")
EMAIL_REPLY_TO = os.getenv("EMAIL_REPLY_TO", "accounting@c-diving.com")
EMAIL_NOTIFICATIONS_ENABLED = os.getenv("EMAIL_NOTIFICATIONS_ENABLED", "true").lower() not in ["0", "false", "no"]

SQL_CONNECTION = (
    os.getenv("PODASHBOARD_SQL")
    or os.getenv("SQLAZURECONNSTR_PODASHBOARD_SQL")
    or os.getenv("CUSTOMCONNSTR_PODASHBOARD_SQL")
    or ""
)


REQUIRED_PO_COLUMNS = [
    "ProjectCode",
    "ProjectName",
    "PONumber",
    "VendorName",
    "PODate",
    "Description",
    "Unit",
    "UnitCost",
    "Qty",
    "LineAmount",
]

DEPARTMENT_OPTIONS = [
    "Engineering",
    "Marine Construction",
    "Commercial Diving",
    "Dredging",
    "Marine Services",
]


REQUEST_ATTACHMENT_ROOT = os.getenv("REQUEST_ATTACHMENT_ROOT", "/home/site/wwwroot/request_attachments")
ALLOWED_ATTACHMENT_EXTENSIONS = {"pdf", "png", "jpg", "jpeg", "doc", "docx", "xls", "xlsx", "csv", "txt", "eml", "msg"}


VALID_ROLES = [
    "Admin",
    "Executive",
    "Project Manager - Dredging Only",
    "Project Manager - Diving",
    "Division Manager - Diving",
    "Purchaser - All Departments",
    "Bookkeeping - All Departments",
    "No Access",
]

# Legacy role names are kept as aliases so an older user row does not break the app
# before Admin has a chance to reassign the user to the new July 1 role package.
ROLE_ALIASES = {
    "Accounting": "Bookkeeping - All Departments",
    "Project Manager": "Project Manager - Diving",
    "Viewer": "Bookkeeping - All Departments",
}

ROLE_GROUP_ALL_PO_VIEW = ["Admin", "Executive", "Purchaser - All Departments", "Bookkeeping - All Departments"]
ROLE_GROUP_NON_DREDGING_VIEW = ["Division Manager - Diving"]
ROLE_GROUP_DREDGING_VIEW = ["Project Manager - Dredging Only"]
ROLE_GROUP_ASSIGNED_NON_DREDGING_VIEW = ["Project Manager - Diving"]
ROLE_GROUP_CAN_CREATE_REQUEST = [
    "Admin",
    "Executive",
    "Project Manager - Dredging Only",
    "Project Manager - Diving",
    "Division Manager - Diving",
    "Purchaser - All Departments",
]
ROLE_GROUP_CAN_REVIEW_REQUESTS = ["Admin", "Executive"]
ROLE_GROUP_ADMIN_ONLY = ["Admin"]

PAGE_ACCESS = {
    "Dashboard": ["Admin", "Executive", "Project Manager - Dredging Only", "Project Manager - Diving", "Division Manager - Diving", "Purchaser - All Departments", "Bookkeeping - All Departments"],
    "My Dashboard": ["Admin", "Executive", "Project Manager - Dredging Only", "Project Manager - Diving", "Division Manager - Diving", "Purchaser - All Departments", "Bookkeeping - All Departments"],
    "Help Center": ["Admin", "Executive", "Project Manager - Dredging Only", "Project Manager - Diving", "Division Manager - Diving", "Purchaser - All Departments", "Bookkeeping - All Departments"],
    "New Purchase Request": ROLE_GROUP_CAN_CREATE_REQUEST,
    "Purchase Requests": ["Admin", "Executive", "Project Manager - Dredging Only", "Project Manager - Diving", "Division Manager - Diving", "Purchaser - All Departments"],
    "Approver Queue": ["Admin", "Executive"],
    "POs & Balances": ["Admin", "Executive", "Project Manager - Dredging Only", "Project Manager - Diving", "Division Manager - Diving", "Purchaser - All Departments", "Bookkeeping - All Departments"],
    "Projects": ["Admin", "Executive", "Project Manager - Dredging Only", "Project Manager - Diving", "Division Manager - Diving", "Purchaser - All Departments", "Bookkeeping - All Departments"],
    "Forecasting": ["Admin", "Executive"],
    "Project PO Setup": ["Admin"],
    "PO Setup Review": ["Admin", "Executive"],
    "PO Maintenance": ["Admin"],
    "PO Summary": ["Admin", "Executive", "Project Manager - Dredging Only", "Project Manager - Diving", "Division Manager - Diving", "Purchaser - All Departments", "Bookkeeping - All Departments"],
    "PO List": ["Admin", "Executive", "Project Manager - Dredging Only", "Project Manager - Diving", "Division Manager - Diving", "Purchaser - All Departments", "Bookkeeping - All Departments"],
    "PO Detail": ["Admin", "Executive", "Project Manager - Dredging Only", "Project Manager - Diving", "Division Manager - Diving", "Purchaser - All Departments", "Bookkeeping - All Departments"],
    "Upload Issued POs": ["Admin"],
    "Expense Upload / PO Matching": ["Admin"],
    "Clear Expense Data": ["Admin"],
    "Expenses": ["Admin", "Executive"],
    "Missing PO Review": ["Admin", "Executive"],
    "Vendors": ["Admin", "Executive", "Project Manager - Dredging Only", "Project Manager - Diving", "Division Manager - Diving", "Purchaser - All Departments", "Bookkeeping - All Departments"],
    "POs in PM Comments": ["Admin", "Executive"],
    "Import History": ["Admin", "Executive", "Purchaser - All Departments", "Bookkeeping - All Departments"],
    "Exceptions": ["Admin", "Executive"],
    "Exports": ["Admin", "Executive"],
    "User Access": ["Admin", "Executive"],
    "Future Pages": ["Admin"],
    "Who Am I": ["Admin", "Executive", "Project Manager - Dredging Only", "Project Manager - Diving", "Division Manager - Diving", "Purchaser - All Departments", "Bookkeeping - All Departments"],
}


# ------------------------------------------------------------
# General helpers
# ------------------------------------------------------------

def h(value):
    if value is None:
        return ""
    return html.escape(str(value))


def currency(value):
    try:
        return "${:,.2f}".format(float(value or 0))
    except Exception:
        return "$0.00"


def select_option(value, selected_value):
    selected = " selected" if str(value or "") == str(selected_value or "") else ""
    return f'<option value="{h(value)}"{selected}>{h(value)}</option>'


def get_sql_connection():
    if not SQL_CONNECTION:
        raise RuntimeError("SQL connection string was not found.")

    connection_string = SQL_CONNECTION

    if "Driver=" not in connection_string and "DRIVER=" not in connection_string:
        connection_string = "Driver={ODBC Driver 18 for SQL Server};" + connection_string

    import pyodbc
    return pyodbc.connect(connection_string, timeout=20)


def clean_text(value):
    if value is None:
        return None
    value = str(value).strip()
    return value if value else None


def clean_decimal(value):
    if value is None:
        return None

    if isinstance(value, Decimal):
        return value

    value = str(value).strip()

    if value == "":
        return None

    value = value.replace("$", "").replace(",", "")

    try:
        return Decimal(value)
    except InvalidOperation:
        return None


def clean_date(value):
    if value is None or value == "":
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    value = str(value).strip()

    if value == "":
        return None

    for fmt in ["%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%Y/%m/%d"]:
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            pass

    return None


def normalize_header(header):
    if header is None:
        return ""
    return str(header).strip()


# ------------------------------------------------------------
# Authentication / role helpers
# ------------------------------------------------------------

def normalize_role(role_name):
    role = clean_text(role_name) or "No Access"
    return ROLE_ALIASES.get(role, role)


def role_is_admin(role_name):
    return normalize_role(role_name) == "Admin"


def role_can_create_purchase_request(role_name):
    return normalize_role(role_name) in ROLE_GROUP_CAN_CREATE_REQUEST


def role_can_review_requests(role_name):
    return normalize_role(role_name) in ROLE_GROUP_CAN_REVIEW_REQUESTS


def role_can_maintain_pos(role_name):
    return normalize_role(role_name) == "Admin"


def get_current_user():
    user_email = request.headers.get("X-MS-CLIENT-PRINCIPAL-NAME", "")
    user_id = request.headers.get("X-MS-CLIENT-PRINCIPAL-ID", "")
    identity_provider = request.headers.get("X-MS-CLIENT-PRINCIPAL-IDP", "")

    user_email = (user_email or "").strip().lower()

    email_domain = ""
    if "@" in user_email:
        email_domain = user_email.split("@")[-1].lower()

    allowed_domain = (ALLOWED_EMAIL_DOMAIN or "").strip().lower()

    return {
        "email": user_email,
        "user_id": user_id,
        "identity_provider": identity_provider,
        "email_domain": email_domain,
        "allowed_domain": allowed_domain,
        "is_authenticated": bool(user_email),
        "is_allowed_domain": bool(email_domain) and email_domain == allowed_domain,
    }


def get_user_access():
    user = get_current_user()

    access = {
        "email": user["email"],
        "display_name": "",
        "role": "No Access",
        "is_active": False,
        "found_in_sql": False,
        "lookup_error": "",
    }

    if not user["email"]:
        return access

    try:
        conn = get_sql_connection()
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT TOP 1
                Email,
                DisplayName,
                RoleName,
                IsActive
            FROM dbo.DashboardUsers
            WHERE LOWER(Email) = LOWER(?);
            """,
            user["email"],
        )

        row = cursor.fetchone()
        conn.close()

        if row:
            access["display_name"] = row.DisplayName or ""
            access["role"] = normalize_role(row.RoleName or "No Access")
            access["is_active"] = bool(row.IsActive)
            access["found_in_sql"] = True

            if not access["is_active"]:
                access["role"] = "No Access"

        return access

    except Exception as e:
        access["lookup_error"] = str(e)
        return access


def load_assignable_users():
    """Return active named users for assignment dropdowns."""
    conn = get_sql_connection()
    cursor = conn.cursor()
    cursor.execute(
        """
        SELECT Email, DisplayName, RoleName
        FROM dbo.DashboardUsers
        WHERE IsActive = 1
        ORDER BY
            CASE WHEN COALESCE(DisplayName, '') = '' THEN Email ELSE DisplayName END;
        """
    )
    rows = cursor.fetchall()
    conn.close()
    return rows



# ------------------------------------------------------------
# Email notification helpers
# ------------------------------------------------------------

def app_url(path="/"):
    """Build an absolute URL for email buttons."""
    path = str(path or "/")
    if path.startswith("http://") or path.startswith("https://"):
        return path
    base = (APP_BASE_URL or (request.url_root.rstrip("/") if request else ""))
    if not base:
        return path
    if not path.startswith("/"):
        path = "/" + path
    return base.rstrip("/") + path


def clean_email(value):
    value = clean_text(value).strip()
    if "@" not in value:
        return ""
    return value.lower()


def load_active_users_by_roles(role_names):
    role_names = [clean_text(r) for r in (role_names or []) if clean_text(r)]
    if not role_names:
        return []
    placeholders = ",".join("?" for _ in role_names)
    conn = get_sql_connection()
    cursor = conn.cursor()
    cursor.execute(
        f"""
        SELECT Email, DisplayName, RoleName
        FROM dbo.DashboardUsers
        WHERE IsActive = 1 AND RoleName IN ({placeholders})
        ORDER BY RoleName, COALESCE(NULLIF(DisplayName, ''), Email);
        """,
        *role_names,
    )
    rows = cursor.fetchall()
    conn.close()
    return rows


def user_display_name(row):
    if not row:
        return ""
    return clean_text(getattr(row, "DisplayName", "")) or clean_text(getattr(row, "Email", ""))


def unique_recipients(*recipient_groups):
    seen = set()
    recipients = []
    for group in recipient_groups:
        if not group:
            continue
        for item in group:
            if hasattr(item, "Email"):
                email = clean_email(getattr(item, "Email", ""))
                name = clean_text(getattr(item, "DisplayName", "")) or email
            elif isinstance(item, (tuple, list)):
                email = clean_email(item[0] if item else "")
                name = clean_text(item[1] if len(item) > 1 else "") or email
            else:
                email = clean_email(item)
                name = email
            if email and email not in seen:
                seen.add(email)
                recipients.append((email, name))
    return recipients


def html_button(label, url):
    return f"""<p style=\"margin:20px 0;\"><a href=\"{h(url)}\" style=\"background:#2563eb;color:#ffffff;text-decoration:none;padding:12px 18px;border-radius:10px;font-weight:700;display:inline-block;\">{h(label)}</a></p>"""


def send_email_notification(to_recipients, subject, html_body, text_body="", attachments=None, cc_recipients=None):
    """Send an HTML email with optional attachments. Safe no-op if SMTP is not configured."""
    recipients = unique_recipients(to_recipients)
    cc = unique_recipients(cc_recipients)
    if not recipients:
        return {"sent": False, "reason": "No recipients"}
    if not EMAIL_NOTIFICATIONS_ENABLED:
        return {"sent": False, "reason": "Email notifications disabled"}
    if not SMTP_HOST or not EMAIL_FROM:
        print(f"EMAIL SKIPPED: SMTP_HOST/EMAIL_FROM not configured. Subject={subject}. To={[r[0] for r in recipients]}")
        return {"sent": False, "reason": "SMTP not configured"}

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = EMAIL_FROM
    msg["To"] = ", ".join(email for email, _name in recipients)
    if cc:
        msg["Cc"] = ", ".join(email for email, _name in cc)
    if EMAIL_REPLY_TO:
        msg["Reply-To"] = EMAIL_REPLY_TO
    msg.set_content(text_body or re.sub(r"<[^>]+>", " ", html_body or ""))
    msg.add_alternative(html_body or text_body or "", subtype="html")

    for att in attachments or []:
        try:
            msg.add_attachment(
                att.get("content", b""),
                maintype=att.get("maintype", "application"),
                subtype=att.get("subtype", "octet-stream"),
                filename=att.get("filename", "attachment"),
            )
        except Exception as exc:
            print(f"EMAIL ATTACHMENT SKIPPED: {exc}")

    all_recipients = [email for email, _ in recipients] + [email for email, _ in cc]
    try:
        if SMTP_USE_TLS:
            context = ssl.create_default_context()
            with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as server:
                server.starttls(context=context)
                if SMTP_USERNAME:
                    server.login(SMTP_USERNAME, SMTP_PASSWORD)
                server.send_message(msg, to_addrs=all_recipients)
        else:
            with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as server:
                if SMTP_USERNAME:
                    server.login(SMTP_USERNAME, SMTP_PASSWORD)
                server.send_message(msg, to_addrs=all_recipients)
        return {"sent": True, "recipients": all_recipients}
    except Exception as exc:
        print(f"EMAIL SEND FAILED: {exc}. Subject={subject}. To={all_recipients}")
        return {"sent": False, "reason": str(exc)}


def wrap_pdf_text(text, width=92):
    text = clean_text(text)
    if not text:
        return [""]
    words = text.replace("\r", "").split()
    lines = []
    line = ""
    for word in words:
        if len(line) + len(word) + 1 > width:
            lines.append(line)
            line = word
        else:
            line = (line + " " + word).strip()
    if line:
        lines.append(line)
    return lines or [""]


def pdf_escape(text):
    return str(text or "").replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")



def simple_pdf_bytes(title, lines):
    """Fallback valid text PDF without external dependencies."""
    lines = [clean_text(x) for x in (lines or [])]
    content_lines = ["BT", "/F1 16 Tf", "72 760 Td", f"({pdf_escape(title)}) Tj", "/F1 9 Tf", "0 -22 Td"]
    for raw in lines:
        for line in wrap_pdf_text(raw, 105):
            content_lines.append(f"({pdf_escape(line)}) Tj")
            content_lines.append("0 -13 Td")
    content_lines.append("ET")
    stream = "\n".join(content_lines).encode("latin-1", errors="replace")
    objects = []
    objects.append(b"1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj\n")
    objects.append(b"2 0 obj << /Type /Pages /Kids [3 0 R] /Count 1 >> endobj\n")
    objects.append(b"3 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >> endobj\n")
    objects.append(b"4 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> endobj\n")
    objects.append(b"5 0 obj << /Length " + str(len(stream)).encode() + b" >> stream\n" + stream + b"\nendstream endobj\n")
    pdf = b"%PDF-1.4\n"
    offsets = [0]
    for obj in objects:
        offsets.append(len(pdf))
        pdf += obj
    xref_offset = len(pdf)
    pdf += f"xref\n0 {len(objects)+1}\n0000000000 65535 f \n".encode()
    for off in offsets[1:]:
        pdf += f"{off:010d} 00000 n \n".encode()
    pdf += f"trailer << /Root 1 0 R /Size {len(objects)+1} >>\nstartxref\n{xref_offset}\n%%EOF\n".encode()
    return pdf


def _coastal_logo_image_reader():
    try:
        import base64
        from io import BytesIO
        from reportlab.lib.utils import ImageReader
        data_uri = globals().get("CE_LOGO_DATA_URI", "")
        if "," not in data_uri:
            return None
        raw = base64.b64decode(data_uri.split(",", 1)[1])
        return ImageReader(BytesIO(raw))
    except Exception:
        return None



def _safe_text(value):
    value = clean_text(value)
    if value is None:
        return ""
    return str(value).replace("\r", " ").replace("\n", " ").strip()


def _safe_multiline(value):
    value = clean_text(value)
    if value is None:
        return ""
    return str(value).replace("\r", "").strip()


def _money(value):
    return currency(value)


def _row_value(row, name, default=""):
    try:
        value = getattr(row, name)
        return default if value is None else value
    except Exception:
        try:
            value = row[name]
            return default if value is None else value
        except Exception:
            return default


def _styled_po_packet_pdf_bytes(po, lines, posted_expenses, packet_type="internal"):
    # PDF layout cleanup: larger fonts, left-aligned line item columns, aligned tables, cleaned ship-to block.
    """Create formal document style PO packet PDFs.

    Selected design: Mockup C for both internal and vendor packets.
    This layout is intentionally PDF-native and reliable: clean letterhead,
    formal purchase-order title block, two-column details, line items, totals,
    and internal/vendor-specific sections.
    """
    try:
        from io import BytesIO
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas
    except Exception:
        title = ("Vendor PO Packet" if packet_type == "vendor" else "Internal PO Packet") + f" - {_row_value(po, 'PONumber')}"
        pdf_lines = [
            "Coastal Engineering Group",
            f"PO Number: {_row_value(po, 'PONumber')}",
            f"Vendor: {_row_value(po, 'VendorName')}",
            f"Project: {_row_value(po, 'ProjectName')}",
            f"Department: {_row_value(po, 'Department')}",
            f"Requestor: {_row_value(po, 'Requestor')}",
            f"PO Date: {_row_value(po, 'PODate')}",
            f"PO Value: {_money(_row_value(po, 'POValue', 0))}",
        ]
        for line in (lines or [])[:80]:
            pdf_lines.append(f"- {_row_value(line, 'LineDescription')} | Qty {_row_value(line, 'Qty')} | Unit {_row_value(line, 'Unit')} | Amount {_money(_row_value(line, 'LineAmount', 0))}")
        return simple_pdf_bytes(title, pdf_lines)

    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)
    W, H = letter
    margin = 36

    navy = colors.HexColor("#061b36")
    blue = colors.HexColor("#0b5dad")
    blue_dark = colors.HexColor("#07447f")
    green = colors.HexColor("#0f766e")
    text = colors.HexColor("#0f172a")
    muted = colors.HexColor("#64748b")
    line_color = colors.HexColor("#cbd5e1")
    light_line = colors.HexColor("#e2e8f0")
    panel = colors.HexColor("#f8fafc")
    soft_blue = colors.HexColor("#eff6ff")

    logo = _coastal_logo_image_reader()
    po_number = _safe_text(_row_value(po, "PONumber")) or "PO"
    vendor = _safe_text(_row_value(po, "VendorName")) or "Vendor TBD"
    project = _safe_text(_row_value(po, "ProjectName")) or "Project TBD"
    project_code = _safe_text(_row_value(po, "ProjectCode")) or ""
    department = _safe_text(_row_value(po, "Department")) or "Department TBD"
    requestor = _safe_text(_row_value(po, "Requestor")) or "Requestor TBD"
    po_date = _safe_text(_row_value(po, "PODate")) or "TBD"
    required_date = _safe_text(_row_value(po, "ExpectedPaymentDate")) or "TBD"
    po_value = _row_value(po, "POValue", 0)
    posted_amount = _row_value(po, "PostedExpenseAmount", 0)
    current_balance = _row_value(po, "RemainingAmount", 0)
    payment_type = _safe_text(_row_value(po, "PaymentType")) or "Net 30"
    payment_schedule = _safe_multiline(_row_value(po, "PaymentSchedule"))
    status = _safe_text(_row_value(po, "POStatus")) or "Open"
    is_vendor = packet_type == "vendor"
    accent = blue if is_vendor else navy
    title = "PURCHASE ORDER"
    subtitle = po_number

    def draw_logo(x, y, w, h):
        if logo:
            try:
                c.drawImage(logo, x, y, width=w, height=h, preserveAspectRatio=True, mask="auto")
                return
            except Exception:
                pass
        c.setFillColor(navy)
        c.setFont("Helvetica-Bold", 19)
        c.drawString(x, y + h - 22, "COASTAL")
        c.setFillColor(blue)
        c.setFont("Helvetica-Bold", 8.5)
        c.drawString(x, y + h - 36, "ENGINEERING GROUP")

    def header():
        c.setFillColor(colors.white)
        c.rect(0, 0, W, H, fill=1, stroke=0)
        draw_logo(margin, H - 76, 175, 50)
        c.setFillColor(navy)
        c.setFont("Helvetica-Bold", 21)
        c.drawRightString(W - margin, H - 47, title)
        c.setFillColor(blue if is_vendor else navy)
        c.setFont("Helvetica-Bold", 11)
        c.drawRightString(W - margin, H - 65, subtitle)
        c.setStrokeColor(line_color)
        c.setLineWidth(1)
        c.line(margin, H - 88, W - margin, H - 88)
        c.setStrokeColor(accent)
        c.setLineWidth(2.2)
        c.line(margin, H - 92, W - margin, H - 92)
        if not is_vendor:
            c.setFillColor(colors.HexColor("#334155"))
            c.setFont("Helvetica-Bold", 7.5)
            c.drawRightString(W - margin, H - 77, "FOR INTERNAL USE ONLY")
        return H - 118

    def footer(page_label=""):
        c.setStrokeColor(line_color)
        c.setLineWidth(0.7)
        c.line(margin, 42, W - margin, 42)
        c.setFillColor(muted)
        c.setFont("Helvetica", 7.4)
        c.drawString(margin, 28, "Coastal Engineering Group")
        c.drawCentredString(W / 2, 28, "Generated " + datetime.now().strftime("%Y-%m-%d %I:%M %p"))
        c.drawRightString(W - margin, 28, page_label)

    def section_title(y, title_text):
        c.setFillColor(accent)
        c.setFont("Helvetica-Bold", 9.5)
        c.drawString(margin, y, _safe_text(title_text).upper())
        c.setStrokeColor(accent)
        c.setLineWidth(0.8)
        c.line(margin, y - 4, W - margin, y - 4)
        return y - 16

    def label_value(x, y, label, value, width=150, label_w=72, max_lines=3):
        c.setFillColor(navy)
        c.setFont("Helvetica-Bold", 7.2)
        c.drawString(x, y, _safe_text(label).upper() + ":")
        c.setFillColor(text)
        c.setFont("Helvetica", 8.2)
        lines_wrapped = []
        for part in _safe_text(value).split("\n"):
            lines_wrapped.extend(wrap_pdf_text(part, max(12, int((width - label_w) / 4.1))))
        yy = y
        for line in (lines_wrapped or [""])[:max_lines]:
            c.drawString(x + label_w, yy, line)
            yy -= 9
        return yy

    def two_column_details(y):
        left_x = margin
        right_x = W / 2 + 8
        col_w = (W - 2 * margin - 16) / 2
        box_h = 116 if is_vendor else 138
        c.setFillColor(colors.white)
        c.setStrokeColor(light_line)
        c.roundRect(left_x, y - box_h, col_w, box_h, 4, fill=1, stroke=1)
        c.roundRect(right_x, y - box_h, col_w, box_h, 4, fill=1, stroke=1)
        c.setFillColor(panel)
        c.roundRect(left_x, y - 22, col_w, 22, 4, fill=1, stroke=0)
        c.roundRect(right_x, y - 22, col_w, 22, 4, fill=1, stroke=0)
        c.setFillColor(navy)
        c.setFont("Helvetica-Bold", 7.5)
        c.drawString(left_x + 10, y - 14, "PO DETAILS")
        c.drawString(right_x + 10, y - 14, "VENDOR / DELIVERY")
        yy = y - 38
        label_value(left_x + 10, yy, "PO Date", po_date, col_w - 20, 92, 1)
        yy -= 15
        label_value(left_x + 10, yy, "Required Date", required_date, col_w - 20, 92, 1)
        yy -= 15
        label_value(left_x + 10, yy, "Payment Terms", payment_type, col_w - 20, 92, 1)
        yy -= 15
        label_value(left_x + 10, yy, "Requestor", requestor, col_w - 20, 92, 1)
        yy -= 15
        label_value(left_x + 10, yy, "Department", department, col_w - 20, 92, 1)
        if not is_vendor:
            yy -= 15
            label_value(left_x + 10, yy, "Status", status, col_w - 20, 92, 1)
        # Vendor / Delivery details are intentionally laid out as two side-by-side blocks
        # so the vendor and ship-to information do not crowd the bottom of the card.
        detail_top = y - 40
        detail_bottom = y - box_h + 14
        inner_x = right_x + 10
        inner_w = col_w - 20
        split_x = inner_x + (inner_w / 2)
        c.setStrokeColor(light_line)
        c.setLineWidth(0.6)
        c.line(split_x, detail_top + 6, split_x, detail_bottom)

        c.setFillColor(navy)
        c.setFont("Helvetica-Bold", 7.4)
        c.drawString(inner_x, detail_top, "VENDOR:")
        c.drawString(split_x + 10, detail_top, "SHIP TO:")

        c.setFillColor(text)
        c.setFont("Helvetica", 8.4)
        vendor_y = detail_top - 13
        vendor_wrap_width = max(18, int(((inner_w / 2) - 10) / 4.2))
        for vendor_line in wrap_pdf_text(vendor, vendor_wrap_width)[:5]:
            c.drawString(inner_x, vendor_y, vendor_line)
            vendor_y -= 10

        ship_lines = ["Coastal Engineering Group"]
        if project_code:
            ship_lines.append(f"Project: {project_code} - {project}")
        else:
            ship_lines.append(f"Project: {project}")
        ship_lines.append(f"Attn: {requestor}")
        ship_y = detail_top - 13
        ship_wrap_width = max(18, int(((inner_w / 2) - 12) / 4.2))
        for ship_line in ship_lines[:4]:
            for wrapped_line in wrap_pdf_text(_safe_text(ship_line), ship_wrap_width)[:2]:
                c.drawString(split_x + 10, ship_y, wrapped_line)
                ship_y -= 10
        return y - box_h - 20

    def table_header(y, cols, widths):
        c.setFillColor(accent)
        c.rect(margin, y - 18, sum(widths), 18, fill=1, stroke=0)
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 7.0)
        xx = margin
        for col, w in zip(cols, widths):
            c.drawString(xx + 4, y - 12, _safe_text(col).upper())
            xx += w
        return y - 18

    def new_page():
        c.showPage()
        yy = header()
        return yy

    def line_items_table(y):
        y = section_title(y, "Line Items")
        if is_vendor:
            cols = ["Line", "Description", "Unit", "Qty", "Unit Cost", "Line Amount"]
            widths = [32, 238, 48, 48, 82, 92]
        else:
            cols = ["Line", "Description", "Unit", "Qty", "Unit Cost", "Line Amount", "Open"]
            widths = [32, 178, 48, 48, 72, 82, 80]
        y = table_header(y, cols, widths)
        c.setFont("Helvetica", 7.5)
        total_width = sum(widths)
        if not lines:
            c.setFillColor(colors.white)
            c.setStrokeColor(light_line)
            c.rect(margin, y - 22, total_width, 22, fill=1, stroke=1)
            c.setFillColor(muted)
            c.drawString(margin + 6, y - 14, "No PO line items available.")
            return y - 22
        for idx, line in enumerate(lines, start=1):
            desc_lines = wrap_pdf_text(_safe_text(_row_value(line, "LineDescription")), 39 if is_vendor else 30)[:3]
            row_h = max(22, 11 + len(desc_lines) * 9)
            if y - row_h < 70:
                footer(f"PO {po_number}")
                y = new_page()
                y = section_title(y, "Line Items Continued")
                y = table_header(y, cols, widths)
                c.setFont("Helvetica", 7.5)
            c.setFillColor(colors.white if idx % 2 else panel)
            c.setStrokeColor(light_line)
            c.rect(margin, y - row_h, total_width, row_h, fill=1, stroke=1)
            vals = [
                str(idx),
                "\n".join(desc_lines),
                _safe_text(_row_value(line, "Unit")),
                _safe_text(_row_value(line, "Qty")),
                _money(_row_value(line, "UnitCost", 0)),
                _money(_row_value(line, "LineAmount", 0)),
            ]
            if not is_vendor:
                vals.append(_money(_row_value(line, "RemainingAmount", 0)))
            xx = margin
            for i, (val, w) in enumerate(zip(vals, widths)):
                c.setFillColor(text)
                if "\n" in val:
                    yy = y - 13
                    for dl in val.split("\n"):
                        c.drawString(xx + 4, yy, dl)
                        yy -= 9
                else:
                    c.drawString(xx + 4, y - 13, val[:48])
                xx += w
            y -= row_h
        return y - 10

    def totals_block(y):
        x = W - margin - 190
        row_h = 18
        labels = [("Subtotal", po_value)]
        if not is_vendor:
            labels.extend([("Posted Expenses", posted_amount), ("Current App Balance", current_balance)])
        else:
            labels.append(("Total PO Value", po_value))
        h = 18 + row_h * len(labels)
        if y - h < 70:
            footer(f"PO {po_number}")
            y = new_page()
        c.setStrokeColor(line_color)
        c.setFillColor(colors.white)
        c.roundRect(x, y - h, 190, h, 4, fill=1, stroke=1)
        c.setFillColor(panel)
        c.roundRect(x, y - 18, 190, 18, 4, fill=1, stroke=0)
        c.setFillColor(navy)
        c.setFont("Helvetica-Bold", 7.5)
        c.drawString(x + 10, y - 12, "TOTALS")
        yy = y - 32
        for label, val in labels:
            is_total = label.lower().startswith("total") or label.lower().startswith("current")
            c.setFillColor(navy if is_total else muted)
            c.setFont("Helvetica-Bold" if is_total else "Helvetica", 8.0)
            c.drawString(x + 10, yy, label)
            c.drawRightString(x + 180, yy, _money(val))
            yy -= row_h
        return y - h - 18

    def notes_block(y):
        if y < 135:
            footer(f"PO {po_number}")
            y = new_page()
        note_w = (W - 2 * margin - 18) / 3
        blocks = [
            ("Please Note", "Reference PO number on all invoices."),
            ("Send Invoices To", "accounting@c-diving.com"),
            ("Payment Terms", "Net 30 from receipt of valid invoice and satisfactory delivery."),
        ]
        for i, (lbl, val) in enumerate(blocks):
            x = margin + i * (note_w + 9)
            c.setFillColor(panel)
            c.setStrokeColor(light_line)
            c.roundRect(x, y - 58, note_w, 58, 4, fill=1, stroke=1)
            c.setFillColor(navy)
            c.setFont("Helvetica-Bold", 7.2)
            c.drawString(x + 9, y - 15, lbl.upper())
            c.setFillColor(text)
            c.setFont("Helvetica", 7.4)
            yy = y - 27
            for ln in wrap_pdf_text(val, int(note_w / 4.2))[:3]:
                c.drawString(x + 9, yy, ln)
                yy -= 8
        return y - 74

    def internal_sections(y):
        if payment_schedule:
            if y < 116:
                footer(f"PO {po_number}")
                y = new_page()
            y = section_title(y, "Payment Schedule")
            c.setFillColor(panel)
            c.setStrokeColor(light_line)
            c.roundRect(margin, y - 48, W - 2 * margin, 48, 4, fill=1, stroke=1)
            c.setFillColor(text)
            c.setFont("Helvetica", 7.8)
            yy = y - 13
            for ln in wrap_pdf_text(payment_schedule, 112)[:4]:
                c.drawString(margin + 10, yy, ln)
                yy -= 9
            y -= 64
        if posted_expenses:
            if y < 144:
                footer(f"PO {po_number}")
                y = new_page()
            y = section_title(y, "Posted Expenses")
            cols = ["Date", "Vendor", "Type", "Amount", "Posted By"]
            widths = [74, 154, 98, 86, 128]
            y = table_header(y, cols, widths)
            c.setFont("Helvetica", 7.2)
            for idx, exp in enumerate(posted_expenses[:34], start=1):
                if y < 62:
                    footer(f"PO {po_number}")
                    y = new_page()
                    y = section_title(y, "Posted Expenses Continued")
                    y = table_header(y, cols, widths)
                    c.setFont("Helvetica", 7.2)
                c.setFillColor(colors.white if idx % 2 else panel)
                c.setStrokeColor(light_line)
                c.rect(margin, y - 18, sum(widths), 18, fill=1, stroke=1)
                vals = [
                    _safe_text(_row_value(exp, "TxDate")),
                    _safe_text(_row_value(exp, "VendorName")),
                    _safe_text(_row_value(exp, "TxType")),
                    _money(_row_value(exp, "PostedAmount", _row_value(exp, "Amount", 0))),
                    _safe_text(_row_value(exp, "PostedBy")),
                ]
                xx = margin
                for i, (val, w) in enumerate(zip(vals, widths)):
                    c.setFillColor(text)
                    c.drawString(xx + 4, y - 12, val[:34])
                    xx += w
                y -= 18
            y -= 10

        audit_entries = []
        audit_entries.append((po_date, "PO Created / Loaded", requestor, f"PO created or loaded for {project}."))
        setup_by = _safe_text(_row_value(po, "SetupUpdatedBy"))
        setup_at = _safe_text(_row_value(po, "SetupUpdatedAt"))
        if setup_by or setup_at:
            audit_entries.append((setup_at or "Not recorded", "PO Setup Updated", setup_by or "System", "Setup, payment schedule, or review status updated."))
        for exp in (posted_expenses or [])[:8]:
            audit_entries.append((
                _safe_text(_row_value(exp, "PostedAt")) or _safe_text(_row_value(exp, "TxDate")),
                "Expense Posted",
                _safe_text(_row_value(exp, "PostedBy")) or _safe_text(_row_value(exp, "ReviewerEmail")) or "System",
                f"{_money(_row_value(exp, 'PostedAmount', _row_value(exp, 'Amount', 0)))} posted to this PO from {_safe_text(_row_value(exp, 'VendorName')) or 'expense row'}."
            ))
        audit_entries.append((datetime.now().strftime("%Y-%m-%d %I:%M %p"), "Packet Generated", "System", "Internal PO packet generated for review."))

        if y < 150:
            footer(f"PO {po_number}")
            y = new_page()
        y = section_title(y, "Audit Trail")
        cols = ["Date / Time", "Action", "User", "Details"]
        widths = [96, 112, 128, 204]
        y = table_header(y, cols, widths)
        c.setFont("Helvetica", 7.0)
        for idx, (dt, action, user_name, detail) in enumerate(audit_entries[:14], start=1):
            detail_lines = wrap_pdf_text(_safe_text(detail), 48)[:3]
            row_h = max(20, 9 + len(detail_lines) * 8)
            if y - row_h < 62:
                footer(f"PO {po_number}")
                y = new_page()
                y = section_title(y, "Audit Trail Continued")
                y = table_header(y, cols, widths)
                c.setFont("Helvetica", 7.0)
            c.setFillColor(colors.white if idx % 2 else panel)
            c.setStrokeColor(light_line)
            c.rect(margin, y - row_h, sum(widths), row_h, fill=1, stroke=1)
            vals = [_safe_text(dt), _safe_text(action), _safe_text(user_name), "\n".join(detail_lines)]
            xx = margin
            for i, (val, w) in enumerate(zip(vals, widths)):
                c.setFillColor(text)
                if "\n" in val:
                    yy = y - 12
                    for dl in val.split("\n"):
                        c.drawString(xx + 4, yy, dl[:62])
                        yy -= 8
                else:
                    c.drawString(xx + 4, y - 12, val[:32])
                xx += w
            y -= row_h
        y -= 10
        return y

    def terms_block(y):
        terms = [
            ("Delivery", "Vendor must deliver goods or perform services by the required date. Delays without written approval may result in cancellation."),
            ("Invoicing & Payment", "Include the PO number on all invoices. Send invoices to accounting@c-diving.com. Unless otherwise agreed, payment terms are Net 30 from receipt of a valid invoice and satisfactory delivery."),
            ("Changes", "No substitutions or changes to quantity or delivery date without written approval from Coastal Engineering."),
            ("Inspection", "All items are subject to inspection. Non-compliant goods or services may be rejected at the vendor's expense."),
            ("Warranties", "Vendor warrants that goods and services are free from defects, conform to specifications, and are fit for their intended use."),
            ("Compliance", "Vendor must comply with all applicable laws and regulations."),
            ("Indemnification", "Vendor agrees to hold Coastal Engineering harmless from any claims or liabilities arising from this Purchase Order."),
            ("PO Cancellation", "Coastal Engineering reserves the right to cancel this PO at any time for undelivered goods or services."),
        ]
        # Keep terms readable and aligned to the same page/table edges.
        # The old version used too many characters per line, which made the terms
        # look cramped and caused long lines to run across the section.
        if y < 230:
            footer(f"PO {po_number}")
            y = new_page()
        y = section_title(y, "Coastal Engineering PO Terms and Conditions")

        box_x = margin
        box_w = W - 2 * margin
        inner_x = box_x + 12
        inner_w = box_w - 24
        max_chars = 104
        row_gap = 8
        line_h = 10.2

        # Pre-wrap so the box height matches the actual terms content.
        prepared = []
        total_h = 18
        for label, body in terms:
            first = f"{label}: {body}"
            wrapped = wrap_pdf_text(first, max_chars)
            prepared.append((label, wrapped))
            total_h += max(16, len(wrapped) * line_h) + row_gap
        total_h = min(total_h + 8, y - 58)

        c.setFillColor(colors.HexColor("#f8fbff"))
        c.setStrokeColor(colors.HexColor("#bfdbfe"))
        c.roundRect(box_x, y - total_h, box_w, total_h, 4, fill=1, stroke=1)

        yy = y - 17
        for idx, (label, wrapped) in enumerate(prepared, start=1):
            # If the terms section gets too long for the page, continue cleanly on a new page.
            needed = max(16, len(wrapped) * line_h) + row_gap
            if yy - needed < 58:
                footer(f"PO {po_number}")
                y = new_page()
                y = section_title(y, "Coastal Engineering PO Terms and Conditions Continued")
                yy = y - 14
                remaining_h = max(120, min(y - 58, 260))
                c.setFillColor(colors.HexColor("#f8fbff"))
                c.setStrokeColor(colors.HexColor("#bfdbfe"))
                c.roundRect(box_x, y - remaining_h, box_w, remaining_h, 4, fill=1, stroke=1)

            bullet_x = inner_x
            text_x = inner_x + 12
            c.setFillColor(accent)
            c.circle(bullet_x + 3, yy - 2, 2.2, fill=1, stroke=0)

            for line_idx, txt in enumerate(wrapped):
                if line_idx == 0 and ":" in txt:
                    label_part, rest = txt.split(":", 1)
                    c.setFillColor(navy)
                    c.setFont("Helvetica-Bold", 8.6)
                    c.drawString(text_x, yy, label_part + ":")
                    label_width = c.stringWidth(label_part + ": ", "Helvetica-Bold", 8.6)
                    c.setFillColor(text)
                    c.setFont("Helvetica", 8.6)
                    c.drawString(text_x + label_width, yy, rest.strip())
                else:
                    c.setFillColor(text)
                    c.setFont("Helvetica", 8.6)
                    c.drawString(text_x, yy, txt)
                yy -= line_h
            yy -= row_gap
        return yy - 6

    y = header()
    y = two_column_details(y)
    y = line_items_table(y)
    y = totals_block(y)
    y = notes_block(y)
    if is_vendor:
        y = terms_block(y)
    else:
        y = internal_sections(y)
    footer(f"PO {po_number}")
    c.save()
    return buf.getvalue()

def po_packet_pdf_attachment(po_number, packet_type="internal"):
    po, lines, posted_expenses = load_po_packet_data(po_number)
    if not po:
        return None
    filename_type = "vendor" if packet_type == "vendor" else "internal"
    return {
        "filename": f"{po.PONumber}_{filename_type}_packet.pdf",
        "content": _styled_po_packet_pdf_bytes(po, lines, posted_expenses, packet_type),
        "maintype": "application",
        "subtype": "pdf",
    }



APPROVAL_THRESHOLD_AMOUNT = Decimal("3000.00")


def ensure_purchase_request_approval_columns(cursor):
    """Add July 1 approval tracking columns if they do not already exist."""
    cursor.execute("""
    IF COL_LENGTH('dbo.PurchaseRequests', 'AdminApprovedBy') IS NULL
        ALTER TABLE dbo.PurchaseRequests ADD AdminApprovedBy NVARCHAR(255) NULL;
    IF COL_LENGTH('dbo.PurchaseRequests', 'AdminApprovedAt') IS NULL
        ALTER TABLE dbo.PurchaseRequests ADD AdminApprovedAt DATETIME2 NULL;
    IF COL_LENGTH('dbo.PurchaseRequests', 'ExecutiveApprovedBy') IS NULL
        ALTER TABLE dbo.PurchaseRequests ADD ExecutiveApprovedBy NVARCHAR(255) NULL;
    IF COL_LENGTH('dbo.PurchaseRequests', 'ExecutiveApprovedAt') IS NULL
        ALTER TABLE dbo.PurchaseRequests ADD ExecutiveApprovedAt DATETIME2 NULL;
    IF COL_LENGTH('dbo.PurchaseRequests', 'ApprovalRequirement') IS NULL
        ALTER TABLE dbo.PurchaseRequests ADD ApprovalRequirement NVARCHAR(100) NULL;
    """)


def requires_executive_approval(amount):
    return clean_decimal(amount) >= APPROVAL_THRESHOLD_AMOUNT


def approval_requirement_label(amount):
    if requires_executive_approval(amount):
        return "Admin + Executive approval required"
    return "Admin approval required"


def purchase_request_approval_roles_for_amount(amount):
    if requires_executive_approval(amount):
        return ["Admin", "Executive"]
    return ["Admin"]


def load_purchase_request_approval_recipients(req):
    return load_active_users_by_roles(purchase_request_approval_roles_for_amount(getattr(req, "EstimatedAmount", 0)))


def send_purchase_request_pending_approval_email(purchase_request_id, remaining_role):
    req = load_purchase_request_for_email(purchase_request_id)
    if not req:
        return
    recipients = load_active_users_by_roles([remaining_role])
    if not recipients:
        return
    review_url = app_url("/purchase-requests")
    body = f"""
    <h2>Purchase request still needs {h(remaining_role)} approval</h2>
    <p>This request is $3,000 or more and requires both Admin and Executive approval. One approval has been recorded; the remaining approval is still needed before the PO is created.</p>
    {purchase_request_summary_html(req)}
    {html_button("Open Purchase Requests", review_url)}
    """
    send_email_notification(recipients, f"Approval still needed: {req.RequestNumber}", body)

def load_purchase_request_for_email(purchase_request_id):
    conn = get_sql_connection()
    cursor = conn.cursor()
    ensure_purchase_request_approval_columns(cursor)
    conn.commit()
    cursor.execute(
        """
        SELECT TOP 1 PurchaseRequestId, RequestNumber, RequestedByEmail, RequestedByName,
               RequestedAt, NeededByDate, VendorName, ProjectCode, ProjectName, Department,
               RequestTitle, RequestDescription, EstimatedAmount, Priority, RequestStatus,
               ReviewerEmail, ReviewedAt, ReviewNotes, ConvertedPONumber, UpdatedAt,
               AdminApprovedBy, AdminApprovedAt, ExecutiveApprovedBy, ExecutiveApprovedAt,
               ApprovalRequirement
        FROM dbo.PurchaseRequests
        WHERE PurchaseRequestId = ?;
        """,
        purchase_request_id,
    )
    row = cursor.fetchone()
    conn.close()
    return row

def purchase_request_summary_html(req):
    return f"""
    <table cellpadding="6" cellspacing="0" style="border-collapse:collapse;border:1px solid #e2e8f0;">
      <tr><th align="left">Request</th><td>{h(req.RequestNumber)}</td></tr>
      <tr><th align="left">Title</th><td>{h(req.RequestTitle)}</td></tr>
      <tr><th align="left">Vendor</th><td>{h(req.VendorName)}</td></tr>
      <tr><th align="left">Project</th><td>{h(req.ProjectName)}</td></tr>
      <tr><th align="left">Department</th><td>{h(req.Department)}</td></tr>
      <tr><th align="left">Amount</th><td>{currency(req.EstimatedAmount)}</td></tr>
      <tr><th align="left">Approval Required</th><td>{h(approval_requirement_label(req.EstimatedAmount))}</td></tr>
      <tr><th align="left">Priority</th><td>{h(req.Priority)}</td></tr>
      <tr><th align="left">Needed By</th><td>{h(req.NeededByDate)}</td></tr>
      <tr><th align="left">Requested By</th><td>{h(req.RequestedByName or req.RequestedByEmail)}</td></tr>
    </table>
    """


def send_purchase_request_submitted_emails(purchase_request_id):
    req = load_purchase_request_for_email(purchase_request_id)
    if not req:
        return

    approvers = load_purchase_request_approval_recipients(req)
    approver_names = ", ".join(user_display_name(a) for a in approvers) or approval_requirement_label(req.EstimatedAmount)
    review_url = app_url("/purchase-requests")
    subject = f"Purchase request approval needed: {req.RequestNumber}"
    body = f"""
    <h2>Purchase request approval needed</h2>
    <p>A purchase request has been submitted and is ready for review.</p>
    <p><strong>July 1 approval rule:</strong> {h(approval_requirement_label(req.EstimatedAmount))}.</p>
    <p>For requests $3,000 and up, Admin and Executive can approve in any order. The PO is created only after both approvals are recorded.</p>
    {purchase_request_summary_html(req)}
    {html_button("Open Purchase Requests", review_url)}
    """
    send_email_notification(approvers, subject, body)

    requester = clean_email(req.RequestedByEmail)
    requester_subject = f"What happens now: {req.RequestNumber}"
    requester_body = f"""
    <h2>Your purchase request was submitted</h2>
    <p>Your request was submitted and routed for approval.</p>
    <p><strong>Approval required:</strong> {h(approval_requirement_label(req.EstimatedAmount))}.</p>
    <p>Approval email was sent to: <strong>{h(approver_names)}</strong>.</p>
    <p>You will receive another email after it is fully approved and converted to a PO.</p>
    {purchase_request_summary_html(req)}
    {html_button("View Purchase Requests", review_url)}
    """
    send_email_notification([requester], requester_subject, requester_body)

def send_purchase_request_approved_email(purchase_request_id, po_number):
    req = load_purchase_request_for_email(purchase_request_id)
    if not req:
        return
    internal_url = app_url("/po-packet-pdf/" + quote_plus(str(po_number)) + "?type=internal")
    vendor_url = app_url("/po-packet-pdf/" + quote_plus(str(po_number)) + "?type=vendor")
    attachments = []
    for packet_type in ["internal", "vendor"]:
        att = po_packet_pdf_attachment(po_number, packet_type)
        if att:
            attachments.append(att)
    body = f"""
    <h2>Purchase request approved and converted to PO</h2>
    <p>Purchase request <strong>{h(req.RequestNumber)}</strong> has been approved and converted to PO <strong>{h(po_number)}</strong>.</p>
    {purchase_request_summary_html(req)}
    {html_button("Open Internal PO Packet PDF", internal_url)}
    <p><a href="{h(vendor_url)}">Open Vendor PO Packet PDF</a></p>
    <p>The internal and vendor-facing PO packet PDFs are attached. The buttons above open the same PDF format as the attachments.</p>
    """
    recipients = unique_recipients([req.RequestedByEmail], [req.ReviewerEmail])
    send_email_notification(recipients, f"Approved PO created: {po_number}", body, attachments=attachments)


def send_po_upload_summary_email(import_result, filename, department, requestor_name, requestor_email=""):
    admins_execs = load_active_users_by_roles(["Admin", "Executive"])
    recipients = unique_recipients(admins_execs, [requestor_email])
    if not recipients:
        return
    upload_url = app_url("/pos-balances")
    body = f"""
    <h2>Issued POs uploaded</h2>
    <p>A new issued PO setup file was uploaded into the Command Center.</p>
    <table cellpadding="6" cellspacing="0" style="border-collapse:collapse;border:1px solid #e2e8f0;">
      <tr><th align="left">File</th><td>{h(filename)}</td></tr>
      <tr><th align="left">Import Batch</th><td>{h(import_result.get('import_batch_id'))}</td></tr>
      <tr><th align="left">Rows</th><td>{h(import_result.get('success_count'))} successful / {h(import_result.get('total_rows'))} total</td></tr>
      <tr><th align="left">Department</th><td>{h(department)}</td></tr>
      <tr><th align="left">Requestor</th><td>{h(requestor_name)}</td></tr>
      <tr><th align="left">Status</th><td>{h(import_result.get('status'))}</td></tr>
    </table>
    {html_button("Open POs & Balances", upload_url)}
    """
    send_email_notification(recipients, "Issued PO upload summary", body)

def role_can_access(role_name, page_name):
    allowed_roles = PAGE_ACCESS.get(page_name, [])
    return normalize_role(role_name) in allowed_roles


def require_page_access(page_name):
    user = get_current_user()
    actual_access = get_user_access()

    if not user["is_authenticated"]:
        return False, "Microsoft login was not detected."

    if not user["is_allowed_domain"]:
        return False, f"Your email domain is not allowed. Expected @{user['allowed_domain']}."

    if not actual_access["found_in_sql"]:
        return False, "Your account has not been added to the dashboard access list."

    if not actual_access["is_active"]:
        return False, "Your dashboard account is inactive."

    # User Access itself remains available to the real Admin/Executive so they can
    # change users or clear View As even while testing a restricted role.
    if page_name == "User Access" and role_can_access(actual_access["role"], page_name):
        return True, ""

    access = get_effective_user_access()
    if not access.get("found_in_sql"):
        return False, "The selected view-as account has not been added to the dashboard access list."

    if not access.get("is_active"):
        return False, "The selected view-as dashboard account is inactive."

    if not role_can_access(access["role"], page_name):
        return False, f"Your effective role, {access['role']}, does not have access to {page_name}."

    return True, ""


def access_denied_response(page_name, reason):
    content = """
    <div class="notice error">Access denied.</div>

    <div class="card">
        <h3>You do not have access to this dashboard</h3>
        <p class="card-subtitle">
            Your Microsoft account is signed in, but it has not been approved for this procurement dashboard.
        </p>

        <p>
            Contact a dashboard Admin if you believe you should have access.
        </p>
    </div>
    """

    return shell(
        title="Access Denied",
        subtitle="Your account is not approved for this dashboard.",
        active="",
        content=content,
    ), 403


# ------------------------------------------------------------
# File upload / import helpers
# ------------------------------------------------------------

def read_uploaded_po_file(uploaded_file):
    filename = uploaded_file.filename or ""

    if filename.lower().endswith(".csv"):
        raw = uploaded_file.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(raw))
        rows = []

        for row in reader:
            rows.append({normalize_header(k): v for k, v in row.items()})

        return rows

    if filename.lower().endswith(".xlsx"):
        from openpyxl import load_workbook

        workbook = load_workbook(uploaded_file, data_only=True)
        sheet = workbook.active

        headers = [normalize_header(cell.value) for cell in sheet[1]]
        rows = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_dict = {}

            for index, value in enumerate(row):
                if index < len(headers):
                    row_dict[headers[index]] = value

            if any(value not in [None, ""] for value in row_dict.values()):
                rows.append(row_dict)

        return rows

    raise ValueError("Unsupported file type. Please upload a .csv file.")


def validate_po_rows(rows, selected_department=None, selected_requestor=None, valid_requestors=None):
    errors = []

    if not rows:
        errors.append("The file has no data rows.")
        return errors

    selected_department = clean_text(selected_department)
    selected_requestor = clean_text(selected_requestor)
    valid_requestors = set(valid_requestors or [])

    if selected_department not in DEPARTMENT_OPTIONS:
        errors.append("Select a department before uploading. Department must be one of: " + ", ".join(DEPARTMENT_OPTIONS) + ".")

    if not selected_requestor:
        errors.append("Select a requestor before uploading.")
    elif valid_requestors and selected_requestor not in valid_requestors:
        errors.append("Selected requestor is not an active user in User Access.")

    actual_columns = set(rows[0].keys())
    missing_columns = [col for col in REQUIRED_PO_COLUMNS if col not in actual_columns]

    if missing_columns:
        errors.append("Missing required columns: " + ", ".join(missing_columns))

    for index, row in enumerate(rows, start=2):
        project_code = clean_text(row.get("ProjectCode"))
        po_number = clean_text(row.get("PONumber"))
        vendor_name = clean_text(row.get("VendorName"))
        project_name = clean_text(row.get("ProjectName"))
        line_amount = clean_decimal(row.get("LineAmount"))

        if not project_code:
            errors.append(f"Row {index}: ProjectCode is required.")

        if not po_number:
            errors.append(f"Row {index}: PONumber is required.")

        if not vendor_name:
            errors.append(f"Row {index}: VendorName is required.")

        if not project_name:
            errors.append(f"Row {index}: ProjectName is required.")

        if line_amount is None:
            errors.append(f"Row {index}: LineAmount is required and must be a number.")

    return errors


def get_or_create_vendor(cursor, vendor_name):
    cursor.execute("SELECT VendorId FROM dbo.Vendors WHERE VendorName = ?", vendor_name)
    row = cursor.fetchone()

    if row:
        return row.VendorId

    cursor.execute(
        """
        INSERT INTO dbo.Vendors (VendorName, IsActive)
        OUTPUT INSERTED.VendorId
        VALUES (?, 1)
        """,
        vendor_name,
    )

    return cursor.fetchone().VendorId


def ensure_project_code_columns(cursor):
    """Add ProjectCode to project/PO tables for the Phase 1 setup template."""
    for table_name in ["Projects", "PurchaseOrders", "IssuedPOLines"]:
        cursor.execute(
            f"""
            IF COL_LENGTH('dbo.{table_name}', 'ProjectCode') IS NULL
            BEGIN
                ALTER TABLE dbo.{table_name} ADD ProjectCode NVARCHAR(100) NULL;
            END
            """
        )


def get_or_create_project(cursor, project_name, department, project_code=None):
    project_code = clean_text(project_code)
    project_name = clean_text(project_name)

    ensure_project_code_columns(cursor)

    if project_code:
        cursor.execute(
            "SELECT ProjectId FROM dbo.Projects WHERE ProjectCode = ? OR ProjectName = ?",
            project_code,
            project_name,
        )
    else:
        cursor.execute("SELECT ProjectId FROM dbo.Projects WHERE ProjectName = ?", project_name)

    row = cursor.fetchone()

    if row:
        cursor.execute(
            """
            UPDATE dbo.Projects
            SET ProjectCode = COALESCE(NULLIF(ProjectCode, ''), ?),
                Department = COALESCE(NULLIF(Department, ''), ?)
            WHERE ProjectId = ?
            """,
            project_code,
            department,
            row.ProjectId,
        )
        return row.ProjectId

    cursor.execute(
        """
        INSERT INTO dbo.Projects (ProjectCode, ProjectName, Department, IsActive)
        OUTPUT INSERTED.ProjectId
        VALUES (?, ?, ?, 1)
        """,
        project_code,
        project_name,
        department,
    )

    return cursor.fetchone().ProjectId


def upsert_purchase_order(
    cursor,
    po_number,
    vendor_id,
    project_id,
    department,
    requestor,
    po_date,
    po_status,
    original_amount,
    revised_amount,
    remaining_amount,
    import_batch_id,
    project_code=None,
):
    cursor.execute(
        "SELECT PurchaseOrderId FROM dbo.PurchaseOrders WHERE PONumber = ?",
        po_number,
    )
    row = cursor.fetchone()

    if row:
        purchase_order_id = row.PurchaseOrderId

        cursor.execute(
            """
            UPDATE dbo.PurchaseOrders
            SET
                VendorId = ?,
                ProjectId = ?,
                ProjectCode = ?,
                Department = ?,
                Requestor = ?,
                PODate = ?,
                POStatus = ?,
                OriginalAmount = ?,
                RevisedAmount = ?,
                RemainingAmount = ?,
                LastImportBatchId = ?,
                UpdatedAt = SYSUTCDATETIME()
            WHERE PurchaseOrderId = ?
            """,
            vendor_id,
            project_id,
            project_code,
            department,
            requestor,
            po_date,
            po_status,
            original_amount,
            revised_amount,
            remaining_amount,
            import_batch_id,
            purchase_order_id,
        )

        return purchase_order_id

    cursor.execute(
        """
        INSERT INTO dbo.PurchaseOrders
            (
                PONumber,
                VendorId,
                ProjectId,
                ProjectCode,
                Department,
                Requestor,
                PODate,
                POStatus,
                OriginalAmount,
                RevisedAmount,
                PaidAmount,
                RemainingAmount,
                LastImportBatchId
            )
        OUTPUT INSERTED.PurchaseOrderId
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, ?, ?)
        """,
        po_number,
        vendor_id,
        project_id,
        project_code,
        department,
        requestor,
        po_date,
        po_status,
        original_amount,
        revised_amount,
        remaining_amount,
        import_batch_id,
    )

    return cursor.fetchone().PurchaseOrderId


def import_po_rows(rows, filename, selected_department=None, selected_requestor=None):
    conn = get_sql_connection()
    cursor = conn.cursor()

    try:
        cursor.execute(
            """
            INSERT INTO dbo.ImportBatches
                (
                    FileName,
                    SourceSystem,
                    UploadedBy,
                    TotalRows,
                    SuccessCount,
                    ErrorCount,
                    ImportStatus
                )
            OUTPUT INSERTED.ImportBatchId
            VALUES (?, 'PO Upload', ?, ?, 0, 0, 'Processing')
            """,
            filename,
            get_current_user()["email"] or "Manual Upload",
            len(rows),
        )

        import_batch_id = cursor.fetchone().ImportBatchId
        success_count = 0
        error_count = 0

        ensure_project_code_columns(cursor)

        po_numbers = sorted(
            set(clean_text(row.get("PONumber")) for row in rows if clean_text(row.get("PONumber")))
        )

        po_totals = {}
        for source_row in rows:
            source_po = clean_text(source_row.get("PONumber"))
            source_amount = clean_decimal(source_row.get("LineAmount")) or Decimal("0")
            if source_po:
                po_totals[source_po] = po_totals.get(source_po, Decimal("0")) + source_amount

        for po_number in po_numbers:
            cursor.execute("DELETE FROM dbo.IssuedPOLines WHERE PONumber = ?", po_number)

        for index, row in enumerate(rows, start=2):
            try:
                project_code = clean_text(row.get("ProjectCode"))
                po_number = clean_text(row.get("PONumber"))
                vendor_name = clean_text(row.get("VendorName"))
                project_name = clean_text(row.get("ProjectName"))
                department = clean_text(selected_department)
                po_date = clean_date(row.get("PODate"))
                po_status = "Open"
                description = clean_text(row.get("Description"))
                unit = clean_text(row.get("Unit"))
                unit_cost = clean_decimal(row.get("UnitCost"))
                qty = clean_decimal(row.get("Qty"))
                line_amount = clean_decimal(row.get("LineAmount"))
                original_amount = po_totals.get(po_number, Decimal("0"))
                revised_amount = original_amount
                remaining_amount = original_amount
                requestor = clean_text(selected_requestor)

                vendor_id = get_or_create_vendor(cursor, vendor_name)
                project_id = get_or_create_project(cursor, project_name, department, project_code)

                purchase_order_id = upsert_purchase_order(
                    cursor=cursor,
                    po_number=po_number,
                    vendor_id=vendor_id,
                    project_id=project_id,
                    department=department,
                    requestor=requestor,
                    po_date=po_date,
                    po_status=po_status,
                    original_amount=original_amount,
                    revised_amount=revised_amount,
                    remaining_amount=remaining_amount,
                    import_batch_id=import_batch_id,
                    project_code=project_code,
                )

                cursor.execute(
                    """
                    INSERT INTO dbo.IssuedPOLines
                        (
                            PurchaseOrderId,
                            ImportBatchId,
                            PONumber,
                            VendorName,
                            ProjectCode,
                            ProjectName,
                            Department,
                            PODate,
                            POStatus,
                            LineDescription,
                            Unit,
                            UnitCost,
                            Qty,
                            LineAmount,
                            OriginalAmount,
                            RevisedAmount,
                            RemainingAmount,
                            Requestor
                        )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    purchase_order_id,
                    import_batch_id,
                    po_number,
                    vendor_name,
                    project_code,
                    project_name,
                    department,
                    po_date,
                    po_status,
                    description,
                    unit,
                    unit_cost,
                    qty,
                    line_amount,
                    original_amount,
                    revised_amount,
                    remaining_amount,
                    requestor,
                )

                success_count += 1

            except Exception as row_error:
                error_count += 1

                cursor.execute(
                    """
                    INSERT INTO dbo.ImportErrors
                        (
                            ImportBatchId,
                            RowNumber,
                            ErrorMessage,
                            RawRow
                        )
                    VALUES (?, ?, ?, ?)
                    """,
                    import_batch_id,
                    index,
                    str(row_error),
                    str(row),
                )

        final_status = "Completed" if error_count == 0 else "Completed With Errors"

        cursor.execute(
            """
            UPDATE dbo.ImportBatches
            SET
                SuccessCount = ?,
                ErrorCount = ?,
                ImportStatus = ?
            WHERE ImportBatchId = ?
            """,
            success_count,
            error_count,
            final_status,
            import_batch_id,
        )

        conn.commit()

        return {
            "import_batch_id": import_batch_id,
            "total_rows": len(rows),
            "success_count": success_count,
            "error_count": error_count,
            "status": final_status,
        }

    except Exception:
        conn.rollback()
        raise

    finally:
        conn.close()


# ------------------------------------------------------------
# Expense upload / PO matching helpers
# ------------------------------------------------------------

REQUIRED_EXPENSE_COLUMNS = [
    "ExpenseId",
    "ProjectName",
    "TxDate",
    "TxType",
    "VendorName",
    "Amount",
    "Description",
    "PMComments",
]


def row_value(row, aliases):
    """Return the first populated value from a row using several possible column names."""
    normalized = {str(k or "").strip().lower(): v for k, v in row.items()}
    for alias in aliases:
        key = alias.strip().lower()
        if key in normalized and normalized[key] not in [None, ""]:
            return normalized[key]
    return None


def compact_match_text(value):
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def ensure_expense_review_tables():
    conn = get_sql_connection()
    cursor = conn.cursor()
    cursor.execute(
        """
        IF OBJECT_ID('dbo.ExpenseUploadBatches', 'U') IS NULL
        BEGIN
            CREATE TABLE dbo.ExpenseUploadBatches (
                ExpenseBatchId INT IDENTITY(1,1) PRIMARY KEY,
                FileName NVARCHAR(255) NOT NULL,
                UploadedAt DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME(),
                UploadedBy NVARCHAR(255) NULL,
                TotalRows INT NOT NULL DEFAULT 0,
                AutoMatchedCount INT NOT NULL DEFAULT 0,
                NeedsReviewCount INT NOT NULL DEFAULT 0,
                NoMatchCount INT NOT NULL DEFAULT 0,
                ImportStatus NVARCHAR(100) NOT NULL DEFAULT 'Completed'
            );
        END;

        IF OBJECT_ID('dbo.ExpenseReviewItems', 'U') IS NULL
        BEGIN
            CREATE TABLE dbo.ExpenseReviewItems (
                ExpenseReviewItemId INT IDENTITY(1,1) PRIMARY KEY,
                ExpenseBatchId INT NOT NULL,
                SourceRowNumber INT NULL,
                ExpenseId NVARCHAR(100) NULL,
                ProjectName NVARCHAR(255) NULL,
                TxDate DATE NULL,
                TxType NVARCHAR(100) NULL,
                VendorName NVARCHAR(255) NULL,
                Description NVARCHAR(MAX) NULL,
                Amount DECIMAL(18,2) NULL,
                PMComments NVARCHAR(MAX) NULL,
                ExtractedPONumber NVARCHAR(100) NULL,
                MatchedPONumber NVARCHAR(100) NULL,
                MatchStatus NVARCHAR(100) NOT NULL DEFAULT 'Needs Review',
                MatchConfidence NVARCHAR(50) NULL,
                MatchReason NVARCHAR(MAX) NULL,
                ReviewDecision NVARCHAR(100) NULL,
                CorrectPONumber NVARCHAR(100) NULL,
                ReviewerNotes NVARCHAR(MAX) NULL,
                CreatedAt DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME(),
                UpdatedAt DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME()
            );
            CREATE INDEX IX_ExpenseReviewItems_Status ON dbo.ExpenseReviewItems(MatchStatus);
            CREATE INDEX IX_ExpenseReviewItems_Project ON dbo.ExpenseReviewItems(ProjectName);
            CREATE INDEX IX_ExpenseReviewItems_PO ON dbo.ExpenseReviewItems(MatchedPONumber);
        END;
        """
    )
    # Add rollout balance-posting/audit columns to existing databases.
    rollout_columns = [
        ("ExpenseReviewItems", "ExpenseUniqueKey", "NVARCHAR(500) NULL"),
        ("ExpenseReviewItems", "ReviewerEmail", "NVARCHAR(255) NULL"),
        ("ExpenseReviewItems", "ReviewedAt", "DATETIME2 NULL"),
        ("ExpenseReviewItems", "PostedToPO", "BIT NOT NULL CONSTRAINT DF_ExpenseReviewItems_PostedToPO DEFAULT 0"),
        ("ExpenseReviewItems", "PostedPONumber", "NVARCHAR(100) NULL"),
        ("ExpenseReviewItems", "PostedAmount", "DECIMAL(18,2) NULL"),
        ("ExpenseReviewItems", "PostedAt", "DATETIME2 NULL"),
        ("ExpenseReviewItems", "PostedBy", "NVARCHAR(255) NULL"),
        ("ExpenseReviewItems", "PostingBatchId", "INT NULL"),
        ("ExpenseReviewItems", "IsDuplicate", "BIT NOT NULL CONSTRAINT DF_ExpenseReviewItems_IsDuplicate DEFAULT 0"),
        ("ExpenseUploadBatches", "DuplicateCount", "INT NOT NULL CONSTRAINT DF_ExpenseUploadBatches_DuplicateCount DEFAULT 0"),
        ("ExpenseUploadBatches", "PostedCount", "INT NOT NULL CONSTRAINT DF_ExpenseUploadBatches_PostedCount DEFAULT 0"),
        ("ExpenseUploadBatches", "PostedAmount", "DECIMAL(18,2) NOT NULL CONSTRAINT DF_ExpenseUploadBatches_PostedAmount DEFAULT 0"),
    ]
    for table_name, column_name, column_type in rollout_columns:
        cursor.execute(f"""
            IF COL_LENGTH('dbo.{table_name}', '{column_name}') IS NULL
            BEGIN
                ALTER TABLE dbo.{table_name} ADD {column_name} {column_type};
            END
        """)

    cursor.execute("""
        IF NOT EXISTS (SELECT 1 FROM sys.indexes WHERE name = 'IX_ExpenseReviewItems_PostedPO' AND object_id = OBJECT_ID('dbo.ExpenseReviewItems'))
        BEGIN
            CREATE INDEX IX_ExpenseReviewItems_PostedPO ON dbo.ExpenseReviewItems(PostedToPO, PostedPONumber);
        END;
        IF NOT EXISTS (SELECT 1 FROM sys.indexes WHERE name = 'IX_ExpenseReviewItems_UniqueKey' AND object_id = OBJECT_ID('dbo.ExpenseReviewItems'))
        BEGIN
            CREATE INDEX IX_ExpenseReviewItems_UniqueKey ON dbo.ExpenseReviewItems(ExpenseUniqueKey);
        END;
    """)

    # Backfill existing reviewed/auto-matched rows so current uploaded expenses begin reducing balances after deployment.
    cursor.execute("""
        UPDATE dbo.ExpenseReviewItems
        SET PostedToPO = 1,
            PostedPONumber = COALESCE(NULLIF(CorrectPONumber, ''), NULLIF(MatchedPONumber, '')),
            PostedAmount = COALESCE(Amount, 0),
            PostedAt = COALESCE(PostedAt, UpdatedAt, SYSUTCDATETIME()),
            PostedBy = COALESCE(NULLIF(ReviewerEmail, ''), 'System Backfill'),
            UpdatedAt = SYSUTCDATETIME()
        WHERE COALESCE(PostedToPO, 0) = 0
          AND COALESCE(NULLIF(CorrectPONumber, ''), NULLIF(MatchedPONumber, '')) IS NOT NULL
          AND (
                ReviewDecision = 'Matched to PO'
                OR (
                    MatchStatus = 'Auto Matched'
                    AND LOWER(COALESCE(TxType, '')) NOT LIKE '%time sheet%'
                    AND LOWER(COALESCE(TxType, '')) NOT LIKE '%labor%'
                    AND LOWER(COALESCE(TxType, '')) NOT LIKE '%per diem%'
                    AND LOWER(COALESCE(TxType, '')) NOT LIKE '%perdiem%'
                    AND LOWER(COALESCE(TxType, '')) NOT LIKE '%mileage%'
                    AND LOWER(COALESCE(TxType, '')) NOT LIKE '%travel stipend%'
                    AND (LOWER(COALESCE(TxType, '')) LIKE '%purchase%' OR LOWER(COALESCE(TxType, '')) LIKE '%disbursement%')
                )
          );
    """)
    conn.commit()
    conn.close()


def expense_unique_key(project_name, tx_date, tx_type, vendor_name, amount, expense_id, pm_comments):
    parts = [
        compact_match_text(project_name),
        str(tx_date or ''),
        compact_match_text(tx_type),
        compact_match_text(vendor_name),
        str(clean_decimal(amount) or Decimal('0')),
        compact_match_text(expense_id),
        compact_match_text(str(pm_comments or '')[:80]),
    ]
    return '|'.join(parts)[:500]


def expense_type_auto_posts(tx_type):
    text = str(tx_type or '').strip().lower()
    if not text:
        return False
    excluded = ['time sheet', 'labor', 'per diem', 'perdiem', 'mileage', 'travel stipend']
    if any(term in text for term in excluded):
        return False
    return any(term in text for term in ['purchase', 'disbursement'])


def sync_expense_posting(cursor, item_id, posted_by=None):
    """Synchronize whether one expense review item reduces an app PO balance.

    POs are uploaded once at project setup. From that point forward, current app
    balance is calculated as issued PO amount minus review items where PostedToPO=1.
    This function is idempotent and prevents the same review row from being counted
    more than once.
    """
    posted_by = posted_by or (get_current_user().get('email') or 'System')
    cursor.execute(
        """
        SELECT ExpenseReviewItemId, TxType, Amount, MatchedPONumber, CorrectPONumber,
               MatchStatus, ReviewDecision, PostedToPO
        FROM dbo.ExpenseReviewItems
        WHERE ExpenseReviewItemId = ?;
        """,
        item_id,
    )
    row = cursor.fetchone()
    if not row:
        return False

    decision = clean_text(row.ReviewDecision) or 'Pending Review'
    match_status = clean_text(row.MatchStatus) or ''
    correct_po = clean_text(row.CorrectPONumber)
    matched_po = clean_text(row.MatchedPONumber)
    po_to_post = correct_po or matched_po

    should_post = False
    if decision == 'Matched to PO' and po_to_post:
        # Manual review decisions are allowed to post, including reimbursables if Accounting deliberately matches them.
        should_post = True
    elif match_status == 'Auto Matched' and po_to_post and expense_type_auto_posts(row.TxType):
        # Safe automatic posting: a non-labor Purchase/Disbursement found an exact PO in PM comments.
        should_post = True

    if should_post:
        cursor.execute(
            """
            UPDATE dbo.ExpenseReviewItems
            SET PostedToPO = 1,
                PostedPONumber = ?,
                PostedAmount = COALESCE(Amount, 0),
                PostedAt = COALESCE(PostedAt, SYSUTCDATETIME()),
                PostedBy = COALESCE(PostedBy, ?),
                UpdatedAt = SYSUTCDATETIME()
            WHERE ExpenseReviewItemId = ?;
            """,
            po_to_post,
            posted_by,
            item_id,
        )
        return True

    cursor.execute(
        """
        UPDATE dbo.ExpenseReviewItems
        SET PostedToPO = 0,
            PostedPONumber = NULL,
            PostedAmount = NULL,
            PostedAt = NULL,
            PostedBy = NULL,
            UpdatedAt = SYSUTCDATETIME()
        WHERE ExpenseReviewItemId = ?;
        """,
        item_id,
    )
    return False


def load_po_match_candidates(cursor):
    cursor.execute(
        """
        SELECT
            po.PONumber,
            COALESCE(v.VendorName, '') AS VendorName,
            COALESCE(p.ProjectName, '') AS ProjectName,
            COALESCE(po.RemainingAmount, po.RevisedAmount, po.OriginalAmount, 0) AS RemainingAmount
        FROM dbo.PurchaseOrders po
        LEFT JOIN dbo.Vendors v ON po.VendorId = v.VendorId
        LEFT JOIN dbo.Projects p ON po.ProjectId = p.ProjectId
        ORDER BY po.PONumber;
        """
    )
    return cursor.fetchall()


def choose_expense_match(project_name, vendor_name, amount, pm_comments, po_candidates):
    comments_key = compact_match_text(pm_comments)
    project_key = compact_match_text(project_name)
    vendor_key = compact_match_text(vendor_name)
    amount_float = float(amount or 0)

    # First: exact PO reference in PM comments.
    for po in po_candidates:
        po_number = po.PONumber or ""
        if po_number and compact_match_text(po_number) and compact_match_text(po_number) in comments_key:
            return {
                "status": "Auto Matched",
                "confidence": "High",
                "extracted_po": po_number,
                "matched_po": po_number,
                "reason": "PO number was found directly in PM Comments.",
            }

    # Second: a project + vendor candidate. This should be reviewed before it affects balances.
    candidate_matches = []
    for po in po_candidates:
        po_project = compact_match_text(po.ProjectName)
        po_vendor = compact_match_text(po.VendorName)
        project_match = bool(project_key and (project_key in po_project or po_project in project_key))
        vendor_match = bool(vendor_key and (vendor_key in po_vendor or po_vendor in vendor_key))
        if project_match and vendor_match:
            candidate_matches.append(po)

    if len(candidate_matches) == 1:
        po = candidate_matches[0]
        remaining = float(po.RemainingAmount or 0)
        amount_note = " Remaining balance appears sufficient." if remaining >= amount_float else " Remaining balance may be insufficient."
        return {
            "status": "Needs Review",
            "confidence": "Medium",
            "extracted_po": "",
            "matched_po": po.PONumber,
            "reason": "One PO matched by project and vendor." + amount_note,
        }

    if len(candidate_matches) > 1:
        return {
            "status": "Needs Review",
            "confidence": "Low",
            "extracted_po": "",
            "matched_po": "",
            "reason": f"{len(candidate_matches)} possible POs matched by project/vendor. Select the correct PO manually.",
        }

    # Fallback: try to extract a PO-looking token from comments so the reviewer can see it.
    extracted = ""
    token_match = re.search(r"\b(?:PO[-\s#:]*)?([A-Z]{0,3}-?\d{2,3}-\d{3}(?:-\d{3})?)\b", str(pm_comments or ""), re.IGNORECASE)
    if token_match:
        extracted = token_match.group(1).upper()

    return {
        "status": "No Match",
        "confidence": "None",
        "extracted_po": extracted,
        "matched_po": "",
        "reason": "No exact PO comment match or unique project/vendor PO match was found.",
    }


def import_expense_rows(rows, filename):
    ensure_expense_review_tables()
    conn = get_sql_connection()
    cursor = conn.cursor()
    try:
        po_candidates = load_po_match_candidates(cursor)
        cursor.execute(
            """
            INSERT INTO dbo.ExpenseUploadBatches (FileName, UploadedBy, TotalRows, ImportStatus)
            OUTPUT INSERTED.ExpenseBatchId
            VALUES (?, ?, ?, 'Completed')
            """,
            filename,
            get_current_user()["email"] or "Manual Upload",
            len(rows),
        )
        batch_id = cursor.fetchone().ExpenseBatchId
        counts = {"Auto Matched": 0, "Needs Review": 0, "No Match": 0}
        duplicate_count = 0
        posted_count = 0
        posted_amount = Decimal("0")

        for index, row in enumerate(rows, start=2):
            expense_id = clean_text(row_value(row, ["ExpenseId", "Expense ID", "Transaction ID", "Tx ID", "Track No."]))
            project_name = clean_text(row_value(row, ["ProjectName", "Project Name", "Project", "Project Short Name", "Project Code"]))
            tx_date = clean_date(row_value(row, ["TxDate", "Tx Date", "Transaction Date", "Date"]));
            tx_type = clean_text(row_value(row, ["TxType", "Tx Type", "Transaction Type", "Type"]));
            # Rollout rule: timesheet/labor expense rows are ignored during upload and do not enter PO matching or reduce balances.
            tx_type_lc = str(tx_type or "").strip().lower()
            if "time sheet" in tx_type_lc or "timesheet" in tx_type_lc:
                continue
            vendor_name = clean_text(row_value(row, ["VendorName", "Vendor Name", "Vendor", "Vendor/Purchaser", "Charger Name", "Purchaser"]));
            description = clean_text(row_value(row, ["Description", "Charge Code", "Memo", "Notes"]));
            amount = clean_decimal(row_value(row, ["Amount", "Transaction Amount", "Cost", "Total"]));
            pm_comments = clean_text(row_value(row, ["PMComments", "PM Comments", "Project Manager Comments", "Comments"]));

            unique_key = expense_unique_key(project_name, tx_date, tx_type, vendor_name, amount, expense_id, pm_comments)
            cursor.execute(
                """
                SELECT TOP 1 ExpenseReviewItemId
                FROM dbo.ExpenseReviewItems
                WHERE ExpenseUniqueKey = ? AND COALESCE(IsDuplicate, 0) = 0;
                """,
                unique_key,
            )
            existing = cursor.fetchone()
            if existing:
                duplicate_count += 1
                continue

            match = choose_expense_match(project_name, vendor_name, amount, pm_comments, po_candidates)
            counts[match["status"]] = counts.get(match["status"], 0) + 1
            review_decision = "Matched to PO" if match["status"] == "Auto Matched" else "Pending Review"

            cursor.execute(
                """
                INSERT INTO dbo.ExpenseReviewItems
                    (ExpenseBatchId, SourceRowNumber, ExpenseId, ExpenseUniqueKey, ProjectName, TxDate, TxType, VendorName,
                     Description, Amount, PMComments, ExtractedPONumber, MatchedPONumber, MatchStatus,
                     MatchConfidence, MatchReason, ReviewDecision, CorrectPONumber)
                OUTPUT INSERTED.ExpenseReviewItemId
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                batch_id,
                index,
                expense_id,
                unique_key,
                project_name,
                tx_date,
                tx_type,
                vendor_name,
                description,
                amount,
                pm_comments,
                match["extracted_po"],
                match["matched_po"],
                match["status"],
                match["confidence"],
                match["reason"],
                review_decision,
                match["matched_po"] if match["status"] == "Auto Matched" else None,
            )
            new_item_id = cursor.fetchone().ExpenseReviewItemId
            if sync_expense_posting(cursor, new_item_id, posted_by=get_current_user()["email"] or "Manual Upload"):
                posted_count += 1
                posted_amount += Decimal(str(amount or 0))

        cursor.execute(
            """
            UPDATE dbo.ExpenseUploadBatches
            SET AutoMatchedCount = ?, NeedsReviewCount = ?, NoMatchCount = ?,
                DuplicateCount = ?, PostedCount = ?, PostedAmount = ?
            WHERE ExpenseBatchId = ?
            """,
            counts.get("Auto Matched", 0),
            counts.get("Needs Review", 0),
            counts.get("No Match", 0),
            duplicate_count,
            posted_count,
            posted_amount,
            batch_id,
        )
        conn.commit()
        return {"batch_id": batch_id, "total_rows": len(rows), "auto": counts.get("Auto Matched", 0), "review": counts.get("Needs Review", 0), "no_match": counts.get("No Match", 0), "duplicates": duplicate_count, "posted": posted_count, "posted_amount": posted_amount}
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def load_expense_upload_page_data(status_filter="All"):
    ensure_expense_review_tables()
    conn = get_sql_connection()
    cursor = conn.cursor()
    cursor.execute(
        """
        SELECT
            COUNT(*) AS TotalRows,
            SUM(CASE WHEN MatchStatus = 'Auto Matched' THEN 1 ELSE 0 END) AS AutoMatched,
            SUM(CASE WHEN MatchStatus = 'Manually Matched' THEN 1 ELSE 0 END) AS ManuallyMatched,
            SUM(CASE WHEN MatchStatus = 'Needs Review' THEN 1 ELSE 0 END) AS NeedsReview,
            SUM(CASE WHEN MatchStatus = 'No Match' THEN 1 ELSE 0 END) AS NoMatch,
            SUM(CASE WHEN ReviewDecision IS NOT NULL AND ReviewDecision <> 'Pending Review' THEN 1 ELSE 0 END) AS ReviewedRows,
            SUM(CASE WHEN COALESCE(PostedToPO, 0) = 1 THEN 1 ELSE 0 END) AS PostedRows,
            SUM(CASE WHEN COALESCE(PostedToPO, 0) = 1 THEN COALESCE(PostedAmount, Amount, 0) ELSE 0 END) AS PostedAmount
        FROM dbo.ExpenseReviewItems;
        """
    )
    stats = cursor.fetchone()

    where = ""
    params = []
    if status_filter and status_filter != "All":
        where = "WHERE MatchStatus = ?"
        params.append(status_filter)

    cursor.execute(
        f"""
        SELECT TOP 250
            ExpenseReviewItemId, ExpenseBatchId, ExpenseId, ProjectName, TxDate, TxType, VendorName,
            Description, Amount, PMComments, ExtractedPONumber, MatchedPONumber, MatchStatus,
            MatchConfidence, MatchReason, ReviewDecision, CorrectPONumber, ReviewerNotes, PostedToPO, PostedPONumber, PostedAmount, CreatedAt, UpdatedAt
        FROM dbo.ExpenseReviewItems
        {where}
        ORDER BY CreatedAt DESC, ExpenseReviewItemId DESC;
        """,
        *params,
    )
    rows = cursor.fetchall()

    cursor.execute(
        """
        SELECT TOP 6 ExpenseBatchId, FileName, UploadedAt, UploadedBy, TotalRows,
               AutoMatchedCount, NeedsReviewCount, NoMatchCount, DuplicateCount, PostedCount, PostedAmount, ImportStatus
        FROM dbo.ExpenseUploadBatches
        ORDER BY UploadedAt DESC;
        """
    )
    batches = cursor.fetchall()

    cursor.execute("SELECT PONumber FROM dbo.PurchaseOrders ORDER BY PONumber;")
    po_numbers = [r.PONumber for r in cursor.fetchall() if r.PONumber]
    conn.close()
    return stats, rows, batches, po_numbers


# ------------------------------------------------------------
# Data loaders
# ------------------------------------------------------------

def load_summary_data():
    conn = get_sql_connection()
    cursor = conn.cursor()
    req_where, req_params = requestor_filter_sql("l")

    cursor.execute(
        f"""
        WITH UniquePOs AS (
            SELECT
                PONumber,
                MAX(VendorName) AS VendorName,
                MAX(ProjectName) AS ProjectName,
                MAX(POStatus) AS POStatus,
                MAX(COALESCE(RevisedAmount, OriginalAmount, 0)) AS POValue,
                MAX(COALESCE(RemainingAmount, 0)) AS RemainingAmount
            FROM dbo.IssuedPOLines l
            WHERE {req_where}
            GROUP BY PONumber
        ),
        LineTotals AS (
            SELECT SUM(COALESCE(LineAmount, 0)) AS TotalLineAmount
            FROM dbo.IssuedPOLines l
            WHERE {req_where}
        )
        SELECT
            COUNT(*) AS TotalPOs,
            SUM(CASE WHEN UPPER(COALESCE(POStatus, '')) = 'OPEN' THEN 1 ELSE 0 END) AS OpenPOs,
            SUM(POValue) AS TotalPOValue,
            (SELECT TotalLineAmount FROM LineTotals) AS TotalLineAmount,
            SUM(RemainingAmount) AS TotalRemainingAmount
        FROM UniquePOs;
        """,
        *req_params,
        *req_params,
    )
    row = cursor.fetchone()

    overall = {
        "total_pos": row.TotalPOs or 0,
        "open_pos": row.OpenPOs or 0,
        "total_po_value": float(row.TotalPOValue or 0),
        "total_line_amount": float(row.TotalLineAmount or 0),
        "total_remaining_amount": float(row.TotalRemainingAmount or 0),
    }

    cursor.execute(
        f"""
        WITH UniquePOs AS (
            SELECT
                PONumber,
                MAX(VendorName) AS VendorName,
                MAX(COALESCE(RevisedAmount, OriginalAmount, 0)) AS POValue,
                MAX(COALESCE(RemainingAmount, 0)) AS RemainingAmount
            FROM dbo.IssuedPOLines l
            WHERE {req_where}
            GROUP BY PONumber
        ),
        VendorLines AS (
            SELECT VendorName, SUM(COALESCE(LineAmount, 0)) AS TotalLineAmount
            FROM dbo.IssuedPOLines l
            WHERE {req_where}
            GROUP BY VendorName
        )
        SELECT
            u.VendorName,
            COUNT(*) AS POCount,
            SUM(u.POValue) AS TotalPOValue,
            COALESCE(MAX(v.TotalLineAmount), 0) AS TotalLineAmount,
            SUM(u.RemainingAmount) AS TotalRemainingAmount
        FROM UniquePOs u
        LEFT JOIN VendorLines v ON u.VendorName = v.VendorName
        GROUP BY u.VendorName
        ORDER BY TotalPOValue DESC;
        """,
        *req_params,
        *req_params,
    )
    vendors = cursor.fetchall()

    cursor.execute(
        f"""
        WITH UniquePOs AS (
            SELECT
                PONumber,
                MAX(ProjectName) AS ProjectName,
                MAX(COALESCE(RevisedAmount, OriginalAmount, 0)) AS POValue,
                MAX(COALESCE(RemainingAmount, 0)) AS RemainingAmount
            FROM dbo.IssuedPOLines l
            WHERE {req_where}
            GROUP BY PONumber
        ),
        ProjectLines AS (
            SELECT ProjectName, SUM(COALESCE(LineAmount, 0)) AS TotalLineAmount
            FROM dbo.IssuedPOLines l
            WHERE {req_where}
            GROUP BY ProjectName
        )
        SELECT
            u.ProjectName,
            COUNT(*) AS POCount,
            SUM(u.POValue) AS TotalPOValue,
            COALESCE(MAX(p.TotalLineAmount), 0) AS TotalLineAmount,
            SUM(u.RemainingAmount) AS TotalRemainingAmount
        FROM UniquePOs u
        LEFT JOIN ProjectLines p ON u.ProjectName = p.ProjectName
        GROUP BY u.ProjectName
        ORDER BY TotalPOValue DESC;
        """,
        *req_params,
        *req_params,
    )
    projects = cursor.fetchall()

    cursor.execute(
        """
        SELECT TOP 10
            ImportBatchId,
            FileName,
            UploadedAt,
            TotalRows,
            SuccessCount,
            ErrorCount,
            ImportStatus
        FROM dbo.ImportBatches
        ORDER BY UploadedAt DESC;
        """
    )
    imports = cursor.fetchall()

    conn.close()
    return overall, vendors, projects, imports

def load_personal_dashboard_data():
    conn = get_sql_connection()
    cursor = conn.cursor()

    req_where, req_params = requestor_filter_sql("l")

    cursor.execute(
        f"""
        WITH UniquePOs AS (
            SELECT
                PONumber,
                MAX(VendorName) AS VendorName,
                MAX(ProjectName) AS ProjectName,
                MAX(Department) AS Department,
                MAX(POStatus) AS POStatus,
                MAX(COALESCE(RevisedAmount, OriginalAmount, 0)) AS POValue,
                SUM(COALESCE(LineAmount, 0)) AS TotalLineAmount,
                MAX(COALESCE(RemainingAmount, 0)) AS RemainingAmount
            FROM dbo.IssuedPOLines l
            WHERE {req_where}
            GROUP BY PONumber
        )
        SELECT
            COUNT(*) AS TotalPOs,
            SUM(CASE WHEN UPPER(COALESCE(POStatus, '')) = 'OPEN' THEN 1 ELSE 0 END) AS OpenPOs,
            SUM(CASE WHEN UPPER(COALESCE(POStatus, '')) IN ('CLOSED', 'COMPLETE', 'COMPLETED') THEN 1 ELSE 0 END) AS ClosedPOs,
            SUM(POValue) AS TotalPOValue,
            SUM(TotalLineAmount) AS TotalLineAmount,
            SUM(RemainingAmount) AS TotalRemainingAmount,
            SUM(CASE WHEN ABS(COALESCE(POValue, 0) - COALESCE(TotalLineAmount, 0)) > 0.01 THEN 1 ELSE 0 END) AS AmountMismatchCount
        FROM UniquePOs;
        """,
        *req_params,
    )
    row = cursor.fetchone()

    overall = {
        "total_pos": row.TotalPOs or 0,
        "open_pos": row.OpenPOs or 0,
        "closed_pos": row.ClosedPOs or 0,
        "total_po_value": row.TotalPOValue or 0,
        "total_line_amount": row.TotalLineAmount or 0,
        "total_remaining_amount": row.TotalRemainingAmount or 0,
        "amount_mismatch_count": row.AmountMismatchCount or 0,
    }

    cursor.execute(
        f"""
        SELECT TOP 5
            VendorName,
            COUNT(DISTINCT PONumber) AS POCount,
            SUM(COALESCE(LineAmount, 0)) AS TotalLineAmount
        FROM dbo.IssuedPOLines l
        WHERE {req_where}
        GROUP BY VendorName
        ORDER BY TotalLineAmount DESC;
        """,
        *req_params,
    )
    top_vendors = cursor.fetchall()

    cursor.execute(
        f"""
        SELECT TOP 5
            ProjectName,
            COUNT(DISTINCT PONumber) AS POCount,
            SUM(COALESCE(LineAmount, 0)) AS TotalLineAmount
        FROM dbo.IssuedPOLines l
        WHERE {req_where}
        GROUP BY ProjectName
        ORDER BY TotalLineAmount DESC;
        """,
        *req_params,
    )
    top_projects = cursor.fetchall()

    cursor.execute(
        """
        SELECT TOP 5
            ImportBatchId,
            FileName,
            UploadedAt,
            TotalRows,
            SuccessCount,
            ErrorCount,
            ImportStatus
        FROM dbo.ImportBatches
        ORDER BY UploadedAt DESC;
        """
    )
    recent_imports = cursor.fetchall()

    cursor.execute(
        """
        SELECT COUNT(*) AS ActiveUserCount
        FROM dbo.DashboardUsers
        WHERE IsActive = 1;
        """
    )
    user_row = cursor.fetchone()
    active_user_count = user_row.ActiveUserCount or 0

    cursor.execute(
        """
        SELECT COUNT(*) AS ErrorCount
        FROM dbo.ImportErrors;
        """
    )
    error_row = cursor.fetchone()
    import_error_count = error_row.ErrorCount or 0

    conn.close()

    return {
        "overall": overall,
        "top_vendors": top_vendors,
        "top_projects": top_projects,
        "recent_imports": recent_imports,
        "active_user_count": active_user_count,
        "import_error_count": import_error_count,
    }


# ------------------------------------------------------------
# Purchase request helpers
# ------------------------------------------------------------

def purchase_request_status_badge(status):
    status = status or "Submitted"
    status_lower = status.lower()
    badge_class = "blue"

    if status_lower in ["submitted", "under review", "needs more info", "pending admin approval", "pending executive approval"]:
        badge_class = "amber"
    elif status_lower in ["approved", "converted to po", "auto approved"]:
        badge_class = "green"
    elif status_lower in ["rejected", "cancelled", "canceled"]:
        badge_class = "red"

    return f'<span class="badge {badge_class}">{h(status)}</span>'


def can_review_purchase_requests(role):
    return role_can_review_requests(role)


def generate_purchase_request_number(cursor):
    today_prefix = datetime.utcnow().strftime("PR-%Y%m%d")

    cursor.execute(
        """
        SELECT COUNT(*) AS RequestCount
        FROM dbo.PurchaseRequests
        WHERE RequestNumber LIKE ?;
        """,
        today_prefix + "-%",
    )

    row = cursor.fetchone()
    next_number = (row.RequestCount or 0) + 1

    return f"{today_prefix}-{next_number:04d}"


def purchase_request_visibility_sql(alias="pr"):
    role = normalize_role(current_data_role())
    email = current_data_email()
    dept_expr = f"COALESCE({alias}.Department, '')"
    req_email_expr = f"COALESCE({alias}.RequestedByEmail, '')"
    if role in ["Admin", "Executive", "Purchaser - All Departments"]:
        return "1=1", []
    if role == "Division Manager - Diving":
        return f"LOWER({dept_expr}) <> 'dredging'", []
    if role == "Project Manager - Dredging Only":
        return f"LOWER({dept_expr}) = 'dredging'", []
    if role == "Project Manager - Diving" and email:
        return f"LOWER({dept_expr}) <> 'dredging' AND LOWER({req_email_expr}) = LOWER(?)", [email]
    return "1=0", []


def load_purchase_request_stats():
    conn = get_sql_connection()
    cursor = conn.cursor()

    req_where, req_params = purchase_request_visibility_sql("pr")
    cursor.execute(
        f"""
        SELECT
            COUNT(*) AS TotalRequests,
            SUM(CASE WHEN RequestStatus = 'Submitted' THEN 1 ELSE 0 END) AS SubmittedRequests,
            SUM(CASE WHEN RequestStatus = 'Under Review' THEN 1 ELSE 0 END) AS UnderReviewRequests,
            SUM(CASE WHEN RequestStatus = 'Needs More Info' THEN 1 ELSE 0 END) AS NeedsMoreInfoRequests,
            SUM(CASE WHEN RequestStatus = 'Pending Admin Approval' THEN 1 ELSE 0 END) AS PendingAdminRequests,
            SUM(CASE WHEN RequestStatus = 'Pending Executive Approval' THEN 1 ELSE 0 END) AS PendingExecutiveRequests,
            SUM(CASE WHEN RequestStatus = 'Approved' THEN 1 ELSE 0 END) AS ApprovedRequests,
            SUM(CASE WHEN RequestStatus = 'Rejected' THEN 1 ELSE 0 END) AS RejectedRequests,
            SUM(CASE WHEN RequestStatus = 'Converted to PO' THEN 1 ELSE 0 END) AS ConvertedRequests,
            SUM(COALESCE(EstimatedAmount, 0)) AS TotalEstimatedAmount
        FROM dbo.PurchaseRequests pr
        WHERE {req_where};
        """,
        *req_params,
    )

    row = cursor.fetchone()
    conn.close()

    return {
        "total_requests": row.TotalRequests or 0,
        "submitted_requests": row.SubmittedRequests or 0,
        "under_review_requests": row.UnderReviewRequests or 0,
        "needs_more_info_requests": row.NeedsMoreInfoRequests or 0,
        "pending_admin_requests": row.PendingAdminRequests or 0,
        "pending_executive_requests": row.PendingExecutiveRequests or 0,
        "approved_requests": row.ApprovedRequests or 0,
        "rejected_requests": row.RejectedRequests or 0,
        "converted_requests": row.ConvertedRequests or 0,
        "total_estimated_amount": row.TotalEstimatedAmount or 0,
    }


def load_purchase_requests(limit=100):
    conn = get_sql_connection()
    cursor = conn.cursor()
    ensure_purchase_request_project_code_column(cursor)
    ensure_purchase_request_approval_columns(cursor)
    conn.commit()

    req_where, req_params = purchase_request_visibility_sql("pr")
    cursor.execute(
        f"""
        SELECT TOP {int(limit)}
            PurchaseRequestId,
            RequestNumber,
            RequestedByEmail,
            RequestedByName,
            RequestedAt,
            NeededByDate,
            VendorName,
            ProjectCode,
            ProjectName,
            Department,
            RequestTitle,
            RequestDescription,
            EstimatedAmount,
            Priority,
            RequestStatus,
            ReviewerEmail,
            ReviewedAt,
            ReviewNotes,
            ConvertedPONumber,
            UpdatedAt,
            AdminApprovedBy,
            AdminApprovedAt,
            ExecutiveApprovedBy,
            ExecutiveApprovedAt,
            ApprovalRequirement
        FROM dbo.PurchaseRequests pr
        WHERE {req_where}
        ORDER BY RequestedAt DESC;
        """,
        *req_params,
    )

    rows = cursor.fetchall()
    conn.close()

    return rows


def create_purchase_request(form, files=None):
    user = get_current_user()
    access = get_user_access()

    request_title = clean_text(form.get("request_title"))
    vendor_name = clean_text(form.get("vendor_name"))
    selected_project_value = clean_text(form.get("project_value") or form.get("project_name"))
    project_name = ""
    project_code = ""
    department = clean_text(form.get("department"))
    needed_by_date = clean_date(form.get("needed_by_date"))
    estimated_amount = clean_decimal(form.get("estimated_amount"))
    priority = clean_text(form.get("priority"))
    request_description = clean_text(form.get("request_description"))

    estimated_purchase_date = clean_text(form.get("estimated_purchase_date"))
    requested_by = clean_text(form.get("requested_by"))
    business_justification = clean_text(form.get("business_justification"))
    payment_type = clean_text(form.get("payment_type"))
    selected_issued_items = clean_text(form.get("selected_issued_items"))
    other_items = clean_text(form.get("other_items"))
    quote_backup = clean_text(form.get("quote_backup"))

    if not request_title:
        raise ValueError("Request Title is required.")

    detail_parts = []
    if request_description:
        detail_parts.append("Description: " + request_description)
    if business_justification:
        detail_parts.append("Business Justification: " + business_justification)
    if estimated_purchase_date:
        detail_parts.append("Estimated Purchase Date: " + estimated_purchase_date)
    if payment_type:
        detail_parts.append("Payment Type: " + payment_type)
    if selected_issued_items:
        detail_parts.append("Selected Issued PO Items: " + selected_issued_items)
    if other_items:
        detail_parts.append("Other Items: " + other_items)
    if quote_backup:
        detail_parts.append("Quote / Backup: " + quote_backup)

    request_description = "\n\n".join(detail_parts) if detail_parts else request_description

    requested_by_name = requested_by or access["display_name"] or user["email"]

    conn = get_sql_connection()
    cursor = conn.cursor()

    try:
        ensure_purchase_request_project_code_column(cursor)
        ensure_purchase_request_approval_columns(cursor)
        project_row = validate_selected_project(cursor, selected_project_value)
        project_code = clean_text(getattr(project_row, "ProjectCode", ""))
        project_name = clean_text(getattr(project_row, "ProjectName", ""))

        request_number = generate_purchase_request_number(cursor)

        cursor.execute(
            """
            INSERT INTO dbo.PurchaseRequests
                (
                    RequestNumber,
                    RequestedByEmail,
                    RequestedByName,
                    NeededByDate,
                    VendorName,
                    ProjectCode,
                    ProjectName,
                    Department,
                    RequestTitle,
                    RequestDescription,
                    EstimatedAmount,
                    Priority,
                    RequestStatus,
                    ApprovalRequirement
                )
            OUTPUT INSERTED.PurchaseRequestId
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'Submitted', ?);
            """,
            request_number,
            user["email"],
            requested_by_name,
            needed_by_date,
            vendor_name,
            project_code,
            project_name,
            department,
            request_title,
            request_description,
            estimated_amount,
            priority,
            approval_requirement_label(estimated_amount),
        )

        purchase_request_id = cursor.fetchone().PurchaseRequestId
        saved_attachments = save_purchase_request_attachments(cursor, purchase_request_id, request_number, files or [])
        conn.commit()

        try:
            send_purchase_request_submitted_emails(purchase_request_id)
        except Exception as email_error:
            print(f"Purchase request email notification failed: {email_error}")

        return {
            "purchase_request_id": purchase_request_id,
            "request_number": request_number,
            "saved_attachments": saved_attachments,
        }

    except Exception:
        conn.rollback()
        raise

    finally:
        conn.close()


def generate_app_po_number(cursor):
    today_prefix = datetime.utcnow().strftime("APP-PO-%Y%m%d")
    cursor.execute(
        """
        SELECT COUNT(*) AS POCount
        FROM dbo.PurchaseOrders
        WHERE PONumber LIKE ?;
        """,
        today_prefix + "-%",
    )
    row = cursor.fetchone()
    next_number = (row.POCount or 0) + 1
    while True:
        po_number = f"{today_prefix}-{next_number:04d}"
        cursor.execute("SELECT COUNT(*) AS ExistingCount FROM dbo.PurchaseOrders WHERE PONumber = ?", po_number)
        existing = cursor.fetchone()
        if not (existing.ExistingCount or 0):
            return po_number
        next_number += 1


def create_or_update_po_from_purchase_request(cursor, purchase_request_id, requested_po_number=None):
    cursor.execute(
        """
        SELECT
            PurchaseRequestId,
            RequestNumber,
            RequestedByEmail,
            RequestedByName,
            NeededByDate,
            VendorName,
            ProjectCode,
            ProjectName,
            Department,
            RequestTitle,
            RequestDescription,
            EstimatedAmount,
            ConvertedPONumber
        FROM dbo.PurchaseRequests
        WHERE PurchaseRequestId = ?;
        """,
        purchase_request_id,
    )
    req = cursor.fetchone()
    if not req:
        raise ValueError("Purchase request was not found.")

    po_number = clean_text(requested_po_number) or clean_text(req.ConvertedPONumber) or generate_app_po_number(cursor)
    vendor_name = clean_text(req.VendorName) or "Vendor TBD"
    project_code = clean_text(getattr(req, "ProjectCode", ""))
    project_name = clean_text(req.ProjectName) or "Project TBD"
    department = clean_text(req.Department)
    requestor = clean_text(req.RequestedByName) or clean_text(req.RequestedByEmail)
    amount = clean_decimal(req.EstimatedAmount)
    today = datetime.utcnow().date()

    vendor_id = get_or_create_vendor(cursor, vendor_name)
    project_row = validate_selected_project(cursor, f"{project_code}||{project_name}")
    project_id = project_row.ProjectId
    project_code = clean_text(getattr(project_row, "ProjectCode", ""))
    project_name = clean_text(getattr(project_row, "ProjectName", "")) or project_name

    cursor.execute(
        """
        INSERT INTO dbo.ImportBatches
            (
                FileName,
                SourceSystem,
                UploadedBy,
                TotalRows,
                SuccessCount,
                ErrorCount,
                ImportStatus
            )
        OUTPUT INSERTED.ImportBatchId
        VALUES (?, 'Purchase Request Auto PO', ?, 1, 1, 0, 'Complete');
        """,
        f"Auto-created from {clean_text(req.RequestNumber) or 'purchase request'}",
        get_current_user()["email"] or "Purchase Request Approval",
    )
    import_batch_id = cursor.fetchone().ImportBatchId

    purchase_order_id = upsert_purchase_order(
        cursor=cursor,
        po_number=po_number,
        vendor_id=vendor_id,
        project_id=project_id,
        department=department,
        requestor=requestor,
        po_date=today,
        po_status="Open",
        original_amount=amount,
        revised_amount=amount,
        remaining_amount=amount,
        import_batch_id=import_batch_id,
        project_code=project_code,
    )

    cursor.execute(
        """
        SELECT COUNT(*) AS LineCount
        FROM dbo.IssuedPOLines
        WHERE PONumber = ?;
        """,
        po_number,
    )
    line_count = cursor.fetchone().LineCount or 0
    if line_count == 0:
        cursor.execute(
            """
            INSERT INTO dbo.IssuedPOLines
                (
                    PurchaseOrderId,
                    ImportBatchId,
                    PONumber,
                    VendorName,
                    ProjectCode,
                    ProjectName,
                    Department,
                    PODate,
                    POStatus,
                    LineDescription,
                    Unit,
                    UnitCost,
                    Qty,
                    LineAmount,
                    OriginalAmount,
                    RevisedAmount,
                    RemainingAmount,
                    Requestor
                )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'Open', ?, 'LS', ?, 1, ?, ?, ?, ?, ?);
            """,
            purchase_order_id,
            import_batch_id,
            po_number,
            vendor_name,
            project_code,
            project_name,
            department,
            today,
            clean_text(req.RequestTitle) or clean_text(req.RequestDescription) or "Purchase request approved item",
            amount,
            amount,
            amount,
            amount,
            amount,
            requestor,
        )


    try:
        ensure_request_attachment_table(cursor)
        cursor.execute(
            """
            UPDATE dbo.PurchaseRequestAttachments
            SET PONumber = ?
            WHERE PurchaseRequestId = ?;
            """,
            po_number,
            purchase_request_id,
        )
    except Exception:
        pass

    return po_number


def update_purchase_request_status(form):
    user = get_current_user()

    purchase_request_id = clean_text(form.get("purchase_request_id"))
    requested_status = clean_text(form.get("request_status"))
    reviewer_email = clean_text(form.get("reviewer_email")) or user["email"]
    review_notes = clean_text(form.get("review_notes"))
    converted_po_number = clean_text(form.get("converted_po_number"))
    created_or_converted_po = ""
    pending_email_role = ""

    valid_statuses = [
        "Submitted",
        "Under Review",
        "Needs More Info",
        "Pending Admin Approval",
        "Pending Executive Approval",
        "Approved",
        "Rejected",
        "Converted to PO",
    ]

    if requested_status not in valid_statuses:
        raise ValueError("Invalid request status.")

    if not purchase_request_id:
        raise ValueError("Purchase Request ID is required.")

    conn = get_sql_connection()
    cursor = conn.cursor()

    try:
        ensure_purchase_request_approval_columns(cursor)
        conn.commit()
        cursor.execute(
            """
            SELECT TOP 1 Department, EstimatedAmount, AdminApprovedBy, ExecutiveApprovedBy, ConvertedPONumber
            FROM dbo.PurchaseRequests
            WHERE PurchaseRequestId = ?;
            """,
            purchase_request_id,
        )
        request_row = cursor.fetchone()
        if not request_row:
            raise ValueError("Purchase request was not found.")

        amount = clean_decimal(getattr(request_row, "EstimatedAmount", 0))
        needs_exec = requires_executive_approval(amount)
        reviewer_role = normalize_role(get_user_access().get("role"))

        if requested_status in ["Approved", "Converted to PO"]:
            if reviewer_role not in ["Admin", "Executive"]:
                raise ValueError("Only Admin and Executive users can approve purchase requests under the July 1 approval workflow.")

            admin_approved = bool(clean_text(getattr(request_row, "AdminApprovedBy", "")))
            executive_approved = bool(clean_text(getattr(request_row, "ExecutiveApprovedBy", "")))

            if reviewer_role == "Admin":
                admin_approved = True
                cursor.execute(
                    """
                    UPDATE dbo.PurchaseRequests
                    SET AdminApprovedBy = ?,
                        AdminApprovedAt = COALESCE(AdminApprovedAt, SYSUTCDATETIME()),
                        ApprovalRequirement = ?
                    WHERE PurchaseRequestId = ?;
                    """,
                    reviewer_email,
                    approval_requirement_label(amount),
                    purchase_request_id,
                )
            elif reviewer_role == "Executive":
                if not needs_exec:
                    raise ValueError("This request is below $3,000 and only requires Admin approval.")
                executive_approved = True
                cursor.execute(
                    """
                    UPDATE dbo.PurchaseRequests
                    SET ExecutiveApprovedBy = ?,
                        ExecutiveApprovedAt = COALESCE(ExecutiveApprovedAt, SYSUTCDATETIME()),
                        ApprovalRequirement = ?
                    WHERE PurchaseRequestId = ?;
                    """,
                    reviewer_email,
                    approval_requirement_label(amount),
                    purchase_request_id,
                )

            if needs_exec:
                if not admin_approved:
                    request_status = "Pending Admin Approval"
                    converted_po_number = ""
                    pending_email_role = "Admin"
                elif not executive_approved:
                    request_status = "Pending Executive Approval"
                    converted_po_number = ""
                    pending_email_role = "Executive"
                else:
                    converted_po_number = create_or_update_po_from_purchase_request(cursor, purchase_request_id, converted_po_number)
                    created_or_converted_po = converted_po_number
                    request_status = "Converted to PO"
            else:
                if reviewer_role != "Admin":
                    raise ValueError("Requests below $3,000 require Admin approval.")
                converted_po_number = create_or_update_po_from_purchase_request(cursor, purchase_request_id, converted_po_number)
                created_or_converted_po = converted_po_number
                request_status = "Converted to PO"

            approval_note = f"{reviewer_role} approval recorded by {reviewer_email}."
            if created_or_converted_po:
                approval_note += f" Fully approved and auto-created PO {created_or_converted_po}."
            elif pending_email_role:
                approval_note += f" Waiting for {pending_email_role} approval."
            if review_notes:
                if approval_note not in review_notes:
                    review_notes = review_notes + "\n" + approval_note
            else:
                review_notes = approval_note
        else:
            request_status = requested_status
            converted_po_number = ""

        cursor.execute(
            """
            UPDATE dbo.PurchaseRequests
            SET
                RequestStatus = ?,
                ReviewerEmail = ?,
                ReviewedAt = SYSUTCDATETIME(),
                ReviewNotes = ?,
                ConvertedPONumber = ?,
                ApprovalRequirement = ?,
                UpdatedAt = SYSUTCDATETIME()
            WHERE PurchaseRequestId = ?;
            """,
            request_status,
            reviewer_email,
            review_notes,
            converted_po_number,
            approval_requirement_label(amount),
            purchase_request_id,
        )

        conn.commit()

    except Exception:
        conn.rollback()
        raise

    finally:
        conn.close()

    if pending_email_role:
        try:
            send_purchase_request_pending_approval_email(purchase_request_id, pending_email_role)
        except Exception as email_error:
            print(f"Purchase request pending approval email failed: {email_error}")

    if created_or_converted_po:
        try:
            send_purchase_request_approved_email(purchase_request_id, created_or_converted_po)
        except Exception as email_error:
            print(f"Purchase request approved email failed: {email_error}")


# ------------------------------------------------------------
# Branding / layout
# ------------------------------------------------------------


CE_LOGO_DATA_URI = "data:image/jpeg;base64,/9j/4AAQSkZJRgABAQEAYABgAAD/2wBDAAMCAgMCAgMDAwMEAwMEBQgFBQQEBQoHBwYIDAoMDAsKCwsNDhIQDQ4RDgsLEBYQERMUFRUVDA8XGBYUGBIUFRT/2wBDAQMEBAUEBQkFBQkUDQsNFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBT/wAARCACVAjQDASIAAhEBAxEB/8QAHwAAAQUBAQEBAQEAAAAAAAAAAAECAwQFBgcICQoL/8QAtRAAAgEDAwIEAwUFBAQAAAF9AQIDAAQRBRIhMUEGE1FhByJxFDKBkaEII0KxwRVS0fAkM2JyggkKFhcYGRolJicoKSo0NTY3ODk6Q0RFRkdISUpTVFVWV1hZWmNkZWZnaGlqc3R1dnd4eXqDhIWGh4iJipKTlJWWl5iZmqKjpKWmp6ipqrKztLW2t7i5usLDxMXGx8jJytLT1NXW19jZ2uHi4+Tl5ufo6erx8vP09fb3+Pn6/8QAHwEAAwEBAQEBAQEBAQAAAAAAAAECAwQFBgcICQoL/8QAtREAAgECBAQDBAcFBAQAAQJ3AAECAxEEBSExBhJBUQdhcRMiMoEIFEKRobHBCSMzUvAVYnLRChYkNOEl8RcYGRomJygpKjU2Nzg5OkNERUZHSElKU1RVVldYWVpjZGVmZ2hpanN0dXZ3eHl6goOEhYaHiImKkpOUlZaXmJmaoqOkpaanqKmqsrO0tba3uLm6wsPExcbHyMnK0tPU1dbX2Nna4uPk5ebn6Onq8vP09fb3+Pn6/9oADAMBAAIRAxEAPwD9UT0460Dkc9aOB3pMA96ADGfelGB9aTrS8fWgAIBoHWkpScDFAAcUgzSgUpz2oAbk55p2M0mD3qjrOt2Ph7S7vUtSu4bCws4mmuLq4cLHEgGSzE9BQBeIAoAz7ivDf+G2/gc2CPifoRB5Hzv/APE0D9tn4H/9FO0L/vt//ia29jU/lf3Ec8e57lmgkCvDD+218D/+inaH/wB9v/8AE01v23fgaoyfifoX/fT/APxNHsan8r+5i9pHue6jjnNDc14WP23fgaR/yU/Qv++3/wDiaQ/tv/A0cf8ACz9C/wC+3/8AiaPY1P5X9w+ePc91z6UnHrzXhY/bf+BpOB8T9C/77f8A+JqT/htf4IYz/wALN0P/AL6f/wCJo9jU/lf3Bzx7nuOMijOOleGP+238Dl6/E/Qx/wACf/4mmJ+3B8DXdVX4maIzMwVQvmEkk4AHydTR7Gp/K/uDnj3PdRyead7VBaXUV7bQzwt5kcqCRGwRlSMg4PtUx6ViWJzzRkijOelHXvzQAucDmkz70Yx1owTQAZxQM9aOnvS/eoAT8aUYpM4ox27UAH48UAZHHWjr9KM8UAGCKCMmjr04oA70ALg0mB+NLkk0hGDQAgGTQRinDjtSE5oAVenNB44pMHvQCBQAYxS596Q5P0oxj3oAU8ikzjjrR9KKAAgjpS84pBxSbs0ALnAzSmk6Dmjg+1AAOvtS9+tIfrR19qADr3pRzSZxwaM/lQAHGaGOBxQetLtHWgBAaT9DS4A6UfhQAnJPNKSegoK0oP4UAIPcUvNIcg0A0ABzmlxntSds0HpkUAH40HkZHWk60u2gA2570ZxRn0oBz0oADQMg0fr9KUsMUAITntQFIoxnpS49aADBopPwooADzQPSlGBxRigBMcdaAcUHP4Uo5HSgBM80Z55pSMUAUAHbnpSZp35VFPKI0LfKAoyxY4AHqTQBDf38Gn2009xMlvbxIZJJpGCrGoGSxJ6ACvyE/bs/bduPjrq914I8H3Tw/Duym23F0hKtrMqnqf8ApiD0H8WM12f7fv7ax+JV1qHw08BXrJ4VtpDFrGr274/tGQHmCNh/yyB+8f4ulfCDWuwcAADgADgCvfweEcLVZrXoeVicRe9ODJo1iljH7qIEcY2CniKLbzFH/wB8CqgzG2QfrUpuMJnvXsqTPKsOkWBASYo/YbBVGXbLJkxR4HQBBUpYyHnmnR253dKTlcaRH9nR1z5af98CoWt0P/LNP++BWtHb8VTuCqHC4JpMq5TFsg5McY/4CKtwhHTiNMj/AGBUIBJ5OaZ5jRvlefb1qk7ES1LMsEbjmOMAc5Kjivv7/gnH+xdH4jubT4seNdLjbS4W3eH9LuYh+/cf8vTqf4R/CPxryD9hb9ka6/aR8YJr+u20sPw60ecG6lYEf2lOpyLeM91H8ZH0r9mdO0+3060gtra3jtbe3RYoYYhtWNAMBQB0AFeRjsXy/u4b9T08Jh/ty+RZijCrnABx0FP6UZJ7UhNfPHrgR3FIPSnZyOlJigAwfwpM4NO/GjAFADeTSj3o6nk0ZH4UABFHagjPSlJ4xQA3p9KU46igYoIz0oAKDxzSgUdKAExk56Uo5oJzR9OlAC5xSY70mRRQAHkUdKAc8UpWgBNtKQTSEke1L170AN2kUo5pQecUHrQAmMd6UCg0A5oATpS9hSd/WlHvQAAAmjJ9KQ8dDSg460AHXrSZHelwDQR+NACcYoHHSjB+lL0FACHrxRQOO1LnPTrQAhbPagHPalx+NHI70AIc5zSk0g557UpGKAEH6UobP0pOKAPSgBTgU3k+1KTg9KPftQAuM0nfpQDRn2oADnNJ0NL35NHNAAcHpQBjrRnFA+lAC5HvRSZ/CigBRk0hzQDxxSkGgAIyBQAR16UAZpMkHFAC9elIc96XI9KbJJsHALZ4wKAI5JvLGM4PvX5k/t+ft3PrVxqXwu+HOolLGImDXtftXwZT/FawMO3ZmH0FdX/wUG/bhXQ4dR+F3w71LOryAw65rds2fsaEc28TD/loe5H3RX5lxxLGoVRhR0Gf1+te3g8I9KtReh5eJxFv3cHr1HQXjWwCxgeWOidgKtJcxXPGdj/3TVLZnNQyErxXuuV9zykn0L97AY1DdqzTeRZ2+cg9s19VfsP/ALG+qftI6zF4h8RCax+GthOPMkOVfVJFPMMR/uD+JvwFfrNb/An4dRwRRp4C8OeXGgRAdNhOABgDJWvNr42nSlypXO6lhJTV3ofz4G7gA4mT86vQahamIu1xGAvU56V/QL/wor4d/wDQheGv/BXD/wDE0x/gR8Osf8iF4bx6f2ZD/wDE1yvMYv7JusG11P5+ZNbt5zsjmRE9SeTTo40mXKMHHqK/fi++DXwz021mubnwR4Yt7WBGklml02FUjQDJYnHAAFfjP+078W9A+Lvxg1PUfCei6foXhOwzYaZDp9skAuI1Y5ncKBksenoMV24bFLES5VE5a2H9ir3PGnTyz6V61+yx+zXrX7UHxKTQbLzLLw9ZMs2tauB8ttDn/Vqe8j9AO3Wue+FXwk8RfHT4gab4P8K23n6leNl5mH7q0hH35pD2VR27nAr9vf2ffgF4f/Z2+G+n+EvDsQeOIebe30gHm3twfvyufc9B2FTi8R7BcsXqy8PR9q+aWx13gLwJovw58KaZ4b8P2EWnaJpsCwWtvEMAKO59SepPcmuiFKOFAFJ96vmW23dnt7C5x70mfQUHgYpcYpAIAaUj1owfwoyFNAAxwaYZkHBbFcH8a/jT4Z+BHw/1Hxb4quvs9jajbFAn+tupT92KMd2J/LrX5O+L/wDgpD8cdd8S6hqOl+ILfw7ps8pa10mGxhlFtH/CpdlJZvU+tdlDCVMQm47I5quIhRspdT9nfNjH8Yo8+P1r8SG/4KJ/tA9B43j/APBVb/8AxFNj/wCCh37QTH/keY8f9gu3/wDiK6P7Oq7XRj9cp+Z+3ayBvunNLtz9a/M39kP/AIKRa5c+MYvDHxh1GG8sNUmWOx8QJAkH2SU8COZUAGxj0bsetfphHIHUHIIPIIPBFcdahOhLlmjppVY1Y80Rw46CkaVEOCcGuZ+I3xH0H4W+D9T8UeJdQTS9F02Myzzyd/RVHdieAO5r8pvip/wU1+LHijxlfXvg3UIvCfhrOyy0+SyinmKDpJIzqfmbrgcCtMPhamJvybImrXhRtzH6/efH13Zo+0RnvX4lt/wUU/aALY/4TaL/AMFVv/8AEUH/AIKG/tA5H/FbxD/uF2//AMRXT/Z1W9rox+uU/M/bRZkbgHNPBP4V+Pfwo/4KafFfwl40sr7xpqMfi7wz9y806OziglCHrJGyqDuXrg8Gv1e8AfELQ/iZ4R0zxN4cv49T0fUohNb3EZ7HqrDswPBHY1y18LUw7XP1NqVeFZe6dNtzR7VGrFjTmk2CuQ6BThOScUwzIW4avmn9vT9pHUP2e/hDb3Hhy7jtfFus3aWmnPJGsgjA+aSTa3BwoPXua/O64/4KHftARIZE8awuUIfYdKt8MAclfudxx+Nd9HBVa8OeNrHLUxMKUuSW5+1f3gMdPWl6fSuB+DfxJg+LXwv8L+L7J1aDVrGO4bb/AAyYxIv4OGH4V3CljXE007M6t9STPPrTvwphyBwK4v4r/F3w58FvBGo+LfFWopp+j2CZdjy8rn7sca/xOx4ApJOTshN21Z2bTIvG7FNE8frivxt+Iv8AwU6+MfifxfqV/wCGdUg8KaBJJiy0r7FDO8UY4Bd2Ukuep7Aniuci/wCCiX7Qbt/yO0X/AIKrf/4ivTWXVn1RxPGU0ftt50fXcKPNVvunNfion/BQ79oLv41gH/cJt/8A4ivZP2Yf+ClnirTPGyaX8XdRi1fw9qUixx6vFbJC+nSE4BdUADRk9T1FOeW14RctGEMZSm7H6kgbvrQzrGPmOKrW97HdWsM0MqSxzKJI5IzlWUjII9iK+LP+CjX7R/xD+Aup+BovA+tppEWprcfag9pHPvKj5fvg4/CuCjSlWmqcd2dVSapxcnsfbYuIx/FQbhF/izX4qP8A8FFPj8OvjaHP/YKt/wD4iqkn/BRD9oBySPHEQ+ml2/8A8RXoPLay6r+vkcf12n2Z+2puEP8AFR5sZ43V+IJ/4KKftBK3/I8Rkf8AYLt//iKmf/go5+0BHaTMPG0W5UJB/su34P8A3xU/2fV7or65T8z9uxzyOlAOOtcr8ONeu/EXgPwxqV7IJru9023uJpAAu52QFjgdOa6mVcg4rzGrOx2imRR1OBTTNGx+8K+Nv+Cif7Uuv/ADwh4c0nwdqCad4u1y6LrO0Ky+Rax/fbawI5OFr4b07/gpD8dtJ1Gzur3xfHe2NvPHLc2x0y3XzoQwMi5C5GVz0rupYOpVh7RWsc88RCEuR7n7YAY6UA5rA8H+MrHxx4V0XxBpkglsNVtI7uBgc/K6g4+o6VuAnvxXC01udI/2xSO4j6nApkjlVOK+Iv8AgpD+0X8QvgPD4DPgbWk0g6pNPHdlraObeFGV++DitKVN1ZqEd2ROahFyZ9vLKjHAOaeOK/Mb9g39sH4t/GP9ouz8MeL/ABMmqaHJplxcNbCxhiJdQNp3KoPFfpsjblBPpVVqMqMuWRNOoqseaI8cUA89KQ80vasDUQ4Jox6Gjp0oHrQAE4HNKDQQTSbKAELHNFLmigAAwKUnPSkJyKco4oAToKAOPegrzQQR3oARvlGe9fKX/BRn44eJfgt8B1uPCtwun6nrd8ultqA5kto2U7mj9GxwD2r6sdvkNfC//BW2Df8AALw0R28QQ/8AoLV04ZKVaKfcyqtqDaPykZpHduWkZmLM7MWZ2J5Yk9ST3qSOJ+jKRVy1sG3sxGQDWkuIB+82qv8AtV9c5PqfN8qexgmMhsV9Efsefsdal+094rF7qKT6f8O9NlH9o6ioKteOP+XaE98/xMOg964n4EfDTR/jf8bPCXgi6vZrG11W4IuZ7UfP5ajLKuehPTPav3O8CeB9E+HPhTTPDfhywh0zQ9PhENvaxLgKB3J7knkk9Sa8zGYpUo8kN2ehhaHM+eWyJvC3hLSvB/h/T9E0XT4dM0iwhWC1tbddqRIBwAP6962VGwU4/LwBSHmvm3qeyKDn2oIyKTGOvFKtAHO+OvAukfETwtqPh3XrU32j6hH5V1beYyeamclSVIOD35rw9P8Agnp+z+oAHw10/A4H7+Yf+z19JfyoGAK0jUnDSLaJcU90eb/CL9nj4efAz+0D4I8K2fh6TUNv2mWAs7yBegLMSQBnoK9H2hRhRgUAEmlwfWplJyd5O40ktENIxTs59qTdg+tBANSMXGaQ5oBxS9aAE5+lIw+X1pRzSPwDzQB+fn/BYGTb8Nfh4/ORrb4/7918TfsbfAzTv2k/jlaeE9b1CfT9Ggs31C7FrxLcIhA8pW/hznlvSvt7/gr5GrfDD4ff9hxv/RdfPH/BLKBY/wBqhiOp0C4/9CWvoaEpQwTlF9zyqsYyxMU0fXZ/4JVfA12J+x+IgDyANXPH/jlc98RP+CXnwj0rwL4gudCbXNL1i1sZbm2vJ9Q89EdFLAMhUZBxg896+7F5FQ3Vul1FJFKiyRyKUdGGQwPBBHpXkLFVk78zO50KbVuU/nCS+geS2UyoWNxEpAPXEqg/yr+i3QGC6DpJH3RaRY/74Fc1/wAKV8CAgjwXoIAOR/xL4+uc+ldgkPlqqKNqqMBQOAPStsXivrNtLWM6FD2CavufI3/BVCZR+yffHH/MXs+P+B1+YXwB+GK/HL4z+GPBEuotpUGrXDCa7RN7LGg3MFH94jgHtX6df8FUISf2Tr0Y/wCYvZf+h18B/sHwiP8Aa8+HXc+fcf8Aos16ODlKOEk4+ZyYhRlXipH3wf8AglR8Dy+Bb+JCB/F/a+M/+OUlx/wSt+CZgkijh8RwyupVJ/7V3GNj0bGznHpX2WCN5NMcBpB9RXj/AFmtvzM7/Y0/5T+ff42/DsfBz4v+K/BCagdWTQ7oQLetH5ZlUjIJXnBA4r9RP+CVz7v2TdL/AOwpd/8AoVfnd+2ZF5v7WnxQPrqaf+gV+jH/AAS2txD+yfpCnvqV2f8Ax+vYxrlLDRlLrY4MOoxrSjE+v46JCrHYe9NYBBkcVwnxr+I1v8KPhX4p8XXUiouk2Es0ZY4zLghB+LEV8+k5OyPVbsrs/Kv/AIKO/FwfE39o+60azlMujeD7cadHg5U3LYaY/h8o/OvAdY8E6rofgTwj4vuo1/sfxM9zHZMM5DQNtcN9eorPhtNa8TXBkcPd67rt4ZGHVpLid+n/AI8B+Ffpv+1h+ytHZ/sNaHoekwrLqvgG2i1CMoOZSF/0kfVsk/hX1SrLDKnS/r+rnhun7eU5mB/wSl+L/wBu8NeJvhhdzgzaVINV01WPJt5DiRR7K+D/AMDr9CEIKjPWvwd/ZX+Kw+EPx/8ABXipZzHp/wBpWyvjnAa1nwh3ewJVv+A1+7dsBIgkDho2AZGB4IPQ14+Ppezq83RnoYSfNTt2LLEAGvhH/grUgb9njRSRnb4hhI/74NfdEgJyK+Hf+CsMYk/Z10jPbxBB/wCgmufCfx4+pvW/hyPzz/ZW+Cun/H/47eHvBmr30+n6ZdiW4uJLYfvJEiAJjB/h3ZxntX6hRf8ABM34CkkDwndjHf8AtKSvy7/Z9+MVx+z58WdG8cWmmRaw9issMtlK+zzI5AA21uzDAxnivuKT/gsNokeSPhZqoP8A2FYf/ia9jF0cTz3p3t6nm4epR5LT3PXNS/4JofA5tOvIrfw1e2tw0LiK4TUHZo3wcEA981+Z2o/sr/GCG61C0h+Gfie6hhuHiinFgwEqK+Fb8QM19nJ/wWE0pzkfCzVD9dVi/wDia9f/AGeP+Ci/gn45+M4/DGoaPeeCNWucDT/7Qu1lhvH7xq6gBX9AetY06mMw8XKUbrzNZxw9WSSdmfSXw8sLiz8EeGLe6ge2uIdMt45YpBhkYRgFSPUGuH+P/wCyt4F/aRl0eTxjbX87aUH+zGyuzBjfw2eDmvYUb5eRtpC3pXiqcoy5ouzPRcU1ZnxtL/wSr+B7yrm38SICe2r/AP2Fflr8avCOn+APjJ408L6QJhpek6k9pbCd98gjAGNzY5PNf0IY3OCT3r8Hf2n7NB+0z8ThwP8AieSf+grXtZfUqVZtSlfQ8zFwhCKsj6X/AGIP2GPhn+0J8C7fxd4ui1iTWX1G5tibK/8AJj2I2F+Xaea+gz/wSk+BjxMps/EW1hg/8Tft/wB8Vd/4JbQhP2UbIjn/AIm96f8Ax+vr9OUrir16sasoqXU6qVKDhFtGJ4c0C18K6FpukWQcWen26WsHmNuYIgwMnucVqySkGML1JxUzKNteJ/tc/GMfA34A+LvE6MF1BbY2mnqTy9zL8iY9xkt/wGuKKdSSS3Z0tqKuflb+298WD8Z/2kvFGowzibR9EY6PYbT8pSP/AFjj6tn8q8d8XeA9W8FW/hu51qz+zw+IdNXVtPBOfNtmYqGPpnB4qX4b+B9X+IXjDw/4TiVzqev38dqZDycyPmRz9ASa/Q//AIKefA/T7H4H+DfEGiWaxDwU8WlSbF5+xuoVR9FZc/8AAq+rdaFF06C2PBVOVRTqs7T/AIJZfFuPxd8CrrwddzebqXhC6MEYc/M1rJ80R/A7h+VfaqyK6givxM/4J/8Axa/4VJ+0rogupvL0bxIv9jXeT8oZjmFv++sD8a/a+JQGKZ5FeDjqXs6z89T1cNU9pTXkPlI8s8V+bv8AwV8bEXwtH/Tzcn/x0V+kcyYiNfnF/wAFeIPNh+FwHX7Tc/8AoIpYH/eIlYn+FI8J/wCCYIDftZWbdxot1/Sv2Uh5jX6V+O3/AATHhWH9q21C/wDQFuufyr9h4j+7X6VrmX8f5IxwX8IlzxSZJNGTilBryzvDbn2pDS8g9aUigBoyaOvelxg0hBzzQAuPaijOKKAExnnpRikIIpwoAAce9KRSZxQD2oAa/KnFfC//AAVtuhB8BPDWRnPiCHg/7pr7pkxsNfBH/BXncfgV4TUd9fj/APQDXVhP48PUyq602fl6dbkKsqsE5P3RVOYvOcsxb3Jqm6lHPXqakiudnB6V9ddPc+Zaa2PoD9gWAr+178PD282f/wBANfuPE3lqM1+IH7A7hv2vPh5jpvn/APQDX7gKhIGa+czGyqpLse5gr+y1JlOaMZpANtOryjvENBAAob2oHA60AJjBo74FAJpN4HUgGgB2e1JvFJuX+8Pzo3r6j86AFB9qU5ApNy/3h+dIXUfxCgBcUZpFYN3zT8CgBD7U1hhTTu/Skf7poA/P/wD4K+yqvwz+Hoz/AMxxv/RdfO3/AAS0uM/tVEH/AKAFz/6EtfQ3/BYCLPw0+HmOv9uP/wCi6+Tv+CenxC8O/Cz9pey1LxTqcej6beaZNYRXk/ESzMQVDt/CDjGT3r6CipSwLSXc8uo4rExuftlGdwGKfj8a8vT9o/4XJx/wsLw76f8AIQSuf+J37Ufw38O/D3xHq0fj/Rt0NhMIRZXiyTmYoRGEUclt2MV4apzbtY9LmVr3Pb2BPao2BU5Ir8H3/aq+NM4gz8U/FI3TRjH25hwZBx+Rr90dDkkm0LSmlkaSR7WJndjks2wZJrqxOElhbczvcwo14178vQ+Tf+CqU4X9k2+Pf+17L/0Ovz8/YRuC37X/AMOv+vif/wBF1+gX/BVW2L/smaht7avZf+h1+dH7HHiPSPAn7T/gPXPEGpQaTpFrcyia8uWxHGWTC7j2yeM16OFUnhJW8/yOOvZV4tn7sxqSTz3NDALKtcFdfHz4c6dKUn8d+Hom64bUY+/PrVGT9pL4XFx/xcLw4Mcn/iYJ0HNeHyS7Hp8y7n49/tg3KR/tT/FFT1/tZf8A0Cv0U/4Jczib9k7R29NSux/4/X5iftMeK9K8e/tF/ELxDoN4uo6Nf6nutbtAQsqquCy57Z6Gv0x/4JXoT+yXpYPB/tS7/wDQ69zGN/Vop+R5tBL28mj7EkcFetfnp/wVe+MMeneH/Cnwxtp8Tavcf2nqKA9LeM/Ip/3m/lX6DSrGkTvI4RIwWdieAB1NfhD+1d8Uz8bf2gfGfisyF9MhuDp2n88C3h+XI+pBP41x4Ci6tW66HRiqihT9T2D9gD4eJ8U/2ktKmliE2l+F4Tq1ySPlEvSEH/gRB/CvbvjT/wAFP4ND8beMPCWleAbTxBodtNNphv59SaP7VwVkO0IRjJIHNd//AME4vgPc+Df2X9T1tJf7L8T+N1kuEvXTe1vDtKQHHtktj6V5fL/wSHlcvJJ8VZ5JXYu7vpaksxOST+JrqdShUrSdd6LRGChVhSSpddz86bh4Wa4WKP7PbuzlIg2fLUkkLnvgYGfav22/YW+Nf/C6v2cvDV/cT+brOkr/AGVqQJy3mxDaGP8AvKAa/N/9rj9iK/8A2XdA8O60viF/E+l6ncPaTTtaiH7NKBuQcdmAbr/dr0H/AIJXfFhfBXxi1fwLeTbLDxXbedbBjhVvIR0+rJx9RXTi4rE4f2tPVI58O3Rreznuz9aEYPXJfFL4V+G/i/4L1Hwr4o0+PUNGvkxIjcPG38Lo38LKeQa6uFT5fzDDV83ft1/tAeKP2d/gqvirwrHZSal/acNmft8fmR7H68etfPU4ylNKG57Mmkm2fI/iH/gkh42TWbxNB8eaHNo4kP2RtRilFx5fYSbVIyOmQeaxm/4JGfE5zx428Lke6z//ABNdj+y1/wAFHPHHxG+Nej+HfiBc+GtI8N3cM7SXiR+QVdQCi7ycDPNfeEfx4+HStg+OfD2R/wBRCP8Axr06mIxdN8smcEKOHmuZH5u3X/BJn4oadp1zPD4s8M31xFGzx2sfnK0pAztBK4BNfFWqXN3olxeRNvs9T02dlOxvmgnibqCO4I61+9uqftBfDWzsbq7l8eeHkhhid3cX6EqNp5wDzX4IeL5F1LWPE1/DJ50F5f3VxFJ/fRpCVb8RXfgq1arzKZzYilSp2cT+gH4Maxfa78KvBmoalcG6v7rSLaWeZhgyOUGWPua71VyK86+AqY+DngNT/wBAO0/9FivRcbelfNT+JnsR2GhfmwOxr8Gv2qiV/ad+J49Nbf8A9BWv3m34f6mvwd/aoYP+078UP+w0/wD6Ctexlf8AEl6Hn474F6n6S/8ABLNw37J9lj/oLXw/8iV9ej7g5r46/wCCWAYfso2Xp/bN7/6HX2CoJ6152J/jT9Ttpfw4+gsrlUJ61+an/BU74lpr/jLwf8NrObfHpyNrOpRqePMb5YVYeoG4/wDAhX6N61rNtoGmXmoXsixWdpC9xNI5wFRVLE/pX8/vxe+MV78VPiv4t8dXErodVvJJoRnlLZfliT8EVa7MvpKdTneyOXF1HGHKup9g/wDBN/4a2+sfF/WvHeogJpng/T22yuPkW4lUkk/7sYNUPjD/AMFMz8TND8Y+D/8AhXVhf+GNUWaxiuZtQdZXjzhJdoTAOQGxmvpj9m79m/Vrf9h3/hFLXWG8M+KPGts19qGpmDzJIhP/AA4/658D0zXkMH/BICO3iVF+JswIGP8AkGLiuhVcPOtKdV7bfIydOrGmo0/mfnZaPPYpDJbTNFdW7LJDMDhkkU7kb6ggV+9P7M/xbtfjT8EPCnjGJwbm8tEjvEByUuEG2RT75Br8g/2s/wBmbUP2WvG2kaNcao2uabqlobm11Iw+Vl1OHjIHGRwfxr6Y/wCCTXxlS31jxR8Lr6bCXQ/tnSwx/i6TIPxw3411Y+MMRQVam72/I58I5Uqrpz6n6aTSAoRX5tf8FebtoJPhYOcGa6P6V+kBKtHlTmvzh/4K+Qq//CrM/wDPxdD/AMdFeNgv48T08R/CkeIf8Ewbsz/tZWw6f8SS6/pX7Hw5AXg9K/F7/gnX4j0TwH+0tHrPiDVrPRNMj0e4jN1fSiOPc2MLk96/VBf2pPhNFGN3xH8NcDr9vSt8fTn7bRX0McJKPs9z1odKTGOeteQt+1n8IAcH4l+Gs/8AX8tdn4F+J3hX4lWdxd+FvEOn6/bW0nlTS6fMJFjfGdpI6HFea4SWrR2qSezOrHzUg4PWlAyKTioKA8mlzTR1pec0ALx6UUYFFACHmlxgc0HrRk0AIOKBkmjr2ozigBJRiM18D/8ABXVx/wAKh8DoTwddHH/ADX3xIcxmvgP/AIK7xs3wk8C4H/Md/wDZDXZg/wCPD1Ma38OR+XNxAHLY9TVN7ZugFacgWBHeRgqrkknoK+y/2Of+CfGp/GVLPxl8Qre50bwQxElppODHd6qOzN3jh/VvpX1FWcKUeabsfP06cqr5YnH/APBNv4T+L/FP7QnhzxhpuizSeFNCeX7bq8vyQAspARCfvtnsucd6/Z4SqAK+Yfip+1j8MP2adJi8J+HrODVdRsEEUGg6GFSCzAGAJXHC/TkmvkLx1/wUJ+LPiu4lXSrnT/Clkx+SGwg8yVR7yNzn6Vz0snxuaP2sY8sejen/AAfwNauZ4TL17KUry7I/Vo3K0CUHvX4yL+1R8Ynk8w/EnXgc9BOMflivQfAv7enxc8LSRrf6rZ+J7VTlotUtwHb/ALaLzXXPhPFpXhOL+9focUeJcJe000fq4Hz0pQO5r5P+En/BQvwR40eDT/FUEngrU3womnbzLORvaQfd/wCBV7N8Xk8f+IfBEV98KfEmjWGrqDLEdStxc2t4pHC7wfl9iPXmvmMTgMRg5qniI8vrt959Dh8ZQxUOehLm9D0ea4SEHcwRACWc8BR3JNfmv+05/wAFK/EuifFG60P4Uz6TP4f0tTb3OpXtsbgXdwD8xjIYfIvTPOTXjf7Sv7UH7Rnl3nw/+JEx8ICbKzw6dZi3+3R+iTg/Mh/2TzXypNPHbxiONQqjgKO1ejhsDGHv1bPsclfFt+7Tuj7Csf8Agpn8aZmYTXPh4YGfl00j/wBnqjqf/BTT43Z/0a/0GLHrppOf/H6+Q4bl/MyTxSzTFgcGu54aj0gjk+sVVvI+sIf+Cnfx26G98PP9dMP/AMXUkv8AwU7+OiqWN74cRQMk/wBmHgf9918jWs5MpjYdelfTH7Fn7J99+0l49jvtUgkh+HujzK+pXRGBeSA5FrGe+f4iOgrOdLD0ouU4rQuFSvUkoxkfoh+w58Qfi98XfAs3jb4lz6db6XqR26LZWdkYZGjHWdyWPDdh6c19Pciqel2Vvpdlb2drDHbWdvGsUMESgLGijCqB2AAq8DmvmKklKTklY96KaVmN3GjG7rT8UmMdKzKPDf2uf2adP/ac+GZ0Ka8Oma3YS/bNIvuqRzgYAcd0bofTrX4xeOPgx488A+I9R8Pa74M1lNQsnMUv2ewluIX9GSRFKsp6g1/QXJk03ylkPzKre7AGvRw2Nnhk4rVHHWw0KzUnufzjy+ANbDc+DdYH10ab/wCIqS08D63FICPCWsIw6MukTAj/AMcr+jQ2cR6xRn/gAqKWyhU8Qxf9+1/wroWZSvflMvqatbmZ+OH7H/7E/iH4/wDiaLU/ElnfeHPAmmzq11cXELQz30ikEQQqwBAyBufGB2ya/ZS2t0ggiiRPLjiUIi56ADAohUIgUKFUdAAAKnzxXDicTPESvI6aNCNCNonl/wC0V8DNL/aF+FWr+C9Yu5bGO7Ky213Dy0E6HMb47gHqO4r8Vfip8BvHXwh8Z3/hfxB4b1Ce9tidl1YWUs9vdR/wyxsqkYPoeR0NfvsxprRRykFkRiOAWUE1rhcbUwycVqmRXw0K9m90fzl3PgjWMnPhLVh7nRpv/iKji8DauWGPCWqn/uDzf/EV/Ro9jC/WGI/VB/hUf9nQg5EEX/ftf8K6f7Sle/KZLBpK12fgp8Jv2fPHnxi8ZWXhnQfDeoW9zOcyXmoWckFvax95HZlHA9Bya/aX9nH4E6Z+zt8KNI8GaXdy362haa5u5eDNM/LsB2Gegr0tIVh+6irnrtUD+VTqwPauXE4yeISTVkbUcPGjdo8F/bV+JOpfDD9n3xPeaHZ3t/r2pRf2bYRWNu8zrJJ8pfCgkAAk5r8cfh98FvFvjvxp4Z8JJ4Y1y1TU76K2nubnTpo0jjLZkdmZQBxnmv6CHRZBhlDD0IzVdrdd2VRQfUKKeHxbw8XGK1YVcOqsk5PYzPDmh23hjRdO0nT0EVjYW8drCijACIoUfyzWoQ3WpIotoyak2iuG51Hin7XHwr/4XT+z74v8NiIPffZTeWJC5YXEXzoF9zgr/wACr8UfCH/Cb+A/Fmh+JNO8M6/Dqmi30V5GU0yfhkb5hnb3GRX9Ccsee2KgFjF/zyj/AO+B/hXdQxcqEHC10zlq4dVZKXYyPBXi6Pxx4T0bxBaxSQW+p2cdysUyFHj3KCVZTyCCcYNeYftbfs6r+0t8G9R8HpqR0rUfNS9sbkjMYnTlVkHdT0Pp1r2uG3VMAAKo6ADAqxgHiuNScJKUToaurM/nq8Y/A3xz4G8Rah4e17wbq66lYyFJlgsJbiF/R0dVIZT1Brmbn4e63H18H6wv10ab/wCIr+jholY5KKx6ZIyaY1pE45ijP/ABXrf2lJqzief9SindSZ/ORbeB9ZRwf+ET1bPr/ZE3H/jle6fs1fsheMP2j/GQ0x9PvvD3hq1KtqurX1q8WyP/AJ5xBgN0jDj0HU1+3jWEIORBF/37X/CpoY1jXaqBBnOFAA/Sl/aU1FqKsP6lG6bZneFvDVn4V0DStI08OLLTrWO0hEhy2xFCjJ7nArZ60Y44prHHSvG3PRGsBvH1r8J/2pfCHiWT9pn4lXEHhvW7i1m1h2jnh02Z0ddq8hguCK/dhiTVZrVGJJjUn1Kg114bEPDyckr3MKtJVVZnyj/wTC06+0f9lXTrbUbG50+5/tW9cw3cLRSAGTg7WANfXAP7s45NQJbBR8qhR6AYqdUIrCpPnm59zWMeWKj2Plr/AIKL+PdU8L/s16ppOh2N/eav4llXSkGn2zzMkTcysdgJX5QRn3r8wP2d/wBnHX/it8Z/CHhu78PatZaRNepNfz3VhLFEltF87gsygc4C4/2q/eORVfhlVgOmRmhI0Q5Cqp9hXZRxbo0nTjHV9TnqUFUmpN7ENlaw20EUNvGIYYEEUaAcKoGAB+FTyN8pAoLDpikC5zXnnUfIX/BSn4J3XxV/Z7m1HSbSS813wxcrqVtFBGXlljPyyooHJ4OcD0r8uPglqfjT4QfGDwj4xtPDHiGM6ZfxvMP7LuBvgchZVPyeh/Sv6AJId3HUGkSyixgxIfqg/wAK76WKdOm6bV0c06ClNTKVhdxajY297BuEFzGs0YZSp2sMjIPTrX5+f8FaPD+qawvwybT9Mv8AUfJuLkyCytZJtgIGM7QcV+jAXAAwMCmSRrJjeitjpuGa56NX2NRTSvY1nDni4n87sng7xJMuH8Ka4yjoG0mc/wDslUp/BuuoCD4S1j/wUT//ABFf0WeRHnHlR/8AfIpkllEf+WMf/fA/wr13msnpynnLL43vzM/nNi8C65LJx4T1nr/0CJv/AIiv1D/4JM+HbzQPhL41S902602STXVZUurZ4GYeX1AYDIr7tis4lP8AqY/++B/hVmOJYgQqKoPJ2gCuStjXWg4ctjppYZUpcyYqH5BS8+lJgeuKK8w7AOaOnSlI/GkJ46YoAMGil470UAB4pME80rdKBwOKAF6U0/NS7qB696AGyf6sivgf/grnOsXwj8Ds5CqNcyWPbCGvvmTBXFeSfG39nfQfj3f+Dl8TO8+j+HtROptpu0bLyQDCK5/uA8kd66MPUVKrGb6GVWPPBxXU+Hf2Dv2I4PFUdn8UfiPpv/EoVvP0PRbxcJNjkXU4P8PdVP1NdT+1b+3Fd+IL28+H3wtnlWFCba91vT1LSzkcGK2CgkKOhcfhX2D8XPhfrXxQsIvC0Wtv4a8HSKBqB0z5by6j7W8ZHEUeOCRz6Vg6Z4a+Cf7K+hIkaaB4SjAH7+7dWupSO5LZYn6V7+FxNF1fb1oOpP7MVsvXv6W9TyMRQqKl7GnLkj1k9/l/mfl1pH7PnxT8RQiTTPAGvXCSHdvlt/KLk9yZCCSfU1Nq37L/AMY9Eg868+HOtpF6xIkp/JWJr9Edc/4KEfB/THMdvq+o6uw72djIV/M8VLo3/BQf4P6gVWfUdV01icZudPk2j6kA19R/a2cNcywr5fSX9fgfOLLcrWksQr+qPyvvtI1HQLn7Nq9hd6TcZx5V/A8Lf+PAZ/CgbkHtX7OWPiX4XfHrSnghvvD3jG2kX5reQxySAehU/MDXzf8AGn/gnBomtpcaj8ONRPhy/OW/se+Jks5D1wjfej/UV24Piai5ezxcHTl+H+a+44sVw7U5faYWamj88ZLvjaOfrXqHwR/ak8bfAbUE/sa/N9opbM+h3zF7Zx32d4291/EGuO+Jnwp8YfCDVjp/i/QbnR5iSI5nG63m90kHytXGrh25r6aoqGOp8srSi/mj56n7bA1Lq8ZI/Vnwr8Wvg3+294Obwzr+n2zak65k0HVyFuYXP8dtJwT7FSD6ivj746/8ErPGnhu9uL/4Y6nF4q0kksuk6nIIL6Ef3Vf7sn1O2vnqxle1mimhkeGeI7o5YmKuh9Qw5FfSHw2/bz+Jnw8ghtNQntfGWmxDAj1fInUegmXn8xXxeK4ar0L1MBK6/lf6P/O3qfXYbiGjW9zGRs+6PkjxB8C/iV4MuZbfW/h/4ksZIjhz/Z0kqD/gcYYH86g0H4SeO/FUyw6P4H8R6jIxwPK0uYDPuzKAPzr9QvDv/BTvwPdxImu+Hdc0iYj5zDsuYvwxzj61uaj/AMFKPhXZ2wOn22v6jJ/zyishCB+LYFeN7DM4vk+rts9f2+Akuf2ysfKH7Pn/AAS48ZeMr+11f4nz/wDCG6ErBm0i3kEmoXK/3WIysQPrkn6V+hXiPxn8Of2VPAGl6TFFb6VZxKttpPh7T1DXV3IThVRByzMerH8TXxZ8Wf8AgpP4w8QRzW3gvR7XwlZkEHUL1hcXePUD7qn35r0H9if9nDXfFXiOL4z/ABLlvNU1STMmjQ6q5eU563DA/d/2Rj3qa2W1KVP6zmcuWK2it2+3l5vWxpRx1OrP2OCXM+sui8z7g0G7udT0ezur2yOm3k0SyS2bPvMJIztJ7kd60gSvWhFC5bGCetOI718a3d3R9GttQ/Sgn8aZ/nFOXgUhjeaVf90185ft6+Mdd8D/ALP93qnh7V7rRdTXU7aNLq0fY+1iQVz6GuZ0L9mH4iapoWmXzftCeMree9tY7gxrbxuqFlDEZz2zivWpYGE6Cr1Kqim2lo3tbsn3OCeJkqrpQg20k9119T616n0NI3I6Vl6DYXGkaFpdjc6hLqlzbwpFLezAB52AALsB0JPNfnX8Lf2tvF3gX9pnxHB4u13UNT+H93rdxozPdHfDp8hkbySp/hxgg+2fSlg8uqY2NV0mvcV/X0DEYyGGcFUXxaenqfpCZcHHpSiYYr5b/bK8c654R1j4QDQ9ZutOh1LxEsFz9lk2rcQlQdreo5r6B8b+K7PwJ4T1jxDf5+xaZbPcyBerBRwo9ycD8aynhJRp0pp3572Xo7GscRFynFqyjbX5XOj83PTmlVtx4r4b+Gfgn4t/taaTJ4+8QfEjVPAXhy/lb+xtI0SIZ8oEgO2cce/U16d8A9T+MHgD4r6n8N/Hi3vi/wANrbG50rxkbYqvGP3Ur9CcHp6g12VssVFTiqsXOG8dfnZtWbXVI5qeMdRxbptRls/81uj6bA29qdzXzn+3d4217wD+z9eav4e1e60TU0vbZBdWb7XClxkZ9CK9p8J6nLf+G/D1zLK0sk9jDJI7HlmKAkn8a8+WGlGhHEX0k2vut/mdarJ1XStqkn950DHA6EVEJcmvzm+Ev7V3izwT+0Tr6+L9b1DU/AN9rc2jtLeNuh0+Yv8Autpx8vpj0r239r/x54h8KfEX4KW2ha7d6Xa6rrXk3kdpJhLmIkYDeo/xr1Z5LXp4iFCTXvJtPportepwwzKlUpSqRT0drfM+r0bdwKd+Bri/iVf3em+APFtxaTvb3MGnXEkMsZwyMFOCD618e/s5fDL4k/HP4Q6N4vuvjp4w0i4vXlRrW3RJEXYxAwx55xXHh8DGtRlXnUUYppapvVpvon2OmriHTqKnGN21fp+p97lwByMfWozMe1ec/CD4d638NvC0ula14y1PxtdPctMNR1RQsqqQMJgdhj9a+Z/jhJ428bfto6X8O9G+IGteDtJu9DF2W0whtrrznafWlh8FHEVZ041FaKbvZ2svK1x1cQ6UIycdW7W06n3AmWOcGn4x2r54+GX7OXjXwb400vWtS+NfijxLp9m7NLpN9bosVwCpGGI5wCQfwrpf2rPEmq+EP2evG2saPqE+m6nbWpkgurdtrxtkcg1n9VhKvCjSqKXNZXs1a7t1Q3XlGlKpONra9D2H8DQWAGM18Q/D74IfFf4ifCbQ/GFl8fPEtpqWoWAvUs5olMSvgkLu644616n+xn8cNe+NHw31H/hKBFJ4h0C/bTbq7gGFudv3ZMdicc104jLfY051KdRTUHaVrq19t0r/ACMaOM9pKMJQceZXW2v3H0M0gpVlwcDk18p/Ffxv4k0v9t34U+HbPXL218P6hp0kl3pkcmIJ2DPgsvc8D8q9B/ax8Qan4X/Z58bappOo3GlahawK0N3bNtkj+cZIP0rL6hNSox5l+9Sa8rtrX7i/rUeWpK3wX/BXPcASRzTTxXxp+wt+0NrfiSzvPh745v7qbxZBGNR0251Fv3t7aOARg9yM8e1dZ8KfiBrusfth/FvQLvWbq60TTrS3a00+R8xQMQNxUds1tVyqtRqVqcmv3a5vVXS0+8zpY6nVhCcV8Tt6M+nGnC00ylucGvnz9r/476z8IPAOm2vhKBLjxn4ku107SxINyxMeDJjuR2ryuP8AY/8AjJHoH9vr8ddePj5YvtBtNubPzMbvKz+nTFTRy+M6SrVqqgpOyvd3tu9Nl5lVMU1N06cHJrfy/wCCfa4cHvTweK8j/Zo8Y+N/Gvw2ik+IXh+50HxTYTNaXXnw+Wt1t6TIPQ166owua86tSdCpKnJptPpqjspzVSCmuo8HA54pN5/uk+9eTftOfG1fgN8JNT8TxRJd6mSttp9tJ9153OFyO4HWvBPC/wCzf8YviJ4WtfF/iP41azoXijUIhd2un2UY+zW24bkVgOnBGQBxXfh8AqtL29aooRbsr3d36LourOOrinCp7KnByla78l8z7VODTCpB4GRXgn7LXxA+IviDQNc0D4l6Pc2viHw/c/Zl1d4DHDqUfZ1PQkdyOua89/bT8S+KU+J3wk8OeHvFepeFYdeuZbW6n058N/DhsdyOadPLpzxbwjmlvrurJXvp5BPGRjQ9uovpps9XY+vAuOoxUnGP618L/Grwl8XP2VfDP/CxNM+MGp+MNO065iS80fXIVCTxswXAx35r7G8JeK4vFPgjR/Eqr9mtb+xjvmRz/q1ZAx/IE1nicF7GnGtTmpwk2rq61XSzSfUujifaTlTnHlktbeXyOi3ccDd9KC3GTx9a+G9F8R/Ej9szx74pk0LxneeAPhlod0bGGTTIx9ovJAcEknr6+g4p194q+I/7HfxR8JWPiPxdd+P/AIZeJbkWH2nU0AubGYnAOR06g+hBrt/seV/Ze0XtbX5Nb7Xte1r26HMswT9/kfJe3Np99t7eZ9utNzQsm44PWua8ezzWvgnxDNbTNFOmm3EkM0ZwysIyVYe/evBP2B/GOv8AjL9nuHUvEOr3es6kNUuYjdXkm+Qqp4GfQV50MJKeGniU9ItK3rf/ACOyVdRrRpW3Tf3H1L+FDZ+tfKGq+P8AxHD+3/pHhlNavF8OyaE0zaWJP3DSbSdxX1rqP25PGWteCv2b9f1jQtSudI1SGeAR3Vq+x1BYAgH3rb+z5+1o0eZXqJNeV3bUz+tR5Kk7fA2vuPoMBs9DUgI6V8Q2vwJ+LE3wxtPF+j/HTxG+tPpy6mmn3camAnbv8st6dua9w/ZK+NepfG34N2Oua3FEmtwTyWN40IwkkiHG8Dtn0rTEZd7Gk61KopqL5Xa6afTdbaMijjPaTVOcHFtXWzv9x7cRxSYpQeAaD81eMeiIaAPxo6UvAFAB1ooAooAMUAEUdcUEkUALn1pM54pOvWlGM5FAAQKztW1W10ayuL2+uIrSztozLNcTOESNByWYnoK0WG4EV478cPgPc/Hi4s9G1zxFdad4EiAku9I0s+XNqMueBLL2jH90dT1row8Kc6iVWXLHq9/uXcyqylGN4K7Pj/8AaR/4KO3Ws3l34d+F92um6WhMU3iOQDzrjsfs6n7q/wC2Rk9sV8Yan4hk1/UJb/UdSk1G+lbL3N5OZJGPuzHNfs14N/Zb+FvgK3ih0fwPo8RQAebPbiaRvcs+cmun1X4Q+DNatTb3vhLQ7iE8FX0+MfyAr7fCZ/gMvXs8NQdurbV36/8ADnymLybE46XPWrfLoj8P48kbgcr6iphOUGM1+qvj79gH4SeMEkex0ibwtesDtn0aYogb1MZyD9OK+PfjT+wH8QfhtFcah4fx430aPLE2SbLyNfVov4v+A5r63BcRYHFWgpcsuz0/HY+VxeQ4vDrmS5l5f5HzTDqM9jexXNnPLaXcZ3JcW8hjkQ+oYcivo/4O/t7fEX4avBaa5OvjfRFwrQag226Rf9iYck/72a+YZFaKR0dWSRCVeNxtZSOoIPINIJsnA617GIwmHxkeXERUl/XzR52HxFfCSvRk0fsN4B+M/wAK/wBq3wrNpaGz1MzJi68O6xGouIzjnCnrjsy/pXy/8ef+CbV3YSz6v8K7r7TASXfw5qMuHT2hlPX/AHW/OviezvrjT7mG6tJ5rW7hO6O4t3KSIfZhyK+n/hB/wUJ8f+AxBYeKYU8baOmF8ydvKvY19pBw/wBGx9a+Pnk+My2bqZZUvH+V/wBW/Jn1EM1wmYQ9nj4Wf8y/q6/E+cPFPg3xD4B1F7DxLol/oN0jbSl9AUUn2f7p/A1mDMi5BDL6qc1+r3hH9rj4J/GrT1sdV1Kxsp5Fw2m+J4FTBPYM42n8DWncfso/AfxiPtUHhPRZfM+bzdOudqt+CtitY8STw3uY2hKL/rvb9TGXD8K75sLWTX9dj8hZwI8lsKPU8VueBfBfiD4i6tFpnhbRb3Xr6Q7QlnEWVfdn+6o+pr9U3/ZN/Z+8F4vL/wAM6LbCP5t+p3eVHvgtiuW8aftr/BP4GWD6X4XSDXLmP5BYeFrZRHnsGlGF/Umj/WGeIfLgaEpSffb8P80XHIoUtcVVSXlucj+zf/wTzs/DV5aeJPiY0GsarEVlt9BgO61t26hpW/5aMPTp9a+2o4BAqqgCqoAUKMAAdABXj/wE1r4jfEK3HjXxvbL4W0+8j/4lXhS25eKI9JrlzyzsOijAH417QEIHPJr8+zTE4jEYhvEzUmu2y8l/XzZ9tgcPRw9JRoxsvxfmCk45pf1puD60oNeOegKQPxppB7U4c0E0AfKH/BSiQxfs1zA/xaxaA+3J5qpof7H+u6v4a0i7T42+PIPOtYJlijuRsUFFIUDHQdK9q/aI+Bll+0F8PH8J3+qXGkQNdR3X2m1iWRwUzgYJA714eP2EdcSBYV+OnjeOFFCIiBAFUDAA+boBX1OExdOODjR9tySUm37vNdO3k+x4dfDzliZVPZ8yaS3t3Pq+xQwx28Lu0jRKqF3+8xAxk+5r8/PgB8EtO+Ow/aT8J6thFuNe8y0uMc21yHm2SD6Hr7E190fD3wjJ4G8HaPocuqXWty6fAsJ1C9P764I/if3Ncj8GP2ftP+DmteOdRstWu9Rk8Vah9vmSeNUFu2XO1SDyPnPX0rjwuLjhaVeMJe8+Xldu0r38tDor4eVepSco6K9/mrHwBr/xL1rXLb4VfDjxmZU8c+CPFyWNx5oOZ7YYEcme/TGe9ff/AO1Nol1rH7PXjuzso2mupNOZlRByQrKzfoDXNfF79kDw38Vvil4f8dvqN5o+saY0TTC0iVlvPLbKb89x0z6V75NAk6kOodSCpVhkEHqCK6sdmNCq8NVoKzhdtdL3vp5NmGGwlWKrU6r0lon5WsfPX7I+ux+Lf2SvDVj4c1aG11iz01rBpQu82lyMgb09jg+9ec+BfHfxV8OftY+H/h14g+Ilv4ytWsZrzUobKxjiWDA+VXwMg9+veuk8T/sI6Wnia/1nwH421/4dm/Yvc2elsGgZickqCRt+ld98BP2YPDPwFbUtQs7m917xLqXF5rmqPvnkGc7R/dGfetqmJwMI16kJc3tL2i4K6b/vPt5bmdOjiZOnCS5eS2qlo0vLz8zgf+Ckpz+zPf476jbdP98V7r8PufBHhc/9Q23/APRYrA+P3wTtPj/8OpvCN9qdzo9vLcRz/araNXcFDkDBIHNd3oOgR6Fo2m6ejtKllbx26uwwWCqBk/lXkTr03gqdFP3lKTfo0rfkejGlNYiVRrRpHwP8Cvg9Y/HTwr+0P4Ovz5Ru/EDy2k+OYLlSTG4/HFebw/E7VfFuu/BbwT4w82Pxt4I8SjTrwSqczwhgI5Mn2AH5V9//AAY/Z7074N6r4wvbLVLrUW8SaidRmW4RVELE/dXHUfWua+KH7IHhz4jfGPQviImp3ejatp8kcs8VrCrR3bIcqWz0PbIr6annGF+s1PaN8jV4u20uTlfyf6I8eeBrqhDkXvJ6rur3PSfilhfh74xz/wBAu4P/AI4a+Lv2Tv2a9X+JPwG0HWbL4reLvC1tPJOBpukzBYIiHIJUY79a+6PFnh5PE/h/V9JeZ7ZNRtpLZpkGWQMMEgeozXyloH/BPq98MaXDpuk/Gnxlp1hCSY7a0jjjjUk5JADd68jL8VTpYapSdXkk2nrHm2T8n3O7FUJTrQmocySfW3Y+nvh/4PuPAPgnTNAu9dv/ABJcWSFW1PU2DXE/JOXPr2r4w+Ovgy4+JP8AwUB0PQrPxBqPha4m8O711TSW23CbecA+h719V/BD4MXPwg8L3ek3Xi3WPGEtxdNcm+1kgyrkY2DBPFcB8Zf2O4vir8UYfHVr441rwpq8Vmtkh0qJCVQdSGJzzSwGJpYbFVJyqaOMkpW6vra36F4mlOrRjFQ2adr9vM2fhL+znrnwz8ZR65d/FHxX4rtlgeI6Zq8weBi38WMdR2qP9taTyv2XvHuct/oYzj/eFYngj9kLWPCXjLSNcn+MHjHWIrCcTNp94V8m4A/hfB6V678YfhZa/F34c694RvL6bT7bVovKe5gUM8YznIB4NZSr0442lWlV50mm2o22faw/Zylh501DlbT0vfofCv8AZfxz8J/skaN4q0f4hG48LLpyb9FsbFEuLa1b5SRJjLYzz9a+r/2QvAvhDwV8D9Fm8H302q2esL/aFzqFx/rZ5m+9uH8ODkYru/h78MNP8EfC3TfAkksmq6ZZ2JsHluUAaeMgg7gOOQa5/wDZ7+AkP7P2iatomna/fatot1dtdWtnexqBZ7uqowPI/wAK6MZmMMXh6lO6i+a6tFLmXnZbrfXv3OfD4OWHqwnZtctnd35X5Xez8jxP40uq/t+/Bod/7MlH5tJXoX7az7f2X/iLj/n1X/0MV0XjH9nqx8YfHTwl8TJdYu7a98OwNBFp8cSmKYEscsxOR97tXR/GD4VW/wAX/htr/hC6vZtMg1aIRvdQIHePDZyAeDWX1yj7TCSvpBLm+Um/yNFh6nLXVvibt91j4z8efDLVrj9mv4P/ABk8GB18beDtKtbiVYh813Zj76nHXaM/hW5+xz44sPit+1F8UfF+mrJHaarpdtOI5Rgo+AHH4Nmvr/4a/Diz+Hnw50Pwis76naaXYrYedcIAZkAwSyjjn0rzv4H/ALKHh/4E+PfE3iLQ9TvZIdaBUadMiiK2XduwjDkjPrXZ/alCeHxFKfxO6g/JyTs/uuvVmH1OrCrSnHb7Xqla55P+3DIvhnxX8GPF1+pXRNL1vZdzYysW4ggmvcvjBovjXx14d0fUPhx49svCNsrm5utQlt1uI7i3K5BBPAx1rsviH8N9A+KXhC98NeJbBNR0i7XDxnhlPZlPZh2NfL//AA7yEEL6Pa/FnxhbeEWYk6Ksg27M/c3bun4fhXPh8RhqtKlGrPklTvvHmTTd9u689DWpRrU6k5U48ynbrZpo6n9iX4neMviPpHjmTxR4iHiaPTNWaxs9QSBY45VXgsu0Dg9a+nAcxc9SK5L4ZfDXw/8ACLwfZeGfDNgtjpdqOFzueRj1d2/iY+tdYeRivJx9anXxM6lGNot6KyX4LTXc9DDU50qMYVHd9T5G/wCCken3UvwR0nUYoWmtdL1q3uLraM7Y89TXqevPr3xc+C+hXnwy8XweHbm5jt5o9U8lZ0WMIA8ZXoCOh9MV6j4k8KaZ4x0G/wBG1mzj1DS76Iw3FtMMq6nt/wDXr5Z/4d8W2jPd2Phf4neK/Dfhy6ZvN0m3kDoFPVQ2R+eK9PC4jD1KEKVafJKm203HmTTtdNfL0OCvSrQrSnTjzKSSetmreZV/ZG+JPjzxR8X/AImeHvEvjIeM9K8PRx21vfQwJHC0u/5mXaPTjqax/wBu7S9X1X4ufBKDQ71dM1mW9lW0vZE3rDISuGK9wPSvpv4K/Ajwt8CvCI0Hwzaukcj+bdXdw2+e5k7u7d/Ydqo/Fb4DWHxT8c+BvEt1qt1YzeFblrmG3gjVknJxwxPIHHatoZhh4Zj9YgrQSaWi19217LTVkTwlaWE9i3eTa6+d9/JHxp8YvBfjq4+NPgj4f/HD4g3Wq+BtanEkV5p8C28EswOFicAcZbAyemc1+gF9o0dt4Su9F06JYYUsXtLWFeAo8sqoH6Vxv7QfwG0b4/8AgyPQ9Tu59Mnt7hbq11C1QNLbuOpGfUV2/hfRLnRPDOmabfalLq91aQJC9/MgSScqMb2A4BOK5cbjo4qhRadpRunFKy/xK2muz9DbDYaVCrUTV4ytZt3fprrp0Pkj/gnFrlnZ+CvGnhKeVYtf03Xp5J7JziXYTjdt6kZBqt/wUV1WDV0+GHhKyIuPEF7r8U8VmnMgQFRuI7A4P5V6Z8W/2KvDXxB8Zy+MNC1vVfAvimfm5v8ARWAWc46smRz6nvS/Bv8AYx8O/DLxn/wmOsa3qvjnxYgxBqOtPu8j3Rcnn37V6X13BrFf2kpvm35Lfatbfa3XucccPifYfU3FW/mv0vfbe56744hdfAGtQ7S0q6ROhA5yREa+av8Agm5qdvffs9XFjDKkl5aazcieBTl48ngkdRmvr1bYMpDgEHqGGc18qeKf2ANGbxjqfiHwN411/wCH0upOXubXSyGhZmOW2jIwPbnFeZg69CWHqYWvPk5mmna6ur6O2vU7sRSqRqwr0481k1a9tzCnu7bVv+CllktnMly1joLJceUdwibaflbHQ13P/BQyNJf2WvEgP3PtFv0/3xXVfs+fsq+Fv2fxqN9ZXN7rviTUuLzWtTbdPIM52r/dGevJzXRfHf4O23x1+Gt/4NvdSudJtruSOQ3Vqgd12HIGDxzXRPG0Pr2HnBvkpcqvbezu3Yyhh6v1arGS96d3btfpc+ZdE+FP7Sfjb4Y6ToifETQdL8M39hFGrQ22LlLZlGF3AA528cGvqD4GfBnSvgX8ONM8J6VI91HbZknu5Rh7iZuXc+mT0FdV4Q8Mx+FfDOk6PHM9zHp9slss0oAZwowCcd62iMnIrgxmY1MSnSVlC97JJX7NnTh8JCi1N3crW1d7CA9BSniijGeteQegHFHHakx+VKOvSgAzRTtw9aKAGn26Up4o6Ggg/WgAA9aG4o5+lKQDQAm4dKD7UEAUgPHrQAoPrSEnNGM+1LjFADQoJoaMHpwfUUv0pc0AeAftAfsaeBfjvFPezWv/AAj/AInI+TW9OQK7nt5qdJB9efQ1+c/xj/ZC+I3wRuZpdS0p9b0ND8mtaShkiI7GRB80Z/Me9fsvkVHLCkqspAIYYYEZDD0Ir6PL89xWBtF+9Hs/0fT8vI8TG5Th8am2uWXdH4HRLvGVYOB6HpTj8tfsH8T/ANjf4VfFKd7nUPDKadqL9b/SH+yyfUgDafyr548Uf8EtbRmlk8P/ABBu4QTmODUbJZAvsXVgT+VffYbifAVY/vLwfmr/AIq58VX4cxcH+7akvu/M+AHZW4YA/WrFlczWjg29xNb/APXGVl/ka+yY/wDglv4rMo8zx5pIjzyVs5Ca9N8E/wDBL/wdpcqTeJvE+qeIsAE21si2kefQkFiR+VdFXiHLqav7S/kk/wDIzp5FjpactvmfAmh6FrfxE1uDSNIsb/xJq0pCpbRbpmz75OFHucV+gP7LP7BNn4AubTxV8QY7fVfESESWmkR4e1sT2Z/+ekn6DtX1F8PfhR4T+FWkrp3hTQbPRLcDDm3j/eSf7zn5m/E116KFXAGPavi804lq4yLpYdckPxf+Xy+8+ry/IqeFaqVnzS/BDYlKjmnluOlAPtR196+KPqQGKDjHtTQOadwKAAUZ5oyKafrQAvSlPXmkIo60ADDFCn8qPrR2oAN2KXcDTQuaMUAOzmmEZ7U6loAaUGOlKOOtHU9aX60AL1pOvFJmgE0ABUUKmKM54oyRxQA7PvSc/hSUhJ/CgB26kPPvSAZ9c0YOaAFwBQDnrR360h4oAdxnpRmk+9RjHegBcnNLzTSe4oBNAC5waQtmgmgACgBNuelLgY96XPpSY5oAMmlyT0pBkmigBTnFJu4pckijH0oAbx+NLQRSDjrzQA7JpD+VGePQ0daAANSk+tN4NKR2oATdTuMU3bmlzigAwaATSAnNLjNAAeKQc0vXtSY54oAXPPSg0DINBoACKKMYooAXpSk4oooAMZpG+XpRRQAgJoJxzRRQAuOM0gNFFIAAyaCccUUUwF7Ug4oooAUmgdOlFFAAcA9BRnmiigBDwaQnmiigBw45pM0UUAKBQOBRRQAds0mMmiigAoFFFAAeaMcUUUAApcZFFFACDrQDk0UUADfL0oB4oooAXPFJRRQAoHelPSiigBoHNKBmiigA6GgjvRRQA00/AIoooAaTjiiiigBTSZoooACe9J70UUAKOlIOaKKAHA4oIyKKKAE6LSgZFFFACEcUgORRRQAE5PSlPNFFAAPSlHXFFFAgKgU3PSiigYp5xR0oooAXtSUUUALnIpoNFFAC0UUUAf/Z"

BRANDED_STYLE = """
<style>
:root {
  --navy: #061b36;
  --blue: #2563eb;
  --green: #16a34a;
  --amber: #f59e0b;
  --red: #dc2626;
  --bg: #f5f7fb;
  --card: #ffffff;
  --text: #0f172a;
  --muted: #64748b;
  --line: #e2e8f0;
  --shadow: 0 14px 30px rgba(15, 23, 42, 0.08);
  --radius: 16px;
}
* { box-sizing: border-box; }
body {
  margin: 0;
  font-family: Inter, ui-sans-serif, system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
  background:
    radial-gradient(circle at 15% 20%, rgba(56, 189, 248, 0.18), transparent 28%),
    radial-gradient(circle at 85% 15%, rgba(37, 99, 235, 0.12), transparent 30%),
    linear-gradient(180deg, #eef7ff 0%, #f5f7fb 44%, #eef6fb 100%);
  color: var(--text);
  min-height: 100vh;
  position: relative;
}
body::before {
  content: "";
  position: fixed;
  inset: 0;
  pointer-events: none;
  z-index: -1;
  background:
    url("data:image/svg+xml,%3Csvg width='1600' height='900' viewBox='0 0 1600 900' xmlns='http://www.w3.org/2000/svg'%3E%3Cpath d='M0 620 C 240 520 360 730 620 620 C 860 520 1000 690 1240 600 C 1420 530 1510 560 1600 520 L1600 900 L0 900 Z' fill='%232563eb' fill-opacity='0.055'/%3E%3Cpath d='M0 690 C 220 590 430 770 680 660 C 920 555 1070 740 1310 650 C 1450 600 1530 610 1600 590 L1600 900 L0 900 Z' fill='%23061b36' fill-opacity='0.045'/%3E%3Cpath d='M0 760 C 260 650 420 820 700 720 C 930 640 1120 790 1360 720 C 1480 685 1540 690 1600 670 L1600 900 L0 900 Z' fill='%2338bdf8' fill-opacity='0.08'/%3E%3C/svg%3E");
  background-size: cover;
  background-position: bottom center;
}
.sidebar {
  width: 270px;
  background: linear-gradient(180deg, var(--navy), #020617);
  color: white;
  padding: 24px 16px;
  position: fixed;
  top: 0;
  left: 0;
  bottom: 0;
  overflow-y: auto;
}
.brand { display:flex; gap:12px; align-items:center; margin-bottom:28px; padding:0 8px; }

.logo {
  width:54px;
  height:54px;
  border-radius:14px;
  display:grid;
  place-items:center;
  background:white;
  overflow:hidden;
  box-shadow:0 10px 18px rgba(0,0,0,.18);
  flex:0 0 auto;
}
.logo img {
  width:100%;
  height:100%;
  object-fit:contain;
  display:block;
}
.brand h1 { font-size:18px; line-height:1; margin:0 0 5px; }
.brand p { margin:0; font-size:12px; color:#bfdbfe; }
.nav-section { margin: 18px 8px 8px; color: #93c5fd; font-size: 11px; font-weight: 900; letter-spacing: .12em; text-transform: uppercase; }
.nav-divider { height: 1px; background: rgba(191, 219, 254, 0.18); margin: 16px 8px 10px; }
.nav-item { display:flex; gap:12px; align-items:center; color:#e2e8f0; text-decoration:none; padding:13px 14px; border-radius:11px; margin:4px 0; font-size:14px; }
.nav-item:hover { background: rgba(255,255,255,.09); }
.nav-item.active { background: linear-gradient(135deg, #2563eb, #1d4ed8); color:white; box-shadow: 0 10px 18px rgba(37,99,235,.25); }
.sync-card { margin-top: 22px; background:rgba(37,99,235,.16); border:1px solid rgba(191,219,254,.15); border-radius:16px; padding:15px; }
.status-dot { width:9px; height:9px; border-radius:99px; background:#22c55e; display:inline-block; margin-right:8px; box-shadow:0 0 0 5px rgba(34,197,94,.12); }
.main { margin-left:270px; padding:26px; position: relative; z-index: 1; }
.topbar { display:flex; justify-content:space-between; gap:20px; align-items:flex-start; margin-bottom:20px; }
.topbar h2 { margin:0; font-size:28px; letter-spacing:-.04em; }
.topbar p { margin:8px 0 0; color:var(--muted); }
.top-actions { display:flex; gap:12px; align-items:center; color:var(--muted); font-size:13px; }
.button, button, a.primary, a.secondary { border:1px solid var(--line); background:white; border-radius:12px; padding:10px 14px; cursor:pointer; font-weight:700; text-decoration:none; color:var(--text); display:inline-flex; align-items:center; justify-content:center; gap:6px; line-height:1.1; white-space:nowrap; }
.primary, a.primary { background: var(--blue); color:white !important; border-color:var(--blue); }
.secondary, a.secondary { background:white; color:var(--text) !important; border-color:var(--line); }
a.primary:visited { color:white !important; }
a.secondary:visited { color:var(--text) !important; }
.grid { display:grid; gap:16px; }
.kpis { grid-template-columns: repeat(3, minmax(0, 1fr)); }
.two { grid-template-columns: 1fr 1fr; }
.card { background:rgba(255, 255, 255, 0.92); backdrop-filter: blur(10px); border:1px solid rgba(226, 232, 240, 0.86); border-radius:var(--radius); box-shadow:var(--shadow); padding:18px; overflow:hidden; margin-bottom:16px; }
.card h3 { margin:0 0 15px; font-size:16px; letter-spacing:-.02em; }
.card-subtitle { margin:-8px 0 15px; color:var(--muted); font-size:13px; }
.kpi { position:relative; min-height:125px; }
.kpi .label { color:var(--muted); font-size:12px; font-weight:700; }
.kpi .value { margin:9px 0 4px; font-size:25px; font-weight:900; letter-spacing:-.04em; }
.kpi .trend { font-size:12px; color:var(--muted); }
.badge { border-radius:999px; font-size:11px; font-weight:800; padding:4px 8px; display:inline-flex; align-items:center; }
.badge.green { background:#dcfce7; color:#166534; }
.badge.blue { background:#dbeafe; color:#1e40af; }
.badge.amber { background:#fef3c7; color:#92400e; }
.badge.red { background:#fee2e2; color:#991b1b; }
.table-wrap { max-height:560px; overflow:auto; border:1px solid var(--line); border-radius:14px; }
table { width:100%; border-collapse:collapse; font-size:13px; }
th, td { text-align:left; border-bottom:1px solid var(--line); padding:10px 8px; vertical-align:top; }
th { color:var(--muted); font-size:11px; text-transform:uppercase; letter-spacing:.04em; background:white; }
.right { text-align:right; }
.column-filter-row th {
  padding:6px 8px 10px;
  background:#f8fafc;
  position:sticky;
  top:0;
  z-index:2;
}
.column-filter-row input,
.column-filter-row select {
  width:100%;
  max-width:none;
  padding:7px 8px;
  border:1px solid var(--line);
  border-radius:9px;
  background:white;
  font-size:12px;
  text-transform:none;
  letter-spacing:normal;
  color:var(--text);
}
.filter-hint {
  display:flex;
  justify-content:space-between;
  gap:12px;
  flex-wrap:wrap;
  align-items:center;
  margin-bottom:12px;
  color:var(--muted);
  font-size:13px;
}
.filter-hint button {
  border:1px solid var(--line);
  background:white;
  border-radius:10px;
  padding:8px 10px;
  font-weight:800;
  cursor:pointer;
}
input[type=file], input[type=text], input[type=date], input[type=number], select, textarea { padding:12px; border:1px solid var(--line); border-radius:12px; background:white; width:100%; max-width:520px; font-family: inherit; }
textarea { min-height: 110px; resize: vertical; }
.notice { padding:13px 15px; border-radius:13px; font-weight:700; margin-bottom:16px; }
.notice.ok { background:#dcfce7; color:#166534; }
.notice.error { background:#fee2e2; color:#991b1b; }
code { background:#f1f5f9; padding:8px 10px; display:block; border-radius:12px; white-space:normal; }
@media (max-width: 1000px) { .sidebar { position:relative; width:100%; bottom:auto; } .main { margin-left:0; } .kpis, .two { grid-template-columns: 1fr; } }


/* Prototype-style purchase request page */
.page-hero {
  display:flex;
  gap:16px;
  align-items:center;
  margin-bottom:16px;
  padding:18px;
  border-radius:18px;
  border:1px solid rgba(226,232,240,.9);
  background:linear-gradient(135deg, rgba(22,163,74,.13), rgba(37,99,235,.10));
  box-shadow:var(--shadow);
}
.page-hero-icon {
  width:54px;
  height:54px;
  border-radius:16px;
  display:grid;
  place-items:center;
  color:white;
  font-size:26px;
  background:linear-gradient(135deg, #16a34a, #2563eb);
}
.page-hero h2 { margin:0 0 5px; font-size:24px; letter-spacing:-.03em; }
.page-hero p { margin:0; color:var(--muted); }
.form-grid { display:grid; grid-template-columns: repeat(2, minmax(0, 1fr)); gap:14px; }
.form-field label { display:block; font-weight:800; font-size:12px; color:var(--muted); margin-bottom:7px; }
.form-field input, .form-field select, .form-field textarea { width:100%; max-width:none; border:1px solid var(--line); border-radius:11px; padding:11px; background:white; font-family:inherit; }
.form-field.full { grid-column:1 / -1; }
.form-field textarea { min-height:88px; resize:vertical; }
.request-actions { display:flex; gap:10px; justify-content:flex-end; margin-top:16px; flex-wrap:wrap; }
.workflow { display:grid; gap:0; }
.workflow-step { display:grid; grid-template-columns:44px 1fr; gap:12px; align-items:center; padding:10px; border:1px solid var(--line); border-radius:14px; background:#fff; }
.workflow-circle { width:36px; height:36px; border-radius:999px; display:grid; place-items:center; font-weight:900; color:white; background:#94a3b8; }
.workflow-step.done .workflow-circle { background:#16a34a; }
.workflow-step.active .workflow-circle { background:#2563eb; }
.workflow-step.warning .workflow-circle { background:#f59e0b; }
.workflow-step.info .workflow-circle { background:#7c3aed; }
.workflow-step.future .workflow-circle { background:#64748b; }
.workflow-text strong { display:block; font-size:14px; }
.workflow-text span { display:block; color:var(--muted); font-size:12px; margin-top:3px; }
.workflow-line { width:2px; height:16px; background:var(--line); margin-left:27px; }
.role-card { display:grid; gap:14px; }
.role-meta { display:flex; gap:12px; flex-wrap:wrap; color:var(--muted); font-size:13px; }
.role-meta span { background:#f8fafc; border:1px solid var(--line); border-radius:999px; padding:7px 10px; }
.issued-items-box { border:1px solid var(--line); border-radius:13px; background:#f8fafc; padding:10px; display:grid; gap:8px; }
.empty-issued-items { color:var(--muted); font-size:13px; padding:10px; }
.other-items-header, .other-item-row { display:grid; grid-template-columns:1fr 90px 130px 90px; gap:8px; align-items:center; }
.other-items-header { color:var(--muted); font-size:11px; font-weight:800; margin-bottom:6px; }
.other-items-box { display:grid; gap:8px; margin-bottom:10px; }
.other-item-row input { width:100%; max-width:none; border:1px solid var(--line); border-radius:10px; padding:9px; }
.match-summary { margin-top:14px; border:1px solid var(--line); background:#f8fafc; border-radius:14px; padding:13px; display:grid; gap:6px; font-size:13px; }
.match-summary strong { font-size:14px; }
.match-summary span { color:var(--muted); }
.form-field input[type="file"] { width:100%; border:1px dashed var(--line); border-radius:11px; padding:10px; background:#f8fafc; }
.validation-banner { background:#fff7ed; border:1px solid #fed7aa; color:#7c2d12; border-radius:13px; padding:12px 14px; margin:12px 0 14px; font-size:13px; }
.submit-status-box { border-radius:13px; padding:12px 14px; margin:0 0 16px; font-size:13px; }
.submit-status-box.success { background:#dcfce7; border:1px solid #bbf7d0; color:#166534; }
.submit-status-box.error { background:#fee2e2; border:1px solid #fecaca; color:#991b1b; }
@media (max-width:820px) { .form-grid, .issued-item-option, .other-items-header, .other-item-row { grid-template-columns:1fr; } }

/* Keep links visually consistent after click; no purple/gray visited-state shift */
a, a:visited { color:#1d4ed8; }
.sidebar a, .sidebar a:visited, a.button, a.button:visited, .button, .button:visited, .status-card, .status-card:visited, .action-card, .action-card:visited, .filter-chip, .filter-chip:visited, .status-pill-row a, .status-pill-row a:visited { color:inherit; }
.po-review-id .po-link, .po-review-id .po-link:visited { color:#1d4ed8; }


/* Consolidated dashboard and forecasting additions */
.action-card {
  display:block;
  text-decoration:none;
  color:inherit;
  border:1px solid var(--line);
  border-radius:16px;
  background:linear-gradient(180deg,#ffffff,#f8fafc);
  padding:16px;
  box-shadow:var(--shadow);
  transition:transform .15s ease, box-shadow .15s ease, border-color .15s ease;
}
.action-card:hover { transform:translateY(-2px); border-color:#93c5fd; box-shadow:0 18px 38px rgba(15,23,42,.12); }
.action-card .icon { width:42px; height:42px; border-radius:14px; display:grid; place-items:center; color:white; font-size:21px; margin-bottom:12px; }
.action-card.blue .icon { background:linear-gradient(135deg,#38bdf8,#2563eb); }
.action-card.green .icon { background:linear-gradient(135deg,#22c55e,#16a34a); }
.action-card.amber .icon { background:linear-gradient(135deg,#fbbf24,#f59e0b); }
.action-card.purple .icon { background:linear-gradient(135deg,#a78bfa,#7c3aed); }
.action-card.red .icon { background:linear-gradient(135deg,#fb7185,#dc2626); }
.action-card strong { display:block; font-size:15px; margin-bottom:4px; }
.action-card span { color:var(--muted); font-size:12px; line-height:1.35; }
.visual-chart-row { align-items:start; }
.mini-chart-card { background:white; border:1px solid var(--line); border-radius:var(--radius); box-shadow:var(--shadow); padding:18px; }
.mini-chart-card h4 { margin:0 0 14px; font-size:15px; }
.mini-bar-row { display:grid; grid-template-columns:155px 1fr 105px; gap:10px; align-items:center; margin:10px 0; font-size:12px; }
.mini-bar-row span { color:var(--text); font-weight:700; overflow:hidden; text-overflow:ellipsis; white-space:nowrap; }
.mini-bar-row div { height:12px; background:#e2e8f0; border-radius:99px; overflow:hidden; }
.mini-bar-row b { display:block; height:100%; border-radius:99px; background:linear-gradient(90deg,#93c5fd,#2563eb); }
.mini-bar-row em { font-style:normal; color:var(--muted); font-weight:800; text-align:right; }
.project-bucket-grid { display:grid; grid-template-columns:repeat(5,minmax(0,1fr)); gap:12px; }
.project-bucket-item { background:#fff; border:1px solid var(--line); border-radius:16px; padding:14px; text-align:center; }
.project-bucket-item .bucket-label { font-size:12px; font-weight:900; margin-bottom:8px; }
.project-bucket-item.red .bucket-label { color:#dc2626; }
.project-bucket-item.orange .bucket-label { color:#f97316; }
.project-bucket-item.amber .bucket-label { color:#f59e0b; }
.project-bucket-item.lime .bucket-label { color:#65a30d; }
.project-bucket-item.green .bucket-label { color:#16a34a; }
.project-bucket-item .donut { width:96px; height:96px; margin:8px auto; border-radius:50%; display:grid; place-items:center; background:conic-gradient(#2563eb calc(var(--pct,0) * 1%), #e5e7eb 0); }
.project-bucket-item .donut > div { width:64px; height:64px; background:white; border-radius:50%; display:grid; place-items:center; line-height:1.05; }
.project-bucket-item .donut strong { display:block; font-size:20px; }
.project-bucket-item .donut span { display:block; font-size:10px; color:var(--muted); font-weight:800; }
.bucket-metric { display:flex; justify-content:space-between; gap:8px; border-top:1px solid var(--line); padding-top:8px; margin-top:8px; font-size:11px; color:var(--muted); }
.bucket-metric strong { color:var(--text); font-size:12px; }
.forecast-row { display:grid; grid-template-columns:repeat(7,minmax(0,1fr)); border:1px solid var(--line); border-radius:16px; overflow:hidden; }
.forecast-bucket { padding:14px; border-right:1px solid var(--line); min-height:145px; background:white; }
.forecast-bucket:last-child { border-right:none; }
.forecast-bucket strong { display:block; font-size:12px; }
.forecast-bucket .amount { color:#1d4ed8; font-size:19px; font-weight:900; margin-top:8px; }
.forecast-bucket .bucket-note { color:var(--muted); font-size:11px; margin-top:4px; }
.forecast-bucket .bars { display:flex; align-items:flex-end; height:54px; margin-top:14px; }
.forecast-bucket .bar { width:100%; min-height:7px; border-radius:6px 6px 0 0; background:linear-gradient(180deg,#93c5fd,#2563eb); }
.clickable-kpi { cursor:pointer; }
.filter-chip-row { display:flex; gap:8px; flex-wrap:wrap; margin:10px 0 14px; }
.filter-chip { border:1px solid var(--line); background:white; color:var(--text); border-radius:999px; padding:7px 10px; font-size:12px; font-weight:800; text-decoration:none; }
.filter-chip:hover { border-color:#93c5fd; }
.status-chip { border-radius:999px; padding:4px 8px; font-size:11px; font-weight:900; display:inline-flex; align-items:center; gap:4px; line-height:1.2; }
.status-chip.submitted, .status-chip.under-review, .status-chip.pending-approval { background:#fef3c7; color:#92400e; }
.status-chip.approved, .status-chip.converted-to-po, .status-chip.complete { background:#dcfce7; color:#166534; }
.status-chip.rejected, .status-chip.needs-info { background:#fee2e2; color:#991b1b; }
.status-chip.needs-payment-schedule { background:#ffedd5; color:#9a3412; }
.status-chip.assigned-to-pm { background:#dbeafe; color:#1e40af; }
.status-chip.in-progress { background:#ede9fe; color:#5b21b6; }
.status-chip.not-required { background:#e2e8f0; color:#334155; }
.status-chip.open { background:#ecfeff; color:#155e75; }
.status-chip.closed { background:#e2e8f0; color:#334155; }
.status-chip.default { background:#f1f5f9; color:#334155; }
@media (max-width:1200px) { .project-bucket-grid, .forecast-row { grid-template-columns:1fr 1fr; } .mini-bar-row { grid-template-columns:1fr; } }
@media (max-width:820px) { .project-bucket-grid, .forecast-row { grid-template-columns:1fr; } }


/* Mobile off-canvas navigation and phone-friendly layout */
.mobile-menu-button,
.mobile-nav-close,
.mobile-menu-overlay {
  display: none;
}

@media (max-width: 820px) {
  html, body {
    max-width: 100%;
    overflow-x: hidden;
  }

  body {
    display: block;
  }

  .mobile-menu-button {
    display: inline-flex;
    align-items: center;
    justify-content: center;
    gap: 8px;
    min-height: 44px;
    border: 1px solid var(--line);
    background: #ffffff;
    color: var(--text);
    border-radius: 12px;
    padding: 10px 13px;
    font-weight: 900;
    box-shadow: 0 8px 18px rgba(15, 23, 42, 0.08);
  }

  .mobile-nav-close {
    display: grid;
    place-items: center;
    position: absolute;
    top: 14px;
    right: 14px;
    width: 42px;
    height: 42px;
    border-radius: 12px;
    border: 1px solid rgba(255,255,255,.22);
    background: rgba(255,255,255,.08);
    color: #ffffff;
    font-size: 28px;
    line-height: 1;
    cursor: pointer;
  }

  .mobile-menu-overlay {
    display: block;
    position: fixed;
    inset: 0;
    background: rgba(2, 6, 23, 0.56);
    opacity: 0;
    pointer-events: none;
    transition: opacity .2s ease;
    z-index: 998;
  }

  body.mobile-nav-open .mobile-menu-overlay {
    opacity: 1;
    pointer-events: auto;
  }

  .sidebar {
    position: fixed;
    top: 0;
    left: 0;
    bottom: 0;
    width: min(88vw, 330px);
    min-height: 100vh;
    z-index: 999;
    transform: translateX(-104%);
    transition: transform .22s ease;
    overflow-y: auto;
    padding: 22px 16px 18px;
    box-shadow: 18px 0 44px rgba(2, 6, 23, 0.35);
  }

  body.mobile-nav-open .sidebar {
    transform: translateX(0);
  }

  .brand {
    padding-right: 48px;
    margin-bottom: 18px;
  }

  .brand h1 {
    font-size: 16px;
  }

  .brand p {
    font-size: 11px;
  }

  .nav-section {
    margin-top: 14px;
  }

  .nav-item {
    min-height: 44px;
    font-size: 15px;
    margin: 5px 0;
  }

  .sync-card {
    margin-top: 18px;
  }

  .main {
    margin-left: 0;
    width: 100%;
    padding: 12px;
  }

  .topbar {
    position: sticky;
    top: 0;
    z-index: 20;
    background: rgba(245, 247, 251, 0.94);
    backdrop-filter: blur(10px);
    border: 1px solid rgba(226, 232, 240, 0.85);
    border-radius: 16px;
    padding: 12px;
    margin-bottom: 12px;
    flex-direction: column;
    gap: 10px;
  }

  .topbar > div:first-of-type {
    width: 100%;
  }

  .topbar h2 {
    font-size: 22px;
    line-height: 1.12;
  }

  .topbar p {
    font-size: 13px;
    margin-top: 5px;
  }

  .top-actions {
    width: 100%;
    flex-direction: column;
    align-items: stretch;
    gap: 6px;
    font-size: 12px;
  }

  .grid,
  .grid.kpis,
  .grid.two,
  .grid.three,
  .grid.four,
  .two,
  .kpis,
  .form-grid,
  .project-bucket-grid,
  .forecast-row,
  .visual-chart-row {
    grid-template-columns: 1fr !important;
  }

  .card,
  .mini-chart-card,
  .action-card,
  .project-bucket-item,
  .forecast-bucket {
    padding: 14px;
    border-radius: 14px;
  }

  .card h3 {
    font-size: 15px;
    margin-bottom: 11px;
  }

  .card-subtitle {
    font-size: 12px;
  }

  .filterbar,
  .filters,
  .search-row,
  .request-actions,
  .role-buttons,
  .filter-chip-row {
    flex-direction: column;
    align-items: stretch;
  }

  .filters select,
  .filters input,
  .search-row input,
  .search-row select,
  .form-field input,
  .form-field select,
  .form-field textarea,
  button,
  .primary,
  .secondary,
  .filter-chip {
    width: 100%;
    min-width: 0;
    min-height: 44px;
    font-size: 15px;
  }

  .form-field textarea {
    min-height: 110px;
  }

  .table-wrap {
    width: 100%;
    max-width: 100%;
    overflow-x: auto;
    -webkit-overflow-scrolling: touch;
    border-radius: 14px;
  }

  .table-wrap table,
  table {
    min-width: 860px;
  }

  th, td {
    padding: 9px 8px;
    font-size: 12px;
  }

  .kpi {
    min-height: 105px;
  }

  .kpi .value {
    font-size: 23px;
  }

  .mini-bar-row,
  .bar-row,
  .waterfall-row,
  .approval-item,
  .detail-grid,
  .other-items-header,
  .other-item-row,
  .issued-item-option {
    grid-template-columns: 1fr !important;
  }

  .forecast-bucket {
    border-right: 0;
    border-bottom: 1px solid var(--line);
    min-height: 116px;
  }

  .forecast-bucket:last-child {
    border-bottom: 0;
  }

  .project-bucket-item .donut {
    width: 86px;
    height: 86px;
  }

  .bucket-metric {
    font-size: 12px;
  }

  .page-hero {
    grid-template-columns: 44px 1fr;
    padding: 14px;
  }

  .page-hero h2 {
    font-size: 20px;
  }

  .page-hero p {
    font-size: 12px;
  }
}


/* Feature phase: packets, dashboard cards, timelines, toasts, empty states */
.app-toast {
  position:fixed; right:24px; bottom:24px; z-index:2000;
  background:#020617; color:white; padding:13px 16px; border-radius:14px;
  box-shadow:0 16px 34px rgba(2,6,23,.25); opacity:0; transform:translateY(10px);
  pointer-events:none; transition:all .2s ease; font-weight:800; font-size:13px;
}
.app-toast.show { opacity:1; transform:translateY(0); }
.app-toast.error { background:#991b1b; }
.status-card-grid { display:grid; grid-template-columns:repeat(6,minmax(0,1fr)); gap:14px; margin-bottom:18px; }
.status-card { text-decoration:none; color:inherit; display:block; background:#fff; border:1px solid var(--line); border-radius:16px; padding:16px; box-shadow:var(--shadow); transition:transform .15s ease, border-color .15s ease; }
.status-card:hover, .status-card.active { transform:translateY(-2px); border-color:#93c5fd; }
.status-card .label { color:var(--muted); font-size:12px; font-weight:900; }
.status-card .value { font-size:25px; font-weight:950; letter-spacing:-.04em; margin:6px 0; }
.status-card .trend { color:var(--muted); font-size:12px; }
.status-card.amber { border-top:4px solid #f59e0b; }
.status-card.green { border-top:4px solid #16a34a; }
.status-card.red { border-top:4px solid #dc2626; }
.status-card.blue { border-top:4px solid #2563eb; }
.status-card.purple { border-top:4px solid #7c3aed; }
.status-card.slate { border-top:4px solid #64748b; }
.empty-state { border:1px dashed #cbd5e1; background:#f8fafc; border-radius:16px; padding:24px; text-align:center; color:var(--muted); }
.empty-state strong { display:block; color:var(--text); font-size:16px; margin-bottom:4px; }
.timeline { border-left:3px solid #dbeafe; padding-left:16px; display:grid; gap:12px; margin-top:14px; }
.timeline-item { position:relative; background:#fff; border:1px solid var(--line); border-radius:14px; padding:12px; }
.timeline-item:before { content:''; position:absolute; left:-26px; top:14px; width:14px; height:14px; border-radius:50%; background:#2563eb; border:3px solid white; box-shadow:0 0 0 2px #bfdbfe; }
.timeline-item strong { display:block; font-size:13px; }
.timeline-item span { display:block; color:var(--muted); font-size:12px; margin-top:3px; }
.packet-header { display:flex; justify-content:space-between; gap:20px; align-items:flex-start; border-bottom:2px solid #e2e8f0; padding-bottom:18px; margin-bottom:18px; }
.packet-logo { width:86px; height:86px; object-fit:contain; }
.packet-title h1 { margin:0; font-size:30px; letter-spacing:-.04em; }
.packet-title p { margin:6px 0 0; color:var(--muted); }
.packet-meta { display:grid; grid-template-columns:repeat(4,minmax(0,1fr)); gap:12px; margin:18px 0; }
.packet-field { background:#f8fafc; border:1px solid var(--line); border-radius:13px; padding:12px; }
.packet-field span { display:block; color:var(--muted); font-size:11px; font-weight:900; text-transform:uppercase; letter-spacing:.05em; }
.packet-field strong { display:block; margin-top:5px; font-size:14px; }
.packet-actions { display:flex; gap:10px; justify-content:flex-end; align-items:center; flex-wrap:wrap; margin-bottom:18px; }
.packet-actions a, .packet-actions button { min-height:42px; }
.approval-action-grid { display:grid; grid-template-columns:repeat(4,minmax(0,1fr)); gap:10px; margin-top:14px; }
.approval-action-grid form { margin:0; }
.approval-action-grid button { width:100%; }
.setup-card-grid { display:grid; grid-template-columns:repeat(3,minmax(0,1fr)); gap:14px; }
.setup-card { background:#fff; border:1px solid var(--line); border-radius:16px; box-shadow:var(--shadow); padding:16px; display:grid; gap:8px; }
.setup-card h4 { margin:0; font-size:15px; }
.setup-card .meta { color:var(--muted); font-size:12px; }
@media print {
  .sidebar, .topbar, .packet-actions, .mobile-menu-overlay, .app-toast { display:none !important; }
  .main { margin:0 !important; width:100% !important; padding:0 !important; }
  .card { box-shadow:none !important; border:0 !important; }
  body { background:white !important; }
}
@media (max-width:1200px) { .status-card-grid, .setup-card-grid, .packet-meta, .approval-action-grid { grid-template-columns:1fr 1fr; } }
@media (max-width:820px) { .status-card-grid, .setup-card-grid, .packet-meta, .approval-action-grid { grid-template-columns:1fr; } .packet-header { flex-direction:column; } }


/* Functional PO setup / missing information review */
.setup-table input, .setup-table select, .setup-table textarea { width:100%; min-width:150px; border:1px solid var(--line); border-radius:10px; padding:8px 9px; background:#fff; font-size:12px; }
.setup-table textarea { min-width:220px; min-height:58px; resize:vertical; }
.setup-table .po-number-cell { min-width:145px; font-weight:900; }
.po-review-id { display:grid; gap:8px; align-items:start; }
.po-review-id .po-link { font-size:15px; font-weight:950; color:#1d4ed8; text-decoration:underline; text-underline-offset:2px; }
.po-review-id .po-status-row { display:block; }
.po-review-id .status-chip { white-space:normal; text-align:left; max-width:120px; justify-content:center; }
.setup-table .payment-schedule-cell { min-width:520px; }
.setup-table .assign-cell { min-width:240px; }
.payment-schedule-builder { display:grid; gap:6px; }
.payment-schedule-row { display:grid; grid-template-columns:82px 128px 115px 1fr; gap:6px; align-items:center; }
.payment-row-label { color:var(--muted); font-size:11px; font-weight:800; text-transform:uppercase; letter-spacing:.04em; }
.payment-schedule-row input { min-width:0; }
.payment-schedule-help { color:var(--muted); font-size:11px; line-height:1.35; }
.inline-actions { display:grid; gap:7px; min-width:110px; }
.inline-actions button { width:100%; padding:8px 9px; border-radius:9px; font-size:12px; }
.action-required-card { border-left:5px solid #f59e0b; }
.info-callout { background:#eff6ff; border:1px solid #bfdbfe; color:#1e3a8a; border-radius:16px; padding:14px 16px; margin-bottom:16px; }
.info-callout strong { display:block; margin-bottom:4px; }
.status-pill-row { display:flex; flex-wrap:wrap; gap:8px; margin:10px 0 0; }
.status-pill-row a { text-decoration:none; display:inline-flex; }
.status-pill-row a.active .status-chip { box-shadow:0 0 0 3px rgba(37,99,235,.15); }
.status-pill-row .status-chip.all { background:#e2e8f0; color:#334155; }
@media (max-width:820px) { .setup-table table { min-width:1550px; } .payment-schedule-row { grid-template-columns:1fr; } }


.project-filter-th { min-width:220px; }
.project-filter-head { display:flex; align-items:center; gap:8px; }
.project-filter-button {
  border:1px solid #cbd5e1;
  background:#f8fafc;
  color:#0f2e5c;
  border-radius:8px;
  width:28px;
  height:28px;
  display:inline-grid;
  place-items:center;
  cursor:pointer;
  font-size:13px;
}
.project-filter-button:hover { background:#eaf4ff; border-color:#93c5fd; }
.filter-icon { font-size:15px; line-height:1; display:inline-block; font-weight:900; }
.project-filter-select {
  display:none;
  margin-top:8px;
  width:100%;
  border:1px solid #cbd5e1;
  border-radius:10px;
  background:white;
  padding:8px 10px;
  min-height:38px;
  font-size:12px;
  text-transform:none;
  letter-spacing:0;
  color:#0f172a;
}
.project-filter-select.show { display:block; }


.vendor-terms-card { page-break-inside:avoid; }
.terms-list { margin:0; padding-left:20px; display:grid; gap:8px; font-size:13px; line-height:1.45; color:#334155; }
.terms-list strong { color:#0f172a; }

/* Expense upload / PO matching */
.expense-upload-layout { display:grid; grid-template-columns:1fr 1fr; gap:16px; align-items:start; margin-bottom:16px; }
.expense-review-table table { min-width:1500px; }
.expense-review-table .comments-cell { max-width:320px; white-space:normal; color:#334155; font-size:12px; line-height:1.35; }
.expense-review-table .match-reason-cell { max-width:280px; white-space:normal; color:#64748b; font-size:12px; line-height:1.35; }
.expense-review-table select, .expense-review-table input, .expense-review-table textarea { width:100%; border:1px solid var(--line); border-radius:10px; padding:8px 9px; background:#fff; font-size:12px; }
.expense-review-table textarea { min-height:58px; resize:vertical; }
.expense-action-cell { min-width:240px; }
.expense-match-form { display:grid; gap:7px; }
.expense-match-form button { width:100%; }
.expense-batch-list { display:grid; gap:8px; }
.expense-batch-item { border:1px solid var(--line); border-radius:13px; padding:10px; background:#f8fafc; display:grid; gap:3px; }
.expense-batch-item strong { font-size:13px; }
.expense-batch-item span { color:var(--muted); font-size:12px; }
.status-chip.auto-matched { background:#dcfce7; color:#166534; }
.status-chip.manually-matched { background:#dbeafe; color:#1e40af; }
.status-chip.needs-review { background:#fef3c7; color:#92400e; }
.status-chip.no-match { background:#fee2e2; color:#991b1b; }
.status-chip.no-po-needed { background:#e2e8f0; color:#334155; }
.status-chip.needs-pm-review { background:#ede9fe; color:#5b21b6; }
@media (max-width:1000px) { .expense-upload-layout { grid-template-columns:1fr; } }


.purchase-review-panel {
    display: grid;
    gap: 7px;
    min-width: 230px;
}
.purchase-review-panel label {
    color: var(--muted);
    font-size: 11px;
    font-weight: 900;
    text-transform: uppercase;
    letter-spacing: .04em;
}
.purchase-review-panel select,
.purchase-review-panel input,
.purchase-review-panel textarea {
    width: 100%;
    border: 1px solid var(--line);
    border-radius: 10px;
    padding: 8px 9px;
    background: #fff;
    font-size: 12px;
}
.purchase-review-panel textarea {
    min-height: 76px;
    resize: vertical;
}
.purchase-review-panel button {
    border: 0;
    background: var(--blue);
    color: #fff;
    border-radius: 10px;
    padding: 9px 10px;
    font-weight: 900;
    cursor: pointer;
}
.notice.info {
    background: #eff6ff;
    border-color: #bfdbfe;
    color: #1e3a8a;
}

.rollout-filter-bar { border:1px solid var(--line); border-radius:16px; background:#fff; padding:14px; margin:0 0 16px; box-shadow:var(--shadow); }
.rollout-filter-fields { display:grid; grid-template-columns:repeat(5, minmax(140px, 1fr)); gap:10px; }
.rollout-filter-fields label { display:grid; gap:5px; color:var(--muted); font-size:11px; font-weight:900; text-transform:uppercase; letter-spacing:.04em; }
.rollout-filter-fields input, .rollout-filter-fields select { border:1px solid var(--line); border-radius:10px; padding:9px 10px; background:#fff; color:var(--text); text-transform:none; letter-spacing:0; font-weight:600; }
.rollout-filter-actions { display:flex; gap:10px; justify-content:flex-end; margin-top:12px; flex-wrap:wrap; }
.data-freshness-banner { margin-bottom:16px; }
.posting-audit-cell { min-width:210px; font-size:12px; color:var(--muted); line-height:1.35; }
.status-chip.posted { background:#dcfce7; color:#166534; }
.status-chip.not-posted { background:#e2e8f0; color:#334155; }
.status-chip.duplicate { background:#fee2e2; color:#991b1b; }
.readiness-list { display:grid; gap:8px; margin-top:10px; }
.readiness-item { display:flex; justify-content:space-between; gap:12px; border:1px solid var(--line); border-radius:12px; padding:10px; background:#f8fafc; }
@media (max-width:1100px) { .rollout-filter-fields { grid-template-columns:repeat(2, minmax(0, 1fr)); } }
@media (max-width:760px) { .rollout-filter-fields { grid-template-columns:1fr; } }

/* Rollout-standard KPI cards: consistent top-page card design across pages */
.grid.kpis .card.kpi, .status-card-grid .status-card {
  border-top: 5px solid #2563eb;
  min-height: 122px;
}
.grid.kpis .card.kpi:nth-child(2), .status-card-grid .status-card.green { border-top-color:#16a34a; }
.grid.kpis .card.kpi:nth-child(3), .status-card-grid .status-card.amber { border-top-color:#f59e0b; }
.grid.kpis .card.kpi:nth-child(4), .status-card-grid .status-card.red { border-top-color:#dc2626; }
.grid.kpis .card.kpi:nth-child(5) { border-top-color:#7c3aed; }
.grid.kpis .card.kpi:nth-child(6) { border-top-color:#0ea5e9; }
.po-link { font-weight:900; color:#1d4ed8; text-decoration:underline; text-underline-offset:2px; }
.po-link:visited { color:#1d4ed8; }
.view-as-banner { margin-bottom:16px; background:#eef6ff; border:1px solid #bfdbfe; border-radius:14px; padding:12px 14px; color:#1e3a8a; }
.vendor-detail-link { font-weight:900; color:#1d4ed8; text-decoration:underline; }
.vendor-detail-link:visited { color:#1d4ed8; }


/* PO packet visual redesign - selected Option 1 internal, Option 2 vendor in Coastal blue */
.po-packet-page { background:#ffffff; border:1px solid #dbe3ee; border-radius:18px; overflow:hidden; box-shadow:0 18px 45px rgba(15, 23, 42, .10); margin-bottom:18px; }
.po-packet-page .packet-card { box-shadow:none; border:1px solid #dbe3ee; border-radius:14px; padding:16px; background:#fff; }
.po-packet-page .packet-section-title { margin:0 0 12px; font-size:15px; font-weight:950; color:#08284d; letter-spacing:.01em; }
.po-packet-page .packet-grid-2 { display:grid; grid-template-columns:1fr 1fr; gap:14px; }
.po-packet-page .packet-grid-3 { display:grid; grid-template-columns:repeat(3,minmax(0,1fr)); gap:12px; }
.po-packet-page .packet-grid-4 { display:grid; grid-template-columns:repeat(4,minmax(0,1fr)); gap:12px; }
.po-packet-page .packet-label { display:block; font-size:10px; font-weight:950; letter-spacing:.08em; color:#64748b; text-transform:uppercase; margin-bottom:4px; }
.po-packet-page .packet-value { display:block; color:#0f172a; font-size:13px; font-weight:850; }
.po-packet-page .packet-muted { color:#64748b; font-size:12px; line-height:1.45; }
.po-packet-page .packet-table { width:100%; border-collapse:collapse; font-size:12px; }
.po-packet-page .packet-table th { background:#06264a; color:#fff; padding:9px 8px; font-size:10px; letter-spacing:.06em; text-transform:uppercase; position:static; }
.po-packet-page .packet-table td { padding:9px 8px; border-bottom:1px solid #e2e8f0; vertical-align:top; }
.po-packet-page .packet-table tr:nth-child(even) td { background:#f8fafc; }
.po-packet-page .packet-total-box { background:#eaf4ff; border:1px solid #b9d7f7; border-radius:13px; padding:14px; }
.po-packet-page .packet-total-box .packet-total-row { display:flex; justify-content:space-between; gap:16px; border-bottom:1px solid rgba(15,23,42,.10); padding:7px 0; font-size:13px; }
.po-packet-page .packet-total-box .packet-total-row:last-child { border-bottom:none; }
.po-packet-page .packet-total-box strong { color:#06264a; }
.po-packet-page .packet-pill { display:inline-flex; align-items:center; gap:7px; border-radius:999px; padding:6px 9px; background:#eaf4ff; color:#0b4f96; font-size:11px; font-weight:900; }
.po-packet-page .packet-actions-inline { display:flex; justify-content:flex-end; gap:8px; flex-wrap:wrap; margin:0 0 14px; }

.internal-packet .internal-hero { background:linear-gradient(135deg,#061b36 0%,#09284d 70%,#0b5dad 100%); color:#fff; padding:24px; position:relative; overflow:hidden; }
.internal-packet .internal-hero:after { content:""; position:absolute; left:-6%; right:-6%; bottom:-52px; height:94px; background:rgba(255,255,255,.98); border-radius:50% 50% 0 0/70% 70% 0 0; }
.internal-packet .internal-hero-top { position:relative; z-index:2; display:flex; justify-content:space-between; align-items:flex-start; gap:18px; }
.internal-packet .internal-logo-card { background:#fff; border-radius:16px; padding:12px 16px; width:260px; min-height:80px; display:flex; align-items:center; box-shadow:0 14px 30px rgba(0,0,0,.20); }
.internal-packet .internal-logo-card img { width:100%; max-height:72px; object-fit:contain; }
.internal-packet .internal-title { text-align:right; }
.internal-packet .internal-title h1 { margin:0; color:#fff; font-size:27px; letter-spacing:.02em; }
.internal-packet .internal-title .po-number { margin-top:6px; color:#58b7ff; font-size:19px; font-weight:950; }
.internal-packet .internal-body { padding:24px; }
.internal-packet .internal-summary-strip { display:grid; grid-template-columns:1.2fr 1.2fr 1fr; gap:14px; margin-bottom:16px; }
.internal-packet .internal-info-card { border-left:4px solid #0b5dad; background:#fff; border-radius:13px; padding:14px; border-top:1px solid #dbe3ee; border-right:1px solid #dbe3ee; border-bottom:1px solid #dbe3ee; }
.internal-packet .internal-blue-panel { background:#06264a; color:#fff; border-radius:13px; padding:14px; }
.internal-packet .internal-blue-panel .packet-label { color:#bfdbfe; }
.internal-packet .internal-blue-panel .packet-value { color:#fff; }
.internal-packet .internal-footer-strip { display:grid; grid-template-columns:repeat(4,1fr); background:#061b36; color:#fff; margin:18px -24px -24px; padding:14px 24px; gap:12px; }
.internal-packet .internal-footer-strip .packet-label { color:#bfdbfe; }
.internal-packet .internal-footer-strip .packet-value { color:#fff; }

.vendor-packet .vendor-hero { background:#fff; padding:24px 24px 18px; border-bottom:5px solid #0b5dad; position:relative; }
.vendor-packet .vendor-hero:before { content:""; position:absolute; right:-100px; top:-140px; width:340px; height:340px; background:radial-gradient(circle,#dbeafe 0%,#eff6ff 50%,transparent 70%); border-radius:50%; }
.vendor-packet .vendor-hero-top { display:flex; justify-content:space-between; align-items:flex-start; gap:18px; position:relative; z-index:2; }
.vendor-packet .vendor-logo { width:285px; max-width:48%; }
.vendor-packet .vendor-logo img { width:100%; max-height:82px; object-fit:contain; object-position:left center; }
.vendor-packet .vendor-title { text-align:right; }
.vendor-packet .vendor-title h1 { margin:0; color:#06264a; font-size:28px; letter-spacing:.01em; }
.vendor-packet .vendor-title .po-number { color:#0b5dad; font-size:18px; font-weight:950; margin-top:6px; }
.vendor-packet .vendor-body { padding:22px 24px 24px; }
.vendor-packet .vendor-date-row { display:grid; grid-template-columns:repeat(3,1fr); gap:12px; margin-bottom:14px; }
.vendor-packet .vendor-date-card { border:1px solid #dbe3ee; border-radius:14px; padding:14px; background:#f8fbff; }
.vendor-packet .vendor-total-badge { background:linear-gradient(135deg,#06264a,#0b5dad); color:#fff; border-radius:14px; padding:14px; text-align:right; box-shadow:0 10px 22px rgba(11,93,173,.18); }
.vendor-packet .vendor-total-badge span { color:#bfdbfe; text-transform:uppercase; font-size:10px; font-weight:950; letter-spacing:.08em; display:block; }
.vendor-packet .vendor-total-badge strong { color:#fff; font-size:21px; display:block; margin-top:4px; }
.vendor-packet .vendor-note-row { display:grid; grid-template-columns:1.1fr 1fr 1fr; gap:12px; margin-top:14px; }
.vendor-packet .vendor-note-card { border:1px solid #dbe3ee; background:#f8fafc; border-radius:14px; padding:14px; min-height:92px; }
.vendor-packet .vendor-wave-footer { margin:22px -24px -24px; padding:18px 24px; color:#fff; background:linear-gradient(135deg,#061b36,#0b5dad); position:relative; overflow:hidden; }
.vendor-packet .vendor-wave-footer:before { content:""; position:absolute; left:-40px; right:-40px; top:-55px; height:92px; background:rgba(255,255,255,.18); border-radius:0 0 50% 50%; }
.vendor-packet .vendor-wave-footer strong { position:relative; z-index:2; display:block; text-align:center; letter-spacing:.04em; }
.vendor-packet .vendor-wave-footer span { position:relative; z-index:2; display:block; text-align:center; color:#bfdbfe; font-size:11px; margin-top:4px; }
.vendor-packet .vendor-terms-card { border:1px solid #bfdbfe; background:#f8fbff; border-radius:14px; padding:16px; margin-top:14px; page-break-inside:avoid; }
.vendor-packet .terms-list { columns:1; column-gap:0; margin:0; padding-left:18px; }
.vendor-packet .terms-list li { break-inside:avoid; margin:0 0 8px; font-size:11px; line-height:1.35; }
.vendor-packet .packet-table th { background:#0b5dad; }

@media (max-width:1100px) {
  .internal-packet .internal-summary-strip,
  .po-packet-page .packet-grid-2,
  .po-packet-page .packet-grid-3,
  .po-packet-page .packet-grid-4,
  .vendor-packet .vendor-date-row,
  .vendor-packet .vendor-note-row { grid-template-columns:1fr; }
  .internal-packet .internal-footer-strip { grid-template-columns:1fr 1fr; }
  .vendor-packet .vendor-hero-top, .internal-packet .internal-hero-top { flex-direction:column; }
  .vendor-packet .vendor-title, .internal-packet .internal-title { text-align:left; }
  .vendor-packet .vendor-logo { max-width:100%; width:300px; }
}
@media print {
  body { background:#fff !important; }
  .po-packet-page { box-shadow:none !important; border:1px solid #cbd5e1 !important; border-radius:0 !important; }
  .po-packet-page .packet-table th { -webkit-print-color-adjust:exact; print-color-adjust:exact; }
  .internal-packet .internal-hero, .internal-packet .internal-footer-strip, .vendor-packet .vendor-total-badge, .vendor-packet .vendor-wave-footer { -webkit-print-color-adjust:exact; print-color-adjust:exact; }
  .po-packet-page .packet-card, .vendor-packet .vendor-terms-card { page-break-inside:avoid; }
}

</style>
"""


def get_view_as_email():
    """Admin/Executive helper for role testing and user-specific data filtering."""
    access = get_user_access()
    if normalize_role(access.get("role")) not in ["Admin", "Executive"]:
        return ""
    return clean_text(request.cookies.get("PO_DASHBOARD_VIEW_AS")) or ""

def current_working_email():
    return get_view_as_email() or get_current_user().get("email") or ""


def lookup_dashboard_user_by_email(email):
    email = clean_text(email)
    if not email:
        return None
    try:
        conn = get_sql_connection()
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT TOP 1 Email, DisplayName, RoleName, IsActive
            FROM dbo.DashboardUsers
            WHERE LOWER(Email) = LOWER(?);
            """,
            email,
        )
        row = cursor.fetchone()
        conn.close()
        return row
    except Exception:
        return None


def get_effective_user_access():
    """Return the access profile the app should behave as.

    When an Admin/Executive uses View As, menus, page-access checks, and data
    filtering should use the selected user's role so role testing actually looks
    different. The real signed-in user is still used for permission to start/clear
    View As and for true Admin-only user-management POST actions.
    """
    actual = get_user_access()
    view_email = get_view_as_email()
    if not view_email:
        return actual

    row = lookup_dashboard_user_by_email(view_email)
    if not row or not getattr(row, "IsActive", 0):
        return {
            "email": view_email,
            "display_name": "",
            "role": "No Access",
            "is_active": False,
            "found_in_sql": bool(row),
            "lookup_error": "View-as user is inactive or was not found.",
            "is_view_as": True,
            "actual_email": actual.get("email", ""),
            "actual_role": actual.get("role", "No Access"),
        }

    return {
        "email": clean_text(getattr(row, "Email", view_email)).lower(),
        "display_name": clean_text(getattr(row, "DisplayName", "")),
        "role": normalize_role(getattr(row, "RoleName", "No Access") or "No Access"),
        "is_active": True,
        "found_in_sql": True,
        "lookup_error": "",
        "is_view_as": True,
        "actual_email": actual.get("email", ""),
        "actual_role": actual.get("role", "No Access"),
    }


def current_data_role():
    return get_effective_user_access().get("role") or "No Access"


def current_data_email():
    return get_effective_user_access().get("email") or current_working_email()


def current_visibility_scope():
    role = normalize_role(current_data_role())
    email = current_data_email()
    if role in ROLE_GROUP_ALL_PO_VIEW:
        return "all", email
    if role in ROLE_GROUP_DREDGING_VIEW:
        return "dredging_only", email
    if role in ROLE_GROUP_NON_DREDGING_VIEW:
        return "non_dredging", email
    if role in ROLE_GROUP_ASSIGNED_NON_DREDGING_VIEW and email:
        return "assigned_non_dredging", email
    return "none", email


def should_filter_pos_to_requestor():
    scope, _email = current_visibility_scope()
    return scope == "assigned_non_dredging"


def requestor_filter_sql(alias):
    """Return the PO/project visibility predicate for the current working user.

    July 1 role security rule:
    - Dredging visibility is based on the Department field only.
    - PM-Dredging sees Department = Dredging only.
    - Division Manager-Diving sees Department <> Dredging only.
    - PM-Diving sees assigned non-Dredging only.

    Some older uploaded POs have Department populated on IssuedPOLines but not
    consistently on PurchaseOrders. For PurchaseOrders queries, this function
    falls back to the related issued PO line department so restricted users do
    not see cross-department POs just because a header field is blank.
    """
    scope, email = current_visibility_scope()
    clean_alias = str(alias or "").strip()

    if clean_alias.lower() in ["po", "purchaseorders"]:
        dept_expr = (
            f"COALESCE(NULLIF({alias}.Department, ''), "
            f"(SELECT TOP 1 lvis.Department FROM dbo.IssuedPOLines lvis "
            f"WHERE lvis.PONumber = {alias}.PONumber AND COALESCE(lvis.Department, '') <> '' "
            f"ORDER BY lvis.IssuedPOLineId), '')"
        )
        req_expr = (
            f"COALESCE(NULLIF({alias}.Requestor, ''), "
            f"(SELECT TOP 1 lvis.Requestor FROM dbo.IssuedPOLines lvis "
            f"WHERE lvis.PONumber = {alias}.PONumber AND COALESCE(lvis.Requestor, '') <> '' "
            f"ORDER BY lvis.IssuedPOLineId), '')"
        )
    else:
        dept_expr = f"COALESCE({alias}.Department, '')"
        req_expr = f"COALESCE({alias}.Requestor, '')"

    if scope == "all":
        return "1=1", []
    if scope == "dredging_only":
        return f"LOWER({dept_expr}) = 'dredging'", []
    if scope == "non_dredging":
        return f"LOWER({dept_expr}) <> 'dredging'", []
    if scope == "assigned_non_dredging":
        return f"LOWER({dept_expr}) <> 'dredging' AND LOWER({req_expr}) = LOWER(?)", [email]
    return "1=0", []



def ensure_purchase_request_project_code_column(cursor):
    cursor.execute(
        """
        IF COL_LENGTH('dbo.PurchaseRequests', 'ProjectCode') IS NULL
        BEGIN
            ALTER TABLE dbo.PurchaseRequests ADD ProjectCode NVARCHAR(100) NULL;
        END
        """
    )


def _dedupe_project_option(options, code, name):
    code = clean_text(code)
    name = clean_text(name)
    value = code or name
    if not value:
        return
    key = value.lower()
    if key in options["seen"]:
        return
    options["seen"].add(key)
    label = (code + " - " if code else "") + (name or "Unnamed Project")
    options["rows"].append({"code": code, "name": name, "value": value, "label": label})


def load_existing_project_options():
    """Return project options for New Purchase Request using the same uploaded setup data as Projects.

    This version applies the July 1 role visibility rules:
    - Admin / Executive / Purchaser: all projects
    - Dredging PM: Dredging projects only
    - Diving PM: assigned non-Dredging projects only using Requestor email in current Phase 1 data
    - Division Manager - Diving: all non-Dredging projects
    - Bookkeeping has no New Purchase Request page access, so it should not reach this function.
    """
    conn = None
    options = {"seen": set(), "rows": []}
    try:
        conn = get_sql_connection()
        cursor = conn.cursor()
        try:
            ensure_project_code_columns(cursor)
            conn.commit()
        except Exception:
            try:
                conn.rollback()
            except Exception:
                pass

        # Primary source: uploaded issued PO setup lines. This is the same data source
        # that makes the Projects tab dropdown work, and it includes Department and Requestor.
        line_where, line_params = requestor_filter_sql("l")
        primary_queries = [
            (
                f"""
                SELECT DISTINCT
                    COALESCE(l.ProjectCode, '') AS ProjectCode,
                    COALESCE(l.ProjectName, '') AS ProjectName
                FROM dbo.IssuedPOLines l
                WHERE COALESCE(l.ProjectCode, l.ProjectName, '') <> ''
                  AND {line_where}
                ORDER BY COALESCE(l.ProjectCode, ''), COALESCE(l.ProjectName, '');
                """,
                line_params,
            )
        ]

        # Secondary source: PO headers joined to Projects. Also has Department/Requestor on po.
        po_where, po_params = requestor_filter_sql("po")
        primary_queries.append(
            (
                f"""
                SELECT DISTINCT
                    COALESCE(po.ProjectCode, pr.ProjectCode, '') AS ProjectCode,
                    COALESCE(pr.ProjectName, '') AS ProjectName
                FROM dbo.PurchaseOrders po
                LEFT JOIN dbo.Projects pr ON po.ProjectId = pr.ProjectId
                WHERE COALESCE(po.ProjectCode, pr.ProjectCode, pr.ProjectName, '') <> ''
                  AND {po_where}
                ORDER BY COALESCE(po.ProjectCode, pr.ProjectCode, ''), COALESCE(pr.ProjectName, '');
                """,
                po_params,
            )
        )

        for sql, params in primary_queries:
            try:
                cursor.execute(sql, *params)
                for r in cursor.fetchall():
                    _dedupe_project_option(options, getattr(r, "ProjectCode", ""), getattr(r, "ProjectName", ""))
            except Exception:
                try:
                    conn.rollback()
                except Exception:
                    pass
                continue

        # Final fallback for all-department users only. Do not use this for PM visibility,
        # because Projects may not yet have Requestor/PM assignment in Phase 1.
        scope, _email = current_visibility_scope()
        if not options["rows"] and scope == "all":
            try:
                cursor.execute(
                    """
                    SELECT DISTINCT
                        COALESCE(ProjectCode, '') AS ProjectCode,
                        COALESCE(ProjectName, '') AS ProjectName
                    FROM dbo.Projects
                    WHERE COALESCE(ProjectCode, ProjectName, '') <> ''
                      AND COALESCE(IsActive, 1) = 1
                    ORDER BY COALESCE(ProjectCode, ''), COALESCE(ProjectName, '');
                    """
                )
                for r in cursor.fetchall():
                    _dedupe_project_option(options, getattr(r, "ProjectCode", ""), getattr(r, "ProjectName", ""))
            except Exception:
                try:
                    conn.rollback()
                except Exception:
                    pass

        conn.close()
        return options["rows"]
    except Exception:
        try:
            if conn:
                conn.close()
        except Exception:
            pass
        return []

def validate_selected_project(cursor, selected_project_value):
    """Validate/link a project selected on a purchase request.

    The New Purchase Request dropdown stores values as either:
      - ProjectCode
      - ProjectName
      - ProjectCode||ProjectName

    Approval/PO creation also calls this validator from the saved request values,
    so it must accept all three shapes.  The prior version treated
    "ProjectCode||ProjectName" as one literal value, which meant the approval
    step could fail before changing the request to Converted to PO.
    """
    selected_project = clean_text(selected_project_value)
    if not selected_project:
        raise ValueError("Please select an existing project.")

    ensure_project_code_columns(cursor)

    project_code = ""
    project_name = ""
    if "||" in selected_project:
        parts = selected_project.split("||", 1)
        project_code = clean_text(parts[0])
        project_name = clean_text(parts[1])
    else:
        project_code = selected_project
        project_name = selected_project

    search_values = []
    for value in [project_code, project_name, selected_project]:
        value = clean_text(value)
        if value and value.lower() not in [v.lower() for v in search_values]:
            search_values.append(value)

    for value in search_values:
        cursor.execute(
            """
            SELECT TOP 1 ProjectId, ProjectCode, ProjectName, Department
            FROM dbo.Projects
            WHERE ProjectCode = ? OR ProjectName = ?
            ORDER BY ProjectId;
            """,
            value,
            value,
        )
        project = cursor.fetchone()
        if project:
            return project

    for value in search_values:
        cursor.execute(
            """
            SELECT TOP 1
                COALESCE(pr.ProjectCode, l.ProjectCode, '') AS ProjectCode,
                COALESCE(pr.ProjectName, l.ProjectName, '') AS ProjectName,
                COALESCE(NULLIF(l.Department, ''), NULLIF(pr.Department, ''), '') AS Department
            FROM dbo.IssuedPOLines l
            LEFT JOIN dbo.PurchaseOrders po ON l.PurchaseOrderId = po.PurchaseOrderId
            LEFT JOIN dbo.Projects pr ON po.ProjectId = pr.ProjectId
            WHERE COALESCE(pr.ProjectCode, l.ProjectCode, '') = ?
               OR COALESCE(pr.ProjectName, l.ProjectName, '') = ?
            ORDER BY l.IssuedPOLineId;
            """,
            value,
            value,
        )
        line_project = cursor.fetchone()
        if line_project:
            project_id = get_or_create_project(
                cursor,
                clean_text(line_project.ProjectName) or project_name,
                clean_text(line_project.Department),
                clean_text(line_project.ProjectCode) or project_code,
            )
            cursor.execute(
                """
                SELECT TOP 1 ProjectId, ProjectCode, ProjectName, Department
                FROM dbo.Projects
                WHERE ProjectId = ?;
                """,
                project_id,
            )
            project = cursor.fetchone()
            if project:
                return project

    # Last safe fallback: if the request already contains a saved project code/name,
    # create the normalized project record from that saved request data so approval
    # does not get stuck after a valid request was submitted.
    fallback_name = project_name if project_name != selected_project else ""
    fallback_code = project_code if project_code != selected_project else ""
    if fallback_name or fallback_code:
        project_id = get_or_create_project(cursor, fallback_name or fallback_code, "", fallback_code)
        cursor.execute(
            """
            SELECT TOP 1 ProjectId, ProjectCode, ProjectName, Department
            FROM dbo.Projects
            WHERE ProjectId = ?;
            """,
            project_id,
        )
        project = cursor.fetchone()
        if project:
            return project

    raise ValueError("Selected project was not found in the same project setup data used by the Projects tab.")



def allowed_attachment(filename):
    if not filename or "." not in filename:
        return False
    ext = filename.rsplit(".", 1)[1].lower()
    return ext in ALLOWED_ATTACHMENT_EXTENSIONS


def ensure_request_attachment_table(cursor):
    cursor.execute(
        """
        IF OBJECT_ID('dbo.PurchaseRequestAttachments', 'U') IS NULL
        BEGIN
            CREATE TABLE dbo.PurchaseRequestAttachments (
                AttachmentId INT IDENTITY(1,1) PRIMARY KEY,
                PurchaseRequestId INT NOT NULL,
                RequestNumber NVARCHAR(100) NULL,
                PONumber NVARCHAR(100) NULL,
                OriginalFileName NVARCHAR(260) NOT NULL,
                StoredFileName NVARCHAR(260) NOT NULL,
                StoragePath NVARCHAR(1000) NOT NULL,
                ContentType NVARCHAR(200) NULL,
                FileSize BIGINT NULL,
                UploadedBy NVARCHAR(255) NULL,
                UploadedAt DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME()
            );
        END;

        IF OBJECT_ID('dbo.PurchaseRequestAttachments', 'U') IS NOT NULL
        BEGIN
            IF COL_LENGTH('dbo.PurchaseRequestAttachments', 'PONumber') IS NULL
                ALTER TABLE dbo.PurchaseRequestAttachments ADD PONumber NVARCHAR(100) NULL;
            IF COL_LENGTH('dbo.PurchaseRequestAttachments', 'RequestNumber') IS NULL
                ALTER TABLE dbo.PurchaseRequestAttachments ADD RequestNumber NVARCHAR(100) NULL;
            IF COL_LENGTH('dbo.PurchaseRequestAttachments', 'UploadedBy') IS NULL
                ALTER TABLE dbo.PurchaseRequestAttachments ADD UploadedBy NVARCHAR(255) NULL;
            IF COL_LENGTH('dbo.PurchaseRequestAttachments', 'UploadedAt') IS NULL
                ALTER TABLE dbo.PurchaseRequestAttachments ADD UploadedAt DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME();
        END;
        """
    )

def save_purchase_request_attachments(cursor, purchase_request_id, request_number, files):
    files = files or []
    saved = []
    if not files:
        return saved
    ensure_request_attachment_table(cursor)
    request_folder = secure_filename(str(request_number or purchase_request_id)) or str(purchase_request_id)
    target_dir = os.path.join(REQUEST_ATTACHMENT_ROOT, request_folder)
    os.makedirs(target_dir, exist_ok=True)
    user_email = get_current_user().get("email") or "Unknown"
    for uploaded in files:
        if not uploaded or not uploaded.filename:
            continue
        if not allowed_attachment(uploaded.filename):
            raise ValueError(f"Attachment type not allowed: {uploaded.filename}")
        original_name = uploaded.filename
        safe_name = secure_filename(original_name)
        stamp = datetime.utcnow().strftime("%Y%m%d%H%M%S%f")
        stored_name = f"{stamp}_{safe_name}"
        storage_path = os.path.join(target_dir, stored_name)
        uploaded.save(storage_path)
        file_size = os.path.getsize(storage_path) if os.path.exists(storage_path) else None
        cursor.execute(
            """
            INSERT INTO dbo.PurchaseRequestAttachments
                (PurchaseRequestId, RequestNumber, OriginalFileName, StoredFileName, StoragePath, ContentType, FileSize, UploadedBy)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?);
            """,
            purchase_request_id,
            request_number,
            original_name,
            stored_name,
            storage_path,
            uploaded.content_type,
            file_size,
            user_email,
        )
        saved.append(original_name)
    return saved


def load_po_attachments(po_number):
    po_number = clean_text(po_number)
    if not po_number:
        return []
    try:
        conn = get_sql_connection()
        cursor = conn.cursor()
        ensure_request_attachment_table(cursor)
        cursor.execute(
            """
            SELECT AttachmentId, RequestNumber, PONumber, OriginalFileName, ContentType, FileSize, UploadedBy, UploadedAt
            FROM dbo.PurchaseRequestAttachments
            WHERE PONumber = ?
            ORDER BY UploadedAt DESC, AttachmentId DESC;
            """,
            po_number,
        )
        rows = cursor.fetchall()
        conn.close()
        return rows
    except Exception:
        return []


def attachment_card(po_number):
    rows = load_po_attachments(po_number)
    table_rows = ""
    for r in rows:
        size = getattr(r, "FileSize", 0) or 0
        size_label = f"{size / 1024:.1f} KB" if size else ""
        table_rows += f"""
        <tr>
            <td><a class="po-link" href="/purchase-request-attachment/{r.AttachmentId}" target="_blank">{h(r.OriginalFileName)}</a></td>
            <td>{h(r.RequestNumber)}</td>
            <td>{h(r.ContentType or '')}</td>
            <td>{h(size_label)}</td>
            <td>{h(r.UploadedBy)}</td>
            <td>{h(r.UploadedAt)}</td>
        </tr>
        """
    if not table_rows:
        table_rows = '<tr><td colspan="6"><div class="empty-state"><strong>No quote or backup files linked.</strong><span>Files uploaded with the purchase request will appear here after the request is converted to a PO.</span></div></td></tr>'
    return f"""
    <div class="card">
        <h3>Quote / Backup Files</h3>
        <div class="table-wrap"><table><tr><th>File</th><th>Request</th><th>Type</th><th>Size</th><th>Uploaded By</th><th>Uploaded At</th></tr>{table_rows}</table></div>
    </div>
    """

def view_as_notice():
    view_email = get_view_as_email()
    if not view_email:
        return ""
    effective = get_effective_user_access()
    return f'<div class="view-as-banner"><strong>Viewing as:</strong> {h(view_email)} · Effective role: {h(effective.get("role", "No Access"))} <a class="button secondary" href="/clear-view-as">Clear view-as</a></div>'


def shell(title, subtitle, active, content):
    access = get_effective_user_access()
    role = access["role"]

    procurement_nav_items = [
        ("My Dashboard", "/my-dashboard", "🏠"),
        ("Help Center", "/help-center", "❓"),
        ("New Purchase Request", "/purchase-request", "📝"),
        ("Purchase Requests", "/purchase-requests", "📋"),
        ("POs & Balances", "/pos-balances", "💳"),
        ("Projects", "/projects", "📁"),
        ("PO Setup Review", "/project-po-setup", "🧭"),
    ]

    accounting_nav_items = [
        ("Upload Issued POs", "/upload-po", "⬆️"),
        ("PO Maintenance", "/po-maintenance", "🛠️"),
        ("Expense Upload / PO Matching", "/expense-upload", "🧾"),
        ("Clear Expense Data", "/admin/clear-expense-data", "🧹"),
        ("Expenses", "/expenses", "📄"),
        ("Missing PO Review", "/missing-po-review", "⚠️"),
        ("Vendors", "/vendors", "🏢"),
        # PM Comment PO Audit remains available by direct URL from reporting, but is hidden from main rollout navigation to reduce overlap.
        ("Import History", "/import-history", "🕘"),
        # Dormant for July 1 rollout: Exceptions and Exports hidden from main navigation.

    ]

    admin_nav_items = [
        ("User Access", "/user-access", "🔐"),
        ("Future Pages", "/future-pages", "🧭"),
        ("Who Am I", "/whoami", "👤"),
    ]

    def build_nav_item(label, href, icon):
        active_class = " active" if active == label else ""
        return f'<a class="nav-item{active_class}" href="{href}"><span>{icon}</span>{h(label)}</a>'

    def build_nav_section(section_title, items):
        section_html = ""
        for label, href, icon in items:
            if role_can_access(role, label):
                section_html += build_nav_item(label, href, icon)
        if not section_html:
            return ""
        return f'<div class="nav-divider"></div><div class="nav-section">{h(section_title)}</div>' + section_html

    nav_sections = [
        build_nav_section("Procurement", procurement_nav_items),
        build_nav_section("Accounting", accounting_nav_items),
        build_nav_section("Admin", admin_nav_items),
    ]
    nav_html = "".join(section for section in nav_sections if section)
    if nav_html.startswith('<div class="nav-divider"></div>'):
        nav_html = nav_html.replace('<div class="nav-divider"></div>', '', 1)

    return f"""
<!DOCTYPE html>
<html>
<head>
    <title>{h(title)}</title>
    {BRANDED_STYLE}
</head>
<body>
    <div class="mobile-menu-overlay" onclick="closeMobileMenu()" aria-hidden="true"></div>
    <aside class="sidebar" id="sidebarNav">
        <div class="brand">
            <div class="logo"><img src="{CE_LOGO_DATA_URI}" alt="Coastal Engineering logo"></div>
            <div>
                <h1>Coastal Engineering</h1>
                <p>Command Center</p>
            </div>
        </div>
        <button type="button" class="mobile-nav-close" onclick="closeMobileMenu()" aria-label="Close menu">&times;</button>
        <nav>{nav_html}</nav>
        <div class="sync-card">
            <div style="font-weight:800; font-size:13px; margin-bottom:10px;">Signed-In Role</div>
            <div><span class="status-dot"></span>{h(role)}</div>
            <div style="margin-top:14px; color:#bfdbfe; font-size:12px;">
                {"Viewing As" if access.get("is_view_as") else "User"}<br>
                <strong style="color:white;">{h(access["email"] or "Not detected")}</strong>
            </div>
            {f'<div style="margin-top:10px; color:#93c5fd; font-size:11px;">Actual signed-in user<br><strong style="color:white;">{h(access.get("actual_email") or "")}</strong></div>' if access.get("is_view_as") else ""}
        </div>
    </aside>

    <main class="main">
        <header class="topbar">
            <button type="button" class="mobile-menu-button" onclick="toggleMobileMenu()" aria-label="Open menu">☰ Menu</button>
            <div>
                <h2>{h(title)}</h2>
                <p>{h(subtitle)}</p>
            </div>
            <div class="top-actions">
                <span>Role: {h(role)}</span>
                <span>Database: {h(SQL_DATABASE_NAME)}</span>
            </div>
        </header>

        {view_as_notice()}
        {content}
    </main>
    <div id="appToast" class="app-toast"></div>
    <script>
        function toggleMobileMenu() {{
            document.body.classList.toggle('mobile-nav-open');
        }}
        function closeMobileMenu() {{
            document.body.classList.remove('mobile-nav-open');
        }}
        document.addEventListener('keydown', function(event) {{
            if (event.key === 'Escape') closeMobileMenu();
        }});
        window.addEventListener('resize', function() {{
            if (window.innerWidth > 820) closeMobileMenu();
        }});
    </script>
</body>
</html>
"""



def status_chip(value):
    text = value or "Unknown"
    cls = str(text).lower().replace(" ", "-").replace("/", "-")
    allowed = {"all", "submitted", "under-review", "needs-more-info", "pending-approval", "approved", "converted-to-po", "rejected", "open", "closed", "needs-pm-info", "needs-forecast-date", "needs-payment-schedule", "assigned-to-pm", "in-progress", "needs-info", "complete", "not-required", "auto-matched", "manually-matched", "needs-review", "no-match", "no-po-needed", "needs-pm-review", "posted", "not-posted", "duplicate"}
    if cls not in allowed:
        cls = "default"
    return f'<span class="status-chip {cls}">{h(text)}</span>'


def percent(value):
    try:
        return "{:.1f}%".format(float(value or 0) * 100)
    except Exception:
        return "0.0%"


def build_simple_filter_bar(filters, action, fields):
    """Reusable rollout filter form for high-volume operational pages."""
    pieces = []
    for name, label, placeholder in fields:
        value = filters.get(name) or ""
        pieces.append(f'<label><span>{h(label)}</span><input name="{h(name)}" value="{h(value)}" placeholder="{h(placeholder)}"></label>')
    return (
        f'<form class="rollout-filter-bar" method="get" action="{h(action)}">'
        f'<div class="rollout-filter-fields">{"".join(pieces)}</div>'
        f'<div class="rollout-filter-actions"><button class="primary" type="submit">Apply Filters</button><a class="button secondary" href="{h(action)}">Clear</a></div>'
        f'</form>'
    )


def add_like_filter(where, params, column_sql, value):
    value = clean_text(value)
    if value:
        where.append(f"LOWER(COALESCE({column_sql}, '')) LIKE LOWER(?)")
        params.append(f"%{value}%")


def latest_data_freshness_banner():
    """Show users whether PO and expense data are current enough to trust."""
    po_bits = "PO upload: no upload found"
    expense_bits = "Expense upload: no upload found"
    try:
        conn = get_sql_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT TOP 1 FileName, UploadedAt, UploadedBy, TotalRows, SuccessCount, ErrorCount, ImportStatus
            FROM dbo.ImportBatches
            ORDER BY UploadedAt DESC;
        """)
        r = cursor.fetchone()
        if r:
            po_bits = f"PO upload: {h(r.FileName)} · {h(r.UploadedAt)} · {h(r.TotalRows)} rows · {h(r.ImportStatus)}"
        try:
            cursor.execute("""
                SELECT TOP 1 FileName, UploadedAt, UploadedBy, TotalRows, AutoMatchedCount, NeedsReviewCount, NoMatchCount, ImportStatus
                FROM dbo.ExpenseUploadBatches
                ORDER BY UploadedAt DESC;
            """)
            e = cursor.fetchone()
            if e:
                expense_bits = f"Expense upload: {h(e.FileName)} · {h(e.UploadedAt)} · {h(e.TotalRows)} rows · {h(e.ImportStatus)}"
        except Exception:
            pass
        conn.close()
    except Exception:
        pass
    return (
        f'<div class="notice info data-freshness-banner">'
        f'<strong>Data freshness:</strong> {po_bits}<br>{expense_bits}<br>'
        f'<small>Rollout balance method: current app balance equals issued PO amount minus posted matched expenses. The app does not write back to Unanet or the ERP.</small>'
        f'</div>'
    )


def posting_status_chip(row):
    posted = bool(getattr(row, "PostedToPO", 0))
    duplicate = bool(getattr(row, "IsDuplicate", 0))
    if posted:
        return status_chip("Posted")
    if duplicate:
        return status_chip("Duplicate")
    return status_chip("Not Posted")


def posting_reason(row):
    if bool(getattr(row, "PostedToPO", 0)):
        if (clean_text(getattr(row, "ReviewDecision", "")) or "") == "Matched to PO":
            return "Posted after manual review."
        return "Auto-posted from a valid PO match."
    if bool(getattr(row, "IsDuplicate", 0)):
        return "Not posted because this looks like a duplicate uploaded expense."
    decision = clean_text(getattr(row, "ReviewDecision", "")) or "Pending Review"
    if decision == "No PO Needed":
        return "Not posted because reviewer marked No PO Needed."
    if decision in ["Needs PM Review", "Hold for More Info", "Pending Review"]:
        return "Not posted because review is not final."
    if not (clean_text(getattr(row, "CorrectPONumber", "")) or clean_text(getattr(row, "MatchedPONumber", "")) or clean_text(getattr(row, "ExtractedPONumber", ""))):
        return "Not posted because no valid PO number is linked."
    return "Not posted because posting rules were not met."


def load_pos_balances_data():
    ensure_expense_review_tables()
    conn = get_sql_connection()
    cursor = conn.cursor()
    req_where, req_params = requestor_filter_sql("l")

    posted_cte = """
        PostedExpenses AS (
            SELECT
                PostedPONumber AS PONumber,
                SUM(COALESCE(PostedAmount, Amount, 0)) AS PostedExpenseAmount,
                COUNT(*) AS PostedExpenseCount
            FROM dbo.ExpenseReviewItems
            WHERE COALESCE(PostedToPO, 0) = 1 AND COALESCE(PostedPONumber, '') <> ''
            GROUP BY PostedPONumber
        )
    """

    unique_po_cte = f"""
        WITH LinePOs AS (
            SELECT
                PONumber,
                MAX(VendorName) AS VendorName,
                MAX(ProjectName) AS ProjectName,
                MAX(Department) AS Department,
                MAX(POStatus) AS POStatus,
                MAX(PODate) AS PODate,
                MAX(COALESCE(RevisedAmount, OriginalAmount, 0)) AS POValue,
                SUM(COALESCE(LineAmount, 0)) AS TotalLineAmount,
                MAX(COALESCE(RemainingAmount, 0)) AS UploadedRemainingAmount,
                COUNT(*) AS LineCount
            FROM dbo.IssuedPOLines l
            WHERE {req_where}
            GROUP BY PONumber
        ),
        {posted_cte},
        UniquePOs AS (
            SELECT
                l.*,
                COALESCE(pe.PostedExpenseAmount, 0) AS PostedExpenseAmount,
                COALESCE(pe.PostedExpenseCount, 0) AS PostedExpenseCount,
                CASE
                    WHEN COALESCE(l.POValue, 0) - COALESCE(pe.PostedExpenseAmount, 0) < 0 THEN 0
                    ELSE COALESCE(l.POValue, 0) - COALESCE(pe.PostedExpenseAmount, 0)
                END AS CurrentAppBalance
            FROM LinePOs l
            LEFT JOIN PostedExpenses pe ON pe.PONumber = l.PONumber
        )
    """

    cursor.execute(
        unique_po_cte + """
        SELECT
            COUNT(*) AS TotalPOs,
            SUM(CASE WHEN UPPER(COALESCE(POStatus, '')) = 'OPEN' THEN 1 ELSE 0 END) AS OpenPOs,
            SUM(POValue) AS TotalPOValue,
            SUM(TotalLineAmount) AS TotalLineAmount,
            SUM(UploadedRemainingAmount) AS UploadedRemainingAmount,
            SUM(PostedExpenseAmount) AS PostedExpenseAmount,
            SUM(CurrentAppBalance) AS TotalRemainingAmount,
            SUM(CASE WHEN ABS(COALESCE(POValue, 0) - COALESCE(TotalLineAmount, 0)) > 0.01 THEN 1 ELSE 0 END) AS AmountMismatchCount
        FROM UniquePOs;
        """,
        *req_params,
    )
    row = cursor.fetchone()
    overall = {
        "total_pos": row.TotalPOs or 0,
        "open_pos": row.OpenPOs or 0,
        "total_po_value": row.TotalPOValue or 0,
        "total_line_amount": row.TotalLineAmount or 0,
        "uploaded_remaining_amount": getattr(row, "UploadedRemainingAmount", 0) or 0,
        "posted_expense_amount": getattr(row, "PostedExpenseAmount", 0) or 0,
        "total_remaining_amount": row.TotalRemainingAmount or 0,
        "amount_mismatch_count": row.AmountMismatchCount or 0,
    }

    cursor.execute(
        unique_po_cte + """
        SELECT
            ProjectName,
            COUNT(*) AS POCount,
            SUM(POValue) AS POValue,
            SUM(TotalLineAmount) AS TotalLineAmount,
            SUM(UploadedRemainingAmount) AS UploadedRemainingAmount,
            SUM(PostedExpenseAmount) AS PostedExpenseAmount,
            SUM(CurrentAppBalance) AS RemainingAmount,
            CASE WHEN SUM(POValue) = 0 THEN 0 ELSE SUM(CurrentAppBalance) / SUM(POValue) END AS PercentOpen
        FROM UniquePOs
        GROUP BY ProjectName
        ORDER BY POValue DESC;
        """,
        *req_params,
    )
    projects = cursor.fetchall()

    cursor.execute(
        unique_po_cte + """
        SELECT TOP 10
            VendorName,
            COUNT(*) AS POCount,
            SUM(POValue) AS POValue,
            SUM(TotalLineAmount) AS TotalLineAmount,
            SUM(UploadedRemainingAmount) AS UploadedRemainingAmount,
            SUM(PostedExpenseAmount) AS PostedExpenseAmount,
            SUM(CurrentAppBalance) AS RemainingAmount
        FROM UniquePOs
        GROUP BY VendorName
        ORDER BY POValue DESC;
        """,
        *req_params,
    )
    vendors = cursor.fetchall()

    cursor.execute(
        unique_po_cte + """
        SELECT
            PONumber,
            VendorName,
            ProjectName,
            Department,
            POStatus,
            PODate,
            LineCount,
            POValue,
            TotalLineAmount,
            UploadedRemainingAmount,
            PostedExpenseAmount,
            PostedExpenseCount,
            CurrentAppBalance AS RemainingAmount,
            CASE WHEN ABS(COALESCE(POValue, 0) - COALESCE(TotalLineAmount, 0)) > 0.01 THEN 1 ELSE 0 END AS AmountMismatch
        FROM UniquePOs
        ORDER BY PODate DESC, PONumber DESC;
        """,
        *req_params,
    )
    pos = cursor.fetchall()

    cursor.execute(
        """
        SELECT TOP 200
            PONumber,
            VendorName,
            ProjectName,
            Department,
            LineDescription,
            Unit,
            UnitCost,
            Qty,
            LineAmount,
            RemainingAmount
        FROM dbo.IssuedPOLines l
        WHERE {req_where}
        ORDER BY CreatedAt DESC, IssuedPOLineId DESC;
        """.format(req_where=req_where),
        *req_params,
    )
    lines = cursor.fetchall()
    conn.close()

    return {
        "overall": overall,
        "projects": projects,
        "vendors": vendors,
        "pos": pos,
        "lines": lines,
    }

def open_balance_bucket_rows(projects):
    definitions = [
        ("Less than 20%", float("-inf"), 0.20, "red"),
        ("20% - 40%", 0.20, 0.40, "orange"),
        ("40% - 60%", 0.40, 0.60, "amber"),
        ("60% - 80%", 0.60, 0.80, "lime"),
        ("80% - 100%", 0.80, float("inf"), "green"),
    ]
    rows = []
    for label, min_value, max_value, tone in definitions:
        rows.append({"label": label, "min": min_value, "max": max_value, "tone": tone, "projects": 0, "open": Decimal("0"), "issued": Decimal("0")})

    for p in projects:
        pct_open = float(p.PercentOpen or 0)
        for bucket in rows:
            if pct_open >= bucket["min"] and pct_open < bucket["max"]:
                bucket["projects"] += 1
                bucket["open"] += Decimal(str(p.RemainingAmount or 0))
                bucket["issued"] += Decimal(str(p.POValue or 0))
                break
    return rows


def render_open_balance_buckets(projects):
    buckets = open_balance_bucket_rows(projects)
    total_projects = sum(b["projects"] for b in buckets) or 1
    html_parts = []
    for bucket in buckets:
        pct_value = min(100, max(0, (bucket["projects"] / total_projects) * 100))
        html_parts.append(f"""
        <div class="project-bucket-item {h(bucket['tone'])}">
            <div class="bucket-label">{h(bucket['label'])}</div>
            <div class="donut" style="--pct:{pct_value};"><div><strong>{bucket['projects']}</strong><span>Projects</span></div></div>
            <div class="bucket-metric"><span>Open Balance</span><strong>{currency(bucket['open'])}</strong></div>
            <div class="bucket-metric"><span>Issued PO</span><strong>{currency(bucket['issued'])}</strong></div>
        </div>
        """)
    return "".join(html_parts)



def parse_payment_schedule_for_forecast(payment_schedule, total_amount):
    """Return dated payment schedule entries from PO Info Review.

    The PO Info Review form saves one line per payment. Lines may be as simple
    as a date, or they may include an amount/percent and a note, separated by
    dashes or other text. If payment dates exist but no amounts are entered,
    the open PO amount is allocated evenly across the dated rows so Forecasting
    bucket totals move with the schedule.
    """
    text = str(payment_schedule or "").strip()
    if not text:
        return []

    total = Decimal(str(total_amount or 0))
    entries = []

    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue

        date_match = re.search(r"(\d{4}-\d{1,2}-\d{1,2}|\d{1,2}/\d{1,2}/\d{2,4})", line)
        schedule_date = clean_date(date_match.group(1)) if date_match else None
        if not schedule_date:
            continue

        # Remove the date from the line before looking for an amount so we do
        # not accidentally treat the year/month/day as the payment amount.
        amount_source = line
        if date_match:
            amount_source = (line[:date_match.start()] + " " + line[date_match.end():]).strip()

        schedule_amount = None
        percent_match = re.search(r"(\d+(?:\.\d+)?)\s*%", amount_source)
        if percent_match:
            try:
                schedule_amount = total * Decimal(percent_match.group(1)) / Decimal("100")
            except Exception:
                schedule_amount = None
        else:
            # Prefer explicit currency-like amounts after the date. This handles
            # entries such as "2026-07-01 - $10,000 - Deposit".
            money_matches = re.findall(r"\$?\s*(\d{1,3}(?:,\d{3})*(?:\.\d+)?|\d+(?:\.\d+)?)", amount_source)
            if money_matches:
                # Avoid tiny row labels like "1" or "2" when a real amount is present.
                candidates = []
                for m in money_matches:
                    val = clean_decimal(m)
                    if val is not None:
                        candidates.append(val)
                larger = [v for v in candidates if abs(v) >= Decimal("10")]
                if larger:
                    schedule_amount = larger[-1]
                elif candidates:
                    schedule_amount = candidates[-1]

        entries.append({
            "date": schedule_date,
            "amount": schedule_amount,
            "description": line,
        })

    if not entries:
        return []

    unspecified = [entry for entry in entries if entry["amount"] is None]
    assigned_total = sum((entry["amount"] or Decimal("0")) for entry in entries)

    if unspecified:
        remaining = total - assigned_total
        if remaining < 0:
            remaining = Decimal("0")
        even_amount = remaining / Decimal(len(unspecified)) if total else Decimal("0")
        for entry in unspecified:
            entry["amount"] = even_amount

    return entries

def load_forecast_data():
    conn = get_sql_connection()
    cursor = conn.cursor()

    cursor.execute(
        """
        SELECT
            RequestNumber AS SourceId,
            'Purchase Request' AS SourceType,
            ProjectName,
            VendorName,
            NeededByDate AS ForecastDate,
            RequestStatus AS Status,
            Priority,
            EstimatedAmount AS Amount,
            RequestTitle AS Description
        FROM dbo.PurchaseRequests
        WHERE EstimatedAmount IS NOT NULL
        ORDER BY NeededByDate ASC, RequestedAt DESC;
        """
    )
    request_rows = cursor.fetchall()

    ensure_po_setup_columns(cursor)
    conn.commit()

    cursor.execute(
        """
        SELECT
            po.PONumber AS SourceId,
            'Open PO Balance' AS SourceType,
            pr.ProjectName,
            v.VendorName,
            po.ExpectedPaymentDate AS ForecastDate,
            po.PaymentType,
            po.PaymentSchedule,
            COALESCE(NULLIF(po.SetupStatus, ''), po.POStatus) AS Status,
            NULL AS Priority,
            COALESCE(
                NULLIF(po.RemainingAmount, 0),
                NULLIF(po.RevisedAmount, 0),
                NULLIF(po.OriginalAmount, 0),
                NULLIF(lines.TotalLineAmount, 0),
                0
            ) AS Amount,
            COALESCE(lines.TotalLineAmount, 0) AS TotalLineAmount,
            CASE
                WHEN COALESCE(po.PaymentSchedule, '') <> '' THEN po.PaymentSchedule
                WHEN po.ExpectedPaymentDate IS NULL THEN 'Open PO balance without expected payment date'
                ELSE 'Open PO balance with expected payment date'
            END AS Description
        FROM dbo.PurchaseOrders po
        LEFT JOIN dbo.Vendors v ON po.VendorId = v.VendorId
        LEFT JOIN dbo.Projects pr ON po.ProjectId = pr.ProjectId
        LEFT JOIN (
            SELECT PONumber, SUM(COALESCE(LineAmount, 0)) AS TotalLineAmount
            FROM dbo.IssuedPOLines
            GROUP BY PONumber
        ) lines ON lines.PONumber = po.PONumber
        WHERE COALESCE(
                NULLIF(po.RemainingAmount, 0),
                NULLIF(po.RevisedAmount, 0),
                NULLIF(po.OriginalAmount, 0),
                NULLIF(lines.TotalLineAmount, 0),
                0
            ) > 0
        ORDER BY
            CASE WHEN po.ExpectedPaymentDate IS NULL AND COALESCE(po.PaymentSchedule, '') = '' THEN 1 ELSE 0 END,
            po.ExpectedPaymentDate ASC,
            COALESCE(
                NULLIF(po.RemainingAmount, 0),
                NULLIF(po.RevisedAmount, 0),
                NULLIF(po.OriginalAmount, 0),
                NULLIF(lines.TotalLineAmount, 0),
                0
            ) DESC;
        """
    )
    po_rows = cursor.fetchall()
    conn.close()

    items = []

    for row in request_rows:
        forecast_date = getattr(row, "ForecastDate", None)
        bucket = forecast_bucket(forecast_date)
        items.append({
            "source_id": row.SourceId,
            "source_type": row.SourceType,
            "project": row.ProjectName,
            "vendor": row.VendorName,
            "forecast_date": forecast_date,
            "status": row.Status,
            "priority": row.Priority,
            "amount": row.Amount or 0,
            "description": row.Description,
            "bucket": bucket,
        })

    for row in po_rows:
        po_amount = Decimal(str(row.Amount or 0))
        schedule_entries = parse_payment_schedule_for_forecast(getattr(row, "PaymentSchedule", ""), po_amount)

        if schedule_entries:
            for idx, entry in enumerate(schedule_entries, start=1):
                forecast_date = entry["date"]
                bucket = forecast_bucket(forecast_date)
                items.append({
                    "source_id": f"{row.SourceId} Pmt {idx}",
                    "source_type": "Scheduled PO Payment",
                    "project": row.ProjectName,
                    "vendor": row.VendorName,
                    "forecast_date": forecast_date,
                    "status": row.Status,
                    "priority": row.Priority,
                    "amount": entry["amount"] or Decimal("0"),
                    "description": entry["description"],
                    "bucket": bucket,
                })
        else:
            forecast_date = getattr(row, "ForecastDate", None)
            bucket = forecast_bucket(forecast_date)
            items.append({
                "source_id": row.SourceId,
                "source_type": row.SourceType,
                "project": row.ProjectName,
                "vendor": row.VendorName,
                "forecast_date": forecast_date,
                "status": row.Status,
                "priority": row.Priority,
                "amount": row.Amount or 0,
                "description": row.Description,
                "bucket": bucket,
            })

    return items


def forecast_bucket(forecast_date):
    if not forecast_date:
        return "Unscheduled"
    if isinstance(forecast_date, datetime):
        forecast_date = forecast_date.date()
    today = date.today()
    delta = (forecast_date - today).days
    if delta < 0:
        return "Past Due"
    if delta <= 7:
        return "Next 7 Days"
    if delta <= 14:
        return "8-14 Days"
    if delta <= 30:
        return "15-30 Days"
    if delta <= 60:
        return "31-60 Days"
    if delta <= 90:
        return "61-90 Days"
    return "90+ Days"


def forecast_bucket_summary(items):
    labels = ["Past Due", "Next 7 Days", "8-14 Days", "15-30 Days", "31-60 Days", "61-90 Days", "90+ Days", "Unscheduled"]
    summary = {label: {"count": 0, "amount": Decimal("0")} for label in labels}
    for item in items:
        label = item["bucket"]
        if label not in summary:
            summary[label] = {"count": 0, "amount": Decimal("0")}
        summary[label]["count"] += 1
        summary[label]["amount"] += Decimal(str(item["amount"] or 0))
    return [(label, summary[label]["count"], summary[label]["amount"]) for label in labels]


def aggregate_items(items, key_name):
    buckets = {}
    for item in items:
        key = item.get(key_name) or "Unassigned"
        if key not in buckets:
            buckets[key] = {"count": 0, "amount": Decimal("0"), "next_date": None}
        buckets[key]["count"] += 1
        buckets[key]["amount"] += Decimal(str(item["amount"] or 0))
        fd = item.get("forecast_date")
        if fd and (buckets[key]["next_date"] is None or fd < buckets[key]["next_date"]):
            buckets[key]["next_date"] = fd
    return sorted(buckets.items(), key=lambda x: x[1]["amount"], reverse=True)


# ------------------------------------------------------------
# Feature phase helpers: PO packets and project setup
# ------------------------------------------------------------

def render_payment_schedule_for_packet(payment_schedule):
    lines = [line.strip() for line in str(payment_schedule or "").splitlines() if line.strip()]
    if not lines:
        return '<div class="empty-state"><strong>No payment schedule entered.</strong><span>Use PO Setup Review to add expected payment dates and schedule details.</span></div>'
    items = "".join(f'<div class="timeline-item"><strong>Payment {idx}</strong><span>{h(line)}</span></div>' for idx, line in enumerate(lines, start=1))
    return f'<div class="timeline">{items}</div>'


def load_po_packet_data(po_number):
    conn = get_sql_connection()
    cursor = conn.cursor()
    req_where, req_params = requestor_filter_sql("po")
    line_req_where, line_req_params = requestor_filter_sql("IssuedPOLines")
    ensure_po_setup_columns(cursor)
    conn.commit()
    cursor.execute(
        """
        SELECT
            po.PONumber,
            v.VendorName,
            pr.ProjectName,
            po.Department,
            po.POStatus,
            po.PODate,
            po.Requestor,
            COALESCE(po.RevisedAmount, po.OriginalAmount, 0) AS POValue,
            COALESCE(lines.TotalLineAmount, COALESCE(po.RevisedAmount, po.OriginalAmount, 0)) AS TotalLineAmount,
            COALESCE(posted.PostedExpenseAmount, 0) AS PostedExpenseAmount,
            COALESCE(posted.PostedExpenseCount, 0) AS PostedExpenseCount,
            CASE WHEN COALESCE(po.RevisedAmount, po.OriginalAmount, 0) - COALESCE(posted.PostedExpenseAmount, 0) < 0 THEN 0 ELSE COALESCE(po.RevisedAmount, po.OriginalAmount, 0) - COALESCE(posted.PostedExpenseAmount, 0) END AS RemainingAmount,
            COALESCE(lines.LineCount, 0) AS LineCount,
            po.PaymentType,
            po.ExpectedPaymentDate,
            po.PaymentSchedule,
            po.SetupStatus,
            po.SetupAssignedTo,
            po.SetupUpdatedBy,
            po.SetupUpdatedAt
        FROM dbo.PurchaseOrders po
        LEFT JOIN dbo.Vendors v ON po.VendorId = v.VendorId
        LEFT JOIN dbo.Projects pr ON po.ProjectId = pr.ProjectId
        LEFT JOIN (
            SELECT PONumber, SUM(COALESCE(LineAmount, 0)) AS TotalLineAmount, COUNT(*) AS LineCount
            FROM dbo.IssuedPOLines
            WHERE PONumber = ? AND {line_req_where}
            GROUP BY PONumber
        ) lines ON lines.PONumber = po.PONumber
        LEFT JOIN (
            SELECT PostedPONumber, SUM(COALESCE(PostedAmount, Amount, 0)) AS PostedExpenseAmount, COUNT(*) AS PostedExpenseCount
            FROM dbo.ExpenseReviewItems
            WHERE COALESCE(PostedToPO, 0) = 1 AND PostedPONumber = ?
            GROUP BY PostedPONumber
        ) posted ON posted.PostedPONumber = po.PONumber
        WHERE po.PONumber = ? AND {req_where};
        """.format(line_req_where=line_req_where, req_where=req_where),
        po_number,
        *line_req_params,
        po_number,
        po_number,
        *req_params,
    )
    po = cursor.fetchone()
    cursor.execute(
        """
        SELECT
            LineDescription,
            Unit,
            UnitCost,
            Qty,
            LineAmount,
            RemainingAmount,
            CreatedAt
        FROM dbo.IssuedPOLines
        WHERE PONumber = ? AND {line_req_where}
        ORDER BY IssuedPOLineId;
        """.format(line_req_where=line_req_where),
        po_number,
        *line_req_params,
    )
    lines = cursor.fetchall()
    cursor.execute(
        """
        SELECT TOP 200 ExpenseReviewItemId, ExpenseId, TxDate, TxType, VendorName, Description,
               Amount, PostedAmount, PostedAt, PostedBy, ReviewDecision, ReviewerEmail, PMComments
        FROM dbo.ExpenseReviewItems
        WHERE COALESCE(PostedToPO, 0) = 1 AND PostedPONumber = ?
        ORDER BY COALESCE(PostedAt, UpdatedAt) DESC, ExpenseReviewItemId DESC;
        """,
        po_number,
    )
    posted_expenses = cursor.fetchall()
    conn.close()
    return po, lines, posted_expenses


def ensure_po_setup_columns(cursor):
    setup_columns = [
        ("PaymentType", "NVARCHAR(100) NULL"),
        ("ExpectedPaymentDate", "DATE NULL"),
        ("PaymentSchedule", "NVARCHAR(MAX) NULL"),
        ("SetupStatus", "NVARCHAR(100) NULL"),
        ("SetupAssignedTo", "NVARCHAR(255) NULL"),
        ("SetupNotes", "NVARCHAR(MAX) NULL"),
        ("SetupUpdatedBy", "NVARCHAR(255) NULL"),
        ("SetupUpdatedAt", "DATETIME2 NULL"),
    ]
    for column_name, column_type in setup_columns:
        cursor.execute(
            f"""
            IF COL_LENGTH('dbo.PurchaseOrders', '{column_name}') IS NULL
            BEGIN
                ALTER TABLE dbo.PurchaseOrders ADD {column_name} {column_type};
            END
            """
        )


def load_project_po_setup_items(status_filter=None, assigned_filter=None):
    conn = get_sql_connection()
    cursor = conn.cursor()
    ensure_po_setup_columns(cursor)
    conn.commit()

    where_clauses = []
    params = []
    req_where, req_params = requestor_filter_sql("po")
    if req_where != "1=1":
        where_clauses.append(req_where)
        params.extend(req_params)

    if status_filter and status_filter != "All":
        where_clauses.append("COALESCE(NULLIF(po.SetupStatus, ''), CASE WHEN COALESCE(po.PaymentSchedule, '') = '' THEN 'Needs Payment Schedule' ELSE 'Complete' END) = ?")
        params.append(status_filter)

    if assigned_filter:
        where_clauses.append("LOWER(COALESCE(po.SetupAssignedTo, '')) = LOWER(?)")
        params.append(assigned_filter)

    if not where_clauses:
        where_clauses.append("(COALESCE(NULLIF(po.SetupStatus, ''), CASE WHEN COALESCE(po.PaymentSchedule, '') = '' THEN 'Needs Payment Schedule' ELSE 'Complete' END) <> 'Complete' OR COALESCE(po.PaymentSchedule, '') = '' OR COALESCE(po.PaymentType, '') = '' OR po.ExpectedPaymentDate IS NULL OR COALESCE(v.VendorName, '') = '')")

    where_sql = " AND ".join(where_clauses)

    cursor.execute(
        f"""
        SELECT TOP 250
            po.PurchaseOrderId,
            po.PONumber,
            v.VendorName,
            pr.ProjectName,
            po.Department,
            po.POStatus,
            po.PODate,
            po.Requestor,
            COALESCE(po.RevisedAmount, po.OriginalAmount, 0) AS POValue,
            po.RemainingAmount,
            po.PaymentType,
            po.ExpectedPaymentDate,
            po.PaymentSchedule,
            COALESCE(NULLIF(po.SetupStatus, ''), CASE WHEN COALESCE(po.PaymentSchedule, '') = '' THEN 'Needs Payment Schedule' ELSE 'Complete' END) AS SetupStatus,
            po.SetupAssignedTo,
            po.SetupNotes,
            po.SetupUpdatedBy,
            po.SetupUpdatedAt,
            CASE WHEN COALESCE(po.PaymentSchedule, '') = '' THEN 1 ELSE 0 END AS MissingPaymentSchedule,
            CASE WHEN COALESCE(po.PaymentType, '') = '' THEN 1 ELSE 0 END AS MissingPaymentType,
            CASE WHEN po.ExpectedPaymentDate IS NULL THEN 1 ELSE 0 END AS MissingExpectedPaymentDate,
            CASE WHEN COALESCE(v.VendorName, '') = '' THEN 1 ELSE 0 END AS MissingVendor
        FROM dbo.PurchaseOrders po
        LEFT JOIN dbo.Vendors v ON po.VendorId = v.VendorId
        LEFT JOIN dbo.Projects pr ON po.ProjectId = pr.ProjectId
        WHERE {where_sql}
        ORDER BY
            CASE COALESCE(NULLIF(po.SetupStatus, ''), CASE WHEN COALESCE(po.PaymentSchedule, '') = '' THEN 'Needs Payment Schedule' ELSE 'Complete' END)
                WHEN 'Needs Payment Schedule' THEN 0
                WHEN 'Assigned to PM' THEN 1
                WHEN 'In Progress' THEN 2
                WHEN 'Needs Info' THEN 3
                ELSE 4
            END,
            pr.ProjectName,
            po.PONumber;
        """,
        *params,
    )
    rows = cursor.fetchall()
    conn.close()
    return rows


def count_po_setup_actions_for_current_user():
    user = get_current_user()
    view_email = current_working_email()
    access = get_user_access()
    conn = get_sql_connection()
    cursor = conn.cursor()
    ensure_po_setup_columns(cursor)
    conn.commit()

    if get_view_as_email():
        cursor.execute(
            """
            SELECT COUNT(*) AS ActionCount
            FROM dbo.PurchaseOrders
            WHERE LOWER(COALESCE(SetupAssignedTo, '')) = LOWER(?)
              AND COALESCE(NULLIF(SetupStatus, ''), 'Needs Payment Schedule') <> 'Complete';
            """,
            view_email,
        )
    elif normalize_role(access["role"]) == "Project Manager - Diving" and user["email"]:
        cursor.execute(
            """
            SELECT COUNT(*) AS ActionCount
            FROM dbo.PurchaseOrders
            WHERE LOWER(COALESCE(SetupAssignedTo, '')) = LOWER(?)
              AND COALESCE(NULLIF(SetupStatus, ''), 'Needs Payment Schedule') <> 'Complete';
            """,
            user["email"],
        )
    elif normalize_role(access["role"]) in ["Admin", "Executive"]:
        cursor.execute(
            """
            SELECT COUNT(*) AS ActionCount
            FROM dbo.PurchaseOrders
            WHERE COALESCE(NULLIF(SetupStatus, ''), CASE WHEN COALESCE(PaymentSchedule, '') = '' THEN 'Needs Payment Schedule' ELSE 'Complete' END) <> 'Complete'
               OR COALESCE(PaymentSchedule, '') = ''
               OR COALESCE(PaymentType, '') = ''
               OR ExpectedPaymentDate IS NULL;
            """
        )
    else:
        conn.close()
        return 0

    row = cursor.fetchone()
    conn.close()
    return int(row.ActionCount or 0) if row else 0


def update_po_setup_info(form):
    user = get_current_user()
    po_number = clean_text(form.get("po_number"))
    if not po_number:
        raise ValueError("PO Number is required.")

    payment_type = clean_text(form.get("payment_type")) or "Single Payment"
    setup_status = clean_text(form.get("setup_status")) or "Needs Payment Schedule"
    setup_assigned_to = clean_text(form.get("setup_assigned_to"))
    vendor_name = clean_text(form.get("vendor_name"))
    action = clean_text(form.get("setup_action")) or "save"

    schedule_lines = []
    first_expected_payment_date = None
    for idx in range(1, 5):
        date_raw = clean_text(form.get(f"payment_{idx}_date"))
        amount_raw = clean_text(form.get(f"payment_{idx}_amount"))
        note_raw = clean_text(form.get(f"payment_{idx}_note"))
        date_value = clean_date(date_raw)
        if date_value and first_expected_payment_date is None:
            first_expected_payment_date = date_value
        if date_raw or amount_raw or note_raw:
            parts = []
            if date_raw:
                parts.append(date_raw)
            if amount_raw:
                parts.append(amount_raw)
            if note_raw:
                parts.append(note_raw)
            schedule_lines.append(" - ".join(parts))

    payment_schedule = "\n".join(schedule_lines) or clean_text(form.get("payment_schedule"))
    expected_payment_date = first_expected_payment_date or clean_date(form.get("expected_payment_date"))

    if action == "assign":
        if not setup_assigned_to:
            raise ValueError("Assigned To is required when assigning this PO.")
        setup_status = "Assigned to PM"

    if action == "complete":
        setup_status = "Complete"

    conn = get_sql_connection()
    cursor = conn.cursor()
    try:
        ensure_po_setup_columns(cursor)
        cursor.execute("""
            SELECT po.PONumber, v.VendorName
            FROM dbo.PurchaseOrders po
            LEFT JOIN dbo.Vendors v ON po.VendorId = v.VendorId
            WHERE po.PONumber = ?;
        """, po_number)
        current_po = cursor.fetchone()
        existing_vendor = clean_text(current_po.VendorName) if current_po else None
        effective_vendor = vendor_name or existing_vendor
        if setup_status == "Complete":
            missing = []
            if not effective_vendor:
                missing.append("vendor")
            if not payment_type:
                missing.append("payment type")
            if not payment_schedule:
                missing.append("payment schedule")
            if not expected_payment_date:
                missing.append("expected payment date")
            if missing:
                raise ValueError("Cannot mark Complete. Missing: " + ", ".join(missing) + ".")
        vendor_id = get_or_create_vendor(cursor, vendor_name) if vendor_name else None
        cursor.execute(
            """
            UPDATE dbo.PurchaseOrders
            SET
                VendorId = COALESCE(?, VendorId),
                PaymentType = ?,
                ExpectedPaymentDate = ?,
                PaymentSchedule = ?,
                SetupStatus = ?,
                SetupAssignedTo = ?,
                SetupUpdatedBy = ?,
                SetupUpdatedAt = SYSUTCDATETIME(),
                UpdatedAt = SYSUTCDATETIME()
            WHERE PONumber = ?;
            """,
            vendor_id,
            payment_type,
            expected_payment_date,
            payment_schedule,
            setup_status,
            setup_assigned_to,
            user["email"] or "Unknown",
            po_number,
        )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def action_form(request_id, status, label, css_class="secondary", extra_fields=""):
    return f"""
    <form method="post">
        <input type="hidden" name="purchase_request_id" value="{h(request_id)}">
        <input type="hidden" name="request_status" value="{h(status)}">
        {extra_fields}
        <input type="hidden" name="review_notes" value="Quick action: {h(label)}">
        <button class="{h(css_class)}" type="submit">{h(label)}</button>
    </form>
    """

# ------------------------------------------------------------
# Routes
# ------------------------------------------------------------

@app.route("/")
def home():
    return redirect("/my-dashboard")



@app.route("/help-center")
def help_center():
    allowed, reason = require_page_access("Help Center")
    if not allowed:
        return access_denied_response("Help Center", reason)

    access = get_effective_user_access()
    role = access.get("role", "")
    can_submit_requests = role_can_access(role, "New Purchase Request")

    purchase_request_section = """
        <div class="help-step">
            <div class="step-number">1</div>
            <div><strong>Open New Purchase Request.</strong><br>Use the left menu and click <strong>New Purchase Request</strong>.</div>
        </div>
        <div class="help-step">
            <div class="step-number">2</div>
            <div><strong>Enter what you need.</strong><br>Add a clear request title, description/scope, project, department, needed-by date, vendor if known, and estimated cost.</div>
        </div>
        <div class="help-step">
            <div class="step-number">3</div>
            <div><strong>Attach backup.</strong><br>Upload a quote, estimate, email backup, or other supporting document when available.</div>
        </div>
        <div class="help-step">
            <div class="step-number">4</div>
            <div><strong>Submit for review.</strong><br>After submitting, the request goes to the proper review queue. You can check status from <strong>Purchase Requests</strong>.</div>
        </div>
    """ if can_submit_requests else """
        <div class="soft-alert">
            Your current role does not include purchase request submission. You can still use this help center to look up POs and view project PO information if your role allows it.
        </div>
    """

    content = f"""
    <style>
      .help-hero {{ background: linear-gradient(135deg, rgba(37,99,235,.12), rgba(14,165,233,.08)); border:1px solid rgba(37,99,235,.18); border-radius:20px; padding:24px; margin-bottom:18px; }}
      .help-grid {{ display:grid; grid-template-columns: repeat(3, minmax(0, 1fr)); gap:16px; }}
      .help-card {{ background:white; border:1px solid #dbe7f3; border-radius:18px; padding:20px; box-shadow:0 12px 28px rgba(15,23,42,.06); }}
      .help-card h3 {{ margin-top:0; color:#0f172a; }}
      .help-step {{ display:flex; gap:12px; padding:12px 0; border-top:1px solid #edf2f7; }}
      .help-step:first-of-type {{ border-top:0; }}
      .step-number {{ flex:0 0 30px; width:30px; height:30px; border-radius:999px; background:#2563eb; color:white; display:flex; align-items:center; justify-content:center; font-weight:800; }}
      .quick-links {{ display:flex; gap:10px; flex-wrap:wrap; margin-top:14px; }}
      .soft-alert {{ background:#fff7ed; border:1px solid #fed7aa; color:#9a3412; border-radius:14px; padding:14px; margin:10px 0; font-weight:700; }}
      .tip-list {{ margin:8px 0 0 18px; padding:0; color:#475569; line-height:1.55; }}
      @media (max-width: 1100px) {{ .help-grid {{ grid-template-columns:1fr; }} }}
    </style>

    <div class="help-hero">
      <h2 style="margin:0 0 8px;">Coastal Command Center Help Center</h2>
      <p style="margin:0; color:#475569; max-width:900px;">Quick how-to guides for the most common July 1 rollout tasks: finding a PO, submitting a purchase request, and viewing PO information by project.</p>
      <div class="quick-links">
        <a class="button" href="/pos-balances">Find a PO</a>
        {('<a class="button" href="/purchase-request">Submit Purchase Request</a>' if can_submit_requests else '')}
        <a class="button" href="/projects">View by Project</a>
      </div>
    </div>

    <div class="help-grid">
      <div class="help-card" id="find-po">
        <h3>🔎 How to find a PO</h3>
        <div class="help-step">
          <div class="step-number">1</div>
          <div><strong>Open POs & Balances.</strong><br>Use the left menu and click <strong>POs & Balances</strong>.</div>
        </div>
        <div class="help-step">
          <div class="step-number">2</div>
          <div><strong>Use the filters/search.</strong><br>Search by PO number, vendor, project, department, or status.</div>
        </div>
        <div class="help-step">
          <div class="step-number">3</div>
          <div><strong>Open the PO packet.</strong><br>Click the PO number to view the full PO packet, line details, balance, vendor information, and related history.</div>
        </div>
        <ul class="tip-list">
          <li>Your role may limit which POs appear.</li>
          <li>Voided POs remain visible but show a $0.00 amount/balance.</li>
        </ul>
      </div>

      <div class="help-card" id="submit-request">
        <h3>📝 How to submit a purchase request</h3>
        {purchase_request_section}
        <ul class="tip-list">
          <li>Use a specific request title so approvers understand what is needed.</li>
          <li>Estimated cost is required for review routing.</li>
          <li>The selected project must already exist in the app/project setup list.</li>
        </ul>
      </div>

      <div class="help-card" id="project-info">
        <h3>📁 How to view PO information by project</h3>
        <div class="help-step">
          <div class="step-number">1</div>
          <div><strong>Open Projects.</strong><br>Use the left menu and click <strong>Projects</strong>.</div>
        </div>
        <div class="help-step">
          <div class="step-number">2</div>
          <div><strong>Select a project.</strong><br>Choose the project from the <strong>Select Project</strong> dropdown.</div>
        </div>
        <div class="help-step">
          <div class="step-number">3</div>
          <div><strong>Review the project PO summary.</strong><br>Use the project view to see project-level PO totals, open POs, vendors, and current balances.</div>
        </div>
        <div class="help-step">
          <div class="step-number">4</div>
          <div><strong>Drill into line items.</strong><br>Scroll to the PO line-item section to see what was uploaded for that project.</div>
        </div>
        <ul class="tip-list">
          <li>Dredging-only users will only see Dredging projects.</li>
          <li>Diving project managers may only see projects assigned to them.</li>
        </ul>
      </div>
    </div>
    """
    return shell("Help Center", "How-to guides for common Command Center tasks.", "Help Center", content)

@app.route("/purchase-request", methods=["GET", "POST"])
def purchase_request():
    allowed, reason = require_page_access("New Purchase Request")
    if not allowed:
        return access_denied_response("New Purchase Request", reason)

    user = get_current_user()
    access = get_user_access()
    display_name = access["display_name"] or user["email"] or "Current User"
    role = access["role"]
    project_options = load_existing_project_options()
    project_select_options = '<option value="">Select existing project</option>' + ''.join(
        f'<option value="{h(opt["value"])}">{h(opt["label"])}</option>' for opt in project_options
    )
    no_project_notice = ""
    if not project_options:
        no_project_notice = '<div class="notice warning"><strong>No projects are available for purchase requests yet.</strong><br>Upload issued POs for the project setup first, then return to this form.</div>'

    message_html = ""

    if request.method == "POST":
        try:
            result = create_purchase_request(request.form, request.files.getlist("quote_files"))
            message_html = f"""
            <div class="submit-status-box success">
                <strong>Purchase request submitted successfully.</strong><br>
                Request Number: {h(result["request_number"])}
            </div>
            """
        except Exception as e:
            message_html = f"""
            <div class="submit-status-box error">
                <strong>Error submitting purchase request.</strong><br>{h(e)}
            </div>
            """

    content = f"""
    {message_html}
    {no_project_notice}

    <form method="post" action="/purchase-request" enctype="multipart/form-data">
        <div class="grid two">
            <div>
                <div class="card">
                    <h3>Request Basics</h3>
                    <p class="card-subtitle">Start with a clear request name and short scope. These details help reviewers understand what is needed.</p>
                    <div class="form-grid">
                        <div class="form-field full">
                            <label>What are you requesting? *</label>
                            <input type="text" name="request_title" placeholder="Example: Pump rental extension for Round Valley" required>
                        </div>
                        <div class="form-field full">
                            <label>Description / Scope *</label>
                            <textarea name="request_description" placeholder="Example: Extend pump rental for two additional weeks due to schedule delay." required></textarea>
                        </div>
                    </div>
                </div>

                <div class="card">
                    <h3>Project Details</h3>
                    <p class="card-subtitle">Tell us which project and department this request belongs to.</p>
                    <div class="form-grid">
                        <div class="form-field">
                            <label>Project *</label>
                            <select name="project_value" required>
                                {project_select_options}
                            </select>
                            <p class="field-help">Projects must already exist from issued PO project setup. This form cannot create new projects.</p>
                        </div>
                        <div class="form-field">
                            <label>Department *</label>
                            <select name="department" required>
                                <option value="">Select department</option>
                                <option value="Engineering">Engineering</option>
                                <option value="Marine Construction">Marine Construction</option>
                                <option value="Commercial Diving">Commercial Diving</option>
                                <option value="Dredging">Dredging</option>
                                <option value="Marine Services">Marine Services</option>
                            </select>
                        </div>
                        <div class="form-field">
                            <label>Needed By *</label>
                            <input type="date" name="needed_by_date" required>
                        </div>
                        <div class="form-field">
                            <label>Requested By</label>
                            <input type="text" name="requested_by" value="{h(display_name)}" readonly>
                        </div>
                    </div>
                </div>

                <div class="card">
                    <h3>Vendor & Cost</h3>
                    <p class="card-subtitle">Vendor can be left blank if the requester does not know it yet. Estimated cost is required for review routing.</p>
                    <div class="form-grid">
                        <div class="form-field">
                            <label>Vendor</label>
                            <input type="text" name="vendor_name" placeholder="United Rentals, Grainger, Home Depot, or leave blank if unknown">
                        </div>
                        <div class="form-field">
                            <label>Estimated Cost *</label>
                            <input type="number" name="estimated_amount" step="0.01" min="0" placeholder="8500" required>
                        </div>
                        <div class="form-field">
                            <label>Priority</label>
                            <select name="priority">
                                <option value="Normal">Normal</option>
                                <option value="Low">Low</option>
                                <option value="High">High</option>
                                <option value="Urgent">Urgent</option>
                                <option value="Critical">Critical</option>
                            </select>
                        </div>
                        <div class="form-field">
                            <label>Upload Quote / Backup</label>
                            <input type="file" name="quote_files" multiple accept=".pdf,.png,.jpg,.jpeg,.doc,.docx,.xls,.xlsx,.csv,.txt,.eml,.msg">
                            <p class="field-help">Files are stored with the request and linked to the PO when approved.</p>
                        </div>
                    </div>
                </div>

                <div class="card">
                    <h3>Review & Submit</h3>
                    <p class="card-subtitle">Your request will be saved and routed for review. Accounting or management may update the status or request more information.</p>
                    <div class="match-summary">
                        <strong>Required field checklist</strong>
                        <span>• What are you requesting?</span>
                        <span>• Description / Scope</span>
                        <span>• Existing project selection</span>
                        <span>• Department</span>
                        <span>• Needed By</span>
                        <span>• Estimated Cost</span>
                    </div>
                    <div class="request-actions">
                        <a class="button" href="/my-dashboard">Cancel</a>
                        <button class="primary" type="submit" {"" if project_options else "disabled"}>Submit Purchase Request</button>
                    </div>
                </div>
            </div>

            <div>
                <div class="card role-card">
                    <h3>Requester</h3>
                    <p class="card-subtitle">Pulled from Microsoft login and dashboard role access.</p>
                    <div class="role-meta">
                        <span><strong>User:</strong> {h(display_name)}</span>
                        <span><strong>Role:</strong> {h(role)}</span>
                        <span><strong>Email:</strong> {h(user["email"])}</span>
                    </div>
                </div>

                <div class="card">
                    <h3>Estimated Approval Route</h3>
                    <p class="card-subtitle">Routing can become more detailed later based on dollar amount, department, and project.</p>
                    <div class="workflow">
                        <div class="workflow-step done"><div class="workflow-circle">1</div><div class="workflow-text"><strong>Submitted</strong><span>Request is created by the requester.</span></div></div>
                        <div class="workflow-line"></div>
                        <div class="workflow-step active"><div class="workflow-circle">2</div><div class="workflow-text"><strong>Review</strong><span>Accounting, Admin, or management validates the request.</span></div></div>
                        <div class="workflow-line"></div>
                        <div class="workflow-step warning"><div class="workflow-circle">3</div><div class="workflow-text"><strong>Decision</strong><span>Approve, reject, or request more information.</span></div></div>
                        <div class="workflow-line"></div>
                        <div class="workflow-step future"><div class="workflow-circle">4</div><div class="workflow-text"><strong>PO Follow-Up</strong><span>Approved requests can later be converted to a PO.</span></div></div>
                    </div>
                </div>

                <div class="card">
                    <h3>Cost Guidance</h3>
                    <table>
                        <tr><th>Estimated Cost</th><th>Likely Review</th></tr>
                        <tr><td>Under $500</td><td><span class="badge green">Quick review</span></td></tr>
                        <tr><td>Under $3,000</td><td><span class="badge blue">Admin approval</span></td></tr>
                        <tr><td>$3,000 and up</td><td><span class="badge purple">Admin + Executive approval</span></td></tr>
                    </table>
                </div>

            </div>
        </div>
    </form>
    """

    return shell("New Purchase Request", "Submit a new purchase request for review before a PO is issued.", "New Purchase Request", content)



@app.route("/purchase-requests", methods=["GET", "POST"])
def purchase_requests():
    allowed, reason = require_page_access("Purchase Requests")
    if not allowed:
        return access_denied_response("Purchase Requests", reason)

    access = get_user_access()
    role = access["role"]

    if request.method == "POST":
        if not can_review_purchase_requests(role):
            return redirect("/purchase-requests?toast=You+do+not+have+permission+to+update+purchase+requests&toast_type=error")
        try:
            update_purchase_request_status(request.form)
            return redirect("/purchase-requests?toast=Purchase+request+updated")
        except Exception as e:
            return redirect("/purchase-requests?toast=" + quote_plus("Error updating request: " + str(e)) + "&toast_type=error")

    try:
        stats = load_purchase_request_stats()
        requests = load_purchase_requests()
        reviewer_users = load_assignable_users()
        selected_status = clean_text(request.args.get("status")) or "All"

        def status_card(label, value, status_filter, tone, trend):
            active = " active" if selected_status == status_filter else ""
            href = "/purchase-requests" if status_filter == "All" else "/purchase-requests?status=" + quote_plus(status_filter)
            return f"""
            <a class="status-card {tone}{active}" href="{href}">
                <div class="label">{h(label)}</div>
                <div class="value">{h(value)}</div>
                <div class="trend">{h(trend)}</div>
            </a>
            """

        manual_review_notice = ""

        dashboard_cards = f"""
        <div class="status-card-grid">
            {status_card("All Requests", stats["total_requests"], "All", "blue", "Full request queue")}
            {status_card("Submitted", stats["submitted_requests"], "Submitted", "amber", "Waiting for action")}
            {status_card("Under Review", stats["under_review_requests"], "Under Review", "purple", "Currently being reviewed")}
            {status_card("Needs More Info", stats.get("needs_more_info_requests", 0), "Needs More Info", "amber", "Returned for clarification")}
            {status_card("Pending Admin", stats.get("pending_admin_requests", 0), "Pending Admin Approval", "amber", "Admin approval needed")}
            {status_card("Pending Executive", stats.get("pending_executive_requests", 0), "Pending Executive Approval", "purple", "Executive approval needed")}
            {status_card("Approved", stats["approved_requests"], "Approved", "green", "Ready for PO action")}
            {status_card("Converted", stats["converted_requests"], "Converted to PO", "green", "Linked to issued PO")}
        </div>
        """

        request_rows = ""
        visible_requests = requests if selected_status == "All" else [r for r in requests if (r.RequestStatus or "Submitted") == selected_status]

        for row in visible_requests:
            description = row.RequestDescription or ""
            if len(description) > 160:
                description = description[:160] + "..."

            status_options = ""
            for status in ["Submitted", "Under Review", "Needs More Info", "Pending Admin Approval", "Pending Executive Approval", "Approved", "Rejected", "Converted to PO"]:
                selected = " selected" if status == row.RequestStatus else ""
                status_options += f'<option value="{h(status)}"{selected}>{h(status)}</option>'

            review_form = ""
            if can_review_purchase_requests(role):
                reviewer_value = clean_text(row.ReviewerEmail) or access["email"]
                reviewer_options = ""
                seen_reviewers = set()
                for reviewer in reviewer_users:
                    reviewer_email = clean_text(reviewer.Email)
                    if not reviewer_email:
                        continue
                    reviewer_label = clean_text(reviewer.DisplayName) or reviewer_email
                    reviewer_role = clean_text(reviewer.RoleName)
                    label = reviewer_label + (f" ({reviewer_role})" if reviewer_role else "")
                    selected = " selected" if reviewer_email.lower() == reviewer_value.lower() else ""
                    reviewer_options += f'<option value="{h(reviewer_email)}"{selected}>{h(label)}</option>'
                    seen_reviewers.add(reviewer_email.lower())
                if reviewer_value and reviewer_value.lower() not in seen_reviewers:
                    reviewer_options = f'<option value="{h(reviewer_value)}" selected>{h(reviewer_value)}</option>' + reviewer_options
                converted_display = "block" if (row.RequestStatus or "Submitted") == "Converted to PO" else "none"
                review_form = f"""
                <form method="post" action="/purchase-requests" class="purchase-review-panel">
                    <input type="hidden" name="purchase_request_id" value="{h(row.PurchaseRequestId)}">
                    <label>Status</label>
                    <select name="request_status" onchange="toggleConvertedPOField(this)">{status_options}</select>
                    <small class="field-help">Selecting Approved records your approval. Under $3,000 converts after Admin approval; $3,000+ converts after both Admin and Executive approvals.</small>
                    <label>Reviewer</label>
                    <select name="reviewer_email">{reviewer_options}</select>
                    <div class="converted-po-field" style="display:{converted_display};">
                        <label>Converted PO Number</label>
                        <input type="text" name="converted_po_number" value="{h(row.ConvertedPONumber)}" placeholder="Leave blank to auto-generate">
                    </div>
                    <label>Review Notes</label>
                    <textarea name="review_notes" placeholder="Add review notes, approval comments, rejection reason, or follow-up needed">{h(row.ReviewNotes)}</textarea>
                    <button type="submit">Save Review</button>
                </form>
                """

            request_rows += f"""
            <tr>
                <td><strong>{h(row.RequestNumber)}</strong><br><span style="color:var(--muted);">{h(row.RequestedAt)}</span></td>
                <td><strong>{h(row.RequestTitle)}</strong><br><span style="color:var(--muted);">{h(description)}</span></td>
                <td><a class="vendor-detail-link" href="/vendors?vendor={quote_plus(str(row.VendorName or ''))}">{h(row.VendorName)}</a></td>
                <td><a class="vendor-detail-link" href="/projects?project={quote_plus(str(row.ProjectName or ''))}">{h(row.ProjectName)}</a></td>
                <td>{h(row.Department)}</td>
                <td>{h(row.NeededByDate)}</td>
                <td class="right">{currency(row.EstimatedAmount)}</td>
                <td>{h(row.Priority)}</td>
                <td>{purchase_request_status_badge(row.RequestStatus)}</td>
                <td>{h(row.RequestedByName or row.RequestedByEmail)}</td>
                <td>{review_form}</td>
            </tr>
            """

        if not request_rows:
            request_rows = '<tr><td colspan="11"><div class="empty-state"><strong>No requests found for this filter.</strong><span>Use another status card or clear the filter.</span></div></td></tr>'

        total_requests = max(stats["total_requests"], 1)
        content = f"""
        {manual_review_notice}
        {dashboard_cards}

        <div class="grid two visual-chart-row" style="margin-bottom:24px;">
            <div class="mini-chart-card">
                <h4>Request Status Mix</h4>
                <div class="mini-bar-row"><span>Submitted</span><div><b style="width:{max(5, stats["submitted_requests"] / total_requests * 100)}%"></b></div><em>{stats["submitted_requests"]}</em></div>
                <div class="mini-bar-row"><span>Under Review</span><div><b style="width:{max(5, stats["under_review_requests"] / total_requests * 100)}%"></b></div><em>{stats["under_review_requests"]}</em></div>
                <div class="mini-bar-row"><span>Needs More Info</span><div><b style="width:{max(5, stats.get("needs_more_info_requests",0) / total_requests * 100)}%"></b></div><em>{stats.get("needs_more_info_requests",0)}</em></div>
                <div class="mini-bar-row"><span>Approved</span><div><b style="width:{max(5, stats["approved_requests"] / total_requests * 100)}%"></b></div><em>{stats["approved_requests"]}</em></div>
                <div class="mini-bar-row"><span>Converted</span><div><b style="width:{max(5, stats["converted_requests"] / total_requests * 100)}%"></b></div><em>{stats["converted_requests"]}</em></div>
            </div>
            <div class="card">
                <h3>Request Workflow Timeline</h3>
                <p class="card-subtitle">Simple manual workflow; automatic approval routing is intentionally not included in this version.</p>
                <div class="timeline">
                    <div class="timeline-item"><strong>Submitted</strong><span>Requester submits required details.</span></div>
                    <div class="timeline-item"><strong>Under Review / Needs More Info</strong><span>Accounting or management reviews and follows up.</span></div>
                    <div class="timeline-item"><strong>Approved</strong><span>Request is ready for PO action.</span></div>
                    <div class="timeline-item"><strong>Converted to PO</strong><span>PO number is linked when issued.</span></div>
                </div>
            </div>
        </div>

        <div class="card">
            <h3>Request Dashboard</h3>
            <p class="card-subtitle">Current status filter: <strong>{h(selected_status)}</strong>. Click a status card above to filter the queue.</p>
            <div class="filter-hint"><span>Filter by request number, title, vendor, project, department, status, requester, or other visible text.</span><button type="button" onclick="clearRequestDashboardFilters()">Clear Column Filters</button></div>
            <div class="table-wrap"><table id="requestDashboardTable"><thead><tr><th>Request #</th><th>Title / Description</th><th>Vendor</th><th>Project</th><th>Department</th><th>Needed By</th><th class="right">Estimate</th><th>Priority</th><th>Status</th><th>Requested By</th><th>Review</th></tr><tr class="column-filter-row"><th><input data-col="0" oninput="filterRequestDashboard()" placeholder="Request"></th><th><input data-col="1" oninput="filterRequestDashboard()" placeholder="Title"></th><th><input data-col="2" oninput="filterRequestDashboard()" placeholder="Vendor"></th><th><input data-col="3" oninput="filterRequestDashboard()" placeholder="Project"></th><th><input data-col="4" oninput="filterRequestDashboard()" placeholder="Dept"></th><th><input data-col="5" oninput="filterRequestDashboard()" placeholder="Date"></th><th><input data-col="6" oninput="filterRequestDashboard()" placeholder="Estimate"></th><th><input data-col="7" oninput="filterRequestDashboard()" placeholder="Priority"></th><th><input data-col="8" oninput="filterRequestDashboard()" placeholder="Status"></th><th><input data-col="9" oninput="filterRequestDashboard()" placeholder="Requester"></th><th><input data-col="10" oninput="filterRequestDashboard()" placeholder="Review"></th></tr></thead><tbody>{request_rows}</tbody></table></div>
        </div>
        <script>
        function filterRequestDashboard() {{
            const table = document.getElementById('requestDashboardTable'); if (!table) return;
            const filters = Array.from(table.querySelectorAll('.column-filter-row input')).map(input => {{ return {{ col: Number(input.dataset.col), value: input.value.trim().toLowerCase() }}; }});
            const rows = Array.from(table.querySelectorAll('tbody tr'));
            rows.forEach(row => {{ const cells = Array.from(row.children); const show = filters.every(filter => {{ if (!filter.value) return true; const cell = cells[filter.col]; return cell && cell.textContent.toLowerCase().includes(filter.value); }}); row.style.display = show ? '' : 'none'; }});
        }}
        function clearRequestDashboardFilters() {{ const table = document.getElementById('requestDashboardTable'); if (!table) return; table.querySelectorAll('.column-filter-row input').forEach(input => input.value = ''); filterRequestDashboard(); }}
        function toggleConvertedPOField(select) {{
            const form = select.closest('form');
            if (!form) return;
            const field = form.querySelector('.converted-po-field');
            if (!field) return;
            const input = field.querySelector('input');
            const show = select.value === 'Converted to PO';
            field.style.display = show ? 'block' : 'none';
            if (!show && input) input.value = '';
        }}
        document.querySelectorAll('.purchase-review-panel select[name="request_status"]').forEach(toggleConvertedPOField);
        </script>
        """

        return shell("Purchase Requests", "Review submitted purchase requests before they become issued POs.", "Purchase Requests", content)

    except Exception as e:
        content = f'<div class="notice error">Error loading purchase requests: {h(e)}</div>'
        return shell("Purchase Requests", "Unable to load purchase requests.", "Purchase Requests", content), 500


@app.route("/my-dashboard")
def my_dashboard():
    allowed, reason = require_page_access("My Dashboard")
    if not allowed:
        return access_denied_response("My Dashboard", reason)

    access = get_user_access()
    role = normalize_role(access["role"])
    display_name = access["display_name"] or access["email"] or "User"

    try:
        data = load_personal_dashboard_data()
        overall = data["overall"]
        pr_stats = load_purchase_request_stats()
        po_setup_action_count = count_po_setup_actions_for_current_user()

        def safe_num(value):
            try:
                return float(value or 0)
            except Exception:
                return 0.0

        total_value = safe_num(overall.get("total_po_value"))
        total_line = safe_num(overall.get("total_line_amount"))
        total_remaining = safe_num(overall.get("total_remaining_amount"))
        committed_amount = max(total_line - total_remaining, 0)
        utilization = (committed_amount / total_line * 100) if total_line else 0
        open_ratio = (safe_num(overall.get("open_pos")) / safe_num(overall.get("total_pos")) * 100) if safe_num(overall.get("total_pos")) else 0
        avg_po_size = (total_value / safe_num(overall.get("total_pos"))) if safe_num(overall.get("total_pos")) else 0

        vendor_rows = ""
        max_vendor_amount = max([safe_num(getattr(row, "TotalLineAmount", 0)) for row in data["top_vendors"]] + [1])
        for row in data["top_vendors"]:
            vendor_name = clean_text(row.VendorName) or "Missing Vendor"
            vendor_url = "/vendors?vendor=" + quote_plus(vendor_name)
            amount = safe_num(row.TotalLineAmount)
            pct = min(100, int((amount / max_vendor_amount) * 100)) if max_vendor_amount else 0
            vendor_rows += f"""
            <tr onclick=\"window.location='{vendor_url}'\" class=\"click-row\">
                <td><a href=\"{vendor_url}\">{h(vendor_name)}</a><div class=\"mini-track\"><span style=\"width:{pct}%\"></span></div></td>
                <td class=\"right\">{row.POCount}</td>
                <td class=\"right strong\">{currency(amount)}</td>
            </tr>"""
        if not vendor_rows:
            vendor_rows = '<tr><td colspan="3">No vendor data found.</td></tr>'

        project_rows = ""
        project_cards = ""
        max_project_amount = max([safe_num(getattr(row, "TotalLineAmount", 0)) for row in data["top_projects"]] + [1])
        donut_colors = ["#16a34a", "#84cc16", "#f59e0b", "#f97316", "#ef4444"]
        for idx, row in enumerate(data["top_projects"]):
            project_name = clean_text(row.ProjectName) or "Unnamed Project"
            project_url = "/projects?project=" + quote_plus(project_name)
            amount = safe_num(row.TotalLineAmount)
            pct = min(100, int((amount / max_project_amount) * 100)) if max_project_amount else 0
            util_pct = min(100, max(6, pct))
            color = donut_colors[idx % len(donut_colors)]
            project_rows += f"""
            <tr onclick=\"window.location='{project_url}'\" class=\"click-row\">
                <td><a href=\"{project_url}\">{h(project_name)}</a><div class=\"mini-track\"><span style=\"width:{pct}%\"></span></div></td>
                <td class=\"right\">{row.POCount}</td>
                <td class=\"right strong\">{currency(amount)}</td>
            </tr>"""
            project_cards += f"""
            <a class=\"util-card\" href=\"{project_url}\">
                <div class=\"donut\" style=\"--p:{util_pct}; --c:{color};\"><span>{row.POCount}<small>POs</small></span></div>
                <div class=\"util-name\">{h(project_name)}</div>
                <div class=\"util-amount\">{currency(amount)}</div>
            </a>"""
        if not project_rows:
            project_rows = '<tr><td colspan="3">No project data found.</td></tr>'
        if not project_cards:
            project_cards = '<div class="empty-state">No project data found.</div>'

        approval_pending = pr_stats.get("pending_admin_requests", 0) + pr_stats.get("pending_executive_requests", 0)

        dashboard_css = """
        <style>
            .dash-hero{display:flex;justify-content:space-between;gap:18px;align-items:flex-start;margin-bottom:18px;}
            .dash-title h2{margin:0 0 6px 0;font-size:30px;letter-spacing:-.02em;}
            .dash-title p{margin:0;color:#64748b;}
            .dash-actions{display:flex;gap:10px;flex-wrap:wrap;justify-content:flex-end;}
            .exec-kpis{display:grid;grid-template-columns:repeat(6,minmax(150px,1fr));gap:14px;margin-bottom:16px;}
            .exec-kpi{display:block;text-decoration:none;color:inherit;border:1px solid #e2e8f0;background:rgba(255,255,255,.9);border-radius:16px;padding:16px;box-shadow:0 10px 24px rgba(15,23,42,.05);transition:.15s ease;}
            .exec-kpi:hover,.click-row:hover,.util-card:hover,.assignment-card:hover,.status-item:hover{transform:translateY(-2px);box-shadow:0 12px 30px rgba(15,23,42,.10);}
            .kpi-top{display:flex;gap:10px;align-items:center;margin-bottom:8px;}
            .kpi-icon{width:38px;height:38px;border-radius:13px;display:grid;place-items:center;color:white;font-weight:800;}
            .kpi-icon.blue{background:#2563eb}.kpi-icon.green{background:#16a34a}.kpi-icon.purple{background:#7c3aed}.kpi-icon.amber{background:#f59e0b}.kpi-icon.red{background:#ef4444}.kpi-icon.slate{background:#475569}
            .exec-kpi .label{font-size:13px;color:#475569;font-weight:700;}.exec-kpi .value{font-size:25px;font-weight:850;color:#0f172a;line-height:1.2;}.exec-kpi .hint{font-size:12px;color:#64748b;margin-top:6px;}
            .dash-grid{display:grid;grid-template-columns:1.1fr 1fr;gap:16px;margin-bottom:16px;}.dash-grid.three{grid-template-columns:1fr 1fr 1fr;}.dash-grid.wide-left{grid-template-columns:1.35fr .85fr;}
            .dashboard-panel{background:rgba(255,255,255,.92);border:1px solid #e2e8f0;border-radius:18px;padding:18px;box-shadow:0 10px 28px rgba(15,23,42,.05);}
            .panel-head{display:flex;justify-content:space-between;gap:12px;align-items:center;margin-bottom:12px;}.panel-head h3{margin:0;font-size:18px;}.panel-head a{font-size:13px;text-decoration:none;font-weight:700;}
            .forecast-strip{display:grid;grid-template-columns:repeat(5,1fr);border:1px solid #e2e8f0;border-radius:14px;overflow:hidden;}.forecast-box{padding:14px;text-align:center;border-right:1px solid #e2e8f0;}.forecast-box:last-child{border-right:0}.forecast-box b{display:block;color:#0f172a;margin-bottom:8px}.forecast-box .money{font-size:18px;font-weight:850;color:#1d4ed8}.spark{height:54px;display:flex;align-items:end;gap:5px;justify-content:center;margin-top:10px}.spark span{width:10px;background:linear-gradient(#60a5fa,#2563eb);border-radius:4px 4px 0 0;opacity:.85}
            .util-grid{display:grid;grid-template-columns:repeat(5,1fr);gap:12px;}.util-card{text-align:center;text-decoration:none;color:inherit;border:1px solid #e2e8f0;border-radius:16px;padding:14px;background:#fff;}.donut{width:86px;height:86px;margin:0 auto 8px;border-radius:50%;background:conic-gradient(var(--c) calc(var(--p)*1%),#e5e7eb 0);display:grid;place-items:center;}.donut span{width:58px;height:58px;border-radius:50%;background:white;display:grid;place-items:center;font-weight:850;line-height:1.0}.donut small{display:block;font-size:10px;color:#64748b;font-weight:600}.util-name{font-size:12px;font-weight:700;color:#334155;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}.util-amount{font-size:14px;font-weight:850;color:#0f172a;margin-top:4px;}
            .click-row{cursor:pointer;transition:.15s ease;}.click-row a{text-decoration:none;font-weight:750}.strong{font-weight:800}.mini-track{height:6px;background:#e2e8f0;border-radius:999px;margin-top:7px;overflow:hidden}.mini-track span{display:block;height:100%;background:linear-gradient(90deg,#60a5fa,#2563eb);border-radius:999px;}
            .status-list{display:grid;gap:10px}.status-item{display:flex;justify-content:space-between;gap:12px;align-items:center;padding:12px;border:1px solid #e2e8f0;border-radius:13px;background:#fff;text-decoration:none;color:inherit}.status-item b{font-size:22px}.status-item .red{color:#dc2626}.status-item .amber{color:#d97706}.status-item .green{color:#16a34a}.status-item .blue{color:#2563eb}
            .assignment-list{display:grid;gap:10px}.assignment-card{display:flex;justify-content:space-between;gap:14px;align-items:center;padding:13px;border-radius:14px;border:1px solid #e2e8f0;background:#fff;text-decoration:none;color:inherit;transition:.15s ease}.assignment-card .title{font-weight:800}.assignment-card .sub{font-size:13px;color:#64748b}.empty-state{padding:22px;text-align:center;color:#64748b;border:1px dashed #cbd5e1;border-radius:14px;}
            @media(max-width:1300px){.exec-kpis{grid-template-columns:repeat(3,1fr)}.dash-grid,.dash-grid.three,.dash-grid.wide-left{grid-template-columns:1fr}.util-grid{grid-template-columns:repeat(2,1fr)}}
        </style>
        """

        def kpi(label, value, icon, color, href, hint):
            return f"""
            <a class=\"exec-kpi\" href=\"{href}\">
                <div class=\"kpi-top\"><div class=\"kpi-icon {color}\">{icon}</div><div class=\"label\">{label}</div></div>
                <div class=\"value\">{value}</div>
                <div class=\"hint\">{hint}</div>
            </a>"""

        kpis = f"""
        <div class=\"exec-kpis\">
            {kpi("Total Open PO Amount", currency(total_remaining), "$", "blue", "/pos-balances#posBalancesPOListTable", "Click for balance drilldown")}
            {kpi("Total Committed", currency(committed_amount), "✓", "purple", "/pos-balances#posBalancesLineTable", "Posted/spent line activity")}
            {kpi("Total PO Value", currency(total_value), "Σ", "green", "/pos-balances#posBalancesPOListTable", "All visible POs")}
            {kpi("POs Outstanding", overall["open_pos"], "PO", "amber", "/pos-balances#posBalancesPOListTable", f"{overall['total_pos']} total visible POs")}
            {kpi("Avg. PO Size", currency(avg_po_size), "Ø", "slate", "/pos-balances#posBalancesPOListTable", "Average visible PO value")}
            {kpi("Open Utilization", f"{utilization:.0f}%", "%", "red" if utilization >= 80 else "green", "/projects", "Committed vs line total")}
        </div>
        """

        forecast = f"""
        <div class=\"dashboard-panel\">
            <div class=\"panel-head\"><h3>Cash Out Forecast / Open PO Exposure</h3><a href=\"/pos-balances\">View PO balances ›</a></div>
            <div class=\"forecast-strip\">
                <div class=\"forecast-box\"><b>Open Now</b><span class=\"money\">{currency(total_remaining)}</span><div class=\"spark\"><span style=\"height:25%\"></span><span style=\"height:42%\"></span><span style=\"height:35%\"></span><span style=\"height:55%\"></span></div></div>
                <div class=\"forecast-box\"><b>Committed</b><span class=\"money\">{currency(committed_amount)}</span><div class=\"spark\"><span style=\"height:45%\"></span><span style=\"height:34%\"></span><span style=\"height:60%\"></span><span style=\"height:72%\"></span></div></div>
                <div class=\"forecast-box\"><b>Line Total</b><span class=\"money\">{currency(total_line)}</span><div class=\"spark\"><span style=\"height:55%\"></span><span style=\"height:68%\"></span><span style=\"height:80%\"></span><span style=\"height:62%\"></span></div></div>
                <div class=\"forecast-box\"><b>Open Ratio</b><span class=\"money\">{open_ratio:.0f}%</span><div class=\"spark\"><span style=\"height:34%\"></span><span style=\"height:48%\"></span><span style=\"height:77%\"></span><span style=\"height:41%\"></span></div></div>
                <div class=\"forecast-box\"><b>Requests</b><span class=\"money\">{pr_stats['total_requests']}</span><div class=\"spark\"><span style=\"height:38%\"></span><span style=\"height:58%\"></span><span style=\"height:46%\"></span><span style=\"height:66%\"></span></div></div>
            </div>
            <p class=\"card-subtitle\" style=\"text-align:center;margin-top:18px;\">Total visible open balance: <strong>{currency(total_remaining)}</strong></p>
        </div>
        """

        project_utilization = f"""
        <div class=\"dashboard-panel\">
            <div class=\"panel-head\"><h3>PO Activity by Project</h3><a href=\"/projects\">View all projects ›</a></div>
            <div class=\"util-grid\">{project_cards}</div>
        </div>
        """

        request_summary = f"""
        <div class=\"dashboard-panel\">
            <div class=\"panel-head\"><h3>Purchase Requests Summary</h3><a href=\"/purchase-requests\">View requests ›</a></div>
            <div class=\"status-list\">
                <a class=\"status-item\" href=\"/purchase-requests?status=Submitted\"><span>New / Submitted</span><b class=\"blue\">{pr_stats['submitted_requests']}</b></a>
                <a class=\"status-item\" href=\"/purchase-requests?status=Pending Admin Approval\"><span>Pending Admin Approval</span><b class=\"amber\">{pr_stats['pending_admin_requests']}</b></a>
                <a class=\"status-item\" href=\"/purchase-requests?status=Pending Executive Approval\"><span>Pending Executive Approval</span><b class=\"amber\">{pr_stats['pending_executive_requests']}</b></a>
                <a class=\"status-item\" href=\"/purchase-requests?status=Converted to PO\"><span>Converted to PO</span><b class=\"green\">{pr_stats['converted_requests']}</b></a>
                <a class=\"status-item\" href=\"/purchase-requests?status=Rejected\"><span>Rejected</span><b class=\"red\">{pr_stats['rejected_requests']}</b></a>
            </div>
        </div>
        """

        top_vendor_panel = f"""
        <div class=\"dashboard-panel\">
            <div class=\"panel-head\"><h3>Top Vendors by PO Amount</h3><a href=\"/vendors\">View all vendors ›</a></div>
            <div class=\"table-wrap\"><table><tr><th>Vendor</th><th class=\"right\">POs</th><th class=\"right\">Amount</th></tr>{vendor_rows}</table></div>
        </div>
        """

        top_project_panel = f"""
        <div class=\"dashboard-panel\">
            <div class=\"panel-head\"><h3>Top Projects by PO Amount</h3><a href=\"/projects\">View all projects ›</a></div>
            <div class=\"table-wrap\"><table><tr><th>Project</th><th class=\"right\">POs</th><th class=\"right\">Amount</th></tr>{project_rows}</table></div>
        </div>
        """

        admin_assignments = f"""
        <div class=\"dashboard-panel\">
            <div class=\"panel-head\"><h3>Admin Assignments & Controls</h3><a href=\"/user-access\">User access ›</a></div>
            <div class=\"assignment-list\">
                <a class=\"assignment-card\" href=\"/purchase-requests\"><div><div class=\"title\">Approval queue</div><div class=\"sub\">Admin and Executive approval work</div></div><strong>{approval_pending}</strong></a>
                <a class=\"assignment-card\" href=\"/project-po-setup?mine=1\"><div><div class=\"title\">PO setup items</div><div class=\"sub\">Payment schedule / planning details</div></div><strong>{po_setup_action_count}</strong></a>
                <a class=\"assignment-card\" href=\"/pos-balances#posBalancesPOListTable\"><div><div class=\"title\">Amount mismatch flags</div><div class=\"sub\">PO header vs line total review</div></div><strong>{overall['amount_mismatch_count']}</strong></a>
                <a class=\"assignment-card\" href=\"/user-access\"><div><div class=\"title\">Active users</div><div class=\"sub\">Users enabled in Command Center</div></div><strong>{data['active_user_count']}</strong></a>
                <a class=\"assignment-card\" href=\"/admin/clear-expense-data\"><div><div class=\"title\">Expense reset tool</div><div class=\"sub\">Admin-only clean upload support</div></div><strong>Open</strong></a>
            </div>
        </div>
        """

        manager_assignments = f"""
        <div class=\"dashboard-panel\">
            <div class=\"panel-head\"><h3>My Assignment Snapshot</h3><a href=\"/purchase-request\">New request ›</a></div>
            <div class=\"assignment-list\">
                <a class=\"assignment-card\" href=\"/purchase-request\"><div><div class=\"title\">Submit a purchase request</div><div class=\"sub\">Start a new request for your visible projects</div></div><strong>New</strong></a>
                <a class=\"assignment-card\" href=\"/purchase-requests\"><div><div class=\"title\">Visible request queue</div><div class=\"sub\">Role-filtered requests</div></div><strong>{pr_stats['total_requests']}</strong></a>
                <a class=\"assignment-card\" href=\"/projects\"><div><div class=\"title\">Visible projects</div><div class=\"sub\">Open project PO information</div></div><strong>{len(data['top_projects'])}</strong></a>
                <a class=\"assignment-card\" href=\"/pos-balances\"><div><div class=\"title\">Visible open POs</div><div class=\"sub\">Role-filtered PO balances</div></div><strong>{overall['open_pos']}</strong></a>
            </div>
        </div>
        """

        bookkeeping_lookup = f"""
        <div class=\"dashboard-panel\">
            <div class=\"panel-head\"><h3>PO Lookup Tools</h3><a href=\"/pos-balances\">Open PO lookup ›</a></div>
            <div class=\"assignment-list\">
                <a class=\"assignment-card\" href=\"/pos-balances\"><div><div class=\"title\">Search POs & balances</div><div class=\"sub\">Amounts, balances, and packet links</div></div><strong>{overall['total_pos']}</strong></a>
                <a class=\"assignment-card\" href=\"/vendors\"><div><div class=\"title\">Vendor lookup</div><div class=\"sub\">Find POs by vendor</div></div><strong>{len(data['top_vendors'])}</strong></a>
                <a class=\"assignment-card\" href=\"/projects\"><div><div class=\"title\">Project lookup</div><div class=\"sub\">View PO information by project</div></div><strong>{len(data['top_projects'])}</strong></a>
            </div>
        </div>
        """

        executive_summary = f"""
        <div class=\"dashboard-panel\">
            <div class=\"panel-head\"><h3>Executive Summary</h3><a href=\"/purchase-requests\">Open approvals ›</a></div>
            <p>Total visible open PO balance is <strong>{currency(total_remaining)}</strong> across <strong>{overall['open_pos']}</strong> open POs.</p>
            <p>Current approval queue contains <strong>{approval_pending}</strong> request(s), with <strong>{pr_stats['pending_executive_requests']}</strong> awaiting Executive approval.</p>
            <p>Top project and vendor rows are clickable for drilldown.</p>
            <div class=\"notice warning\" style=\"margin-top:12px;\">Recommendation: review high-utilization projects and any pending approval items before approving additional spend.</div>
        </div>
        """

        title_by_role = {
            "Admin": "Admin Command Center Dashboard",
            "Executive": "Executive Command Center Dashboard",
            "Project Manager - Dredging Only": "Dredging Project Manager Dashboard",
            "Project Manager - Diving": "Diving Project Manager Dashboard",
            "Division Manager - Diving": "Diving Division Manager Dashboard",
            "Purchaser - All Departments": "Purchasing Dashboard",
            "Bookkeeping - All Departments": "Bookkeeping PO Lookup Dashboard",
        }
        dashboard_title = title_by_role.get(role, "Command Center Dashboard")

        po_setup_action_card = ""
        if po_setup_action_count:
            po_setup_action_card = f"""
            <div class=\"dashboard-panel\" style=\"border-color:#f59e0b;background:#fffbeb;margin-bottom:16px;\">
                <div class=\"panel-head\"><h3>Actions Required</h3><a href=\"/project-po-setup?mine=1\">Open tasks ›</a></div>
                <p>You have <strong>{po_setup_action_count}</strong> PO information item(s) that need payment schedule or planning details.</p>
            </div>
            """

        if role == "Executive":
            main_content = f"""
            {kpis}
            <div class=\"dash-grid wide-left\">{forecast}{project_utilization}</div>
            <div class=\"dash-grid three\">{top_vendor_panel}{request_summary}{executive_summary}</div>
            <div class=\"dash-grid two\">{top_project_panel}{project_utilization}</div>
            """
        elif role == "Admin":
            main_content = f"""
            {kpis}
            <div class=\"dash-grid wide-left\">{admin_assignments}{request_summary}</div>
            <div class=\"dash-grid two\">{forecast}{project_utilization}</div>
            <div class=\"dash-grid two\">{top_vendor_panel}{top_project_panel}</div>
            """
        elif role in ["Project Manager - Dredging Only", "Project Manager - Diving", "Division Manager - Diving"]:
            main_content = f"""
            {kpis}
            <div class=\"dash-grid wide-left\">{manager_assignments}{request_summary}</div>
            <div class=\"dash-grid two\">{top_project_panel}{project_utilization}</div>
            <div class=\"dash-grid two\">{forecast}{top_vendor_panel}</div>
            """
        elif role == "Purchaser - All Departments":
            main_content = f"""
            {kpis}
            <div class=\"dash-grid wide-left\">{manager_assignments}{request_summary}</div>
            <div class=\"dash-grid two\">{top_vendor_panel}{top_project_panel}</div>
            """
        else:
            main_content = f"""
            {kpis}
            <div class=\"dash-grid wide-left\">{bookkeeping_lookup}{top_vendor_panel}</div>
            <div class=\"dash-grid two\">{top_project_panel}{forecast}</div>
            """

        content = f"""
        {dashboard_css}
        <div class=\"dash-hero\">
            <div class=\"dash-title\"><h2>{h(dashboard_title)}</h2><p>Welcome, {h(display_name)}. This view is filtered for <strong>{h(role)}</strong>.</p></div>
            <div class=\"dash-actions\"><a class=\"button\" href=\"/help-center\">Help Center</a><a class=\"button primary\" href=\"/pos-balances\">Open PO Drilldown</a></div>
        </div>
        {po_setup_action_card}
        {main_content}
        """

        return shell("My Dashboard", f"Personalized Command Center dashboard for {role}.", "My Dashboard", content)

    except Exception as e:
        content = f'<div class="notice error">Error loading personal dashboard: {h(e)}</div>'
        return shell("My Dashboard", "Unable to load personalized dashboard.", "My Dashboard", content), 500

@app.route("/po-summary")
def po_summary():
    allowed, reason = require_page_access("PO Summary")
    if not allowed:
        return access_denied_response("PO Summary", reason)

    try:
        overall, vendors, projects, imports = load_summary_data()

        vendor_rows = ""
        for row in vendors:
            vendor_rows += f"""
            <tr>
                <td>{h(row.VendorName)}</td>
                <td>{row.POCount}</td>
                <td class="right">{currency(row.TotalPOValue)}</td>
                <td class="right">{currency(row.TotalLineAmount)}</td>
                <td class="right">{currency(row.TotalRemainingAmount)}</td>
            </tr>
            """

        project_rows = ""
        for row in projects:
            project_rows += f"""
            <tr>
                <td>{h(row.ProjectName)}</td>
                <td>{row.POCount}</td>
                <td class="right">{currency(row.TotalPOValue)}</td>
                <td class="right">{currency(row.TotalLineAmount)}</td>
                <td class="right">{currency(row.TotalRemainingAmount)}</td>
            </tr>
            """

        import_rows = ""
        for row in imports:
            import_rows += f"""
            <tr>
                <td>{row.ImportBatchId}</td>
                <td>{h(row.FileName)}</td>
                <td>{h(row.UploadedAt)}</td>
                <td>{row.TotalRows}</td>
                <td>{row.SuccessCount}</td>
                <td>{row.ErrorCount}</td>
                <td>{h(row.ImportStatus)}</td>
            </tr>
            """

        content = f"""
        <div class="grid kpis">
            <div class="card kpi"><div class="label">Total Unique POs</div><div class="value">{overall["total_pos"]}</div><div class="trend">Grouped by PO number</div></div>
            <div class="card kpi"><div class="label">Open POs</div><div class="value">{overall["open_pos"]}</div><div class="trend">Status = Open</div></div>
            <div class="card kpi"><div class="label">Total PO Value</div><div class="value">{currency(overall["total_po_value"])}</div><div class="trend">Unique PO totals</div></div>
            <div class="card kpi"><div class="label">Line Item Total</div><div class="value">{currency(overall["total_line_amount"])}</div><div class="trend">Sum of line amounts</div></div>
            <div class="card kpi"><div class="label">Remaining</div><div class="value">{currency(overall["total_remaining_amount"])}</div><div class="trend">Unique PO remaining</div></div>
        </div>

        <div class="grid two">
            <div class="card"><h3>POs by Vendor</h3><p class="card-subtitle">Vendor-level committed value and remaining balance.</p><div class="table-wrap"><table><tr><th>Vendor</th><th>PO Count</th><th class="right">PO Value</th><th class="right">Line Amount</th><th class="right">Remaining</th></tr>{vendor_rows}</table></div></div>
            <div class="card"><h3>POs by Project</h3><p class="card-subtitle">Project-level committed value and remaining balance.</p><div class="table-wrap"><table><tr><th>Project</th><th>PO Count</th><th class="right">PO Value</th><th class="right">Line Amount</th><th class="right">Remaining</th></tr>{project_rows}</table></div></div>
        </div>

        <div class="card"><h3>Recent Import Batches</h3><p class="card-subtitle">Latest PO upload activity.</p><div class="table-wrap"><table><tr><th>Batch ID</th><th>File Name</th><th>Uploaded At</th><th>Total Rows</th><th>Success</th><th>Errors</th><th>Status</th></tr>{import_rows}</table></div></div>
        """

        return shell("PO Summary", "Live issued PO summary grouped by vendor, project, and import batch.", "PO Summary", content)

    except Exception as e:
        content = f'<div class="notice error">Error loading PO summary: {h(e)}</div>'
        return shell("PO Summary", "Unable to load summary.", "PO Summary", content), 500


@app.route("/po-list")
def po_list():
    allowed, reason = require_page_access("PO List")
    if not allowed:
        return access_denied_response("PO List", reason)

    try:
        conn = get_sql_connection()
        cursor = conn.cursor()

        cursor.execute(
            """
            WITH POList AS (
                SELECT
                    PONumber,
                    MAX(VendorName) AS VendorName,
                    MAX(ProjectName) AS ProjectName,
                    MAX(Department) AS Department,
                    MAX(POStatus) AS POStatus,
                    MAX(PODate) AS PODate,
                    COUNT(*) AS LineCount,
                    MAX(COALESCE(RevisedAmount, OriginalAmount, 0)) AS POValue,
                    SUM(COALESCE(LineAmount, 0)) AS TotalLineAmount,
                    MAX(COALESCE(RemainingAmount, 0)) AS RemainingAmount
                FROM dbo.IssuedPOLines
                GROUP BY PONumber
            )
            SELECT
                PONumber,
                VendorName,
                ProjectName,
                Department,
                POStatus,
                PODate,
                LineCount,
                POValue,
                TotalLineAmount,
                RemainingAmount,
                CASE WHEN ABS(COALESCE(POValue, 0) - COALESCE(TotalLineAmount, 0)) > 0.01 THEN 1 ELSE 0 END AS AmountMismatch
            FROM POList
            ORDER BY PODate DESC, PONumber DESC;
            """
        )

        pos = cursor.fetchall()
        conn.close()

        po_rows = ""

        for row in pos:
            status_text = str(row.POStatus or "")
            status_lower = status_text.lower()

            status_class = "blue"
            if status_lower == "open":
                status_class = "green"
            elif status_lower in ["closed", "complete", "completed"]:
                status_class = "blue"
            elif status_lower in ["cancelled", "canceled"]:
                status_class = "red"
            elif status_lower in ["pending", "draft"]:
                status_class = "amber"

            mismatch_badge = ""
            if row.AmountMismatch:
                mismatch_badge = '<span class="badge amber">Check totals</span>'

            po_url = "/po-packet/" + quote_plus(str(row.PONumber or "")) + "?type=internal"

            po_rows += f"""
            <tr>
                <td><a href="{po_url}">{h(row.PONumber)}</a></td>
                <td><a class="vendor-detail-link" href="/vendors?vendor={quote_plus(str(row.VendorName or ''))}">{h(row.VendorName)}</a></td>
                <td><a class="vendor-detail-link" href="/projects?project={quote_plus(str(row.ProjectName or ''))}">{h(row.ProjectName)}</a></td>
                <td>{h(row.Department)}</td>
                <td><span class="badge {status_class}">{h(status_text)}</span></td>
                <td>{h(row.PODate)}</td>
                <td class="right">{row.LineCount}</td>
                <td class="right">{currency(row.POValue)}</td>
                <td class="right">{currency(row.TotalLineAmount)}</td>
                <td class="right">{currency(row.RemainingAmount)}</td>
                <td>{mismatch_badge}</td>
            </tr>
            """

        if not po_rows:
            po_rows = '<tr><td colspan="11">No issued POs found yet.</td></tr>'

        content = f"""
        <div class="card">
            <h3>Issued PO List</h3>
            <p class="card-subtitle">Browse all issued POs imported into the dashboard. Click a PO number to open the full PO packet.</p>
            <div class="filter-hint">
                <span>Use the filters below each column heading to narrow the issued PO list.</span>
                <button type="button" onclick="clearPOListFilters()">Clear Filters</button>
            </div>
            <div class="table-wrap">
                <table id="issuedPOListTable">
                    <thead>
                        <tr><th>PO Number</th><th>Vendor</th><th>Project</th><th>Department</th><th>Status</th><th>PO Date</th><th class="right">Lines</th><th class="right">PO Value</th><th class="right">Line Total</th><th class="right">Remaining</th><th>Flag</th><th>Packets</th></tr>
                        <tr class="column-filter-row">
                            <th><input data-col="0" oninput="filterIssuedPOList()" placeholder="Filter PO"></th>
                            <th><input data-col="1" oninput="filterIssuedPOList()" placeholder="Filter vendor"></th>
                            <th><input data-col="2" oninput="filterIssuedPOList()" placeholder="Filter project"></th>
                            <th><input data-col="3" oninput="filterIssuedPOList()" placeholder="Filter dept"></th>
                            <th><input data-col="4" oninput="filterIssuedPOList()" placeholder="Filter status"></th>
                            <th><input data-col="5" oninput="filterIssuedPOList()" placeholder="Filter date"></th>
                            <th><input data-col="6" oninput="filterIssuedPOList()" placeholder="Lines"></th>
                            <th><input data-col="7" oninput="filterIssuedPOList()" placeholder="PO value"></th>
                            <th><input data-col="8" oninput="filterIssuedPOList()" placeholder="Line total"></th>
                            <th><input data-col="9" oninput="filterIssuedPOList()" placeholder="Remaining"></th>
                            <th><input data-col="10" oninput="filterIssuedPOList()" placeholder="Flag"></th>
                        </tr>
                    </thead>
                    <tbody>
                        {po_rows}
                    </tbody>
                </table>
            </div>
        </div>
        <script>
        function filterIssuedPOList() {{
            const table = document.getElementById('issuedPOListTable');
            if (!table) return;
            const filters = Array.from(table.querySelectorAll('.column-filter-row input')).map(input => {{
                return {{ col: Number(input.dataset.col), value: input.value.trim().toLowerCase() }};
            }});
            const rows = Array.from(table.querySelectorAll('tbody tr'));
            rows.forEach(row => {{
                const cells = Array.from(row.children);
                const show = filters.every(filter => {{
                    if (!filter.value) return true;
                    const cell = cells[filter.col];
                    return cell && cell.textContent.toLowerCase().includes(filter.value);
                }});
                row.style.display = show ? '' : 'none';
            }});
        }}
        function clearPOListFilters() {{
            const table = document.getElementById('issuedPOListTable');
            if (!table) return;
            table.querySelectorAll('.column-filter-row input').forEach(input => input.value = '');
            filterIssuedPOList();
        }}
        </script>
        """

        return shell("PO List", "Browse issued purchase orders and open PO detail records.", "PO List", content)

    except Exception as e:
        content = f'<div class="notice error">Error loading PO list: {h(e)}</div>'
        return shell("PO List", "Unable to load issued PO list.", "PO List", content), 500


@app.route("/po-detail", methods=["GET"])
def po_detail():
    allowed, reason = require_page_access("PO Detail")
    if not allowed:
        return access_denied_response("PO Detail", reason)

    po_number = clean_text(request.args.get("po_number"))
    search_value = h(po_number or "")

    search_form = f"""
    <div class="card">
        <h3>Search Purchase Order</h3>
        <p class="card-subtitle">Enter a PO number to view its line items and totals.</p>
        <form method="get" action="/po-detail">
            <p><input type="text" name="po_number" value="{search_value}" placeholder="Example: 26-204-002" required></p>
            <p><button class="primary" type="submit">Search PO</button></p>
        </form>
    </div>
    """

    if not po_number:
        content = search_form + """
        <div class="card"><h3>PO Detail</h3><p class="card-subtitle">Search for a PO number to see vendor, project, totals, and line items.</p></div>
        """
        return shell("PO Detail", "Search and review issued PO line items.", "PO Detail", content)

    try:
        conn = get_sql_connection()
        cursor = conn.cursor()
        req_where, req_params = requestor_filter_sql("IssuedPOLines")

        cursor.execute(
            """
            SELECT TOP 1
                PONumber,
                VendorName,
                ProjectName,
                Department,
                PODate,
                POStatus,
                OriginalAmount,
                RevisedAmount,
                RemainingAmount,
                Requestor
            FROM dbo.IssuedPOLines
            WHERE PONumber = ? AND {req_where}
            ORDER BY CreatedAt DESC;
            """.format(req_where=req_where),
            po_number,
            *req_params,
        )

        header = cursor.fetchone()

        if not header:
            conn.close()
            content = search_form + f'<div class="notice error">No PO found for PO number: {h(po_number)}</div>'
            return shell("PO Detail", "PO number was not found.", "PO Detail", content)

        cursor.execute(
            """
            SELECT LineDescription, Unit, UnitCost, Qty, LineAmount
            FROM dbo.IssuedPOLines
            WHERE PONumber = ? AND {req_where}
            ORDER BY IssuedPOLineId;
            """.format(req_where=req_where),
            po_number,
            *req_params,
        )
        lines = cursor.fetchall()

        cursor.execute(
            """
            SELECT
                COUNT(*) AS LineCount,
                SUM(COALESCE(LineAmount, 0)) AS TotalLineAmount,
                MAX(COALESCE(OriginalAmount, 0)) AS OriginalAmount,
                MAX(COALESCE(RevisedAmount, OriginalAmount, 0)) AS RevisedAmount,
                MAX(COALESCE(RemainingAmount, 0)) AS RemainingAmount
            FROM dbo.IssuedPOLines
            WHERE PONumber = ? AND {req_where};
            """.format(req_where=req_where),
            po_number,
            *req_params,
        )
        totals = cursor.fetchone()
        conn.close()

        line_rows = ""
        for row in lines:
            line_rows += f"""
            <tr><td>{h(row.LineDescription)}</td><td>{h(row.Unit)}</td><td class="right">{currency(row.UnitCost)}</td><td class="right">{h(row.Qty)}</td><td class="right">{currency(row.LineAmount)}</td></tr>
            """

        content = search_form + f"""
        <div class="grid kpis">
            <div class="card kpi"><div class="label">PO Number</div><div class="value" style="font-size:22px;">{h(header.PONumber)}</div><div class="trend">{h(header.POStatus)}</div></div>
            <div class="card kpi"><div class="label">Original Amount</div><div class="value">{currency(totals.OriginalAmount)}</div><div class="trend">Original issued value</div></div>
            <div class="card kpi"><div class="label">Revised Amount</div><div class="value">{currency(totals.RevisedAmount)}</div><div class="trend">Current approved value</div></div>
            <div class="card kpi"><div class="label">Line Item Total</div><div class="value">{currency(totals.TotalLineAmount)}</div><div class="trend">{totals.LineCount} line item(s)</div></div>
            <div class="card kpi"><div class="label">Remaining</div><div class="value">{currency(totals.RemainingAmount)}</div><div class="trend">Current PO balance</div></div>
        </div>
        <div class="card"><h3>PO Header</h3><table><tr><th>Vendor</th><td>{h(header.VendorName)}</td></tr><tr><th>Project</th><td>{h(header.ProjectName)}</td></tr><tr><th>Department</th><td>{h(header.Department)}</td></tr><tr><th>PO Date</th><td>{h(header.PODate)}</td></tr><tr><th>Status</th><td>{h(header.POStatus)}</td></tr><tr><th>Requestor</th><td>{h(header.Requestor)}</td></tr></table></div>
        <div class="card"><h3>Line Items</h3><p class="card-subtitle">Issued PO line items imported from the upload document.</p><div class="table-wrap"><table><tr><th>Description</th><th>Unit</th><th class="right">Unit Cost</th><th class="right">Qty</th><th class="right">Line Amount</th></tr>{line_rows}</table></div></div>
        """

        return shell("PO Detail", f"Line item detail for PO {po_number}.", "PO Detail", content)

    except Exception as e:
        content = search_form + f'<div class="notice error">Error loading PO detail: {h(e)}</div>'
        return shell("PO Detail", "Unable to load PO detail.", "PO Detail", content), 500




@app.route("/projects")
def projects_page():
    allowed, reason = require_page_access("Projects")
    if not allowed:
        return access_denied_response("Projects", reason)

    selected_project = clean_text(request.args.get("project")) or ""
    try:
        conn = get_sql_connection()
        cursor = conn.cursor()
        ensure_project_code_columns(cursor)
        conn.commit()
        req_where_po, req_params_po = requestor_filter_sql("po")
        req_where_line, req_params_line = requestor_filter_sql("l")

        cursor.execute(
            f"""
            SELECT DISTINCT
                COALESCE(pr.ProjectCode, '') AS ProjectCode,
                COALESCE(pr.ProjectName, l.ProjectName, '') AS ProjectName
            FROM dbo.IssuedPOLines l
            LEFT JOIN dbo.PurchaseOrders po ON l.PurchaseOrderId = po.PurchaseOrderId
            LEFT JOIN dbo.Projects pr ON po.ProjectId = pr.ProjectId
            WHERE {req_where_line}
            ORDER BY COALESCE(pr.ProjectCode, ''), COALESCE(pr.ProjectName, l.ProjectName, '');
            """,
            *req_params_line,
        )
        projects = cursor.fetchall()

        if not selected_project and projects:
            first = projects[0]
            selected_project = clean_text(first.ProjectCode) or clean_text(first.ProjectName) or ""

        options = ""
        for pr in projects:
            value = clean_text(pr.ProjectCode) or clean_text(pr.ProjectName) or ""
            label = (clean_text(pr.ProjectCode) + " - " if clean_text(pr.ProjectCode) else "") + (clean_text(pr.ProjectName) or "Unnamed Project")
            sel = " selected" if value == selected_project else ""
            options += f'<option value="{h(value)}"{sel}>{h(label)}</option>'

        project_filter_sql = "1=0"
        params_po = []
        params_line = []
        if selected_project:
            project_filter_sql = "(LOWER(COALESCE(pr.ProjectCode, '')) = LOWER(?) OR LOWER(COALESCE(pr.ProjectName, l.ProjectName, '')) = LOWER(?))"
            params_po = [selected_project, selected_project]
            params_line = [selected_project, selected_project]

        cursor.execute(
            f"""
            WITH PostedExpenses AS (
                SELECT PostedPONumber AS PONumber, SUM(COALESCE(PostedAmount, Amount, 0)) AS PostedExpenseAmount
                FROM dbo.ExpenseReviewItems
                WHERE COALESCE(PostedToPO, 0) = 1 AND COALESCE(PostedPONumber, '') <> ''
                GROUP BY PostedPONumber
            ), LineRollup AS (
                SELECT PONumber, SUM(COALESCE(LineAmount, 0)) AS LineAmount, COUNT(*) AS LineCount
                FROM dbo.IssuedPOLines
                GROUP BY PONumber
            )
            SELECT
                po.PONumber,
                COALESCE(v.VendorName, 'Missing Vendor') AS VendorName,
                COALESCE(pr.ProjectCode, '') AS ProjectCode,
                COALESCE(pr.ProjectName, '') AS ProjectName,
                po.Department,
                po.POStatus,
                po.PODate,
                po.Requestor,
                COALESCE(po.RevisedAmount, po.OriginalAmount, lr.LineAmount, 0) AS POValue,
                COALESCE(pe.PostedExpenseAmount, 0) AS PostedExpenseAmount,
                CASE WHEN COALESCE(po.RevisedAmount, po.OriginalAmount, lr.LineAmount, 0) - COALESCE(pe.PostedExpenseAmount, 0) < 0 THEN 0 ELSE COALESCE(po.RevisedAmount, po.OriginalAmount, lr.LineAmount, 0) - COALESCE(pe.PostedExpenseAmount, 0) END AS CurrentAppBalance,
                COALESCE(lr.LineCount, 0) AS LineCount
            FROM dbo.PurchaseOrders po
            LEFT JOIN dbo.Vendors v ON po.VendorId = v.VendorId
            LEFT JOIN dbo.Projects pr ON po.ProjectId = pr.ProjectId
            LEFT JOIN LineRollup lr ON lr.PONumber = po.PONumber
            LEFT JOIN PostedExpenses pe ON pe.PONumber = po.PONumber
            LEFT JOIN dbo.IssuedPOLines l ON l.PONumber = po.PONumber
            WHERE {project_filter_sql} AND {req_where_po}
            GROUP BY po.PONumber, v.VendorName, pr.ProjectCode, pr.ProjectName, po.Department, po.POStatus, po.PODate, po.Requestor, po.RevisedAmount, po.OriginalAmount, lr.LineAmount, pe.PostedExpenseAmount, lr.LineCount
            ORDER BY po.PONumber;
            """,
            *(params_po + req_params_po),
        )
        po_rows_raw = cursor.fetchall()

        cursor.execute(
            f"""
            SELECT TOP 1000
                l.PONumber,
                l.VendorName,
                COALESCE(pr.ProjectCode, '') AS ProjectCode,
                COALESCE(pr.ProjectName, l.ProjectName, '') AS ProjectName,
                l.Department,
                l.LineDescription,
                l.Unit,
                l.UnitCost,
                l.Qty,
                l.LineAmount
            FROM dbo.IssuedPOLines l
            LEFT JOIN dbo.PurchaseOrders po ON l.PurchaseOrderId = po.PurchaseOrderId
            LEFT JOIN dbo.Projects pr ON po.ProjectId = pr.ProjectId
            WHERE {project_filter_sql} AND {req_where_line}
            ORDER BY l.PONumber, l.IssuedPOLineId;
            """,
            *(params_line + req_params_line),
        )
        line_rows_raw = cursor.fetchall()
        conn.close()

        po_rows = ""
        for row in po_rows_raw:
            po_rows += f"""
            <tr>
                <td><a class="po-link" href="/po-packet/{quote_plus(str(row.PONumber or ''))}?type=internal">{h(row.PONumber)}</a></td>
                <td>{h(row.VendorName)}</td>
                <td>{status_chip(row.POStatus or 'Open')}</td>
                <td>{h(row.PODate)}</td>
                <td>{h(row.Requestor)}</td>
                <td class="right">{currency(row.POValue)}</td>
                <td class="right">{currency(row.PostedExpenseAmount)}</td>
                <td class="right">{currency(row.CurrentAppBalance)}</td>
                <td class="right">{h(row.LineCount)}</td>
            </tr>
            """
        if not po_rows:
            po_rows = '<tr><td colspan="9"><div class="empty-state"><strong>No POs found for this project.</strong></div></td></tr>'

        line_rows = ""
        for row in line_rows_raw:
            line_rows += f"""
            <tr>
                <td><a class="po-link" href="/po-packet/{quote_plus(str(row.PONumber or ''))}?type=internal">{h(row.PONumber)}</a></td>
                <td>{h(row.VendorName)}</td>
                <td>{h(row.Department)}</td>
                <td>{h(row.LineDescription)}</td>
                <td>{h(row.Unit)}</td>
                <td class="right">{currency(row.UnitCost)}</td>
                <td class="right">{h(row.Qty)}</td>
                <td class="right">{currency(row.LineAmount)}</td>
            </tr>
            """
        if not line_rows:
            line_rows = '<tr><td colspan="8"><div class="empty-state"><strong>No line items found for this project.</strong></div></td></tr>'

        content = f"""
        <div class="card">
            <h3>Select Project</h3>
            <form method="get" action="/projects" class="filters">
                <select name="project" onchange="this.form.submit()">{options}</select>
                <noscript><button class="primary" type="submit">Open Project</button></noscript>
            </form>
        </div>
        <div class="card">
            <h3>Project POs</h3>
            <p class="card-subtitle">POs for the selected project. Click a PO number to open the full PO packet.</p>
            <div class="filter-hint"><span>Filter the PO table below.</span><button type="button" onclick="clearPOListFilters('projectPOTable')">Clear Filters</button></div>
            <div class="table-wrap"><table id="projectPOTable"><thead><tr><th>PO</th><th>Vendor</th><th>Status</th><th>PO Date</th><th>Requestor</th><th class="right">PO Value</th><th class="right">Posted Expenses</th><th class="right">Current Balance</th><th class="right">Lines</th></tr><tr class="column-filter-row"><th><input data-col="0" oninput="filterPOListTable('projectPOTable')" placeholder="PO"></th><th><input data-col="1" oninput="filterPOListTable('projectPOTable')" placeholder="Vendor"></th><th><input data-col="2" oninput="filterPOListTable('projectPOTable')" placeholder="Status"></th><th><input data-col="3" oninput="filterPOListTable('projectPOTable')" placeholder="Date"></th><th><input data-col="4" oninput="filterPOListTable('projectPOTable')" placeholder="Requestor"></th><th><input data-col="5" oninput="filterPOListTable('projectPOTable')" placeholder="Value"></th><th><input data-col="6" oninput="filterPOListTable('projectPOTable')" placeholder="Posted"></th><th><input data-col="7" oninput="filterPOListTable('projectPOTable')" placeholder="Balance"></th><th><input data-col="8" oninput="filterPOListTable('projectPOTable')" placeholder="Lines"></th></tr></thead><tbody>{po_rows}</tbody></table></div>
        </div>
        <div class="card">
            <h3>Project PO Line Items</h3>
            <p class="card-subtitle">Line-item drilldown for the selected project.</p>
            <div class="filter-hint"><span>Filter the line-item table below.</span><button type="button" onclick="clearPOListFilters('projectLineTable')">Clear Filters</button></div>
            <div class="table-wrap"><table id="projectLineTable"><thead><tr><th>PO</th><th>Vendor</th><th>Department</th><th>Description</th><th>Unit</th><th class="right">Unit Cost</th><th class="right">Qty</th><th class="right">Line Amount</th></tr><tr class="column-filter-row"><th><input data-col="0" oninput="filterPOListTable('projectLineTable')" placeholder="PO"></th><th><input data-col="1" oninput="filterPOListTable('projectLineTable')" placeholder="Vendor"></th><th><input data-col="2" oninput="filterPOListTable('projectLineTable')" placeholder="Dept"></th><th><input data-col="3" oninput="filterPOListTable('projectLineTable')" placeholder="Description"></th><th><input data-col="4" oninput="filterPOListTable('projectLineTable')" placeholder="Unit"></th><th><input data-col="5" oninput="filterPOListTable('projectLineTable')" placeholder="Unit cost"></th><th><input data-col="6" oninput="filterPOListTable('projectLineTable')" placeholder="Qty"></th><th><input data-col="7" oninput="filterPOListTable('projectLineTable')" placeholder="Amount"></th></tr></thead><tbody>{line_rows}</tbody></table></div>
        </div>
        <script>
        function filterPOListTable(tableId) {{ const table = document.getElementById(tableId); if (!table) return; const filters = Array.from(table.querySelectorAll('.column-filter-row input')).map(input => {{ return {{ col: Number(input.dataset.col), value: input.value.trim().toLowerCase() }}; }}); Array.from(table.querySelectorAll('tbody tr')).forEach(row => {{ const cells = Array.from(row.children); const show = filters.every(filter => !filter.value || ((cells[filter.col] || {{textContent:''}}).textContent.toLowerCase().includes(filter.value))); row.style.display = show ? '' : 'none'; }}); }}
        function clearPOListFilters(tableId) {{ const table = document.getElementById(tableId); if (!table) return; table.querySelectorAll('.column-filter-row input').forEach(input => input.value = ''); filterPOListTable(tableId); }}
        </script>
        """
        return shell("Projects", "Review project POs and line-item detail.", "Projects", content)
    except Exception as e:
        return shell("Projects", "Unable to load projects.", "Projects", f'<div class="notice error">Error loading Projects: {h(e)}</div>'), 500

@app.route("/pos-balances")
def pos_balances():
    allowed, reason = require_page_access("POs & Balances")
    if not allowed:
        return access_denied_response("POs & Balances", reason)

    try:
        data = load_pos_balances_data()
        overall = data["overall"]

        project_rows = ""
        max_project_value = max([float(row.POValue or 0) for row in data["projects"]] or [1])
        for row in data["projects"]:
            bar_width = 0 if max_project_value == 0 else max(5, float(row.POValue or 0) / max_project_value * 100)
            project_rows += f"""
            <tr>
                <td><a class="vendor-detail-link" href="/projects?project={quote_plus(str(row.ProjectName or ''))}"><strong>{h(row.ProjectName)}</strong></a></td>
                <td class="right">{row.POCount}</td>
                <td class="right">{currency(row.POValue)}</td>
                <td class="right">{currency(row.TotalLineAmount)}</td>
                <td class="right">{currency(getattr(row, 'PostedExpenseAmount', 0))}</td>
                <td class="right">{currency(row.RemainingAmount)}</td>
                <td class="right">{percent(row.PercentOpen)}</td>
                <td><div class="bar-track"><div class="bar-fill" style="width:{bar_width}%"></div></div></td>
            </tr>
            """
        if not project_rows:
            project_rows = '<tr><td colspan="7">No project PO data found.</td></tr>'

        vendor_bar_rows = ""
        max_vendor_value = max([float(row.POValue or 0) for row in data["vendors"]] or [1])
        for row in data["vendors"]:
            bar_width = 0 if max_vendor_value == 0 else max(5, float(row.POValue or 0) / max_vendor_value * 100)
            vendor_bar_rows += f"""
            <div class="mini-bar-row"><span><a class="vendor-detail-link" href="/vendors?vendor={quote_plus(str(row.VendorName or ''))}">{h(row.VendorName)}</a></span><div><b style="width:{bar_width}%"></b></div><em>{currency(row.POValue)}</em></div>
            """
        if not vendor_bar_rows:
            vendor_bar_rows = '<p class="card-subtitle">No vendor data found.</p>'

        po_rows = ""
        for row in data["pos"]:
            status_text = row.POStatus or "Unknown"
            flag = '<span class="badge amber">Mismatch</span>' if row.AmountMismatch else '<span class="badge green">OK</span>'
            po_url = "/po-packet/" + quote_plus(str(row.PONumber or "")) + "?type=internal"
            internal_packet_url = "/po-packet/" + quote_plus(str(row.PONumber or "")) + "?type=internal"
            vendor_packet_url = "/po-packet/" + quote_plus(str(row.PONumber or "")) + "?type=vendor"
            po_rows += f"""
            <tr>
                <td><a href="{po_url}">{h(row.PONumber)}</a></td>
                <td><a class="vendor-detail-link" href="/vendors?vendor={quote_plus(str(row.VendorName or ''))}">{h(row.VendorName)}</a></td>
                <td><a class="vendor-detail-link" href="/projects?project={quote_plus(str(row.ProjectName or ''))}">{h(row.ProjectName)}</a></td>
                <td>{h(row.Department)}</td>
                <td>{status_chip(status_text)}</td>
                <td>{h(row.PODate)}</td>
                <td class="right">{row.LineCount}</td>
                <td class="right">{currency(row.POValue)}</td>
                <td class="right">{currency(row.TotalLineAmount)}</td>
                <td class="right">{currency(getattr(row, 'PostedExpenseAmount', 0))}<br><small>{int(getattr(row, 'PostedExpenseCount', 0) or 0)} expense(s)</small></td>
                <td class="right">{currency(row.RemainingAmount)}</td>
                <td>{flag}</td>
                <td><a class="secondary" href="{internal_packet_url}">Internal</a><br><a class="secondary" href="{vendor_packet_url}">Vendor</a></td>
            </tr>
            """
        if not po_rows:
            po_rows = '<tr><td colspan="12">No issued POs found.</td></tr>'

        line_rows = ""
        for row in data["lines"]:
            line_rows += f"""
            <tr>
                <td><a class="po-link" href="/po-packet/{quote_plus(str(row.PONumber or ''))}?type=internal">{h(row.PONumber)}</a></td>
                <td><a class="vendor-detail-link" href="/vendors?vendor={quote_plus(str(row.VendorName or ''))}">{h(row.VendorName)}</a></td>
                <td><a class="vendor-detail-link" href="/projects?project={quote_plus(str(row.ProjectName or ''))}">{h(row.ProjectName)}</a></td>
                <td>{h(row.Department)}</td>
                <td>{h(row.LineDescription)}</td>
                <td>{h(row.Unit)}</td>
                <td class="right">{currency(row.UnitCost)}</td>
                <td class="right">{h(row.Qty)}</td>
                <td class="right">{currency(row.LineAmount)}</td>
            </tr>
            """
        if not line_rows:
            line_rows = '<tr><td colspan="9">No line items found.</td></tr>'

        content = f"""
        <div class="grid kpis">
            <a class="card kpi action-card blue" href="#posBalancesPOListTable"><div class="label">Issued PO Count</div><div class="value">{overall['total_pos']}</div><div class="trend">Unique issued POs</div></a>
            <a class="card kpi action-card green" href="#posBalancesPOListTable"><div class="label">Open POs</div><div class="value">{overall['open_pos']}</div><div class="trend">Currently open</div></a>
            <div class="card kpi"><div class="label">Issued PO Amount</div><div class="value">{currency(overall['total_po_value'])}</div><div class="trend">Revised/original value</div></div>
            <div class="card kpi"><div class="label">Posted Expenses</div><div class="value">{currency(overall.get('posted_expense_amount', 0))}</div><div class="trend">Matched expenses reducing balances</div></div>
            <a class="card kpi action-card amber" href="/project-po-setup"><div class="label">Current App Balance</div><div class="value">{currency(overall['total_remaining_amount'])}</div><div class="trend">Issued POs minus posted matched expenses</div></a>
            <div class="card kpi"><div class="label">Review Flags</div><div class="value">{overall['amount_mismatch_count']}</div><div class="trend">PO value vs. line total</div></div>
        </div>


        <div class="card project-bucket-card">
            <h3>Project Open Balance Buckets</h3>
            <p class="card-subtitle">Projects grouped by remaining open PO balance percentage. Green means stronger remaining balance; red means lower remaining balance.</p>
            <div class="project-bucket-grid">{render_open_balance_buckets(data['projects'])}</div>
        </div>

        <div class="grid two visual-chart-row" style="margin-bottom:24px;">
            <div class="card">
                <h3>PO Exposure by Project</h3>
                <p class="card-subtitle">Issued value, line totals, and remaining balances by project.</p>
                <div class="table-wrap"><table><tr><th>Project</th><th class="right">POs</th><th class="right">Issued</th><th class="right">Line Total</th><th class="right">Remaining</th><th class="right">% Open</th><th>Scale</th></tr>{project_rows}</table></div>
            </div>
            <div class="mini-chart-card">
                <h4>Top Vendors by Issued PO Amount</h4>
                {vendor_bar_rows}
            </div>
        </div>

        <div class="card">
            <h3>Issued PO List</h3>
            <p class="card-subtitle">Consolidated list formerly shown on the separate PO List page. Click a PO number for full PO detail.</p>
            <div class="filter-hint"><span>Use the filters below each column heading to narrow the PO list.</span><button type="button" onclick="clearPOListFilters('posBalancesPOListTable')">Clear Filters</button></div>
            <div class="table-wrap">
                <table id="posBalancesPOListTable">
                    <thead>
                        <tr><th>PO Number</th><th>Vendor</th><th>Project</th><th>Department</th><th>Status</th><th>PO Date</th><th class="right">Lines</th><th class="right">PO Value</th><th class="right">Line Total</th><th class="right">Remaining</th><th>Flag</th><th>Packets</th></tr>
                        <tr class="column-filter-row">
                            <th><input data-col="0" oninput="filterPOListTable('posBalancesPOListTable')" placeholder="Filter PO"></th>
                            <th><input data-col="1" oninput="filterPOListTable('posBalancesPOListTable')" placeholder="Filter vendor"></th>
                            <th><input data-col="2" oninput="filterPOListTable('posBalancesPOListTable')" placeholder="Filter project"></th>
                            <th><input data-col="3" oninput="filterPOListTable('posBalancesPOListTable')" placeholder="Filter dept"></th>
                            <th><input data-col="4" oninput="filterPOListTable('posBalancesPOListTable')" placeholder="Filter status"></th>
                            <th><input data-col="5" oninput="filterPOListTable('posBalancesPOListTable')" placeholder="Filter date"></th>
                            <th><input data-col="6" oninput="filterPOListTable('posBalancesPOListTable')" placeholder="Lines"></th>
                            <th><input data-col="7" oninput="filterPOListTable('posBalancesPOListTable')" placeholder="PO value"></th>
                            <th><input data-col="8" oninput="filterPOListTable('posBalancesPOListTable')" placeholder="Line total"></th>
                            <th><input data-col="9" oninput="filterPOListTable('posBalancesPOListTable')" placeholder="Remaining"></th>
                            <th><input data-col="10" oninput="filterPOListTable('posBalancesPOListTable')" placeholder="Flag"></th>
                            <th><input data-col="11" oninput="filterPOListTable('posBalancesPOListTable')" placeholder="Packets"></th>
                        </tr>
                    </thead>
                    <tbody>{po_rows}</tbody>
                </table>
            </div>
        </div>

        <div class="card">
            <h3>Recent PO Line Items</h3>
            <p class="card-subtitle">Line-level detail from issued PO imports. This brings PO Detail visibility into the combined page.</p>
            <div class="table-wrap"><table id="posBalancesLineTable"><tr><th>PO</th><th>Vendor</th><th>Project</th><th>Department</th><th>Description</th><th>Unit</th><th class="right">Unit Cost</th><th class="right">Qty</th><th class="right">Line Amount</th></tr>{line_rows}</table></div>
        </div>

        <script>
        function filterPOListTable(tableId) {{
            const table = document.getElementById(tableId);
            if (!table) return;
            const filters = Array.from(table.querySelectorAll('.column-filter-row input')).map(input => {{ return {{ col: Number(input.dataset.col), value: input.value.trim().toLowerCase() }}; }});
            const rows = Array.from(table.querySelectorAll('tbody tr'));
            rows.forEach(row => {{
                const cells = Array.from(row.children);
                const show = filters.every(filter => {{
                    if (!filter.value) return true;
                    const cell = cells[filter.col];
                    return cell && cell.textContent.toLowerCase().includes(filter.value);
                }});
                row.style.display = show ? '' : 'none';
            }});
        }}
        function clearPOListFilters(tableId) {{
            const table = document.getElementById(tableId);
            if (!table) return;
            table.querySelectorAll('.column-filter-row input').forEach(input => input.value = '');
            filterPOListTable(tableId);
        }}
        </script>
        """

        return shell("POs & Balances", "Issued PO value, open balances, project buckets, vendor exposure, and line details in one view.", "POs & Balances", content)

    except Exception as e:
        content = f'<div class="notice error">Error loading POs & Balances: {h(e)}</div>'
        return shell("POs & Balances", "Unable to load consolidated PO view.", "POs & Balances", content), 500


@app.route("/forecasting")
def forecasting():
    allowed, reason = require_page_access("Forecasting")
    if not allowed:
        return access_denied_response("Forecasting", reason)

    try:
        items = load_forecast_data()
        selected_bucket = clean_text(request.args.get("bucket")) or "All Buckets"
        visible_items = items if selected_bucket == "All Buckets" else [item for item in items if item["bucket"] == selected_bucket]
        bucket_summary = forecast_bucket_summary(items)
        total_forecast = sum(Decimal(str(item["amount"] or 0)) for item in items)
        scheduled_total = sum(Decimal(str(item["amount"] or 0)) for item in items if item["bucket"] != "Unscheduled")
        unscheduled_total = total_forecast - scheduled_total
        past_due_count = len([item for item in items if item["bucket"] == "Past Due"])
        next_30_total = sum(Decimal(str(item["amount"] or 0)) for item in items if item["bucket"] in ["Past Due", "Next 7 Days", "8-14 Days", "15-30 Days"])

        max_bucket = max([float(amount or 0) for _, _, amount in bucket_summary] or [1])
        forecast_bucket_cards = ""
        for label, count, amount in bucket_summary:
            bar_height = 8 if max_bucket == 0 else max(8, float(amount or 0) / max_bucket * 100)
            forecast_bucket_cards += f"""
            <a class="forecast-bucket" href="/forecasting?bucket={quote_plus(label)}" style="text-decoration:none;color:inherit;">
                <strong>{h(label)}</strong>
                <div class="amount">{currency(amount)}</div>
                <div class="bucket-note">{count} item(s)</div>
                <div class="bars"><span class="bar" style="height:{bar_height}%"></span></div>
            </a>
            """

        project_rows = ""
        for project, values in aggregate_items(items, "project")[:8]:
            project_rows += f"<tr><td>{h(project)}</td><td class='right'>{values['count']}</td><td>{h(values['next_date'])}</td><td class='right'>{currency(values['amount'])}</td></tr>"
        if not project_rows:
            project_rows = '<tr><td colspan="4">No forecast project data found.</td></tr>'

        vendor_rows = ""
        for vendor, values in aggregate_items(items, "vendor")[:8]:
            vendor_rows += f"<tr><td>{h(vendor)}</td><td class='right'>{values['count']}</td><td>{h(values['next_date'])}</td><td class='right'>{currency(values['amount'])}</td></tr>"
        if not vendor_rows:
            vendor_rows = '<tr><td colspan="4">No forecast vendor data found.</td></tr>'

        detail_rows = ""
        for item in visible_items:
            detail_rows += f"""
            <tr>
                <td>{h(item['source_type'])}</td>
                <td>{h(item['source_id'])}</td>
                <td>{h(item['project'])}</td>
                <td>{h(item['vendor'])}</td>
                <td>{h(item['forecast_date'])}</td>
                <td>{status_chip(item['status'])}</td>
                <td>{h(item['bucket'])}</td>
                <td>{h(item['description'])}</td>
                <td class="right">{currency(item['amount'])}</td>
            </tr>
            """
        if not detail_rows:
            detail_rows = '<tr><td colspan="9">No forecast detail found for this filter.</td></tr>'

        content = f"""
        <div class="grid kpis">
            <a class="card kpi action-card blue" href="/forecasting?bucket=All+Buckets"><div class="label">Total Forecast Exposure</div><div class="value">{currency(total_forecast)}</div><div class="trend">Requests + open PO balances</div></a>
            <a class="card kpi action-card green" href="/forecasting?bucket=Next+7+Days"><div class="label">Scheduled Forecast</div><div class="value">{currency(scheduled_total)}</div><div class="trend">Items with dates</div></a>
            <a class="card kpi action-card amber" href="/forecasting?bucket=Unscheduled"><div class="label">Unscheduled</div><div class="value">{currency(unscheduled_total)}</div><div class="trend">Needs forecast date</div></a>
            <a class="card kpi action-card red" href="/forecasting?bucket=Past+Due"><div class="label">Past Due Items</div><div class="value">{past_due_count}</div><div class="trend">Needed-by date has passed</div></a>
            <a class="card kpi action-card purple" href="/forecasting?bucket=15-30+Days"><div class="label">Next 30 Days</div><div class="value">{currency(next_30_total)}</div><div class="trend">Past due through 30 days</div></a>
            <div class="card kpi"><div class="label">Detail Items</div><div class="value">{len(items)}</div><div class="trend">Forecast rows</div></div>
        </div>

        <div class="card">
            <h3>Forecast Buckets</h3>
            <p class="card-subtitle">Expected purchase/payment timing from purchase requests and currently open PO balances. Click a bucket to filter the detail list.</p>
            <div class="forecast-row">{forecast_bucket_cards}</div>
        </div>

        <div class="grid two">
            <div class="card"><h3>Forecast by Project</h3><div class="table-wrap"><table><tr><th>Project</th><th class="right">Items</th><th>Next Date</th><th class="right">Amount</th></tr>{project_rows}</table></div></div>
            <div class="card"><h3>Top Vendors by Forecast Amount</h3><div class="table-wrap"><table><tr><th>Vendor</th><th class="right">Items</th><th>Next Date</th><th class="right">Amount</th></tr>{vendor_rows}</table></div></div>
        </div>

        <div class="card" id="forecast-detail-section">
            <h3>Forecast Detail</h3>
            <p class="card-subtitle">Current filter: <strong>{h(selected_bucket)}</strong>. Add expected payment or due dates later to make the forecast stronger.</p>
            <div class="filter-chip-row">
                <a class="filter-chip" href="/forecasting?bucket=All+Buckets">All Buckets</a>
                {''.join(f'<a class="filter-chip" href="/forecasting?bucket={quote_plus(label)}">{h(label)}</a>' for label, _, _ in bucket_summary)}
            </div>
            <div class="table-wrap"><table><tr><th>Source</th><th>ID</th><th>Project</th><th>Vendor</th><th>Forecast Date</th><th>Status</th><th>Bucket</th><th>Description</th><th class="right">Amount</th></tr>{detail_rows}</table></div>
        </div>
        """

        return shell("Forecasting", "Upcoming purchase request needs, open PO exposure, and unscheduled forecast gaps.", "Forecasting", content)

    except Exception as e:
        content = f'<div class="notice error">Error loading forecasting page: {h(e)}</div>'
        return shell("Forecasting", "Unable to load forecasting.", "Forecasting", content), 500



@app.route("/approver-queue", methods=["GET", "POST"])
def approver_queue():
    allowed, reason = require_page_access("Approver Queue")
    if not allowed:
        return access_denied_response("Approver Queue", reason)

    access = get_user_access()
    role = access["role"]

    if request.method == "POST":
        if not can_review_purchase_requests(role):
            return redirect("/approver-queue?toast=You+do+not+have+permission+to+update+approval+items&toast_type=error")
        try:
            update_purchase_request_status(request.form)
            return redirect("/approver-queue?toast=Approver+queue+updated")
        except Exception as e:
            return redirect("/approver-queue?toast=" + quote_plus("Error updating approval item: " + str(e)) + "&toast_type=error")

    try:
        rows = [row for row in load_purchase_requests(200) if (row.RequestStatus or "Submitted") in ["Submitted", "Under Review", "Needs More Info", "Pending Admin Approval", "Pending Executive Approval"]]
        selected_id = clean_text(request.args.get("request_id"))
        selected = None
        if selected_id:
            for row in rows:
                if str(row.PurchaseRequestId) == selected_id:
                    selected = row
                    break
        if selected is None and rows:
            selected = rows[0]

        list_items = ""
        total_open_amount = Decimal("0")
        for row in rows:
            total_open_amount += Decimal(str(row.EstimatedAmount or 0))
            active_class = " active" if selected and row.PurchaseRequestId == selected.PurchaseRequestId else ""
            list_items += f"""
            <a class="approval-item{active_class}" href="/approver-queue?request_id={row.PurchaseRequestId}">
                <strong>{h(row.RequestNumber)}</strong>
                <span>{h(row.RequestTitle)}<br>{h(row.ProjectName)} / {h(row.VendorName)}<br>{purchase_request_status_badge(row.RequestStatus)}</span>
                <em>{currency(row.EstimatedAmount)}</em>
            </a>
            """
        if not list_items:
            list_items = '<div class="empty-state"><strong>No open approval items.</strong><span>Submitted, Under Review, and Needs More Info requests will appear here.</span></div>'

        detail_html = ""
        if selected:
            status_options = ""
            for status in ["Submitted", "Under Review", "Needs More Info", "Pending Admin Approval", "Pending Executive Approval", "Approved", "Rejected", "Converted to PO"]:
                selected_option = " selected" if status == selected.RequestStatus else ""
                status_options += f'<option value="{h(status)}"{selected_option}>{h(status)}</option>'
            approve_form = action_form(selected.PurchaseRequestId, "Approved", "Approve", "primary")
            reject_form = action_form(selected.PurchaseRequestId, "Rejected", "Reject", "secondary")
            more_info_form = action_form(selected.PurchaseRequestId, "Needs More Info", "Request Info", "secondary")
            review_form = action_form(selected.PurchaseRequestId, "Under Review", "Mark Reviewing", "secondary")
            detail_html = f"""
            <div class="card">
                <h3>Approval Detail</h3>
                <div class="detail-grid">
                    <div><span>Request</span><strong>{h(selected.RequestNumber)}</strong></div>
                    <div><span>Status</span><strong>{status_chip(selected.RequestStatus)}</strong></div>
                    <div><span>Project</span><strong>{h(selected.ProjectName)}</strong></div>
                    <div><span>Vendor</span><strong>{h(selected.VendorName)}</strong></div>
                    <div><span>Needed By</span><strong>{h(selected.NeededByDate)}</strong></div>
                    <div><span>Estimate</span><strong>{currency(selected.EstimatedAmount)}</strong></div>
                </div>
                <p class="card-subtitle" style="margin-top:14px;">{h(selected.RequestDescription)}</p>
                <div class="approval-action-grid">{review_form}{more_info_form}{approve_form}{reject_form}</div>
                <div class="timeline">
                    <div class="timeline-item"><strong>Submitted</strong><span>{h(selected.RequestedByName or selected.RequestedByEmail)} submitted this request.</span></div>
                    <div class="timeline-item"><strong>Current Status</strong><span>{h(selected.RequestStatus or 'Submitted')}</span></div>
                    <div class="timeline-item"><strong>Last Review Note</strong><span>{h(selected.ReviewNotes or 'No review notes yet.')}</span></div>
                </div>
                <form method="post" action="/approver-queue" style="margin-top:16px;">
                    <input type="hidden" name="purchase_request_id" value="{h(selected.PurchaseRequestId)}">
                    <div class="form-grid">
                        <div class="form-field"><label>Status</label><select name="request_status">{status_options}</select></div>
                        <div class="form-field"><label>Converted PO Number</label><input type="text" name="converted_po_number" value="{h(selected.ConvertedPONumber)}" placeholder="PO number if converted"></div>
                        <div class="form-field full"><label>Review Notes</label><textarea name="review_notes" placeholder="Review notes">{h(selected.ReviewNotes)}</textarea></div>
                    </div>
                    <div class="request-actions"><button class="primary" type="submit">Save Detailed Update</button></div>
                </form>
            </div>
            """
        else:
            detail_html = '<div class="card"><h3>Approval Detail</h3><div class="empty-state"><strong>No request selected.</strong><span>Open approval items will appear when requests are submitted.</span></div></div>'

        content = f"""
        <div class="grid kpis">
            <div class="card kpi"><div class="label">Open Approval Items</div><div class="value">{len(rows)}</div><div class="trend">Submitted / under review / needs info</div></div>
            <div class="card kpi"><div class="label">Open Approval Value</div><div class="value">{currency(total_open_amount)}</div><div class="trend">Estimated request value</div></div>
            <a class="card kpi action-card blue" href="/purchase-requests?status=Submitted"><div class="label">Submitted</div><div class="value">View</div><div class="trend">Filter request dashboard</div></a>
            <a class="card kpi action-card amber" href="/purchase-requests?status=Needs+More+Info"><div class="label">Needs Info</div><div class="value">View</div><div class="trend">Follow-up requests</div></a>
        </div>
        <div class="grid two">
            <div class="card"><h3>Open Approver Queue</h3><p class="card-subtitle">Requests waiting for review or currently under review.</p><div class="approval-list">{list_items}</div></div>
            {detail_html}
        </div>
        """
        return shell("Approver Queue", "Approve, reject, or request more information for open purchase requests.", "Approver Queue", content)

    except Exception as e:
        content = f'<div class="notice error">Error loading approver queue: {h(e)}</div>'
        return shell("Approver Queue", "Unable to load approval queue.", "Approver Queue", content), 500



def posted_expense_card(rows):
    table_rows = ""
    for r in rows:
        table_rows += f"""
        <tr>
            <td><strong>{h(r.ExpenseId or r.ExpenseReviewItemId)}</strong><br><small>{h(r.TxDate)}</small></td>
            <td>{h(r.TxType)}</td>
            <td>{h(r.VendorName)}</td>
            <td>{h(r.Description)}<br><small>{h(r.PMComments)}</small></td>
            <td class="right">{currency(r.PostedAmount or r.Amount)}</td>
            <td>{h(r.ReviewDecision)}<br><small>{h(r.PostedBy or r.ReviewerEmail)}</small></td>
            <td>{h(r.PostedAt)}</td>
        </tr>
        """
    if not table_rows:
        table_rows = '<tr><td colspan="7"><div class="empty-state"><strong>No posted expenses yet.</strong><span>Matched expenses will appear here once they reduce this PO balance.</span></div></td></tr>'
    return f"""
    <div class="card">
        <h3>Posted Expenses Reducing This PO</h3>
        <p class="card-subtitle">These reviewed expense rows reduce the current app PO balance. They do not write back to Unanet or the ERP.</p>
        <div class="table-wrap"><table>
            <tr><th>Expense</th><th>Type</th><th>Vendor / Purchaser</th><th>Description / PM Comments</th><th class="right">Posted Amount</th><th>Decision / Posted By</th><th>Posted At</th></tr>
            {table_rows}
        </table></div>
    </div>
    """


@app.route("/po-packet-pdf/<path:po_number>")
def po_packet_pdf(po_number):
    allowed, reason = require_page_access("POs & Balances")
    if not allowed:
        return access_denied_response("POs & Balances", reason)

    packet_type = clean_text(request.args.get("type")) or "internal"
    if packet_type not in ["internal", "vendor"]:
        packet_type = "internal"

    try:
        att = po_packet_pdf_attachment(po_number, packet_type)
    except Exception as exc:
        # If styled PDF rendering hits an environment-specific issue, return a
        # valid fallback PDF instead of an Internal Server Error.
        title = ("Vendor PO Packet" if packet_type == "vendor" else "Internal PO Packet") + " - " + clean_text(po_number)
        att = {
            "filename": f"{clean_text(po_number)}_{packet_type}_packet.pdf",
            "content": simple_pdf_bytes(title, [
                "Coastal Engineering Group",
                "The styled PDF renderer encountered an issue. This fallback confirms the PO exists while the styled renderer is being corrected.",
                f"PO Number: {clean_text(po_number)}",
            ]),
        }
    if not att:
        content = '<div class="notice error">PO PDF was not found.</div>'
        return shell("PO Packet PDF", "Unable to generate this PO packet PDF.", "POs & Balances", content), 404

    response = Response(att["content"], mimetype="application/pdf")
    response.headers["Content-Disposition"] = f'inline; filename="{att["filename"]}"'
    response.headers["Cache-Control"] = "no-store"
    return response


@app.route("/po-packet/<path:po_number>")
def po_packet(po_number):
    allowed, reason = require_page_access("POs & Balances")
    if not allowed:
        return access_denied_response("POs & Balances", reason)

    packet_type = clean_text(request.args.get("type")) or "internal"
    if packet_type not in ["internal", "vendor"]:
        packet_type = "internal"

    try:
        po, lines, posted_expenses = load_po_packet_data(po_number)
        if not po:
            content = '<div class="notice error">PO was not found.</div>'
            return shell("PO Packet", "Unable to find this PO.", "POs & Balances", content), 404

        internal_link = "/po-packet/" + quote_plus(str(po.PONumber or "")) + "?type=internal"
        vendor_link = "/po-packet/" + quote_plus(str(po.PONumber or "")) + "?type=vendor"

        internal_rows = ""
        vendor_rows = ""
        for idx, line in enumerate(lines, start=1):
            internal_rows += f"""
            <tr>
                <td>{idx}</td>
                <td>{h(line.LineDescription)}</td>
                <td>{h(line.Unit)}</td>
                <td class="right">{h(line.Qty)}</td>
                <td class="right">{currency(line.UnitCost)}</td>
                <td class="right">{currency(line.LineAmount)}</td>
                <td class="right">{currency(line.RemainingAmount)}</td>
            </tr>
            """
            vendor_rows += f"""
            <tr>
                <td>{idx}</td>
                <td>{h(line.LineDescription)}</td>
                <td>{h(line.Unit)}</td>
                <td class="right">{h(line.Qty)}</td>
                <td class="right">{currency(line.UnitCost)}</td>
                <td class="right">{currency(line.LineAmount)}</td>
            </tr>
            """

        if not internal_rows:
            internal_rows = '<tr><td colspan="7"><div class="empty-state"><strong>No line items found.</strong><span>This packet has no imported line detail yet.</span></div></td></tr>'
            vendor_rows = '<tr><td colspan="6"><div class="empty-state"><strong>No line items found.</strong><span>This purchase order has no line detail yet.</span></div></td></tr>'

        internal_pdf_link = "/po-packet-pdf/" + quote_plus(str(po.PONumber or "")) + "?type=internal"
        vendor_pdf_link = "/po-packet-pdf/" + quote_plus(str(po.PONumber or "")) + "?type=vendor"

        packet_actions = f"""
        <div class="packet-actions">
            <a class="button secondary" href="/pos-balances">Back to POs &amp; Balances</a>
            <a class="button secondary" href="{internal_link}">Internal Web Packet</a>
            <a class="button secondary" href="{vendor_link}">Vendor Web Packet</a>
            <a class="button primary" href="{internal_pdf_link}" target="_blank" rel="noopener">Internal PDF</a>
            <a class="button primary" href="{vendor_pdf_link}" target="_blank" rel="noopener">Vendor PDF</a>
            <button class="button secondary" onclick="window.print()">Print This Page</button>
        </div>
        """

        payment_schedule_html = render_payment_schedule_for_packet(getattr(po, 'PaymentSchedule', '') or '')
        generated_at = datetime.now().strftime("%b %-d, %Y %-I:%M %p") if os.name != "nt" else datetime.now().strftime("%b %#d, %Y %#I:%M %p")
        ship_to_html = """
            <span class="packet-label">Ship To</span>
            <span class="packet-value">Coastal Engineering Group</span>
            <div class="packet-muted">Attn: Project Receiver<br>Reference the PO number on all deliveries.</div>
        """
        vendor_terms_html = """
        <div class="vendor-terms-card">
            <h3 class="packet-section-title">Coastal Engineering PO Terms and Conditions</h3>
            <ul class="terms-list">
                <li><strong>Delivery / Performance:</strong> Vendor must deliver goods or perform services by the required date. Delays without written approval may result in cancellation.</li>
                <li><strong>Invoicing &amp; Payment:</strong> Include the PO number on all invoices. Send invoices to accounting@c-diving.com. Unless otherwise agreed, payment terms are Net 30 from receipt of a valid invoice and satisfactory delivery.</li>
                <li><strong>Changes:</strong> No substitutions or changes to quantity or delivery date without written approval from Coastal Engineering.</li>
                <li><strong>Inspection:</strong> All items are subject to inspection. Non-compliant goods or services may be rejected at the vendor's expense.</li>
                <li><strong>Warranties:</strong> Vendor warrants that goods and services are free from defects, conform to specifications, and are fit for their intended use.</li>
                <li><strong>Compliance:</strong> Vendor must comply with all applicable laws and regulations.</li>
                <li><strong>Indemnification:</strong> Vendor agrees to hold Coastal Engineering harmless from any claims or liabilities arising from this Purchase Order.</li>
                <li><strong>PO Cancellation:</strong> Coastal Engineering reserves the right to cancel this PO at any time for undelivered goods or services.</li>
            </ul>
        </div>
        """

        if packet_type == "vendor":
            content = f"""
            {packet_actions}
            <div class="po-packet-page vendor-packet">
                <div class="vendor-hero">
                    <div class="vendor-hero-top">
                        <div class="vendor-logo"><img src="{CE_LOGO_DATA_URI}" alt="Coastal Engineering Group logo"></div>
                        <div class="vendor-title">
                            <h1>PURCHASE ORDER</h1>
                            <div class="po-number">{h(po.PONumber)}</div>
                        </div>
                    </div>
                </div>
                <div class="vendor-body">
                    <div class="vendor-date-row">
                        <div class="vendor-date-card"><span class="packet-label">PO Date</span><strong class="packet-value">{h(po.PODate)}</strong></div>
                        <div class="vendor-date-card"><span class="packet-label">Required / Expected Date</span><strong class="packet-value">{h(getattr(po, 'ExpectedPaymentDate', '') or 'Per project schedule')}</strong></div>
                        <div class="vendor-total-badge"><span>Total Amount</span><strong>{currency(po.POValue)}</strong></div>
                    </div>
                    <div class="packet-grid-2">
                        <div class="packet-card">
                            <h3 class="packet-section-title">Vendor</h3>
                            <span class="packet-value">{h(po.VendorName)}</span>
                            <div class="packet-muted">Please reference PO {h(po.PONumber)} on all invoices and correspondence.</div>
                        </div>
                        <div class="packet-card">
                            <h3 class="packet-section-title">Deliver To</h3>
                            {ship_to_html}
                        </div>
                    </div>
                    <div style="height:14px"></div>
                    <div class="packet-card">
                        <h3 class="packet-section-title">Line Items</h3>
                        <div class="table-wrap"><table class="packet-table"><tr><th>Line</th><th>Description</th><th>Unit</th><th class="right">Qty</th><th class="right">Unit Cost</th><th class="right">Line Amount</th></tr>{vendor_rows}</table></div>
                    </div>
                    <div class="vendor-note-row">
                        <div class="vendor-note-card"><h3 class="packet-section-title">Please Note</h3><div class="packet-muted">Reference the PO number on all invoices and deliveries. Deliver to the site contact upon arrival.</div></div>
                        <div class="vendor-note-card"><h3 class="packet-section-title">Send Invoices To</h3><div class="packet-muted"><strong>accounting@c-diving.com</strong><br>Include PO number and invoice detail.</div></div>
                        <div class="vendor-note-card"><h3 class="packet-section-title">Payment Terms</h3><div class="packet-muted">Net 30 from receipt of valid invoice and satisfactory delivery unless otherwise agreed.</div></div>
                    </div>
                    {vendor_terms_html}
                    <div class="vendor-wave-footer"><strong>COASTAL ENGINEERING PO TERMS &amp; CONDITIONS</strong><span>Vendor-facing purchase order packet</span></div>
                </div>
            </div>
            """
            return shell("Vendor PO Packet", h(po.PONumber), "POs & Balances", content)

        posted_card_html = posted_expense_card(posted_expenses)
        attachment_html = attachment_card(po.PONumber)
        content = f"""
        {packet_actions}
        <div class="po-packet-page internal-packet">
            <div class="internal-hero">
                <div class="internal-hero-top">
                    <div class="internal-logo-card"><img src="{CE_LOGO_DATA_URI}" alt="Coastal Engineering Group logo"></div>
                    <div class="internal-title"><h1>PURCHASE ORDER</h1><div class="po-number">{h(po.PONumber)}</div><div class="packet-muted" style="color:#bfdbfe;margin-top:8px;">Internal packet • for review and reconciliation</div></div>
                </div>
            </div>
            <div class="internal-body">
                <div class="internal-summary-strip">
                    <div class="internal-info-card">
                        <h3 class="packet-section-title">Vendor</h3>
                        <span class="packet-value">{h(po.VendorName)}</span>
                        <div class="packet-muted">Project vendor / supplier on this PO.</div>
                    </div>
                    <div class="internal-info-card">
                        <h3 class="packet-section-title">Ship To</h3>
                        {ship_to_html}
                    </div>
                    <div class="internal-blue-panel">
                        <span class="packet-label">PO Total</span><span class="packet-value">{currency(po.POValue)}</span>
                        <span class="packet-label" style="margin-top:10px;">Posted Expenses</span><span class="packet-value">{currency(getattr(po, 'PostedExpenseAmount', 0))}</span>
                        <span class="packet-label" style="margin-top:10px;">Current App Balance</span><span class="packet-value">{currency(po.RemainingAmount)}</span>
                    </div>
                </div>
                <div class="packet-grid-4">
                    <div class="packet-card"><span class="packet-label">PO Date</span><span class="packet-value">{h(po.PODate)}</span></div>
                    <div class="packet-card"><span class="packet-label">Expected Payment Date</span><span class="packet-value">{h(getattr(po, 'ExpectedPaymentDate', '') or 'Not set')}</span></div>
                    <div class="packet-card"><span class="packet-label">Requestor</span><span class="packet-value">{h(getattr(po, 'Requestor', '') or '')}</span></div>
                    <div class="packet-card"><span class="packet-label">Department</span><span class="packet-value">{h(po.Department)}</span></div>
                    <div class="packet-card"><span class="packet-label">Project</span><span class="packet-value">{h(po.ProjectName)}</span></div>
                    <div class="packet-card"><span class="packet-label">Status</span><span class="packet-value">{status_chip(po.POStatus)}</span></div>
                    <div class="packet-card"><span class="packet-label">Payment Type</span><span class="packet-value">{h(getattr(po, 'PaymentType', '') or 'Not set')}</span></div>
                    <div class="packet-card"><span class="packet-label">Line Count</span><span class="packet-value">{h(po.LineCount)}</span></div>
                </div>
                <div style="height:14px"></div>
                <div class="packet-grid-2">
                    <div class="packet-card"><h3 class="packet-section-title">Payment Schedule</h3>{payment_schedule_html}</div>
                    <div class="packet-total-box">
                        <div class="packet-total-row"><span>Total Line Amount</span><strong>{currency(po.TotalLineAmount)}</strong></div>
                        <div class="packet-total-row"><span>Posted Expenses</span><strong>{currency(getattr(po, 'PostedExpenseAmount', 0))}</strong></div>
                        <div class="packet-total-row"><span>Current App Balance</span><strong>{currency(po.RemainingAmount)}</strong></div>
                        <div class="packet-total-row"><span>Packet Generated</span><strong>{h(generated_at)}</strong></div>
                    </div>
                </div>
                <div style="height:14px"></div>
                <div class="packet-card">
                    <h3 class="packet-section-title">PO Line Items</h3>
                    <div class="table-wrap"><table class="packet-table"><tr><th>Line</th><th>Description</th><th>Unit</th><th class="right">Qty</th><th class="right">Unit Cost</th><th class="right">Line Amount</th><th class="right">Remaining</th></tr>{internal_rows}</table></div>
                </div>
                <div class="internal-footer-strip">
                    <div><span class="packet-label">Internal Status</span><span class="packet-value">{h(po.POStatus)}</span></div>
                    <div><span class="packet-label">Setup Updated By</span><span class="packet-value">{h(getattr(po, 'SetupUpdatedBy', '') or 'Not recorded')}</span></div>
                    <div><span class="packet-label">Setup Updated At</span><span class="packet-value">{h(getattr(po, 'SetupUpdatedAt', '') or 'Not recorded')}</span></div>
                    <div><span class="packet-label">Packet Type</span><span class="packet-value">Internal Only</span></div>
                </div>
            </div>
        </div>
        {attachment_html}
        {posted_card_html}
        """
        return shell("Internal PO Packet", h(po.PONumber), "POs & Balances", content)

    except Exception as e:
        content = f'<div class="notice error">Error loading PO packet: {h(e)}</div>'
        return shell("PO Packet", "Unable to load PO packet.", "POs & Balances", content), 500


@app.route("/project-po-setup", methods=["GET", "POST"])
def project_po_setup():
    allowed, reason = require_page_access("PO Setup Review")
    if not allowed:
        return access_denied_response("PO Setup Review", reason)

    user = get_current_user()

    if request.method == "POST":
        try:
            update_po_setup_info(request.form)
            return redirect("/project-po-setup?toast=" + quote_plus("PO setup information updated."))
        except Exception as e:
            return redirect("/project-po-setup?toast=" + quote_plus("Error updating PO setup: " + str(e)) + "&toast_type=error")

    try:
        status_filter = clean_text(request.args.get("status")) or "All"
        assigned_filter = None
        if request.args.get("mine") == "1":
            assigned_filter = user["email"]

        rows = load_project_po_setup_items(status_filter=status_filter, assigned_filter=assigned_filter)
        assignable_users = load_assignable_users()

        needs_schedule = sum(1 for r in rows if getattr(r, "MissingPaymentSchedule", 0))
        needs_type = sum(1 for r in rows if getattr(r, "MissingPaymentType", 0))
        needs_date = sum(1 for r in rows if getattr(r, "MissingExpectedPaymentDate", 0))
        needs_vendor = sum(1 for r in rows if getattr(r, "MissingVendor", 0))
        assigned = sum(1 for r in rows if (r.SetupAssignedTo or ""))
        total_amount = sum(Decimal(str(r.RemainingAmount or 0)) for r in rows)

        status_links = ""
        for label in ["All", "Needs Payment Schedule", "Assigned to PM", "In Progress", "Needs Info", "Complete"]:
            active_class = " active" if status_filter == label else ""
            href = "/project-po-setup" if label == "All" else "/project-po-setup?status=" + quote_plus(label)
            status_links += f'<a class="{active_class.strip()}" href="{href}">{status_chip(label)}</a>'

        project_names = []
        for r in rows:
            project_name = clean_text(getattr(r, "ProjectName", ""))
            if project_name and project_name not in project_names:
                project_names.append(project_name)
        project_filter_options = '<option value="">All projects</option>'
        for project_name in sorted(project_names):
            project_filter_options += f'<option value="{h(project_name)}">{h(project_name)}</option>'

        table_rows = ""
        for r in rows:
            packet_url = "/po-packet/" + quote_plus(str(r.PONumber or "")) + "?type=internal"
            selected_status = clean_text(r.SetupStatus) or "Needs Payment Schedule"
            selected_type = clean_text(r.PaymentType) or "Single Payment"
            multi_payment_types = {"Multiple Payments", "Deposit + Final", "Progress Payments", "Monthly", "Milestone", "Retainage"}
            payment_type_options = ""
            for opt in ["Single Payment", "Multiple Payments", "Deposit + Final", "Progress Payments", "Monthly", "Milestone", "Retainage", "Other"]:
                sel = " selected" if opt == selected_type else ""
                payment_type_options += f'<option value="{h(opt)}"{sel}>{h(opt)}</option>'

            status_options = ""
            for opt in ["Needs Payment Schedule", "Assigned to PM", "In Progress", "Needs Info", "Complete", "Not Required"]:
                sel = " selected" if opt == selected_status else ""
                status_options += f'<option value="{h(opt)}"{sel}>{h(opt)}</option>'

            missing_bits = []
            if getattr(r, "MissingVendor", 0):
                missing_bits.append("Vendor")
            if getattr(r, "MissingPaymentSchedule", 0):
                missing_bits.append("Payment schedule")
            if getattr(r, "MissingPaymentType", 0):
                missing_bits.append("Payment type")
            if getattr(r, "MissingExpectedPaymentDate", 0):
                missing_bits.append("Expected payment date")
            missing_html = ", ".join(missing_bits) if missing_bits else "No required info missing"

            expected_payment_date = "" if not r.ExpectedPaymentDate else str(r.ExpectedPaymentDate)[:10]

            existing_schedule_lines = [line.strip() for line in str(r.PaymentSchedule or "").splitlines() if line.strip()]
            requires_multiple = selected_type in multi_payment_types or len(existing_schedule_lines) > 1
            schedule_inputs = '<input type="hidden" name="expected_payment_date" value="' + h(expected_payment_date) + '">'
            for idx in range(1, 5):
                existing_line = existing_schedule_lines[idx - 1] if idx - 1 < len(existing_schedule_lines) else ""
                row_style = "" if idx == 1 or requires_multiple else "display:none;"
                date_value = expected_payment_date if idx == 1 and expected_payment_date else ""
                schedule_inputs += f'''
                    <div class="payment-schedule-row" data-payment-row="{idx}" style="{row_style}">
                        <div class="payment-row-label">Payment {idx}</div>
                        <input type="date" name="payment_{idx}_date" value="{h(date_value if idx == 1 else '')}" aria-label="Payment {idx} date">
                        <input type="text" name="payment_{idx}_amount" placeholder="Amount/%" aria-label="Payment {idx} amount or percent">
                        <input type="text" name="payment_{idx}_note" value="{h(existing_line)}" placeholder="Milestone or terms" aria-label="Payment {idx} note">
                    </div>
                '''
            assigned_options = '<option value="">Unassigned</option>'
            selected_assigned = (clean_text(r.SetupAssignedTo) or "").lower()
            for u in assignable_users:
                user_email = (clean_text(u.Email) or "").lower()
                if not user_email:
                    continue
                display = clean_text(u.DisplayName) or user_email
                role = clean_text(u.RoleName)
                label_text = display + (f" ({role})" if role else "")
                sel = " selected" if user_email == selected_assigned else ""
                assigned_options += f'<option value="{h(user_email)}"{sel}>{h(label_text)}</option>'

            table_rows += f"""
            <tr>
                <form method="post" action="/project-po-setup">
                    <input type="hidden" name="po_number" value="{h(r.PONumber)}">
                    <td class="po-number-cell"><div class="po-review-id"><a class="po-link" href="{packet_url}">{h(r.PONumber)}</a><span class="po-status-row">{status_chip(selected_status)}</span></div></td>
                    <td>{h(r.ProjectName)}<br><small>{h(r.Department)}</small></td>
                    <td>{('<input type="text" name="vendor_name" value="" placeholder="Enter vendor" aria-label="Vendor name">' if getattr(r, "MissingVendor", 0) else h(r.VendorName))}<br><small>{currency(r.POValue)}</small></td>
                    <td>{h(r.Requestor or '')}</td>
                    <td><select name="payment_type" onchange="togglePaymentScheduleRows(this)">{payment_type_options}</select></td>
                    <td class="payment-schedule-cell"><div class="payment-schedule-builder">{schedule_inputs}</div></td>
                    <td class="assign-cell"><select name="setup_assigned_to">{assigned_options}</select></td>
                    <td><select name="setup_status">{status_options}</select></td>
                    <td class="inline-actions">
                        <button class="primary" name="setup_action" value="save" type="submit">Update</button>
                        <button class="secondary" name="setup_action" value="assign" type="submit">Assign</button>
                        <button class="secondary" name="setup_action" value="complete" type="submit">Complete</button>
                    </td>
                </form>
            </tr>
            """

        if not table_rows:
            table_rows = '<tr><td colspan="9"><div class="empty-state"><strong>No PO setup items found.</strong><span>POs missing payment schedules, payment type, or expected payment dates will appear here.</span></div></td></tr>'

        mine_link = ""
        if user["email"]:
            mine_link = f'<a class="button secondary" href="/project-po-setup?mine=1">My Assigned PO Info Tasks</a>'

        content = f"""
<div class="card">
            <div style="display:flex; justify-content:space-between; gap:12px; align-items:flex-start; flex-wrap:wrap;">
                <div>
                    <h3>PO Setup Review Table</h3>
                    <p class="card-subtitle">Update missing payment schedule information directly, or assign the PO to a project manager so it appears as an action item on their dashboard.</p>
                </div>
                <div>{mine_link}</div>
            </div>
            <div class="table-wrap setup-table">
                <table id="poInfoReviewTable">
                    <thead>
                        <tr>
                            <th>PO</th>
                            <th class="project-filter-th">
                                <div class="project-filter-head"><span>Project</span><button class="project-filter-button" type="button" onclick="togglePOProjectFilter()" title="Filter by project" aria-label="Filter by project"><span class="filter-icon">▼</span></button></div>
                                <select id="poProjectFilter" class="project-filter-select" onchange="filterPOInfoReviewByProject()">{project_filter_options}</select>
                            </th>
                            <th>Vendor / Amount</th><th>Requestor</th><th>Payment Type</th><th>Payment Schedule</th><th>Assigned To</th><th>Status</th><th>Actions</th>
                        </tr>
                    </thead>
                    <tbody>
                    {table_rows}
                    </tbody>
                </table>
            </div>
        </div>
        """
        content += """
        <script>
        function poSetupPaymentTypeRequiresMultiple(value) {
            return ["Multiple Payments", "Deposit + Final", "Progress Payments", "Monthly", "Milestone", "Retainage"].includes(value || "");
        }
        function togglePaymentScheduleRows(selectEl) {
            const rowContainer = selectEl.closest("tr") || selectEl.closest("form") || document;
            const requiresMultiple = poSetupPaymentTypeRequiresMultiple(selectEl.value);
            rowContainer.querySelectorAll("[data-payment-row]").forEach(function(row) {
                const idx = Number(row.getAttribute("data-payment-row") || "1");
                const show = idx === 1 || requiresMultiple;
                row.style.display = show ? "grid" : "none";
                row.querySelectorAll("input").forEach(function(input) {
                    if (idx > 1 && !show) input.value = "";
                });
            });
        }
        function togglePOProjectFilter() {
            const select = document.getElementById("poProjectFilter");
            if (!select) return;
            select.classList.toggle("show");
            if (select.classList.contains("show")) select.focus();
        }
        function filterPOInfoReviewByProject() {
            const table = document.getElementById("poInfoReviewTable");
            const select = document.getElementById("poProjectFilter");
            if (!table || !select) return;
            const selectedProject = (select.value || "").trim().toLowerCase();
            table.querySelectorAll("tbody tr").forEach(function(row) {
                const cells = row.querySelectorAll("td");
                if (!cells.length) return;
                const projectText = (cells[1] ? cells[1].innerText : "").toLowerCase();
                row.style.display = (!selectedProject || projectText.includes(selectedProject)) ? "" : "none";
            });
        }
        document.addEventListener("DOMContentLoaded", function() {
            document.querySelectorAll('select[name="payment_type"]').forEach(togglePaymentScheduleRows);
            filterPOInfoReviewByProject();
        });
        </script>
        """
        return shell("PO Setup Review", "Update missing PO planning details and assign follow-up work.", "PO Setup Review", content)

    except Exception as e:
        content = f'<div class="notice error">Error loading PO Info Review: {h(e)}</div>'
        return shell("PO Setup Review", "Unable to load setup queue.", "PO Setup Review", content), 500


@app.route("/expenses")
def expenses_page():
    allowed, reason = require_page_access("Expenses")
    if not allowed:
        return access_denied_response("Expenses", reason)
    try:
        ensure_expense_review_tables()
        filters = {
            "project": clean_text(request.args.get("project")) or "",
            "vendor": clean_text(request.args.get("vendor")) or "",
            "type": clean_text(request.args.get("type")) or "",
            "status": clean_text(request.args.get("status")) or "",
            "decision": clean_text(request.args.get("decision")) or "",
            "po": clean_text(request.args.get("po")) or "",
            "search": clean_text(request.args.get("search")) or "",
        }
        where = []
        params = []
        add_like_filter(where, params, "ProjectName", filters["project"])
        add_like_filter(where, params, "VendorName", filters["vendor"])
        add_like_filter(where, params, "TxType", filters["type"])
        add_like_filter(where, params, "MatchStatus", filters["status"])
        add_like_filter(where, params, "ReviewDecision", filters["decision"])
        if filters["po"]:
            where.append("LOWER(COALESCE(CorrectPONumber, MatchedPONumber, ExtractedPONumber, '')) LIKE LOWER(?)")
            params.append("%" + filters["po"] + "%")
        if filters["search"]:
            where.append("LOWER(CONCAT(COALESCE(ExpenseId,''), ' ', COALESCE(ProjectName,''), ' ', COALESCE(TxType,''), ' ', COALESCE(VendorName,''), ' ', COALESCE(Description,''), ' ', COALESCE(PMComments,''))) LIKE LOWER(?)")
            params.append("%" + filters["search"] + "%")
        where_sql = "WHERE " + " AND ".join(where) if where else ""

        conn = get_sql_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT
                COUNT(*) AS TotalRows,
                SUM(CASE WHEN MatchStatus IN ('Auto Matched','Manually Matched') THEN 1 ELSE 0 END) AS MatchedRows,
                SUM(CASE WHEN MatchStatus IN ('Needs Review','No Match') THEN 1 ELSE 0 END) AS ReviewRows,
                SUM(CASE WHEN COALESCE(PostedToPO, 0) = 1 THEN 1 ELSE 0 END) AS PostedRows,
                SUM(COALESCE(Amount, 0)) AS TotalAmount,
                SUM(CASE WHEN COALESCE(PostedToPO, 0) = 1 THEN COALESCE(PostedAmount, Amount, 0) ELSE 0 END) AS PostedAmount
            FROM dbo.ExpenseReviewItems;
        """)
        stats = cursor.fetchone()
        cursor.execute(
            f"""
            SELECT TOP 750 ExpenseReviewItemId, ExpenseId, ProjectName, TxDate, TxType, VendorName,
                   Description, Amount, PMComments, ExtractedPONumber, MatchedPONumber, MatchStatus,
                   MatchConfidence, ReviewDecision, CorrectPONumber, PostedToPO, PostedPONumber,
                   PostedAmount, PostedAt, PostedBy, ReviewerEmail, ReviewedAt, IsDuplicate
            FROM dbo.ExpenseReviewItems
            {where_sql}
            ORDER BY TxDate DESC, ExpenseReviewItemId DESC;
            """,
            *params,
        )
        rows = cursor.fetchall()
        conn.close()

        table_rows = ""
        for r in rows:
            status_class = str(r.MatchStatus or "").lower().replace(" ", "-")
            po_number = r.CorrectPONumber or r.MatchedPONumber or r.ExtractedPONumber or ""
            posted_po = r.PostedPONumber or po_number
            po_link = f'<a href="/po-packet/{quote_plus(str(posted_po))}?type=internal">{h(posted_po)}</a>' if posted_po else ""
            unpost_button = ""
            if getattr(r, "PostedToPO", 0):
                unpost_button = (
                    f'<form method="post" action="/expense-review/unpost" onsubmit="return confirm(\'Reverse this posted expense from the PO balance?\');">'
                    f'<input type="hidden" name="item_id" value="{h(r.ExpenseReviewItemId)}">'
                    f'<button class="secondary" type="submit">Unpost</button></form>'
                )
            table_rows += f"""
            <tr>
                <td><span class="status-chip {h(status_class)}">{h(r.MatchStatus)}</span><br>{posting_status_chip(r)}<br><small>{h(r.MatchConfidence)}</small></td>
                <td><strong>{h(r.ExpenseId or r.ExpenseReviewItemId)}</strong><br><small>{h(r.TxDate)}</small></td>
                <td>{h(r.ProjectName)}</td>
                <td>{h(r.TxType)}</td>
                <td>{h(r.VendorName)}</td>
                <td class="right">{currency(r.Amount)}</td>
                <td>{h(r.Description)}<br><small>{h(r.PMComments)}</small></td>
                <td>{po_link}</td>
                <td>{h(r.ReviewDecision)}</td>
                <td class="posting-audit-cell">{h(posting_reason(r))}<br><strong>{currency(getattr(r, 'PostedAmount', 0)) if getattr(r, 'PostedToPO', 0) else ''}</strong><br><small>{h(getattr(r, 'PostedBy', '') or '')} {h(getattr(r, 'PostedAt', '') or '')}</small>{unpost_button}</td>
            </tr>
            """
        if not table_rows:
            table_rows = '<tr><td colspan="10"><div class="empty-state"><strong>No expenses found.</strong><span>Adjust filters or upload a Unanet expense file.</span></div></td></tr>'

        filter_bar = build_simple_filter_bar(filters, "/expenses", [
            ("project", "Project", "Project name/code"),
            ("vendor", "Vendor", "Vendor or purchaser"),
            ("type", "Type", "Purchase, Disbursement..."),
            ("status", "Match Status", "Auto Matched, No Match..."),
            ("decision", "Decision", "Matched to PO, No PO Needed..."),
            ("po", "PO", "PO number"),
            ("search", "Search", "Description or comments"),
        ])
        content = f"""
        <div class="grid kpis">
            <a class="card kpi status-card" href="/expense-upload"><div class="label">Expense Rows</div><div class="value">{int(stats.TotalRows or 0)}</div><div class="trend">Imported from Unanet uploads</div></a>
            <a class="card kpi status-card" href="/expenses"><div class="label">Expense Amount</div><div class="value">{currency(stats.TotalAmount)}</div><div class="trend">Total uploaded expense value</div></a>
            <a class="card kpi status-card" href="/pos-in-pm-comments"><div class="label">Matched / Linked</div><div class="value">{int(stats.MatchedRows or 0)}</div><div class="trend">Auto or manually matched rows</div></a>
            <a class="card kpi status-card" href="/missing-po-review"><div class="label">Needs Review</div><div class="value">{int(stats.ReviewRows or 0)}</div><div class="trend">Rows needing PO decision</div></a>
            <a class="card kpi status-card" href="/pos-balances"><div class="label">Posted to POs</div><div class="value">{currency(stats.PostedAmount)}</div><div class="trend">{int(stats.PostedRows or 0)} rows reducing balances</div></a>
        </div>
        {filter_bar}
        <div class="card">
            <h3>Expenses</h3>
            <p class="card-subtitle">Imported Unanet expense rows with match status, posting status, and audit detail. Posted rows reduce Current App Balance.</p>
            <div class="table-wrap"><table id="expensesTable">
                <tr><th>Status</th><th>Expense</th><th>Project</th><th>Type</th><th>Vendor / Purchaser</th><th class="right">Amount</th><th>Description / PM Comments</th><th>PO</th><th>Decision</th><th>Posting Audit</th></tr>
                {table_rows}
            </table></div>
        </div>
        """
        return shell("Expenses", "Imported Unanet expense rows and PO match status.", "Expenses", content)
    except Exception as e:
        return shell("Expenses", "Unable to load expenses.", "Expenses", f'<div class="notice error">Error loading Expenses: {h(e)}</div>'), 500


@app.route("/missing-po-review")
def missing_po_review():
    allowed, reason = require_page_access("Missing PO Review")
    if not allowed:
        return access_denied_response("Missing PO Review", reason)
    try:
        ensure_expense_review_tables()
        filters = {
            "project": clean_text(request.args.get("project")) or "",
            "vendor": clean_text(request.args.get("vendor")) or "",
            "type": clean_text(request.args.get("type")) or "",
            "status": clean_text(request.args.get("status")) or "",
            "decision": clean_text(request.args.get("decision")) or "",
            "search": clean_text(request.args.get("search")) or "",
        }
        where = ["(COALESCE(ReviewDecision, 'Pending Review') IN ('Pending Review','Needs PM Review','Hold for More Info') OR MatchStatus IN ('Needs Review','No Match'))"]
        params = []
        add_like_filter(where, params, "ProjectName", filters["project"])
        add_like_filter(where, params, "VendorName", filters["vendor"])
        add_like_filter(where, params, "TxType", filters["type"])
        add_like_filter(where, params, "MatchStatus", filters["status"])
        add_like_filter(where, params, "ReviewDecision", filters["decision"])
        if filters["search"]:
            where.append("LOWER(CONCAT(COALESCE(ExpenseId,''), ' ', COALESCE(ProjectName,''), ' ', COALESCE(VendorName,''), ' ', COALESCE(Description,''), ' ', COALESCE(PMComments,''), ' ', COALESCE(MatchReason,''))) LIKE LOWER(?)")
            params.append("%" + filters["search"] + "%")
        where_sql = "WHERE " + " AND ".join(where)
        conn = get_sql_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT PONumber FROM dbo.PurchaseOrders WHERE PONumber IS NOT NULL ORDER BY PONumber")
        po_numbers = [row.PONumber for row in cursor.fetchall()]
        cursor.execute(
            f"""
            SELECT TOP 500 ExpenseReviewItemId, ExpenseId, ProjectName, TxDate, TxType, VendorName,
                   Description, Amount, PMComments, ExtractedPONumber, MatchedPONumber, MatchStatus,
                   MatchConfidence, MatchReason, ReviewDecision, CorrectPONumber, ReviewerNotes,
                   PostedToPO, PostedPONumber, PostedAmount, PostedAt, PostedBy, ReviewerEmail, ReviewedAt, IsDuplicate
            FROM dbo.ExpenseReviewItems
            {where_sql}
            ORDER BY CASE WHEN MatchStatus = 'No Match' THEN 0 WHEN MatchStatus = 'Needs Review' THEN 1 ELSE 2 END,
                     ABS(COALESCE(Amount,0)) DESC, ExpenseReviewItemId DESC;
            """,
            *params,
        )
        rows = cursor.fetchall()
        cursor.execute(
            """
            SELECT COUNT(*) AS TotalOpen,
                   SUM(CASE WHEN MatchStatus = 'No Match' THEN 1 ELSE 0 END) AS NoMatch,
                   SUM(CASE WHEN MatchStatus = 'Needs Review' THEN 1 ELSE 0 END) AS NeedsReview,
                   SUM(COALESCE(Amount, 0)) AS OpenAmount
            FROM dbo.ExpenseReviewItems
            WHERE COALESCE(ReviewDecision, 'Pending Review') IN ('Pending Review','Needs PM Review','Hold for More Info')
               OR MatchStatus IN ('Needs Review','No Match');
            """
        )
        stats = cursor.fetchone()
        conn.close()

        po_options = "".join(f'<option value="{h(po)}">' for po in po_numbers)
        row_html = ""
        for r in rows:
            suggested = r.MatchedPONumber or r.ExtractedPONumber or ""
            status_class = str(r.MatchStatus or "").lower().replace(" ", "-")
            row_html += f"""
            <tr>
                <td><span class="status-chip {h(status_class)}">{h(r.MatchStatus)}</span><br>{posting_status_chip(r)}<br><small>{h(r.MatchConfidence)}</small></td>
                <td><strong>{h(r.ExpenseId or r.ExpenseReviewItemId)}</strong><br><small>{h(r.TxDate)}</small></td>
                <td>{h(r.ProjectName)}</td>
                <td>{h(r.VendorName)}<br><small>{h(r.TxType)}</small></td>
                <td class="right">{currency(r.Amount)}</td>
                <td>{h(r.Description)}<br><small>{h(r.PMComments)}</small></td>
                <td>{h(suggested)}<br><small>{h(r.MatchReason)}</small></td>
                <td class="posting-audit-cell">{h(posting_reason(r))}<br><small>{h(r.PostedPONumber or '')} {currency(r.PostedAmount) if r.PostedToPO else ''}</small></td>
                <td class="expense-action-cell">
                    <form class="expense-match-form" method="post" action="/expense-review/update">
                        <input type="hidden" name="item_id" value="{h(r.ExpenseReviewItemId)}">
                        <input list="poNumberList" name="correct_po_number" value="{h(r.CorrectPONumber or suggested)}" placeholder="Correct PO number">
                        <select name="review_decision">
                            {select_option('Pending Review', r.ReviewDecision or 'Pending Review')}
                            {select_option('Matched to PO', r.ReviewDecision or 'Pending Review')}
                            {select_option('No PO Needed', r.ReviewDecision or 'Pending Review')}
                            {select_option('Needs PM Review', r.ReviewDecision or 'Pending Review')}
                            {select_option('Hold for More Info', r.ReviewDecision or 'Pending Review')}
                        </select>
                        <textarea name="reviewer_notes" placeholder="Reviewer notes">{h(r.ReviewerNotes)}</textarea>
                        <button class="primary" type="submit">Save Decision</button>
                    </form>
                </td>
            </tr>
            """
        if not row_html:
            row_html = '<tr><td colspan="9"><div class="empty-state"><strong>No missing PO review items.</strong><span>Rows needing PO decisions will appear here after expense uploads.</span></div></td></tr>'
        filter_bar = build_simple_filter_bar(filters, "/missing-po-review", [("project", "Project", "Project name/code"), ("vendor", "Vendor", "Vendor or purchaser"), ("type", "Type", "Purchase, Disbursement..."), ("status", "Match Status", "No Match, Needs Review..."), ("decision", "Decision", "Pending Review..."), ("search", "Search", "Description or comments")])
        content = f"""
        <div class="notice info"><strong>Missing PO Review:</strong> Use this page to review expenses that may need a PO or corrected PO reference. Rows only reduce PO balances after they are posted as valid matches.</div>
        <div class="grid kpis">
            <a class="card kpi status-card" href="/missing-po-review"><div class="label">Open Review Items</div><div class="value">{int(stats.TotalOpen or 0)}</div><div class="trend">Rows needing PO decision</div></a>
            <a class="card kpi status-card" href="/missing-po-review?status=No+Match"><div class="label">No Match</div><div class="value">{int(stats.NoMatch or 0)}</div><div class="trend">No clear PO found</div></a>
            <a class="card kpi status-card" href="/missing-po-review?status=Needs+Review"><div class="label">Needs Review</div><div class="value">{int(stats.NeedsReview or 0)}</div><div class="trend">Suggested match needs validation</div></a>
            <a class="card kpi status-card" href="/missing-po-review"><div class="label">Review Amount</div><div class="value">{currency(stats.OpenAmount)}</div><div class="trend">Open review dollar value</div></a>
        </div>
        {filter_bar}
        <div class="card" id="review-table">
            <h3>Missing PO Review</h3>
            <p class="card-subtitle">PMs can help identify the correct PO; Accounting/Admin controls upload and posting oversight.</p>
            <datalist id="poNumberList">{po_options}</datalist>
            <div class="table-wrap expense-review-table"><table>
                <tr><th>Status</th><th>Expense</th><th>Project</th><th>Vendor / Purchaser</th><th class="right">Amount</th><th>Description / PM Comments</th><th>Suggested PO</th><th>Posting</th><th>Review Action</th></tr>
                {row_html}
            </table></div>
        </div>
        """
        return shell("Missing PO Review", "Expense rows that need a PO decision.", "Missing PO Review", content)
    except Exception as e:
        return shell("Missing PO Review", "Unable to load missing PO review.", "Missing PO Review", f'<div class="notice error">Error loading Missing PO Review: {h(e)}</div>'), 500



@app.route("/vendors")
def vendors_page():
    allowed, reason = require_page_access("Vendors")
    if not allowed:
        return access_denied_response("Vendors", reason)
    try:
        ensure_expense_review_tables()
        conn = get_sql_connection()
        cursor = conn.cursor()
        req_where_po, req_params_po = requestor_filter_sql("po")
        req_where_line, req_params_line = requestor_filter_sql("l")

        selected_vendor = clean_text(request.args.get("vendor")) or ""
        selected_vendor_param = selected_vendor

        # Vendor rollup is the top-level vendor directory. It uses visible PO data and
        # expense rows tied to visible posted PO numbers so role restrictions still apply.
        cursor.execute(
            f"""
            WITH VisiblePOs AS (
                SELECT
                    po.PurchaseOrderId,
                    po.PONumber,
                    COALESCE(v.VendorName, 'Missing Vendor') AS VendorName,
                    COALESCE(po.RevisedAmount, po.OriginalAmount, 0) AS POAmount,
                    COALESCE(po.RemainingAmount, COALESCE(po.RevisedAmount, po.OriginalAmount, 0)) AS RemainingAmount,
                    COALESCE(po.POStatus, 'Open') AS POStatus
                FROM dbo.PurchaseOrders po
                LEFT JOIN dbo.Vendors v ON po.VendorId = v.VendorId
                WHERE {req_where_po}
            ), PostedExpenses AS (
                SELECT
                    COALESCE(NULLIF(LTRIM(RTRIM(e.VendorName)), ''), 'Missing Vendor') AS VendorName,
                    COUNT(*) AS ExpenseRows,
                    SUM(COALESCE(e.PostedAmount, e.Amount, 0)) AS ExpenseAmount,
                    MAX(e.PostedAt) AS LastPostedAt
                FROM dbo.ExpenseReviewItems e
                INNER JOIN VisiblePOs vp ON vp.PONumber = e.PostedPONumber
                WHERE COALESCE(e.PostedToPO, 0) = 1
                GROUP BY COALESCE(NULLIF(LTRIM(RTRIM(e.VendorName)), ''), 'Missing Vendor')
            ), POVendors AS (
                SELECT
                    VendorName,
                    COUNT(DISTINCT PONumber) AS POCount,
                    SUM(COALESCE(POAmount, 0)) AS POAmount,
                    SUM(COALESCE(RemainingAmount, 0)) AS RemainingAmount,
                    SUM(CASE WHEN LOWER(COALESCE(POStatus,'')) IN ('open','issued','active') THEN 1 ELSE 0 END) AS OpenPOCount,
                    SUM(CASE WHEN LOWER(COALESCE(POStatus,'')) = 'voided' THEN 1 ELSE 0 END) AS VoidedPOCount
                FROM VisiblePOs
                GROUP BY VendorName
            )
            SELECT
                COALESCE(p.VendorName, e.VendorName) AS VendorName,
                COALESCE(p.POCount, 0) AS POCount,
                COALESCE(p.OpenPOCount, 0) AS OpenPOCount,
                COALESCE(p.VoidedPOCount, 0) AS VoidedPOCount,
                COALESCE(p.POAmount, 0) AS POAmount,
                COALESCE(p.RemainingAmount, 0) AS RemainingAmount,
                COALESCE(e.ExpenseRows, 0) AS ExpenseRows,
                COALESCE(e.ExpenseAmount, 0) AS ExpenseAmount,
                e.LastPostedAt
            FROM POVendors p
            FULL OUTER JOIN PostedExpenses e ON p.VendorName = e.VendorName
            ORDER BY COALESCE(p.POAmount, 0) DESC, COALESCE(e.ExpenseAmount, 0) DESC, COALESCE(p.VendorName, e.VendorName);
            """,
            *req_params_po,
        )
        rows = cursor.fetchall()

        # Default to the largest visible vendor so the page immediately feels like a vendor profile.
        if not selected_vendor and rows:
            selected_vendor = clean_text(rows[0].VendorName)
            selected_vendor_param = selected_vendor

        vendor_options = ""
        for r in rows:
            name = clean_text(r.VendorName) or "Missing Vendor"
            sel = " selected" if selected_vendor and name.lower() == selected_vendor.lower() else ""
            vendor_options += f'<option value="{h(name)}"{sel}>{h(name)}</option>'

        vendor_po_rows = []
        vendor_project_rows = []
        vendor_line_rows = []
        vendor_tx_rows = []
        vendor_profile = None

        if selected_vendor:
            cursor.execute(
                f"""
                WITH PostedExpenses AS (
                    SELECT PostedPONumber AS PONumber,
                           SUM(COALESCE(PostedAmount, Amount, 0)) AS PostedExpenseAmount,
                           COUNT(*) AS PostedExpenseRows,
                           MAX(PostedAt) AS LastPostedAt
                    FROM dbo.ExpenseReviewItems
                    WHERE COALESCE(PostedToPO, 0) = 1 AND COALESCE(PostedPONumber, '') <> ''
                    GROUP BY PostedPONumber
                ), LineRollup AS (
                    SELECT PurchaseOrderId,
                           PONumber,
                           COUNT(*) AS LineCount,
                           SUM(COALESCE(LineAmount, 0)) AS LineAmount
                    FROM dbo.IssuedPOLines
                    GROUP BY PurchaseOrderId, PONumber
                )
                SELECT
                    po.PONumber,
                    COALESCE(pr.ProjectCode, '') AS ProjectCode,
                    COALESCE(pr.ProjectName, (SELECT TOP 1 l2.ProjectName FROM dbo.IssuedPOLines l2 WHERE l2.PONumber = po.PONumber AND COALESCE(l2.ProjectName,'') <> '' ORDER BY l2.IssuedPOLineId), '') AS ProjectName,
                    COALESCE(NULLIF(po.Department, ''), (SELECT TOP 1 l3.Department FROM dbo.IssuedPOLines l3 WHERE l3.PONumber = po.PONumber AND COALESCE(l3.Department,'') <> '' ORDER BY l3.IssuedPOLineId), '') AS Department,
                    COALESCE(po.POStatus, 'Open') AS POStatus,
                    po.PODate,
                    po.Requestor,
                    COALESCE(po.RevisedAmount, po.OriginalAmount, lr.LineAmount, 0) AS POValue,
                    COALESCE(pe.PostedExpenseAmount, 0) AS PostedExpenseAmount,
                    CASE WHEN COALESCE(po.RevisedAmount, po.OriginalAmount, lr.LineAmount, 0) - COALESCE(pe.PostedExpenseAmount, 0) < 0 THEN 0 ELSE COALESCE(po.RevisedAmount, po.OriginalAmount, lr.LineAmount, 0) - COALESCE(pe.PostedExpenseAmount, 0) END AS CurrentAppBalance,
                    COALESCE(lr.LineCount, 0) AS LineCount,
                    COALESCE(pe.PostedExpenseRows, 0) AS PostedExpenseRows,
                    pe.LastPostedAt
                FROM dbo.PurchaseOrders po
                LEFT JOIN dbo.Vendors v ON po.VendorId = v.VendorId
                LEFT JOIN dbo.Projects pr ON po.ProjectId = pr.ProjectId
                LEFT JOIN LineRollup lr ON lr.PurchaseOrderId = po.PurchaseOrderId OR lr.PONumber = po.PONumber
                LEFT JOIN PostedExpenses pe ON pe.PONumber = po.PONumber
                WHERE LOWER(COALESCE(v.VendorName, 'Missing Vendor')) = LOWER(?) AND {req_where_po}
                GROUP BY po.PONumber, pr.ProjectCode, pr.ProjectName, po.Department, po.POStatus, po.PODate, po.Requestor, po.RevisedAmount, po.OriginalAmount, lr.LineAmount, lr.LineCount, pe.PostedExpenseAmount, pe.PostedExpenseRows, pe.LastPostedAt
                ORDER BY CurrentAppBalance DESC, POValue DESC, po.PONumber;
                """,
                selected_vendor_param,
                *req_params_po,
            )
            vendor_po_rows = cursor.fetchall()

            cursor.execute(
                f"""
                SELECT TOP 500
                    l.PONumber,
                    COALESCE(pr.ProjectCode, '') AS ProjectCode,
                    COALESCE(pr.ProjectName, l.ProjectName, '') AS ProjectName,
                    l.Department,
                    l.LineDescription,
                    l.Unit,
                    l.UnitCost,
                    l.Qty,
                    l.LineAmount
                FROM dbo.IssuedPOLines l
                LEFT JOIN dbo.PurchaseOrders po ON l.PurchaseOrderId = po.PurchaseOrderId
                LEFT JOIN dbo.Vendors v ON po.VendorId = v.VendorId
                LEFT JOIN dbo.Projects pr ON po.ProjectId = pr.ProjectId
                WHERE LOWER(COALESCE(v.VendorName, l.VendorName, 'Missing Vendor')) = LOWER(?) AND {req_where_line}
                ORDER BY l.PONumber, l.IssuedPOLineId;
                """,
                selected_vendor_param,
                *req_params_line,
            )
            vendor_line_rows = cursor.fetchall()

            cursor.execute(
                f"""
                WITH VisibleVendorPOs AS (
                    SELECT po.PONumber
                    FROM dbo.PurchaseOrders po
                    LEFT JOIN dbo.Vendors v ON po.VendorId = v.VendorId
                    WHERE LOWER(COALESCE(v.VendorName, 'Missing Vendor')) = LOWER(?) AND {req_where_po}
                )
                SELECT TOP 100
                    e.TxDate,
                    e.TxType,
                    e.ProjectName,
                    e.Description,
                    e.PMComments,
                    e.Amount,
                    e.PostedAmount,
                    e.PostedPONumber,
                    e.PostedAt,
                    e.PostedBy,
                    e.ReviewDecision,
                    e.MatchStatus
                FROM dbo.ExpenseReviewItems e
                INNER JOIN VisibleVendorPOs vp ON vp.PONumber = e.PostedPONumber
                WHERE COALESCE(e.PostedToPO, 0) = 1
                ORDER BY COALESCE(e.PostedAt, e.TxDate) DESC, e.ExpenseReviewItemId DESC;
                """,
                selected_vendor_param,
                *req_params_po,
            )
            vendor_tx_rows = cursor.fetchall()

            # Project rollup is built from the selected vendor POs to avoid repeating
            # visibility logic in a second complex SQL query.
            project_rollup = {}
            for p in vendor_po_rows:
                pcode = clean_text(getattr(p, "ProjectCode", ""))
                pname = clean_text(getattr(p, "ProjectName", "")) or "Missing Project"
                key = pcode or pname
                if key not in project_rollup:
                    project_rollup[key] = {"code": pcode, "name": pname, "po_count": 0, "po_value": 0.0, "posted": 0.0, "balance": 0.0}
                project_rollup[key]["po_count"] += 1
                project_rollup[key]["po_value"] += float(getattr(p, "POValue", 0) or 0)
                project_rollup[key]["posted"] += float(getattr(p, "PostedExpenseAmount", 0) or 0)
                project_rollup[key]["balance"] += float(getattr(p, "CurrentAppBalance", 0) or 0)
            vendor_project_rows = sorted(project_rollup.values(), key=lambda x: x["po_value"], reverse=True)

            vendor_profile = {
                "po_count": len(vendor_po_rows),
                "open_po_count": sum(1 for p in vendor_po_rows if str(getattr(p, "POStatus", "") or "").lower() in ["open", "issued", "active"]),
                "voided_po_count": sum(1 for p in vendor_po_rows if str(getattr(p, "POStatus", "") or "").lower() == "voided"),
                "po_value": sum(float(getattr(p, "POValue", 0) or 0) for p in vendor_po_rows),
                "posted": sum(float(getattr(p, "PostedExpenseAmount", 0) or 0) for p in vendor_po_rows),
                "balance": sum(float(getattr(p, "CurrentAppBalance", 0) or 0) for p in vendor_po_rows),
                "projects": len(vendor_project_rows),
                "line_count": sum(int(getattr(p, "LineCount", 0) or 0) for p in vendor_po_rows),
                "last_posted": max([clean_text(getattr(p, "LastPostedAt", "")) for p in vendor_po_rows if clean_text(getattr(p, "LastPostedAt", ""))] or [""]),
            }

        conn.close()

        total_vendors = len(rows)
        total_po_amount = sum(float(r.POAmount or 0) for r in rows)
        total_remaining = sum(float(r.RemainingAmount or 0) for r in rows)
        total_expense_amount = sum(float(r.ExpenseAmount or 0) for r in rows)
        total_open_pos = sum(int(r.OpenPOCount or 0) for r in rows)

        table_rows = ""
        top_rows = ""
        max_po = max([float(r.POAmount or 0) for r in rows] or [1])
        for idx, r in enumerate(rows):
            vendor_name = clean_text(r.VendorName) or "Missing Vendor"
            table_rows += f"""
            <tr>
                <td><a class="vendor-detail-link" href="/vendors?vendor={quote_plus(vendor_name)}">{h(vendor_name)}</a></td>
                <td class="right">{int(r.POCount or 0)}</td>
                <td class="right">{int(r.OpenPOCount or 0)}</td>
                <td class="right">{currency(r.POAmount)}</td>
                <td class="right">{currency(r.RemainingAmount)}</td>
                <td class="right">{int(r.ExpenseRows or 0)}</td>
                <td class="right">{currency(r.ExpenseAmount)}</td>
            </tr>
            """
            if idx < 8:
                width = 0 if max_po == 0 else max(5, float(r.POAmount or 0) / max_po * 100)
                top_rows += f"<a class='bar-row clickable-row' href='/vendors?vendor={quote_plus(vendor_name)}'><strong>{h(vendor_name)}</strong><div class='bar-track'><div class='bar-fill' style='width:{width:.1f}%'></div></div><div class='right'>{currency(r.POAmount)}</div></a>"
        if not table_rows:
            table_rows = '<tr><td colspan="7"><div class="empty-state"><strong>No vendor data found.</strong><span>Vendors will appear after PO or expense uploads.</span></div></td></tr>'
            top_rows = '<p class="card-subtitle">No vendor data found.</p>'

        vendor_po_html = ""
        vendor_project_html = ""
        vendor_line_html = ""
        vendor_tx_html = ""
        selected_vendor_section = ""

        if selected_vendor:
            for p in vendor_po_rows:
                project_value = clean_text(getattr(p, "ProjectCode", "")) or clean_text(getattr(p, "ProjectName", ""))
                project_label = (clean_text(getattr(p, "ProjectCode", "")) + " - " if clean_text(getattr(p, "ProjectCode", "")) else "") + (clean_text(getattr(p, "ProjectName", "")) or "Missing Project")
                vendor_po_html += f"""
                <tr>
                    <td><a class="po-link" href="/po-packet/{quote_plus(str(p.PONumber or ''))}?type=internal">{h(p.PONumber)}</a></td>
                    <td><a class="project-link" href="/projects?project={quote_plus(project_value)}">{h(project_label)}</a></td>
                    <td>{h(p.Department)}</td>
                    <td>{status_chip(p.POStatus or 'Open')}</td>
                    <td class="right">{currency(p.POValue)}</td>
                    <td class="right">{currency(p.PostedExpenseAmount)}</td>
                    <td class="right">{currency(p.CurrentAppBalance)}</td>
                    <td class="right">{int(p.LineCount or 0)}</td>
                    <td><a class="pill-link" href="/po-packet-pdf/{quote_plus(str(p.PONumber or ''))}?type=vendor">Vendor PDF</a></td>
                </tr>
                """
            if not vendor_po_html:
                vendor_po_html = '<tr><td colspan="9"><div class="empty-state"><strong>No visible POs found for this vendor.</strong></div></td></tr>'

            max_project_value = max([float(p["po_value"] or 0) for p in vendor_project_rows] or [1])
            for prj in vendor_project_rows[:12]:
                value = prj["code"] or prj["name"]
                label = (prj["code"] + " - " if prj["code"] else "") + prj["name"]
                pct = 0 if max_project_value == 0 else max(5, float(prj["po_value"] or 0) / max_project_value * 100)
                vendor_project_html += f"""
                <a class="vendor-project-card" href="/projects?project={quote_plus(value)}">
                    <div><strong>{h(label)}</strong><span>{int(prj['po_count'])} PO(s)</span></div>
                    <div class="mini-meter"><div style="width:{pct:.1f}%"></div></div>
                    <div class="vendor-project-amounts"><span>PO {currency(prj['po_value'])}</span><span>Balance {currency(prj['balance'])}</span></div>
                </a>
                """
            if not vendor_project_html:
                vendor_project_html = '<p class="card-subtitle">No project activity found for this vendor.</p>'

            for line in vendor_line_rows:
                project_value = clean_text(getattr(line, "ProjectCode", "")) or clean_text(getattr(line, "ProjectName", ""))
                project_label = (clean_text(getattr(line, "ProjectCode", "")) + " - " if clean_text(getattr(line, "ProjectCode", "")) else "") + (clean_text(getattr(line, "ProjectName", "")) or "Missing Project")
                vendor_line_html += f"""
                <tr>
                    <td><a class="po-link" href="/po-packet/{quote_plus(str(line.PONumber or ''))}?type=internal">{h(line.PONumber)}</a></td>
                    <td><a class="project-link" href="/projects?project={quote_plus(project_value)}">{h(project_label)}</a></td>
                    <td>{h(line.Department)}</td>
                    <td>{h(line.LineDescription)}</td>
                    <td>{h(line.Unit)}</td>
                    <td class="right">{currency(line.UnitCost)}</td>
                    <td class="right">{h(line.Qty)}</td>
                    <td class="right">{currency(line.LineAmount)}</td>
                </tr>
                """
            if not vendor_line_html:
                vendor_line_html = '<tr><td colspan="8"><div class="empty-state"><strong>No line items found for this vendor.</strong></div></td></tr>'

            for tx in vendor_tx_rows:
                amount = tx.PostedAmount if tx.PostedAmount is not None else tx.Amount
                vendor_tx_html += f"""
                <tr>
                    <td>{h(tx.TxDate)}</td>
                    <td><a class="po-link" href="/po-packet/{quote_plus(str(tx.PostedPONumber or ''))}?type=internal">{h(tx.PostedPONumber)}</a></td>
                    <td>{h(tx.ProjectName)}</td>
                    <td>{h(tx.TxType)}</td>
                    <td>{h(tx.Description or tx.PMComments)}</td>
                    <td class="right">{currency(amount)}</td>
                    <td>{h(tx.PostedAt)}</td>
                </tr>
                """
            if not vendor_tx_html:
                vendor_tx_html = '<tr><td colspan="7"><div class="empty-state"><strong>No posted transactions found for this vendor.</strong><span>Posted expense activity will appear after expense uploads are matched to POs.</span></div></td></tr>'

            vp = vendor_profile or {}
            selected_vendor_section = f"""
            <div class="vendor-profile-hero card">
                <div>
                    <div class="eyebrow">Vendor View</div>
                    <h2>{h(selected_vendor)}</h2>
                    <p class="card-subtitle">A consolidated view of visible purchase orders, projects, balances, line items, and posted transactions for this vendor.</p>
                </div>
                <div class="vendor-hero-actions">
                    <a class="button-secondary" href="/vendors">All Vendors</a>
                    <a class="button-secondary" href="/pos-balances?vendor={quote_plus(selected_vendor)}">PO Drilldown</a>
                </div>
            </div>
            <div class="grid kpis vendor-kpis">
                <a class="card kpi status-card" href="#vendor-pos"><div class="label">Visible POs</div><div class="value">{int(vp.get('po_count',0))}</div><div class="trend">{int(vp.get('open_po_count',0))} open / {int(vp.get('voided_po_count',0))} voided</div></a>
                <a class="card kpi status-card" href="#vendor-pos"><div class="label">PO Value</div><div class="value">{currency(vp.get('po_value',0))}</div><div class="trend">Issued/committed vendor amount</div></a>
                <a class="card kpi status-card" href="#vendor-transactions"><div class="label">Posted Spend</div><div class="value">{currency(vp.get('posted',0))}</div><div class="trend">Matched transactions against POs</div></a>
                <a class="card kpi status-card" href="#vendor-pos"><div class="label">Current Balance</div><div class="value">{currency(vp.get('balance',0))}</div><div class="trend">Remaining visible PO balance</div></a>
                <a class="card kpi status-card" href="#vendor-projects"><div class="label">Projects</div><div class="value">{int(vp.get('projects',0))}</div><div class="trend">Projects with vendor POs</div></a>
                <a class="card kpi status-card" href="#vendor-lines"><div class="label">Line Items</div><div class="value">{int(vp.get('line_count',0))}</div><div class="trend">Detailed PO scope rows</div></a>
            </div>
            <div class="grid two">
                <div class="card" id="vendor-projects"><h3>Projects Using This Vendor</h3><p class="card-subtitle">Click a project to open the project view.</p><div class="vendor-project-grid">{vendor_project_html}</div></div>
                <div class="card"><h3>Vendor Snapshot</h3><div class="snapshot-list">
                    <div><span>Vendor</span><strong>{h(selected_vendor)}</strong></div>
                    <div><span>Open POs</span><strong>{int(vp.get('open_po_count',0))}</strong></div>
                    <div><span>Posted Spend</span><strong>{currency(vp.get('posted',0))}</strong></div>
                    <div><span>Last Posted Transaction</span><strong>{h(vp.get('last_posted') or 'Not posted yet')}</strong></div>
                </div><p class="card-subtitle">Use this page when you know the vendor and need to see related POs, projects, and transaction history.</p></div>
            </div>
            <div class="card" id="vendor-pos"><h3>Vendor Purchase Orders</h3><p class="card-subtitle">Click a PO number to open the full PO packet, or download the vendor-facing PDF.</p><div class="table-wrap"><table><tr><th>PO</th><th>Project</th><th>Department</th><th>Status</th><th class="right">PO Value</th><th class="right">Posted</th><th class="right">Balance</th><th class="right">Lines</th><th>Packet</th></tr>{vendor_po_html}</table></div></div>
            <div class="card" id="vendor-transactions"><h3>Posted Transactions Against This Vendor's POs</h3><p class="card-subtitle">These are uploaded expenses that have been posted against the vendor's visible POs.</p><div class="table-wrap"><table><tr><th>Date</th><th>PO</th><th>Project</th><th>Type</th><th>Description / Comments</th><th class="right">Amount</th><th>Posted At</th></tr>{vendor_tx_html}</table></div></div>
            <div class="card" id="vendor-lines"><h3>Vendor PO Line Items</h3><p class="card-subtitle">Detailed scope/line item rows for visible POs tied to this vendor.</p><div class="table-wrap"><table><tr><th>PO</th><th>Project</th><th>Department</th><th>Description</th><th>Unit</th><th class="right">Unit Cost</th><th class="right">Qty</th><th class="right">Line Amount</th></tr>{vendor_line_html}</table></div></div>
            """

        content = f"""
        <div class="grid kpis">
            <div class="card kpi"><div class="label">Vendors</div><div class="value">{total_vendors}</div><div class="trend">Visible vendors across POs and posted expenses</div></div>
            <a class="card kpi status-card" href="/pos-balances"><div class="label">Open POs</div><div class="value">{int(total_open_pos)}</div><div class="trend">Vendor PO lookup</div></a>
            <div class="card kpi"><div class="label">Issued PO Amount</div><div class="value">{currency(total_po_amount)}</div><div class="trend">Visible vendor commitment value</div></div>
            <div class="card kpi"><div class="label">Remaining Balance</div><div class="value">{currency(total_remaining)}</div><div class="trend">Visible vendor balance</div></div>
            <div class="card kpi"><div class="label">Posted Spend</div><div class="value">{currency(total_expense_amount)}</div><div class="trend">Posted expense value</div></div>
        </div>
        <div class="card vendor-search-card">
            <form method="get" action="/vendors" class="inline-form vendor-select-form">
                <label><strong>Select Vendor</strong><select name="vendor"><option value="">Choose vendor...</option>{vendor_options}</select></label>
                <button class="primary" type="submit">Open Vendor View</button>
                <a class="button-secondary" href="/vendors">Reset</a>
            </form>
        </div>
        {selected_vendor_section}
        <div class="grid two">
            <div class="card"><h3>Top Vendors by PO Amount</h3><p class="card-subtitle">Click a vendor to open its vendor view.</p><div class="bar-chart vendor-bar-chart">{top_rows}</div></div>
            <div class="card"><h3>Vendor Directory</h3><p class="card-subtitle">Vendor-level totals across visible POs and posted expenses.</p><div class="table-wrap"><table><tr><th>Vendor / Purchaser</th><th class="right">POs</th><th class="right">Open</th><th class="right">PO Amount</th><th class="right">Remaining</th><th class="right">Posted Rows</th><th class="right">Posted Spend</th></tr>{table_rows}</table></div></div>
        </div>
        """
        return shell("Vendors", "Vendor view with PO balances, projects, line items, and posted transactions.", "Vendors", content)
    except Exception as e:
        return shell("Vendors", "Unable to load vendors.", "Vendors", f'<div class="notice error">Error loading Vendors: {h(e)}</div>'), 500


@app.route("/pos-in-pm-comments")
def pos_in_pm_comments_page():
    allowed, reason = require_page_access("POs in PM Comments")
    if not allowed:
        return access_denied_response("POs in PM Comments", reason)
    try:
        ensure_expense_review_tables()
        filters = {"project": clean_text(request.args.get("project")) or "", "vendor": clean_text(request.args.get("vendor")) or "", "po": clean_text(request.args.get("po")) or "", "status": clean_text(request.args.get("status")) or "", "decision": clean_text(request.args.get("decision")) or "", "search": clean_text(request.args.get("search")) or ""}
        where = ["(COALESCE(ExtractedPONumber, '') <> '' OR COALESCE(MatchedPONumber, '') <> '' OR PMComments LIKE '%PO%')", "COALESCE(PMComments, '') <> ''"]
        params = []
        add_like_filter(where, params, "ProjectName", filters["project"])
        add_like_filter(where, params, "VendorName", filters["vendor"])
        add_like_filter(where, params, "MatchStatus", filters["status"])
        add_like_filter(where, params, "ReviewDecision", filters["decision"])
        if filters["po"]:
            where.append("LOWER(COALESCE(CorrectPONumber, MatchedPONumber, ExtractedPONumber, '')) LIKE LOWER(?)")
            params.append("%" + filters["po"] + "%")
        if filters["search"]:
            where.append("LOWER(CONCAT(COALESCE(ExpenseId,''), ' ', COALESCE(ProjectName,''), ' ', COALESCE(VendorName,''), ' ', COALESCE(PMComments,''))) LIKE LOWER(?)")
            params.append("%" + filters["search"] + "%")
        where_sql = "WHERE " + " AND ".join(where)
        conn = get_sql_connection(); cursor = conn.cursor()
        cursor.execute(f"""SELECT TOP 500 ExpenseReviewItemId, ExpenseId, ProjectName, TxDate, TxType, VendorName, Amount, PMComments, ExtractedPONumber, MatchedPONumber, CorrectPONumber, MatchStatus, MatchConfidence, ReviewDecision, PostedToPO, PostedPONumber, PostedAmount, PostedAt, PostedBy, IsDuplicate FROM dbo.ExpenseReviewItems {where_sql} ORDER BY TxDate DESC, ExpenseReviewItemId DESC;""", *params)
        rows = cursor.fetchall(); conn.close()
        total_amount = sum(float(r.Amount or 0) for r in rows)
        unique_pos = len(set(str(r.CorrectPONumber or r.MatchedPONumber or r.ExtractedPONumber or '').strip() for r in rows if str(r.CorrectPONumber or r.MatchedPONumber or r.ExtractedPONumber or '').strip()))
        table_rows = ""
        for r in rows:
            po_number = r.CorrectPONumber or r.MatchedPONumber or r.ExtractedPONumber or ""
            po_link = f'<a href="/po-packet/{quote_plus(str(po_number))}?type=internal">{h(po_number)}</a>' if po_number else ""
            status_class = str(r.MatchStatus or "").lower().replace(" ", "-")
            table_rows += f"""<tr><td><span class="status-chip {h(status_class)}">{h(r.MatchStatus)}</span><br>{posting_status_chip(r)}<br><small>{h(r.MatchConfidence)}</small></td><td><strong>{h(r.ExpenseId or r.ExpenseReviewItemId)}</strong><br><small>{h(r.TxDate)}</small></td><td>{h(r.ProjectName)}</td><td>{h(r.VendorName)}<br><small>{h(r.TxType)}</small></td><td class="right">{currency(r.Amount)}</td><td>{h(r.PMComments)}</td><td>{po_link}</td><td>{h(r.ReviewDecision)}</td><td class="posting-audit-cell">{h(posting_reason(r))}<br><small>{h(r.PostedBy or '')} {h(r.PostedAt or '')}</small></td></tr>"""
        if not table_rows:
            table_rows = '<tr><td colspan="9"><div class="empty-state"><strong>No PO references found in PM comments.</strong><span>Adjust filters or upload expenses with PM comments.</span></div></td></tr>'
        filter_bar = build_simple_filter_bar(filters, "/pos-in-pm-comments", [("project", "Project", "Project name/code"), ("vendor", "Vendor", "Vendor or purchaser"), ("po", "PO", "PO number"), ("status", "Match Status", "Auto Matched..."), ("decision", "Decision", "Matched to PO..."), ("search", "Search", "PM comments")])
        content = f"""
        <div class="notice info"><strong>PM Comment PO Audit:</strong> This read-only page explains where PO references were found in PM Comments. Use Expense Upload / PO Matching or Missing PO Review to change decisions.</div>
        <div class="grid kpis"><div class="card kpi"><div class="label">Rows Found</div><div class="value">{len(rows)}</div><div class="trend">PM comments with PO-like references</div></div><div class="card kpi"><div class="label">Unique POs</div><div class="value">{unique_pos}</div><div class="trend">Referenced or matched POs</div></div><div class="card kpi"><div class="label">Referenced Amount</div><div class="value">{currency(total_amount)}</div><div class="trend">Filtered expense rows</div></div><a class="card kpi status-card" href="/missing-po-review"><div class="label">Review Queue</div><div class="value">Open</div><div class="trend">Validate uncertain matches</div></a></div>
        {filter_bar}
        <div class="card"><h3>PM Comment PO Audit</h3><p class="card-subtitle">Read-only audit view of expense rows where PO numbers or PO-like references were found in PM Comments.</p><div class="table-wrap"><table><tr><th>Status</th><th>Expense</th><th>Project</th><th>Vendor / Purchaser</th><th class="right">Amount</th><th>PM Comments</th><th>PO Reference</th><th>Decision</th><th>Posting</th></tr>{table_rows}</table></div></div>
        """
        return shell("POs in PM Comments", "Rows where PO references were found in PM Comments.", "POs in PM Comments", content)
    except Exception as e:
        return shell("POs in PM Comments", "Unable to load POs in PM Comments.", "POs in PM Comments", f'<div class="notice error">Error loading POs in PM Comments: {h(e)}</div>'), 500



@app.route("/purchase-request-attachment/<int:attachment_id>")
def purchase_request_attachment(attachment_id):
    allowed, reason = require_page_access("POs & Balances")
    if not allowed:
        return access_denied_response("POs & Balances", reason)
    try:
        conn = get_sql_connection()
        cursor = conn.cursor()
        ensure_request_attachment_table(cursor)
        cursor.execute(
            """
            SELECT StoragePath, OriginalFileName
            FROM dbo.PurchaseRequestAttachments
            WHERE AttachmentId = ?;
            """,
            attachment_id,
        )
        row = cursor.fetchone()
        conn.close()
        if not row or not os.path.exists(row.StoragePath):
            return shell("Attachment", "File not found.", "POs & Balances", '<div class="notice error">Attachment file was not found.</div>'), 404
        return send_from_directory(os.path.dirname(row.StoragePath), os.path.basename(row.StoragePath), as_attachment=False, download_name=row.OriginalFileName)
    except Exception as e:
        return shell("Attachment", "Unable to open attachment.", "POs & Balances", f'<div class="notice error">Error opening attachment: {h(e)}</div>'), 500

@app.route("/download-expense-upload-template.csv")
def download_expense_upload_template():
    allowed, reason = require_page_access("Expense Upload / PO Matching")
    if not allowed:
        return access_denied_response("Expense Upload / PO Matching", reason)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(REQUIRED_EXPENSE_COLUMNS)
    writer.writerow(["EXP-1001", "Example Project", "2026-06-22", "Purchase", "Example Vendor", "1250.00", "Materials or service description", "PM comment with PO-XX-XXX-001 if available"])
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=unanet_expense_upload_template.csv"},
    )


@app.route("/expense-review/update", methods=["POST"])
def expense_review_update():
    allowed, reason = require_page_access("Expense Upload / PO Matching")
    if not allowed:
        allowed, reason = require_page_access("Missing PO Review")
    if not allowed:
        return access_denied_response("Missing PO Review", reason)
    ensure_expense_review_tables()
    item_id = request.form.get("item_id")
    decision = clean_text(request.form.get("review_decision")) or "Pending Review"
    correct_po = clean_text(request.form.get("correct_po_number"))
    notes = clean_text(request.form.get("reviewer_notes"))
    new_status = None
    if decision == "Matched to PO" and correct_po:
        new_status = "Manually Matched"
    elif decision == "No PO Needed":
        new_status = "No PO Needed"
    elif decision == "Needs PM Review":
        new_status = "Needs PM Review"
    elif decision == "Hold for More Info":
        new_status = "Needs Review"
    else:
        new_status = "Needs Review"

    conn = get_sql_connection()
    cursor = conn.cursor()
    reviewer_email = get_current_user()["email"] or "Manual Review"
    cursor.execute(
        """
        UPDATE dbo.ExpenseReviewItems
        SET ReviewDecision = ?, CorrectPONumber = ?, ReviewerNotes = ?, MatchStatus = ?,
            ReviewerEmail = ?, ReviewedAt = SYSUTCDATETIME(), UpdatedAt = SYSUTCDATETIME()
        WHERE ExpenseReviewItemId = ?;
        """,
        decision,
        correct_po,
        notes,
        new_status,
        reviewer_email,
        item_id,
    )
    sync_expense_posting(cursor, item_id, posted_by=reviewer_email)
    conn.commit()
    conn.close()
    return redirect(request.headers.get("Referer") or "/expense-upload?saved=1#review-table")


@app.route("/expense-review/unpost", methods=["POST"])
def expense_review_unpost():
    allowed, reason = require_page_access("Expense Upload / PO Matching")
    if not allowed:
        return access_denied_response("Expense Upload / PO Matching", reason)
    ensure_expense_review_tables()
    item_id = request.form.get("item_id")
    reviewer_email = get_current_user()["email"] or "Manual Review"
    conn = get_sql_connection()
    cursor = conn.cursor()
    cursor.execute(
        """
        UPDATE dbo.ExpenseReviewItems
        SET PostedToPO = 0,
            PostedPONumber = NULL,
            PostedAmount = NULL,
            PostedAt = NULL,
            PostedBy = NULL,
            ReviewerNotes = CONCAT(COALESCE(ReviewerNotes, ''), CASE WHEN COALESCE(ReviewerNotes, '') = '' THEN '' ELSE CHAR(10) END, 'Posting reversed by ', ?, ' at ', CONVERT(NVARCHAR(30), SYSUTCDATETIME(), 126)),
            UpdatedAt = SYSUTCDATETIME()
        WHERE ExpenseReviewItemId = ?;
        """,
        reviewer_email,
        item_id,
    )
    conn.commit()
    conn.close()
    return redirect(request.headers.get("Referer") or "/expenses")


def load_expense_reset_counts(cursor):
    """Return current expense row/batch counts for the Admin reset confirmation page."""
    counts = {
        "review_items": 0,
        "posted_items": 0,
        "posted_amount": 0,
        "batches": 0,
        "legacy_expense_lines": 0,
    }
    cursor.execute('''
        IF OBJECT_ID('dbo.ExpenseReviewItems', 'U') IS NOT NULL
            SELECT COUNT(*) AS Cnt,
                   SUM(CASE WHEN COALESCE(PostedToPO, 0) = 1 THEN 1 ELSE 0 END) AS PostedCnt,
                   SUM(CASE WHEN COALESCE(PostedToPO, 0) = 1 THEN COALESCE(PostedAmount, Amount, 0) ELSE 0 END) AS PostedAmt
            FROM dbo.ExpenseReviewItems;
        ELSE
            SELECT 0 AS Cnt, 0 AS PostedCnt, 0 AS PostedAmt;
    ''')
    row = cursor.fetchone()
    if row:
        counts["review_items"] = int(row.Cnt or 0)
        counts["posted_items"] = int(row.PostedCnt or 0)
        counts["posted_amount"] = float(row.PostedAmt or 0)

    cursor.execute('''
        IF OBJECT_ID('dbo.ExpenseUploadBatches', 'U') IS NOT NULL
            SELECT COUNT(*) AS Cnt FROM dbo.ExpenseUploadBatches;
        ELSE
            SELECT 0 AS Cnt;
    ''')
    row = cursor.fetchone()
    if row:
        counts["batches"] = int(row.Cnt or 0)

    cursor.execute('''
        IF OBJECT_ID('dbo.ExpenseLines', 'U') IS NOT NULL
            SELECT COUNT(*) AS Cnt FROM dbo.ExpenseLines;
        ELSE
            SELECT 0 AS Cnt;
    ''')
    row = cursor.fetchone()
    if row:
        counts["legacy_expense_lines"] = int(row.Cnt or 0)
    return counts


@app.route("/admin/clear-expense-data", methods=["GET", "POST"])
def admin_clear_expense_data():
    allowed, reason = require_page_access("Clear Expense Data")
    if not allowed:
        return access_denied_response("Clear Expense Data", reason)

    actual_access = get_user_access()
    if normalize_role(actual_access.get("role")) != "Admin":
        return access_denied_response("Clear Expense Data", "Only a real Admin account can clear expense data.")

    try:
        ensure_expense_review_tables()
        conn = get_sql_connection()
        cursor = conn.cursor()
        counts = load_expense_reset_counts(cursor)
        message_html = ""

        if request.method == "POST":
            action = clean_text(request.form.get("action"))
            confirm_text = clean_text(request.form.get("confirm_text"))
            if action == "cancel":
                conn.close()
                return redirect("/expense-upload")
            if confirm_text != "CLEAR EXPENSES":
                message_html = '<div class="notice error">Confirmation did not match. Type CLEAR EXPENSES exactly to continue.</div>'
            else:
                cleared_by = get_current_user().get("email") or actual_access.get("email") or "Admin"
                before_counts = dict(counts)
                cursor.execute('''
                    IF OBJECT_ID('dbo.ExpenseReviewItems', 'U') IS NOT NULL
                        DELETE FROM dbo.ExpenseReviewItems;
                ''')
                cursor.execute('''
                    IF OBJECT_ID('dbo.ExpenseUploadBatches', 'U') IS NOT NULL
                        DELETE FROM dbo.ExpenseUploadBatches;
                ''')
                cursor.execute('''
                    IF OBJECT_ID('dbo.ExpenseLines', 'U') IS NOT NULL
                        DELETE FROM dbo.ExpenseLines;
                ''')
                # Do not touch dbo.ImportBatches here. In the live database this table is
                # used by issued PO uploads and does not always have an ImportType column.
                # The expense reset is intentionally limited to the expense-specific tables
                # below so issued PO/project setup history remains intact.
                conn.commit()
                counts = load_expense_reset_counts(cursor)
                message_html = f'''
                <div class="notice success">
                    Expense data cleared by {h(cleared_by)}. Removed {h(before_counts["review_items"])} review rows, {h(before_counts["batches"])} upload batches, and {h(before_counts["legacy_expense_lines"])} legacy expense rows. PO setup, issued POs, projects, vendors, users, and purchase requests were not deleted.
                </div>
                '''

        conn.close()

        content = f'''
        {message_html}
        <div class="notice error">
            <strong>ALERT: YOU'RE ABOUT TO CLEAR ALL EXPENSE DATA.</strong><br>
            This will delete prior expense upload rows, review/matching decisions, and posted expense rows that reduce PO balances. It will not delete issued POs, projects, vendors, purchase requests, users, or PO maintenance history.
        </div>

        <div class="card">
            <h3>Current Expense Data</h3>
            <div class="status-card-grid">
                <div class="status-card blue"><div class="label">Expense Review Rows</div><div class="value">{h(counts["review_items"])}</div><div class="trend">Rows from uploaded expense files</div></div>
                <div class="status-card green"><div class="label">Posted to PO</div><div class="value">{h(counts["posted_items"])}</div><div class="trend">Rows currently reducing PO balances</div></div>
                <div class="status-card amber"><div class="label">Posted Amount</div><div class="value">{currency(counts["posted_amount"])}</div><div class="trend">Will be removed from balance calculations</div></div>
                <div class="status-card slate"><div class="label">Upload Batches</div><div class="value">{h(counts["batches"])}</div><div class="trend">Expense upload history rows</div></div>
            </div>
        </div>

        <div class="card">
            <h3>Confirm Clear Expense Data</h3>
            <p class="card-subtitle">Use this only when you are intentionally clearing prior test/upload expense data before a clean upload.</p>
            <form method="post">
                <div class="form-field full">
                    <label>Type CLEAR EXPENSES to continue</label>
                    <input name="confirm_text" autocomplete="off" placeholder="CLEAR EXPENSES">
                </div>
                <div class="request-actions" style="justify-content:flex-start; gap:10px;">
                    <a class="button secondary" href="/expense-upload">Go Back</a>
                    <button class="danger" type="submit" name="action" value="clear">Continue to Delete</button>
                </div>
            </form>
        </div>
        '''
        return shell("Clear Expense Data", "Admin-only reset for expense upload testing.", "Clear Expense Data", content)
    except Exception as e:
        return shell("Clear Expense Data", "Unable to clear expense data.", "Clear Expense Data", f'<div class="notice error">Error loading expense reset page: {h(e)}</div>'), 500


@app.route("/expense-upload", methods=["GET", "POST"])
def expense_upload():
    allowed, reason = require_page_access("Expense Upload / PO Matching")
    if not allowed:
        return access_denied_response("Expense Upload / PO Matching", reason)

    message_html = ""
    try:
        ensure_expense_review_tables()
        if request.method == "POST":
            uploaded_file = request.files.get("expense_file")
            if not uploaded_file or uploaded_file.filename == "":
                message_html = '<div class="notice error">Please choose a Unanet expense file to upload.</div>'
            else:
                rows = read_uploaded_po_file(uploaded_file)
                result = import_expense_rows(rows, uploaded_file.filename)
                message_html = f"""
                <div class="notice success">
                    Expense upload processed. Batch {h(result['batch_id'])}: timesheet rows ignored, {h(result['auto'])} auto matched, {h(result['review'])} need review, {h(result['no_match'])} no match.
                </div>
                """
        elif request.args.get("saved"):
            message_html = '<div class="notice success">Expense review action saved.</div>'

        status_filter = request.args.get("status", "All")
        stats, rows, batches, po_numbers = load_expense_upload_page_data(status_filter)
        total_rows = stats.TotalRows or 0
        auto_matched = stats.AutoMatched or 0
        manually_matched = stats.ManuallyMatched or 0
        needs_review = stats.NeedsReview or 0
        no_match = stats.NoMatch or 0
        reviewed_rows = stats.ReviewedRows or 0

        def status_card(label, value, status, tone, sub):
            active_cls = " active" if status_filter == status else ""
            href = "/expense-upload" if status == "All" else "/expense-upload?status=" + quote_plus(status)
            return f'<a class="status-card {tone}{active_cls}" href="{href}"><div class="label">{h(label)}</div><div class="value">{h(value)}</div><div class="trend">{h(sub)}</div></a>'

        kpi_html = f"""
        <div class="status-card-grid">
            {status_card('All Rows', total_rows, 'All', 'blue', 'Uploaded expense rows')}
            {status_card('Auto Matched', auto_matched, 'Auto Matched', 'green', 'PO found in comments')}
            {status_card('Manual Matches', manually_matched, 'Manually Matched', 'blue', 'Reviewed and linked')}
            {status_card('Needs Review', needs_review, 'Needs Review', 'amber', 'Possible match or unclear')}
            {status_card('No Match', no_match, 'No Match', 'red', 'Needs PO decision')}
            {status_card('Reviewed', reviewed_rows, 'All', 'slate', 'Rows with decisions')}
        </div>
        """

        batch_html = ""
        if batches:
            batch_html = "".join(
                f"""
                <div class="expense-batch-item">
                    <strong>Batch {h(b.ExpenseBatchId)} · {h(b.FileName)}</strong>
                    <span>{h(b.UploadedAt)} · {h(b.TotalRows)} rows · {h(b.AutoMatchedCount)} auto · {h(b.NeedsReviewCount)} review · {h(b.NoMatchCount)} no match</span>
                </div>
                """ for b in batches
            )
        else:
            batch_html = '<div class="empty-state"><strong>No expense uploads yet.</strong><span>Upload a Unanet expense file to begin matching expenses to POs.</span></div>'

        po_options = "".join(f'<option value="{h(po)}"></option>' for po in po_numbers)
        decision_options = ["Pending Review", "Matched to PO", "No PO Needed", "Needs PM Review", "Hold for More Info"]
        table_rows = ""
        if rows:
            for r in rows:
                current_po = r.CorrectPONumber or r.MatchedPONumber or r.ExtractedPONumber or ""
                decision_select = "".join(f'<option value="{h(opt)}" {"selected" if (r.ReviewDecision or "Pending Review") == opt else ""}>{h(opt)}</option>' for opt in decision_options)
                table_rows += f"""
                <tr>
                    <td>{status_chip(r.MatchStatus)}<div class="muted">{h(r.MatchConfidence or '')}</div></td>
                    <td><strong>{h(r.ExpenseId or 'Expense')}</strong><div class="muted">{h(r.TxDate or '')} · {h(r.TxType or '')}</div></td>
                    <td>{h(r.ProjectName)}</td>
                    <td>{h(r.VendorName)}</td>
                    <td class="right">{currency(r.Amount)}</td>
                    <td class="comments-cell">{h(r.PMComments or r.Description or '')}</td>
                    <td>{h(r.ExtractedPONumber or '')}</td>
                    <td><strong>{h(r.MatchedPONumber or '')}</strong><div class="match-reason-cell">{h(r.MatchReason or '')}</div></td>
                    <td class="expense-action-cell">
                        <form class="expense-match-form" method="post" action="/expense-review/update">
                            <input type="hidden" name="item_id" value="{h(r.ExpenseReviewItemId)}">
                            <select name="review_decision">{decision_select}</select>
                            <input name="correct_po_number" list="poNumberList" placeholder="Correct PO number" value="{h(current_po)}">
                            <textarea name="reviewer_notes" placeholder="Reviewer notes">{h(r.ReviewerNotes or '')}</textarea>
                            <button type="submit" class="primary mini">Save Review</button>
                        </form>
                    </td>
                </tr>
                """
        else:
            table_rows = '<tr><td colspan="9"><div class="empty-state"><strong>No expense rows found.</strong><span>Upload a file or clear the status filter.</span></div></td></tr>'

        content = f"""
        {message_html}
        {kpi_html}
        <div class="expense-upload-layout">
            <div class="card">
                <h3>Upload Unanet Expense File</h3>
                <p class="card-subtitle">Upload a CSV or Excel file exported from Unanet. Timesheet rows are ignored. Purchase/disbursement rows are checked for PO numbers in PM Comments and suggested matches by project/vendor.</p>
                <form method="post" enctype="multipart/form-data">
                    <div class="form-field full">
                        <label>Unanet expense file</label>
                        <input type="file" name="expense_file" accept=".csv,.xlsx" required>
                    </div>
                    <div class="request-actions" style="justify-content:flex-start;">
                        <button class="primary" type="submit">Upload and Match Expenses</button>
                        <a class="button" href="/download-expense-upload-template.csv">Download CSV Template</a>
                    </div>
                </form>
            </div>
            <div class="card">
                <h3>Recent Expense Uploads</h3>
                <p class="card-subtitle">Latest expense batches and match results.</p>
                <div class="expense-batch-list">{batch_html}</div>
            </div>
        </div>
        <div class="card" id="review-table">
            <h3>Expense Review Queue</h3>
            <p class="card-subtitle">Use this as the active work queue for uploaded non-timesheet expenses. Confirm PO matches, correct uncertain rows, or mark a row as no PO required / PM review needed. Posted matched rows reduce Current App Balance.</p>
            <datalist id="poNumberList">{po_options}</datalist>
            <div class="table-wrap expense-review-table">
                <table>
                    <tr>
                        <th>Status</th><th>Expense</th><th>Project</th><th>Vendor / Purchaser</th><th class="right">Amount</th><th>PM Comments / Description</th><th>Extracted PO</th><th>Suggested / Matched PO</th><th>Review Action</th>
                    </tr>
                    {table_rows}
                </table>
            </div>
        </div>
        """
        return shell("Expense Upload / PO Matching", "Upload Unanet expenses, review matches, and link transactions to issued POs.", "Expense Upload / PO Matching", content)
    except Exception as e:
        content = f'<div class="notice error">Error loading Expense Upload / PO Matching: {h(e)}</div>'
        return shell("Expense Upload / PO Matching", "Unable to load expense matching page.", "Expense Upload / PO Matching", content), 500



@app.route("/download-issued-po-template.xlsx")
def download_issued_po_template_xlsx():
    allowed, reason = require_page_access("Upload Issued POs")
    if not allowed:
        return access_denied_response("Upload Issued POs", reason)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Issued PO Setup"
    lists = wb.create_sheet("Lists")

    headers = REQUIRED_PO_COLUMNS
    ws.append(headers)
    ws.append(["25-209", "Round Valley", "25-209-PO-001", "Example Vendor", "Marine Construction", "2026-07-01", "Ashley Marion", "Example line item", "EA", 1000, 2, 2000])
    ws.append(["25-209", "Round Valley", "25-209-PO-001", "Example Vendor", "Marine Construction", "2026-07-01", "Ashley Marion", "Second line item", "LS", 500, 1, 500])

    lists["A1"] = "Department Options"
    for idx, dept in enumerate(DEPARTMENT_OPTIONS, start=2):
        lists[f"A{idx}"] = dept

    header_fill = PatternFill("solid", fgColor="0B1F3A")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2EC")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)

    widths = [14, 24, 18, 26, 24, 14, 22, 36, 12, 14, 10, 14]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"

    dv = DataValidation(type="list", formula1="=Lists!$A$2:$A$6", allow_blank=False)
    ws.add_data_validation(dv)
    dv.add("E2:E1000")

    ws["A6"] = "Template notes"
    ws["A6"].fill = PatternFill("solid", fgColor="EAF2FF")
    ws["A6"].font = Font(color="0B1F3A", bold=True)
    ws.merge_cells("A6:L6")
    notes = [
        "ProjectCode is required and should match the project code used internally, such as 25-209.",
        "Department must be selected from the dropdown.",
        "Original/Revised/Remaining Amount and PO Status are intentionally removed. The app calculates PO amount from line amounts and marks uploaded POs Open.",
        "Use one row per PO line item. Repeating the same PO number creates multiple line items under the same PO.",
    ]
    for row_idx, note in enumerate(notes, start=7):
        ws[f"A{row_idx}"] = note
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=12)
        ws[f"A{row_idx}"].alignment = Alignment(wrap_text=True)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=issued_po_project_setup_template.xlsx"},
    )

@app.route("/download-issued-po-template.csv")
def download_issued_po_template():
    allowed, reason = require_page_access("Upload Issued POs")
    if not allowed:
        return access_denied_response("Upload Issued POs", reason)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(REQUIRED_PO_COLUMNS)

    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={
            "Content-Disposition": "attachment; filename=issued_po_project_setup_template.csv"
        },
    )


@app.route("/upload-po", methods=["GET", "POST"])
def upload_po():
    allowed, reason = require_page_access("Upload Issued POs")
    if not allowed:
        return access_denied_response("Upload Issued POs", reason)

    message_html = ""
    result_html = ""
    errors_html = ""

    assignable_users = load_assignable_users()
    requestor_options = []
    for u in assignable_users:
        display = clean_text(getattr(u, "DisplayName", "") or getattr(u, "Email", ""))
        role = clean_text(getattr(u, "RoleName", ""))
        email = clean_text(getattr(u, "Email", ""))
        if display:
            requestor_options.append((display, role, email))
    valid_requestors = [x[0] for x in requestor_options]

    selected_department = clean_text(request.form.get("default_department")) if request.method == "POST" else ""
    selected_requestor = clean_text(request.form.get("default_requestor")) if request.method == "POST" else ""

    if request.method == "POST":
        uploaded_file = request.files.get("po_file")

        if not uploaded_file or uploaded_file.filename == "":
            message_html = '<div class="notice error">No file selected.</div>'
        elif not uploaded_file.filename.lower().endswith(".csv"):
            message_html = '<div class="notice error">Please upload a CSV file. The issued PO setup upload is CSV-only.</div>'
        else:
            try:
                rows = read_uploaded_po_file(uploaded_file)
                validation_errors = validate_po_rows(rows, selected_department, selected_requestor, valid_requestors)

                if validation_errors:
                    message_html = '<div class="notice error">The file could not be imported because validation errors were found.</div>'
                    error_items = "".join(f"<li>{h(error)}</li>" for error in validation_errors)
                    errors_html = f'<div class="card"><h3>Validation Errors</h3><ul>{error_items}</ul></div>'
                else:
                    result = import_po_rows(rows, uploaded_file.filename, selected_department, selected_requestor)
                    selected_requestor_email = ""
                    for display, role, email in requestor_options:
                        if display == selected_requestor:
                            selected_requestor_email = email
                            break
                    try:
                        send_po_upload_summary_email(result, uploaded_file.filename, selected_department, selected_requestor, selected_requestor_email)
                    except Exception as email_error:
                        print(f"PO upload summary email failed: {email_error}")
                    message_html = '<div class="notice ok">Issued PO import completed.</div>'
                    result_html = f"""
                    <div class="card"><h3>Import Result</h3><table><tr><th>Import Batch ID</th><td>{result["import_batch_id"]}</td></tr><tr><th>Total Rows</th><td>{result["total_rows"]}</td></tr><tr><th>Success Count</th><td>{result["success_count"]}</td></tr><tr><th>Error Count</th><td>{result["error_count"]}</td></tr><tr><th>Status</th><td>{h(result["status"])}</td></tr><tr><th>Department Applied</th><td>{h(selected_department)}</td></tr><tr><th>Requestor Applied</th><td>{h(selected_requestor)}</td></tr></table></div>
                    """

            except Exception as e:
                message_html = '<div class="notice error">Import failed.</div>'
                errors_html = f'<div class="card"><h3>Error Details</h3><p>{h(e)}</p></div>'

    department_options_html = "".join(
        f'<option value="{h(dept)}" {"selected" if dept == selected_department else ""}>{h(dept)}</option>'
        for dept in DEPARTMENT_OPTIONS
    )
    requestor_options_html = "".join(
        f'<option value="{h(display)}" {"selected" if display == selected_requestor else ""}>{h(display)}{(" · " + h(role)) if role else ""}</option>'
        for display, role, email in requestor_options
    )

    content = f"""
    {message_html}{result_html}{errors_html}
    <div class="grid two">
        <div class="card">
            <h3>Select Issued PO File</h3>
            <p class="card-subtitle">Upload the first-time project setup PO file as CSV. Department and requestor are selected here and applied to every row in the upload.</p>
            <form method="post" enctype="multipart/form-data">
                <div class="form-grid">
                    <div class="form-field">
                        <label>Department *</label>
                        <select name="default_department" required>
                            <option value="">Select department</option>
                            {department_options_html}
                        </select>
                    </div>
                    <div class="form-field">
                        <label>Requestor *</label>
                        <select name="default_requestor" required>
                            <option value="">Select requestor</option>
                            {requestor_options_html}
                        </select>
                    </div>
                    <div class="form-field full">
                        <label>Issued PO CSV *</label>
                        <input type="file" name="po_file" accept=".csv" required>
                    </div>
                </div>
                <p class="field-help">The selected department and requestor will be stamped onto all POs and PO line items in this upload.</p>
                <p><button class="primary" type="submit">Upload Issued POs</button></p>
            </form>
        </div>
        <div class="card">
            <h3>Project Setup CSV Template</h3>
            <p class="card-subtitle">Download the CSV template for the Phase 1 first-time project PO setup load.</p>
            <p><a class="button primary" href="/download-issued-po-template.csv">Download CSV Template</a></p>
        </div>
    </div>
    """

    return shell("Upload Issued POs", "Import first-time project setup POs and line items into Azure SQL.", "Upload Issued POs", content)

@app.route("/import-history")
def import_history():
    allowed, reason = require_page_access("Import History")
    if not allowed:
        return access_denied_response("Import History", reason)

    try:
        conn = get_sql_connection()
        cursor = conn.cursor()

        cursor.execute(
            """
            SELECT TOP 50
                ImportBatchId,
                FileName,
                SourceSystem,
                UploadedBy,
                UploadedAt,
                TotalRows,
                SuccessCount,
                ErrorCount,
                ImportStatus,
                ErrorMessage
            FROM dbo.ImportBatches
            ORDER BY UploadedAt DESC;
            """
        )
        batches = cursor.fetchall()

        cursor.execute(
            """
            SELECT TOP 100
                e.ImportErrorId,
                e.ImportBatchId,
                b.FileName,
                e.RowNumber,
                e.ErrorMessage,
                e.RawRow,
                e.CreatedAt
            FROM dbo.ImportErrors e
            LEFT JOIN dbo.ImportBatches b ON e.ImportBatchId = b.ImportBatchId
            ORDER BY e.CreatedAt DESC;
            """
        )
        errors = cursor.fetchall()
        conn.close()

        batch_rows = ""
        for row in batches:
            status_badge = "green"
            if row.ErrorCount and row.ErrorCount > 0:
                status_badge = "amber"
            if row.ImportStatus and "fail" in row.ImportStatus.lower():
                status_badge = "red"

            batch_rows += f"""
            <tr><td>{row.ImportBatchId}</td><td>{h(row.FileName)}</td><td>{h(row.UploadedAt)}</td><td>{h(row.SourceSystem)}</td><td>{h(row.UploadedBy)}</td><td>{row.TotalRows}</td><td>{row.SuccessCount}</td><td>{row.ErrorCount}</td><td><span class="badge {status_badge}">{h(row.ImportStatus)}</span></td></tr>
            """

        if not batch_rows:
            batch_rows = '<tr><td colspan="9">No import batches found yet.</td></tr>'

        error_rows = ""
        for row in errors:
            raw_row = row.RawRow or ""
            if len(raw_row) > 300:
                raw_row = raw_row[:300] + "..."

            error_rows += f"""
            <tr><td>{row.ImportErrorId}</td><td>{row.ImportBatchId}</td><td>{h(row.FileName)}</td><td>{h(row.RowNumber)}</td><td>{h(row.ErrorMessage)}</td><td>{h(raw_row)}</td><td>{h(row.CreatedAt)}</td></tr>
            """

        if not error_rows:
            error_rows = '<tr><td colspan="7">No import errors found.</td></tr>'

        content = f"""
        <div class="card"><h3>Import Batches</h3><p class="card-subtitle">Latest uploaded PO files and processing results.</p><div class="table-wrap"><table><tr><th>Batch ID</th><th>File Name</th><th>Uploaded At</th><th>Source</th><th>Uploaded By</th><th>Total Rows</th><th>Success</th><th>Errors</th><th>Status</th></tr>{batch_rows}</table></div></div>
        <div class="card"><h3>Recent Import Errors</h3><p class="card-subtitle">Rows that failed validation or import processing.</p><div class="table-wrap"><table><tr><th>Error ID</th><th>Batch ID</th><th>File Name</th><th>Row Number</th><th>Error Message</th><th>Raw Row</th><th>Created At</th></tr>{error_rows}</table></div></div>
        """

        return shell("Import History", "Review uploaded files, row counts, import status, and row-level errors.", "Import History", content)

    except Exception as e:
        content = f'<div class="notice error">Error loading import history: {h(e)}</div>'
        return shell("Import History", "Unable to load import history.", "Import History", content), 500


@app.route("/exceptions")
def exceptions():
    allowed, reason = require_page_access("Exceptions")
    if not allowed:
        return access_denied_response("Exceptions", reason)

    try:
        conn = get_sql_connection()
        cursor = conn.cursor()

        cursor.execute(
            """
            WITH POList AS (
                SELECT
                    PONumber,
                    MAX(VendorName) AS VendorName,
                    MAX(ProjectName) AS ProjectName,
                    MAX(Department) AS Department,
                    MAX(POStatus) AS POStatus,
                    COUNT(*) AS LineCount,
                    MAX(COALESCE(RevisedAmount, OriginalAmount, 0)) AS POValue,
                    SUM(COALESCE(LineAmount, 0)) AS TotalLineAmount,
                    MAX(COALESCE(RemainingAmount, 0)) AS RemainingAmount
                FROM dbo.IssuedPOLines
                GROUP BY PONumber
            )
            SELECT 'Amount Mismatch' AS ExceptionType, PONumber, VendorName, ProjectName, Department, POStatus,
                   'Line total does not match revised/original PO value.' AS Message, POValue, TotalLineAmount, RemainingAmount
            FROM POList
            WHERE ABS(COALESCE(POValue, 0) - COALESCE(TotalLineAmount, 0)) > 0.01
            UNION ALL
            SELECT 'Closed With Remaining Balance' AS ExceptionType, PONumber, VendorName, ProjectName, Department, POStatus,
                   'PO appears closed but still has remaining balance.' AS Message, POValue, TotalLineAmount, RemainingAmount
            FROM POList
            WHERE UPPER(COALESCE(POStatus, '')) IN ('CLOSED', 'COMPLETE', 'COMPLETED') AND COALESCE(RemainingAmount, 0) > 0.01
            UNION ALL
            SELECT 'Open With Zero Remaining' AS ExceptionType, PONumber, VendorName, ProjectName, Department, POStatus,
                   'PO appears open but has zero remaining balance.' AS Message, POValue, TotalLineAmount, RemainingAmount
            FROM POList
            WHERE UPPER(COALESCE(POStatus, '')) = 'OPEN' AND COALESCE(RemainingAmount, 0) = 0
            UNION ALL
            SELECT 'Missing Department' AS ExceptionType, PONumber, VendorName, ProjectName, Department, POStatus,
                   'PO is missing a department, which may affect role-based filtering.' AS Message, POValue, TotalLineAmount, RemainingAmount
            FROM POList
            WHERE Department IS NULL OR LTRIM(RTRIM(Department)) = ''
            UNION ALL
            SELECT 'Missing Revised Amount' AS ExceptionType, PONumber, VendorName, ProjectName, Department, POStatus,
                   'PO is missing RevisedAmount. OriginalAmount is being used as fallback.' AS Message, POValue, TotalLineAmount, RemainingAmount
            FROM POList
            WHERE PONumber IN (
                SELECT PONumber FROM dbo.IssuedPOLines GROUP BY PONumber HAVING MAX(RevisedAmount) IS NULL
            )
            ORDER BY ExceptionType, PONumber;
            """
        )

        rows = cursor.fetchall()
        conn.close()

        exception_rows = ""
        count_by_type = {}

        for row in rows:
            count_by_type[row.ExceptionType] = count_by_type.get(row.ExceptionType, 0) + 1
            po_url = "/po-packet/" + quote_plus(str(row.PONumber or "")) + "?type=internal"
            badge_class = "red" if row.ExceptionType == "Amount Mismatch" else "amber"

            exception_rows += f"""
            <tr><td><span class="badge {badge_class}">{h(row.ExceptionType)}</span></td><td><a href="{po_url}">{h(row.PONumber)}</a></td><td>{h(row.VendorName)}</td><td>{h(row.ProjectName)}</td><td>{h(row.Department)}</td><td>{h(row.POStatus)}</td><td>{h(row.Message)}</td><td class="right">{currency(row.POValue)}</td><td class="right">{currency(row.TotalLineAmount)}</td><td class="right">{currency(row.RemainingAmount)}</td></tr>
            """

        if not exception_rows:
            exception_rows = '<tr><td colspan="9">No exceptions found. Your issued PO data looks clean.</td></tr>'

        kpi_cards = ""
        if count_by_type:
            for exception_type, count in sorted(count_by_type.items()):
                kpi_cards += f'<div class="card kpi"><div class="label">{h(exception_type)}</div><div class="value">{count}</div><div class="trend">Exception count</div></div>'
        else:
            kpi_cards = '<div class="card kpi"><div class="label">Exceptions</div><div class="value">0</div><div class="trend"><span class="badge green">Clean</span></div></div>'

        content = f"""
        <div class="grid kpis">{kpi_cards}</div>
        <div class="card"><h3>Data Quality Exceptions</h3><p class="card-subtitle">Review issued POs that may need correction before expense tracking begins.</p><div class="table-wrap"><table><tr><th>Type</th><th>PO Number</th><th>Vendor</th><th>Project</th><th>Department</th><th>Status</th><th>Message</th><th class="right">PO Value</th><th class="right">Line Total</th><th class="right">Remaining</th></tr>{exception_rows}</table></div></div>
        """

        return shell("Exceptions", "Data quality checks for issued purchase orders.", "Exceptions", content)

    except Exception as e:
        content = f'<div class="notice error">Error loading exceptions: {h(e)}</div>'
        return shell("Exceptions", "Unable to load exceptions.", "Exceptions", content), 500


@app.route("/exports")
def exports():
    allowed, reason = require_page_access("Exports")
    if not allowed:
        return access_denied_response("Exports", reason)

    content = """
    <div class="grid two">
        <div class="card"><h3>PO List Export</h3><p class="card-subtitle">Download one row per issued PO with totals and status.</p><p><a class="button primary" href="/export-po-list.csv">Download PO List CSV</a></p></div>
        <div class="card"><h3>Issued Line Items Export</h3><p class="card-subtitle">Download all issued PO line items from the upload data.</p><p><a class="button primary" href="/export-issued-lines.csv">Download Line Items CSV</a></p></div>
    </div>
    """

    return shell("Exports", "Download procurement dashboard data as CSV files.", "Exports", content)


@app.route("/export-po-list.csv")
def export_po_list_csv():
    allowed, reason = require_page_access("Exports")
    if not allowed:
        return access_denied_response("Exports", reason)

    conn = get_sql_connection()
    cursor = conn.cursor()

    cursor.execute(
        """
        WITH POList AS (
            SELECT
                PONumber,
                MAX(VendorName) AS VendorName,
                MAX(ProjectName) AS ProjectName,
                MAX(Department) AS Department,
                MAX(POStatus) AS POStatus,
                MAX(PODate) AS PODate,
                COUNT(*) AS LineCount,
                MAX(COALESCE(RevisedAmount, OriginalAmount, 0)) AS POValue,
                SUM(COALESCE(LineAmount, 0)) AS TotalLineAmount,
                MAX(COALESCE(RemainingAmount, 0)) AS RemainingAmount
            FROM dbo.IssuedPOLines
            GROUP BY PONumber
        )
        SELECT PONumber, VendorName, ProjectName, Department, POStatus, PODate, LineCount, POValue, TotalLineAmount, RemainingAmount
        FROM POList
        ORDER BY PODate DESC, PONumber DESC;
        """
    )

    rows = cursor.fetchall()
    conn.close()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["PONumber", "VendorName", "ProjectName", "Department", "POStatus", "PODate", "LineCount", "POValue", "TotalLineAmount", "RemainingAmount"])

    for row in rows:
        writer.writerow([row.PONumber, row.VendorName, row.ProjectName, row.Department, row.POStatus, row.PODate, row.LineCount, row.POValue, row.TotalLineAmount, row.RemainingAmount])

    return Response(output.getvalue(), mimetype="text/csv", headers={"Content-Disposition": "attachment; filename=po_list_export.csv"})


@app.route("/export-issued-lines.csv")
def export_issued_lines_csv():
    allowed, reason = require_page_access("Exports")
    if not allowed:
        return access_denied_response("Exports", reason)

    conn = get_sql_connection()
    cursor = conn.cursor()

    cursor.execute(
        """
        SELECT PONumber, VendorName, ProjectName, Department, PODate, POStatus, LineDescription, Unit, UnitCost, Qty, LineAmount, OriginalAmount, RevisedAmount, RemainingAmount, Requestor, CreatedAt
        FROM dbo.IssuedPOLines
        ORDER BY PONumber, IssuedPOLineId;
        """
    )

    rows = cursor.fetchall()
    conn.close()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["PONumber", "VendorName", "ProjectName", "Department", "PODate", "POStatus", "Description", "Unit", "UnitCost", "Qty", "LineAmount", "OriginalAmount", "RevisedAmount", "RemainingAmount", "Requestor", "CreatedAt"])

    for row in rows:
        writer.writerow([row.PONumber, row.VendorName, row.ProjectName, row.Department, row.PODate, row.POStatus, row.LineDescription, row.Unit, row.UnitCost, row.Qty, row.LineAmount, row.OriginalAmount, row.RevisedAmount, row.RemainingAmount, row.Requestor, row.CreatedAt])

    return Response(output.getvalue(), mimetype="text/csv", headers={"Content-Disposition": "attachment; filename=issued_po_lines_export.csv"})


@app.route("/set-view-as", methods=["POST"])
def set_view_as():
    access = get_user_access()
    if normalize_role(access.get("role")) not in ["Admin", "Executive"]:
        return access_denied_response("User Access", "Only Admin or Executive users can use View As.")
    email = clean_text(request.form.get("view_as_email")) or ""
    resp = make_response(redirect("/user-access"))
    if email:
        resp.set_cookie("PO_DASHBOARD_VIEW_AS", email.lower(), max_age=8*60*60, httponly=True, samesite="Lax")
    else:
        resp.delete_cookie("PO_DASHBOARD_VIEW_AS")
    return resp

@app.route("/clear-view-as")
def clear_view_as():
    resp = make_response(redirect(request.headers.get("Referer") or "/user-access"))
    resp.delete_cookie("PO_DASHBOARD_VIEW_AS")
    return resp


# ------------------------------------------------------------
# PO Maintenance / correction workflow
# ------------------------------------------------------------

def ensure_po_maintenance_schema(cursor):
    """Create audit table and ensure project code columns exist for safe PO corrections."""
    ensure_project_code_columns(cursor)
    cursor.execute(
        """
        IF OBJECT_ID('dbo.POChangeAudit', 'U') IS NULL
        BEGIN
            CREATE TABLE dbo.POChangeAudit (
                POChangeAuditId INT IDENTITY(1,1) PRIMARY KEY,
                ChangeType NVARCHAR(100) NOT NULL,
                OldPONumber NVARCHAR(100) NULL,
                NewPONumber NVARCHAR(100) NULL,
                OldProjectCode NVARCHAR(100) NULL,
                NewProjectCode NVARCHAR(100) NULL,
                OldProjectName NVARCHAR(255) NULL,
                NewProjectName NVARCHAR(255) NULL,
                OldVendorName NVARCHAR(255) NULL,
                NewVendorName NVARCHAR(255) NULL,
                OldDepartment NVARCHAR(100) NULL,
                NewDepartment NVARCHAR(100) NULL,
                OldRequestor NVARCHAR(255) NULL,
                NewRequestor NVARCHAR(255) NULL,
                Reason NVARCHAR(MAX) NOT NULL,
                ChangedBy NVARCHAR(255) NULL,
                ChangedAt DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME()
            );
        END;
        """
    )


def sql_col_exists(cursor, table_name, column_name):
    cursor.execute("SELECT COL_LENGTH(?, ?) AS ColLen", f"dbo.{table_name}", column_name)
    row = cursor.fetchone()
    return bool(row and row.ColLen is not None)


def sql_table_exists(cursor, table_name):
    cursor.execute("SELECT OBJECT_ID(?, 'U') AS ObjId", f"dbo.{table_name}")
    row = cursor.fetchone()
    return bool(row and row.ObjId)


def update_if_table_column(cursor, table_name, column_name, new_value, where_column, old_value):
    if sql_table_exists(cursor, table_name) and sql_col_exists(cursor, table_name, column_name) and sql_col_exists(cursor, table_name, where_column):
        cursor.execute(f"UPDATE dbo.{table_name} SET {column_name} = ? WHERE {where_column} = ?", new_value, old_value)


def load_active_dashboard_users_for_select(cursor):
    try:
        cursor.execute("SELECT Email, DisplayName FROM dbo.DashboardUsers WHERE IsActive = 1 ORDER BY DisplayName, Email")
        return cursor.fetchall()
    except Exception:
        return []


def po_maintenance_row_summary(cursor, po_number):
    cursor.execute(
        """
        SELECT TOP 1
            po.PurchaseOrderId,
            po.PONumber,
            po.ProjectCode,
            COALESCE(p.ProjectName, l.ProjectName, '') AS ProjectName,
            COALESCE(v.VendorName, l.VendorName, '') AS VendorName,
            COALESCE(po.Department, l.Department, '') AS Department,
            COALESCE(po.Requestor, l.Requestor, '') AS Requestor,
            COALESCE(po.POStatus, l.POStatus, 'Open') AS POStatus,
            COALESCE(po.PODate, l.PODate) AS PODate,
            COALESCE(SUM(l.LineAmount), po.RevisedAmount, po.OriginalAmount, 0) AS LineTotal,
            COUNT(l.IssuedPOLineId) AS LineCount
        FROM dbo.PurchaseOrders po
        LEFT JOIN dbo.Projects p ON p.ProjectId = po.ProjectId
        LEFT JOIN dbo.IssuedPOLines l ON l.PONumber = po.PONumber
        LEFT JOIN dbo.Vendors v ON v.VendorId = po.VendorId
        WHERE po.PONumber = ?
        GROUP BY po.PurchaseOrderId, po.PONumber, po.ProjectCode, p.ProjectName, l.ProjectName, v.VendorName, l.VendorName,
                 po.Department, l.Department, po.Requestor, l.Requestor, po.POStatus, l.POStatus, po.PODate, l.PODate,
                 po.RevisedAmount, po.OriginalAmount
        """,
        po_number,
    )
    return cursor.fetchone()


def update_existing_po_values(cursor, old_po, new_po, project_code, project_name, vendor_name, department, requestor, reason, changed_by):
    old_po = clean_text(old_po)
    new_po = clean_text(new_po)
    project_code = clean_text(project_code)
    project_name = clean_text(project_name)
    vendor_name = clean_text(vendor_name)
    department = clean_text(department)
    requestor = clean_text(requestor)
    reason = clean_text(reason)
    changed_by = clean_text(changed_by)

    if not old_po:
        raise ValueError("Original PO number is required.")
    if not new_po:
        raise ValueError("PO number is required.")
    if not reason:
        raise ValueError("A reason for the change is required.")

    current = po_maintenance_row_summary(cursor, old_po)
    if not current:
        raise ValueError("The selected PO was not found.")
    if new_po.lower() != old_po.lower():
        cursor.execute("SELECT COUNT(*) AS ExistingCount FROM dbo.PurchaseOrders WHERE LOWER(PONumber) = LOWER(?)", new_po)
        if cursor.fetchone().ExistingCount:
            raise ValueError("That new PO number already exists. Choose a unique PO number.")

    # Ensure vendor/project records exist or get updated when possible.
    vendor_id = get_or_create_vendor(cursor, vendor_name or current.VendorName or "Unknown Vendor")
    project_id = get_or_create_project(cursor, project_name or current.ProjectName or "Unassigned Project", department or current.Department or "", project_code or current.ProjectCode or None)

    # Primary PO record update.
    cursor.execute(
        """
        UPDATE dbo.PurchaseOrders
        SET PONumber = ?, VendorId = ?, ProjectId = ?, ProjectCode = ?, Department = ?, Requestor = ?, UpdatedAt = SYSUTCDATETIME()
        WHERE PONumber = ?;
        """,
        new_po,
        vendor_id,
        project_id,
        project_code,
        department,
        requestor,
        old_po,
    )

    # Line items hold denormalized PO/project/vendor fields.
    cursor.execute(
        """
        UPDATE dbo.IssuedPOLines
        SET PONumber = ?, VendorName = ?, ProjectCode = ?, ProjectName = ?, Department = ?, Requestor = ?
        WHERE PONumber = ?;
        """,
        new_po,
        vendor_name,
        project_code,
        project_name,
        department,
        requestor,
        old_po,
    )

    # Related records that may point to this PO number.
    for table_name, columns in {
        "ExpenseReviewItems": ["CorrectPONumber", "MatchedPONumber", "PostedPONumber", "ExtractedPONumber"],
        "PurchaseRequests": ["ConvertedPONumber"],
        "PurchaseRequestAttachments": ["PONumber"],
    }.items():
        for column_name in columns:
            update_if_table_column(cursor, table_name, column_name, new_po, column_name, old_po)

    cursor.execute(
        """
        INSERT INTO dbo.POChangeAudit
            (ChangeType, OldPONumber, NewPONumber, OldProjectCode, NewProjectCode, OldProjectName, NewProjectName,
             OldVendorName, NewVendorName, OldDepartment, NewDepartment, OldRequestor, NewRequestor, Reason, ChangedBy)
        VALUES ('PO Maintenance Update', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
        """,
        old_po,
        new_po,
        current.ProjectCode,
        project_code,
        current.ProjectName,
        project_name,
        current.VendorName,
        vendor_name,
        current.Department,
        department,
        current.Requestor,
        requestor,
        reason,
        changed_by,
    )


def bulk_update_project_code(cursor, project_name, new_project_code, reason, changed_by):
    project_name = clean_text(project_name)
    new_project_code = clean_text(new_project_code)
    reason = clean_text(reason)
    changed_by = clean_text(changed_by)
    if not project_name:
        raise ValueError("Project name is required for bulk project code update.")
    if not new_project_code:
        raise ValueError("New project code is required.")
    if not reason:
        raise ValueError("A reason for the bulk project code update is required.")

    cursor.execute("SELECT TOP 1 ProjectCode FROM dbo.Projects WHERE ProjectName = ?", project_name)
    row = cursor.fetchone()
    old_code = row.ProjectCode if row else ""

    cursor.execute("UPDATE dbo.Projects SET ProjectCode = ? WHERE ProjectName = ?", new_project_code, project_name)
    cursor.execute("UPDATE dbo.PurchaseOrders SET ProjectCode = ?, UpdatedAt = SYSUTCDATETIME() WHERE ProjectId IN (SELECT ProjectId FROM dbo.Projects WHERE ProjectName = ?)", new_project_code, project_name)
    cursor.execute("UPDATE dbo.IssuedPOLines SET ProjectCode = ? WHERE ProjectName = ?", new_project_code, project_name)
    cursor.execute(
        """
        INSERT INTO dbo.POChangeAudit
            (ChangeType, OldProjectCode, NewProjectCode, OldProjectName, NewProjectName, Reason, ChangedBy)
        VALUES ('Bulk Project Code Update', ?, ?, ?, ?, ?, ?);
        """,
        old_code,
        new_project_code,
        project_name,
        project_name,
        reason,
        changed_by,
    )

def void_existing_po(cursor, po_number, reason, changed_by):
    """Void a PO while keeping it visible in PO lists.

    Voiding keeps the PO/line records for audit/history, but sets the financial
    value to zero and marks the status as Voided so dashboards and packets no
    longer show an available amount. Linked expense review rows are preserved.
    """
    po_number = clean_text(po_number)
    reason = clean_text(reason)
    changed_by = clean_text(changed_by)
    if not po_number:
        raise ValueError("PO number is required to void a PO.")
    if not reason:
        raise ValueError("A reason is required to void a PO.")

    current = po_maintenance_row_summary(cursor, po_number)
    if not current:
        raise ValueError("The selected PO was not found.")
    if clean_text(getattr(current, "POStatus", "")).lower() == "voided":
        raise ValueError("This PO is already voided.")

    old_value = getattr(current, "LineTotal", 0) or 0
    audit_reason = f"VOID PO - previous displayed amount {currency(old_value)}. {reason}"

    # Header-level values. Keep the PO visible, but make its value/balance zero.
    for col in ["OriginalAmount", "RevisedAmount", "RemainingAmount"]:
        update_if_table_column(cursor, "PurchaseOrders", col, 0, "PONumber", po_number)
    update_if_table_column(cursor, "PurchaseOrders", "POStatus", "Voided", "PONumber", po_number)
    update_if_table_column(cursor, "PurchaseOrders", "SetupStatus", "Voided", "PONumber", po_number)

    if sql_table_exists(cursor, "PurchaseOrders") and sql_col_exists(cursor, "PurchaseOrders", "UpdatedAt"):
        cursor.execute("UPDATE dbo.PurchaseOrders SET UpdatedAt = SYSUTCDATETIME() WHERE PONumber = ?", po_number)

    # Denormalized upload/line values. This makes existing PO lists and packets
    # show 0.00 even when they calculate from line totals.
    for col in ["LineAmount", "OriginalAmount", "RevisedAmount", "RemainingAmount"]:
        update_if_table_column(cursor, "IssuedPOLines", col, 0, "PONumber", po_number)
    update_if_table_column(cursor, "IssuedPOLines", "POStatus", "Voided", "PONumber", po_number)

    cursor.execute(
        """
        INSERT INTO dbo.POChangeAudit
            (ChangeType, OldPONumber, NewPONumber, OldProjectCode, NewProjectCode, OldProjectName, NewProjectName,
             OldVendorName, NewVendorName, OldDepartment, NewDepartment, OldRequestor, NewRequestor, Reason, ChangedBy)
        VALUES ('PO Voided', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
        """,
        po_number,
        po_number,
        current.ProjectCode,
        current.ProjectCode,
        current.ProjectName,
        current.ProjectName,
        current.VendorName,
        current.VendorName,
        current.Department,
        current.Department,
        current.Requestor,
        current.Requestor,
        audit_reason,
        changed_by,
    )



@app.route("/po-maintenance", methods=["GET", "POST"])
def po_maintenance():
    allowed, reason = require_page_access("PO Maintenance")
    if not allowed:
        return access_denied_response("PO Maintenance", reason)

    message_html = ""
    selected_po = clean_text(request.args.get("po"))
    q = clean_text(request.args.get("q"))

    try:
        conn = get_sql_connection()
        cursor = conn.cursor()
        ensure_po_maintenance_schema(cursor)

        if request.method == "POST":
            action = clean_text(request.form.get("action"))
            try:
                if action == "update_po":
                    update_existing_po_values(
                        cursor,
                        request.form.get("old_po_number"),
                        request.form.get("new_po_number"),
                        request.form.get("project_code"),
                        request.form.get("project_name"),
                        request.form.get("vendor_name"),
                        request.form.get("department"),
                        request.form.get("requestor"),
                        request.form.get("reason"),
                        get_current_user().get("email") or "Unknown",
                    )
                    conn.commit()
                    selected_po = clean_text(request.form.get("new_po_number")) or selected_po
                    message_html = '<div class="notice ok">PO maintenance update saved.</div>'
                elif action == "bulk_project_code":
                    bulk_update_project_code(
                        cursor,
                        request.form.get("bulk_project_name"),
                        request.form.get("bulk_project_code"),
                        request.form.get("bulk_reason"),
                        get_current_user().get("email") or "Unknown",
                    )
                    conn.commit()
                    message_html = '<div class="notice ok">Project code update saved for matching project records.</div>'
                elif action == "void_po":
                    void_existing_po(
                        cursor,
                        request.form.get("void_po_number"),
                        request.form.get("void_reason"),
                        get_current_user().get("email") or "Unknown",
                    )
                    conn.commit()
                    selected_po = clean_text(request.form.get("void_po_number")) or selected_po
                    message_html = '<div class="notice ok">PO voided. It will stay visible in PO lists, but its amount and balance are now $0.00.</div>'
                else:
                    message_html = '<div class="notice error">Unknown PO maintenance action.</div>'
            except Exception as e:
                conn.rollback()
                message_html = f'<div class="notice error">{h(e)}</div>'

        # Search/list POs.
        search_where = ""
        search_params = []
        if q:
            like = f"%{q}%"
            search_where = "WHERE po.PONumber LIKE ? OR v.VendorName LIKE ? OR p.ProjectName LIKE ? OR po.ProjectCode LIKE ?"
            search_params = [like, like, like, like]
        cursor.execute(
            f"""
            SELECT TOP 100
                po.PONumber,
                po.ProjectCode,
                COALESCE(p.ProjectName, '') AS ProjectName,
                COALESCE(v.VendorName, '') AS VendorName,
                po.Department,
                po.Requestor,
                po.POStatus,
                COALESCE(SUM(l.LineAmount), po.RevisedAmount, po.OriginalAmount, 0) AS POValue,
                COUNT(l.IssuedPOLineId) AS LineCount
            FROM dbo.PurchaseOrders po
            LEFT JOIN dbo.Projects p ON p.ProjectId = po.ProjectId
            LEFT JOIN dbo.Vendors v ON v.VendorId = po.VendorId
            LEFT JOIN dbo.IssuedPOLines l ON l.PONumber = po.PONumber
            {search_where}
            GROUP BY po.PONumber, po.ProjectCode, p.ProjectName, v.VendorName, po.Department, po.Requestor, po.POStatus, po.RevisedAmount, po.OriginalAmount
            ORDER BY po.PONumber;
            """,
            *search_params,
        )
        pos = cursor.fetchall()

        # Project names for bulk update.
        cursor.execute("SELECT DISTINCT ProjectName FROM dbo.Projects WHERE ProjectName IS NOT NULL AND LTRIM(RTRIM(ProjectName)) <> '' ORDER BY ProjectName")
        project_names = [r.ProjectName for r in cursor.fetchall()]

        users = load_active_dashboard_users_for_select(cursor)

        selected = po_maintenance_row_summary(cursor, selected_po) if selected_po else None
        line_rows = []
        expense_rows = []
        audit_rows = []
        if selected:
            cursor.execute(
                """
                SELECT TOP 200 IssuedPOLineId, LineDescription, Unit, UnitCost, Qty, LineAmount, ProjectCode, ProjectName, VendorName
                FROM dbo.IssuedPOLines
                WHERE PONumber = ?
                ORDER BY IssuedPOLineId;
                """,
                selected.PONumber,
            )
            line_rows = cursor.fetchall()
            try:
                ensure_expense_review_tables()
                cursor.execute(
                    """
                    SELECT TOP 100 ExpenseReviewItemId, TxDate, TxType, VendorName, Amount, ReviewDecision, PostedToPO, PostedAmount, PostedAt, PostedBy
                    FROM dbo.ExpenseReviewItems
                    WHERE CorrectPONumber = ? OR MatchedPONumber = ? OR PostedPONumber = ?
                    ORDER BY COALESCE(PostedAt, UpdatedAt, CreatedAt) DESC;
                    """,
                    selected.PONumber,
                    selected.PONumber,
                    selected.PONumber,
                )
                expense_rows = cursor.fetchall()
            except Exception:
                expense_rows = []
            cursor.execute(
                """
                SELECT TOP 25 ChangeType, OldPONumber, NewPONumber, OldProjectCode, NewProjectCode, Reason, ChangedBy, ChangedAt
                FROM dbo.POChangeAudit
                WHERE OldPONumber = ? OR NewPONumber = ? OR OldProjectName = ? OR NewProjectName = ?
                ORDER BY ChangedAt DESC;
                """,
                selected.PONumber,
                selected.PONumber,
                selected.ProjectName,
                selected.ProjectName,
            )
            audit_rows = cursor.fetchall()

        conn.close()

    except Exception as e:
        return shell("PO Maintenance", "Edit uploaded POs safely with audit tracking.", "PO Maintenance", f'<div class="notice error">Error loading PO Maintenance: {h(e)}</div>')

    po_list_rows = ""
    for po in pos:
        po_list_rows += f"""
        <tr>
            <td><a class="po-link" href="/po-maintenance?po={quote_plus(str(po.PONumber or ''))}">{h(po.PONumber)}</a></td>
            <td>{h(po.ProjectCode)}</td>
            <td>{h(po.ProjectName)}</td>
            <td>{h(po.VendorName)}</td>
            <td>{h(po.Requestor)}</td>
            <td>{status_chip(po.POStatus or 'Open')}</td>
            <td class="right">{currency(0 if clean_text(po.POStatus).lower() == "voided" else po.POValue)}</td>
            <td class="right">{int(po.LineCount or 0)}</td>
        </tr>
        """
    if not po_list_rows:
        po_list_rows = '<tr><td colspan="8">No POs found.</td></tr>'

    dept_options = ''.join(f'<option value="{h(d)}">{h(d)}</option>' for d in DEPARTMENT_OPTIONS)
    requestor_options = ''.join(f'<option value="{h(u.Email)}">{h(u.DisplayName or u.Email)} · {h(u.Email)}</option>' for u in users)

    selected_html = ""
    if selected:
        # Rebuild selected option markup with current values selected.
        dept_options_selected = ''.join(f'<option value="{h(d)}" {"selected" if clean_text(selected.Department)==d else ""}>{h(d)}</option>' for d in DEPARTMENT_OPTIONS)
        requestor_options_selected = ''.join(f'<option value="{h(u.Email)}" {"selected" if clean_text(selected.Requestor).lower()==str(u.Email).lower() else ""}>{h(u.DisplayName or u.Email)} · {h(u.Email)}</option>' for u in users)
        if selected.Requestor and all(str(selected.Requestor).lower() != str(u.Email).lower() for u in users):
            requestor_options_selected = f'<option value="{h(selected.Requestor)}" selected>{h(selected.Requestor)}</option>' + requestor_options_selected

        line_table_rows = ''.join(f"<tr><td>{h(r.LineDescription)}</td><td>{h(r.Unit)}</td><td>{h(r.Qty)}</td><td class='right'>{currency(r.UnitCost)}</td><td class='right'>{currency(r.LineAmount)}</td></tr>" for r in line_rows) or '<tr><td colspan="5">No line items found.</td></tr>'
        expense_table_rows = ''.join(f"<tr><td>{h(r.TxDate)}</td><td>{h(r.TxType)}</td><td>{h(r.VendorName)}</td><td class='right'>{currency(r.Amount)}</td><td>{h(r.ReviewDecision)}</td><td>{'Posted' if r.PostedToPO else 'Not Posted'}</td><td class='right'>{currency(r.PostedAmount)}</td></tr>" for r in expense_rows) or '<tr><td colspan="7">No linked expense review rows found.</td></tr>'
        audit_table_rows = ''.join(f"<tr><td>{h(r.ChangedAt)}</td><td>{h(r.ChangeType)}</td><td>{h(r.OldPONumber)} → {h(r.NewPONumber)}</td><td>{h(r.OldProjectCode)} → {h(r.NewProjectCode)}</td><td>{h(r.ChangedBy)}</td><td>{h(r.Reason)}</td></tr>" for r in audit_rows) or '<tr><td colspan="6">No audit changes recorded yet.</td></tr>'
        selected_html = f"""
        <div class="card">
            <h3>Edit Selected PO</h3>
            <p class="card-subtitle">Use this for setup corrections such as PO number cleanup, missing project code, vendor/requestor corrections, or department cleanup. A reason is required and an audit record is saved.</p>
            <form method="post" class="form-grid">
                <input type="hidden" name="action" value="update_po">
                <input type="hidden" name="old_po_number" value="{h(selected.PONumber)}">
                <div class="form-field"><label>PO Number</label><input name="new_po_number" value="{h(selected.PONumber)}" required></div>
                <div class="form-field"><label>Project Code</label><input name="project_code" value="{h(selected.ProjectCode)}" placeholder="Example: 26-018"></div>
                <div class="form-field"><label>Project Name</label><input name="project_name" value="{h(selected.ProjectName)}" required></div>
                <div class="form-field"><label>Vendor</label><input name="vendor_name" value="{h(selected.VendorName)}" required></div>
                <div class="form-field"><label>Department</label><select name="department"><option value="">Select department</option>{dept_options_selected}</select></div>
                <div class="form-field"><label>Requestor</label><select name="requestor"><option value="">Select requestor</option>{requestor_options_selected}</select></div>
                <div class="form-field full"><label>Reason for Change</label><textarea name="reason" required placeholder="Example: Project setup cleanup; added missing project code; corrected vendor PO number."></textarea></div>
                <div class="form-field full"><button class="primary" type="submit">Save PO Maintenance Update</button> <a class="button secondary" href="/po-packet/{quote_plus(str(selected.PONumber or ''))}?type=internal">Open PO Packet</a></div>
            </form>
        </div>
        <div class="card danger-zone">
            <h3>Void PO</h3>
            <p class="card-subtitle">Use this when a PO should remain in the records but no longer carry any value. The PO status will become Voided, header amounts and line amounts will be set to $0.00, and an audit record will be saved. Linked expense history is preserved.</p>
            <form method="post" class="form-grid" onsubmit="return confirm('Void this PO and set its amount to $0.00? This keeps the PO visible but removes its value from balances.');">
                <input type="hidden" name="action" value="void_po">
                <input type="hidden" name="void_po_number" value="{h(selected.PONumber)}">
                <div class="form-field full"><label>Reason for Void</label><textarea name="void_reason" required placeholder="Example: PO cancelled; duplicate upload; vendor PO replaced by another PO."></textarea></div>
                <div class="form-field full"><button class="secondary" type="submit">Void PO and Set Amount to $0.00</button></div>
            </form>
        </div>
        <div class="grid two">
            <div class="card"><h3>Linked Line Items</h3><div class="table-wrap"><table><tr><th>Description</th><th>Unit</th><th>Qty</th><th class="right">Unit Cost</th><th class="right">Line Amount</th></tr>{line_table_rows}</table></div></div>
            <div class="card"><h3>Linked Posted / Reviewed Expenses</h3><div class="table-wrap"><table><tr><th>Date</th><th>Type</th><th>Vendor</th><th class="right">Amount</th><th>Decision</th><th>Posting</th><th class="right">Posted Amount</th></tr>{expense_table_rows}</table></div></div>
        </div>
        <div class="card"><h3>PO Change Audit</h3><div class="table-wrap"><table><tr><th>Changed At</th><th>Type</th><th>PO Change</th><th>Project Code Change</th><th>Changed By</th><th>Reason</th></tr>{audit_table_rows}</table></div></div>
        """
    else:
        selected_html = '<div class="card"><h3>Select a PO to edit</h3><p class="card-subtitle">Click a PO number in the list below to open its maintenance form.</p></div>'

    project_options = ''.join(f'<option value="{h(name)}">{h(name)}</option>' for name in project_names)
    content = f"""
    {message_html}
    <div class="notice info"><strong>PO Maintenance</strong><br>Use this page for controlled corrections to existing uploaded/app-created POs. You can edit setup details or void a PO. Voided POs stay visible in PO lists but show $0.00 for amount/balance.</div>
    {selected_html}
    <div class="card">
        <h3>Bulk Project Code Update</h3>
        <p class="card-subtitle">Use this when a whole project was uploaded without a project code. This updates the project code on matching project, PO, and line item records.</p>
        <form method="post" class="form-grid">
            <input type="hidden" name="action" value="bulk_project_code">
            <div class="form-field"><label>Project</label><select name="bulk_project_name" required><option value="">Select project</option>{project_options}</select></div>
            <div class="form-field"><label>New Project Code</label><input name="bulk_project_code" placeholder="Example: 26-018" required></div>
            <div class="form-field full"><label>Reason</label><textarea name="bulk_reason" required placeholder="Example: Added missing project code after initial PO upload."></textarea></div>
            <div class="form-field full"><button class="primary" type="submit">Update Project Code</button></div>
        </form>
    </div>
    <div class="card">
        <h3>Find Existing PO</h3>
        <form method="get" class="search-row"><input name="q" value="{h(q)}" placeholder="Search PO, project, vendor, or project code"><button class="secondary" type="submit">Search</button><a class="button secondary" href="/po-maintenance">Clear</a></form>
        <div class="table-wrap"><table><tr><th>PO</th><th>Project Code</th><th>Project</th><th>Vendor</th><th>Requestor</th><th>Status</th><th class="right">Value</th><th class="right">Lines</th></tr>{po_list_rows}</table></div>
    </div>
    """
    return shell("PO Maintenance", "Safely edit PO numbers and project codes with an audit trail.", "PO Maintenance", content)


@app.route("/future-pages")
def future_pages():
    allowed, reason = require_page_access("Future Pages")
    if not allowed:
        return access_denied_response("Future Pages", reason)

    future_page_rows = [
        {
            "name": "Approver Queue",
            "url": "/approver-queue",
            "status": "Hidden for July 1",
            "phase": "Phase 2 candidate",
            "purpose": "Separate approval workspace. For July 1, approvals are handled from Purchase Requests so users have one place to review requests.",
        },
        {
            "name": "Forecasting",
            "url": "/forecasting",
            "status": "Hidden for July 1",
            "phase": "Phase 2 / Command Center expansion",
            "purpose": "Future cash-out, backlog, AP/AR, labor, and project forecasting views once the underlying data model is ready.",
        },
        {
            "name": "POs in PM Comments",
            "url": "/pos-in-pm-comments",
            "status": "Hidden from sidebar",
            "phase": "Audit / research tool",
            "purpose": "Read-only audit view for expenses where PO numbers were found in PM comments. Kept by direct URL to reduce overlap with Expense Upload / PO Matching.",
        },
        {
            "name": "Exceptions",
            "url": "/exceptions",
            "status": "Hidden for July 1",
            "phase": "Phase 2 candidate",
            "purpose": "Future issue/exception center for over-budget POs, expired dates, unmatched transactions, and other cleanup queues.",
        },
        {
            "name": "Exports",
            "url": "/exports",
            "status": "Hidden for July 1",
            "phase": "Phase 2 candidate",
            "purpose": "Export center for PO lists and issued line data. Kept out of the main navigation until export rules and permissions are finalized.",
        },
    ]

    rows = "".join(
        f"""
        <tr>
            <td><strong>{h(item['name'])}</strong><br><span class=\"muted\">{h(item['url'])}</span></td>
            <td><span class=\"badge gray\">{h(item['status'])}</span></td>
            <td>{h(item['phase'])}</td>
            <td>{h(item['purpose'])}</td>
            <td><a class=\"button secondary\" href=\"{h(item['url'])}\">Open Direct URL</a></td>
        </tr>
        """
        for item in future_page_rows
    )

    content = f"""
    <div class=\"page-heading\">
        <div>
            <h2 style=\"margin:0 0 8px;\">Future Pages</h2>
            <p class=\"muted\">Admin-only parking lot for pages that exist in the app but are hidden from the main July 1 rollout navigation.</p>
        </div>
    </div>

    <div class=\"notice info\">
        <strong>Admin only:</strong> This page is a safe place to keep track of dormant, hidden, or future Command Center pages without exposing them to the broader team.
    </div>

    <div class=\"card\">
        <h3>Hidden / Future Pages</h3>
        <div class=\"table-wrap\">
            <table>
                <tr><th>Page</th><th>Status</th><th>Planned Phase</th><th>Why it is hidden</th><th>Admin Link</th></tr>
                {rows}
            </table>
        </div>
    </div>

    <div class=\"card\">
        <h3>Recommended Rule</h3>
        <p>Pages should stay here until the workflow, permissions, training material, and data source are ready. When a page is ready, we can move it into the normal sidebar for the roles that should use it.</p>
    </div>
    """
    return shell("Future Pages", "Admin-only list of hidden and future Command Center pages.", "Future Pages", content)

@app.route("/user-access", methods=["GET", "POST"])
def user_access():
    allowed, reason = require_page_access("User Access")
    if not allowed:
        return access_denied_response("User Access", reason)

    message_html = ""

    if request.method == "POST" and normalize_role(get_user_access().get("role")) != "Admin":
        message_html = '<div class="notice error">Only Admin users can add or update user access. Executive users can use View As only.</div>'
    elif request.method == "POST":
        email = clean_text(request.form.get("email"))
        display_name = clean_text(request.form.get("display_name"))
        role_name = clean_text(request.form.get("role_name"))
        is_active_raw = clean_text(request.form.get("is_active"))

        is_active = 1 if is_active_raw == "1" else 0

        if email:
            email = email.lower()

        if not email or "@" not in email:
            message_html = '<div class="notice error">Email is required.</div>'
        elif role_name not in VALID_ROLES:
            message_html = '<div class="notice error">Invalid role selected.</div>'
        else:
            try:
                conn = get_sql_connection()
                cursor = conn.cursor()

                cursor.execute(
                    """
                    IF EXISTS (SELECT 1 FROM dbo.DashboardUsers WHERE LOWER(Email) = LOWER(?))
                    BEGIN
                        UPDATE dbo.DashboardUsers
                        SET DisplayName = ?, RoleName = ?, IsActive = ?, UpdatedAt = SYSUTCDATETIME()
                        WHERE LOWER(Email) = LOWER(?);
                    END
                    ELSE
                    BEGIN
                        INSERT INTO dbo.DashboardUsers (Email, DisplayName, RoleName, IsActive)
                        VALUES (?, ?, ?, ?);
                    END
                    """,
                    email,
                    display_name,
                    role_name,
                    is_active,
                    email,
                    email,
                    display_name,
                    role_name,
                    is_active,
                )

                conn.commit()
                conn.close()
                message_html = '<div class="notice ok">User access was saved.</div>'

            except Exception as e:
                message_html = f'<div class="notice error">Error saving user access: {h(e)}</div>'

    try:
        conn = get_sql_connection()
        cursor = conn.cursor()

        cursor.execute(
            """
            SELECT DashboardUserId, Email, DisplayName, RoleName, IsActive, CreatedAt, UpdatedAt
            FROM dbo.DashboardUsers
            ORDER BY Email;
            """
        )

        users = cursor.fetchall()
        conn.close()

        role_options = ""
        for role in VALID_ROLES:
            role_options += f'<option value="{h(role)}">{h(role)}</option>'

        user_rows = ""
        for row in users:
            active_badge = '<span class="badge green">Active</span>' if row.IsActive else '<span class="badge red">Disabled</span>'
            row_role_options = ""
            for role in VALID_ROLES:
                selected = "selected" if str(row.RoleName or "") == role else ""
                row_role_options += f'<option value="{h(role)}" {selected}>{h(role)}</option>'
            active_selected = "selected" if row.IsActive else ""
            disabled_selected = "" if row.IsActive else "selected"
            user_rows += f"""
            <tr>
                <td>
                    <input type="hidden" name="email" value="{h(row.Email)}" form="user-edit-{h(row.DashboardUserId)}">
                    <strong>{h(row.Email)}</strong>
                    <div class="small-muted">Login email is the unique user key.</div>
                </td>
                <td><input type="text" name="display_name" value="{h(row.DisplayName)}" form="user-edit-{h(row.DashboardUserId)}"></td>
                <td><select name="role_name" form="user-edit-{h(row.DashboardUserId)}">{row_role_options}</select></td>
                <td>
                    {active_badge}<br>
                    <select name="is_active" form="user-edit-{h(row.DashboardUserId)}">
                        <option value="1" {active_selected}>Active</option>
                        <option value="0" {disabled_selected}>Disabled</option>
                    </select>
                </td>
                <td>{h(row.UpdatedAt)}</td>
                <td>
                    <form id="user-edit-{h(row.DashboardUserId)}" method="post" action="/user-access"></form>
                    <button class="secondary" type="submit" form="user-edit-{h(row.DashboardUserId)}">Save Changes</button>
                </td>
            </tr>"""

        if not user_rows:
            user_rows = '<tr><td colspan="6">No users found.</td></tr>'

        content = f"""
        {message_html}
        <div class="card">
            <h3>View As User</h3>
            <p class="card-subtitle">Admin/Executive users can view user-specific task context while keeping the admin/executive sidebar and permissions.</p>
            <form method="post" action="/set-view-as" class="rollout-filter-bar">
                <div class="rollout-filter-fields">
                    <label>View as<select name="view_as_email"><option value="">Actual signed-in user</option>{''.join(f'<option value="{h(u.Email)}" {"selected" if get_view_as_email().lower() == str(u.Email).lower() else ""}>{h(u.DisplayName or u.Email)} · {h(u.RoleName)}</option>' for u in users if u.IsActive)}</select></label>
                </div>
                <div class="rollout-filter-actions"><button class="primary" type="submit">Apply View As</button><a class="button secondary" href="/clear-view-as">Clear</a></div>
            </form>
        </div>
        <div class="card">
            <h3>Add New User Access</h3>
            <p class="card-subtitle">Admins can add new users here. To correct spelling, change access level, or disable someone, use the editable Current Dashboard Users table below. Executive users can use View As but cannot change access.</p>
            <form method="post" action="/user-access">
                <p><label>Email</label><br><input type="text" name="email" placeholder="person@c-diving.com" required></p>
                <p><label>Display Name</label><br><input type="text" name="display_name" placeholder="Person Name"></p>
                <p><label>Role</label><br><select name="role_name" required>{role_options}</select></p>
                <p><label>Status</label><br><select name="is_active"><option value="1">Active</option><option value="0">Disabled</option></select></p>
                <p><button class="primary" type="submit">Save User Access</button></p>
            </form>
        </div>
        <div class="card"><h3>Current Dashboard Users</h3><p class="card-subtitle">Edit display name spelling, access level, or disabled status directly from this table. Save each row after making changes.</p><div class="table-wrap"><table><tr><th>Email</th><th>Display Name</th><th>Role</th><th>Status</th><th>Updated At</th><th>Action</th></tr>{user_rows}</table></div></div>
        <div class="card"><h3>Role Guide</h3><table><tr><th>Role</th><th>Access</th></tr><tr><td>Admin</td><td>Everything, including user management, PO Maintenance, PO number corrections, project code corrections, uploads, approvals, and voiding POs.</td></tr><tr><td>Executive</td><td>Can view all departments and review/approve purchase requests, but cannot edit users, void POs, or maintain PO numbers/project codes.</td></tr><tr><td>Project Manager - Dredging Only</td><td>Can view all Dredging POs and submit Dredging purchase requests.</td></tr><tr><td>Project Manager - Diving</td><td>Can view assigned non-Dredging POs and submit requests for assigned non-Dredging projects.</td></tr><tr><td>Division Manager - Diving</td><td>Can view all non-Dredging POs and approve non-Dredging purchase requests.</td></tr><tr><td>Purchaser - All Departments</td><td>Can view all departments and submit purchase requests, but cannot approve, void, or maintain POs.</td></tr><tr><td>Bookkeeping - All Departments</td><td>Read-only PO lookup across all departments, including amounts, balances, and PO packet downloads. No purchase request access.</td></tr><tr><td>No Access</td><td>Blocked from dashboard access.</td></tr></table></div>
        """

        return shell("User Access", "Add users, correct names, change roles, and disable access.", "User Access", content)

    except Exception as e:
        content = f'<div class="notice error">Error loading user access: {h(e)}</div>'
        return shell("User Access", "Unable to load user access.", "User Access", content), 500


@app.route("/whoami")
def whoami():
    allowed, reason = require_page_access("Who Am I")
    if not allowed:
        return access_denied_response("Who Am I", reason)

    user = get_current_user()
    access = get_user_access()
    principal = request.headers.get("X-MS-CLIENT-PRINCIPAL", "")
    principal_preview = principal[:500]
    if len(principal) > 500:
        principal_preview += "..."

    auth_status_badge = '<span class="badge green">Authenticated</span>' if user["is_authenticated"] else '<span class="badge amber">Not Detected</span>'
    domain_status_badge = '<span class="badge green">Allowed Domain</span>' if user["is_allowed_domain"] else '<span class="badge amber">Domain Not Confirmed</span>'
    sql_status_badge = '<span class="badge green">Found</span>' if access["found_in_sql"] else '<span class="badge amber">Not Found</span>'
    active_badge = '<span class="badge green">Active</span>' if access["is_active"] else '<span class="badge red">Inactive / No Access</span>'

    content = f"""
    <div class="grid two">
        <div class="card">
            <h3>Signed-In User</h3>
            <p class="card-subtitle">This page reads the Microsoft login headers provided by Azure App Service Authentication.</p>
            <table>
                <tr><th>Authentication Status</th><td>{auth_status_badge}</td></tr>
                <tr><th>Email / User Principal Name</th><td>{h(user["email"])}</td></tr>
                <tr><th>Email Domain</th><td>{h(user["email_domain"])}</td></tr>
                <tr><th>Allowed Domain Setting</th><td>{h(user["allowed_domain"])}</td></tr>
                <tr><th>Domain Check</th><td>{domain_status_badge}</td></tr>
                <tr><th>Identity Provider</th><td>{h(user["identity_provider"])}</td></tr>
                <tr><th>Azure User ID</th><td>{h(user["user_id"])}</td></tr>
            </table>
        </div>
        <div class="card">
            <h3>Dashboard Access</h3>
            <p class="card-subtitle">This is the SQL-backed dashboard permission result.</p>
            <table>
                <tr><th>Found In DashboardUsers</th><td>{sql_status_badge}</td></tr>
                <tr><th>Display Name</th><td>{h(access["display_name"])}</td></tr>
                <tr><th>Role</th><td><span class="badge blue">{h(access["role"])}</span></td></tr>
                <tr><th>Status</th><td>{active_badge}</td></tr>
                <tr><th>Lookup Error</th><td>{h(access["lookup_error"])}</td></tr>
            </table>
        </div>
    </div>
    <div class="card">
        <h3>Raw Azure Authentication Headers</h3>
        <p class="card-subtitle">Useful for troubleshooting.</p>
        <table>
            <tr><th>Header</th><th>Value</th></tr>
            <tr><td>X-MS-CLIENT-PRINCIPAL-NAME</td><td>{h(user["email"])}</td></tr>
            <tr><td>X-MS-CLIENT-PRINCIPAL-ID</td><td>{h(user["user_id"])}</td></tr>
            <tr><td>X-MS-CLIENT-PRINCIPAL-IDP</td><td>{h(user["identity_provider"])}</td></tr>
            <tr><td>X-MS-CLIENT-PRINCIPAL</td><td>{h(principal_preview)}</td></tr>
        </table>
    </div>
    """

    return shell("Who Am I", "View the signed-in Microsoft user and SQL-backed dashboard role.", "Who Am I", content)


@app.route("/access-denied")
def access_denied():
    return access_denied_response("Unknown", "Access denied.")


@app.route("/health")
def health():
    return jsonify({"status": "ok", "environment": APP_ENVIRONMENT, "sql_server": SQL_SERVER_NAME, "database": SQL_DATABASE_NAME, "connection_string_found": bool(SQL_CONNECTION)})


@app.route("/db-test")
def db_test():
    if not SQL_CONNECTION:
        return jsonify({"status": "error", "step": "connection_string", "message": "SQL connection string was not found."}), 500

    try:
        conn = get_sql_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT DB_NAME() AS DatabaseName, GETUTCDATE() AS ServerTime")
        row = cursor.fetchone()
        conn.close()
        return jsonify({"status": "success", "database": row.DatabaseName, "server_time_utc": str(row.ServerTime)})
    except Exception as e:
        return jsonify({"status": "error", "step": "connect_to_sql", "message": str(e)}), 500


if __name__ == "__main__":
    port = int(os.environ.get("PORT", os.environ.get("HTTP_PLATFORM_PORT", 8000)))
    app.run(host="0.0.0.0", port=port)
